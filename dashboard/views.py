# dashboard/views.py
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.conf import settings
from bson import ObjectId
from datetime import datetime, timedelta
from collections import Counter
from django.contrib import messages
import json
import random
from django.contrib.auth import login, logout, authenticate
from django.contrib.sessions.models import Session
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt, csrf_protect, ensure_csrf_cookie
from django.views.decorators.http import require_http_methods
from django.core.mail import send_mail
from django.contrib.auth.hashers import check_password, make_password
from django.contrib.auth import get_user_model
from django.contrib.auth import update_session_auth_hash
import qrcode
from io import BytesIO
from django.core.files.base import ContentFile
import base64
import logging
from dashboard.models import UserSession, SessionLog
from .models import Utilisateur, UserSession
import json
import re
from datetime import datetime, timedelta
from django.utils import timezone
from django.db import models
from .models import ChatbotConversation, ChatbotMessage
from dashboard.models import Notification
from django.contrib.auth import get_user_model
from datetime import datetime
import io
import os
import urllib.request
from django.utils.timezone import is_naive, make_aware
import pytz
from bson import ObjectId
from datetime import datetime
from django.conf import settings
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, HRFlowable,
    Image as RLImage, PageBreak,
)



# ──── Helpers session MongoDB (remplacent Django ORM auth) ────────────────────
from functools import wraps as _wraps

def session_required(view_func):
    """Remplace @session_required — verifie la session MongoDB."""
    @_wraps(view_func)
    def _wrapper(request, *args, **kwargs):
        if not request.session.get("user_id"):
            from django.conf import settings as _s
            login_url = getattr(_s, "LOGIN_URL", "/employe/login/")
            from django.shortcuts import redirect as _r
            return _r(f"{login_url}?next={request.path}")
        return view_func(request, *args, **kwargs)
    return _wrapper

def staff_required(view_func):
    """Acces reserve aux admins (is_staff = True en session)."""
    @_wraps(view_func)
    def _wrapper(request, *args, **kwargs):
        from django.shortcuts import redirect as _r
        if not request.session.get("user_id"):
            return _r("/login/")
        if not request.session.get("is_staff"):
            return _r("employe_espace")
        return view_func(request, *args, **kwargs)
    return _wrapper

def get_session_user(request):
    """Retourne un dict avec les infos utilisateur depuis la session."""
    uid = request.session.get("user_id", "")
    return {
        "id":             uid,
        "username":       request.session.get("username", ""),
        "is_staff":       request.session.get("is_staff", False),
        "is_superuser":   request.session.get("is_superuser", False),
        "is_authenticated": bool(uid),
        "first_name":     request.session.get("prenom", ""),
        "last_name":      request.session.get("nom", ""),
        "email":          request.session.get("email", ""),
    }
# ─────────────────────────────────────────────────────────────────────────────
db = settings.MONGO_DB
logger = logging.getLogger(__name__)

# ====================== HELPER : VÉRIFICATION DISPONIBILITÉ ======================
# ====================== HELPER : VÉRIFICATION DISPONIBILITÉ ======================
def check_ressource_disponibilite(ressource_id, ressource_type, date_debut, date_fin, exclude_resa_id=None):
    """
    Vérifie si une ressource (salle ou matériel) est disponible sur un créneau.
    Accepte les IDs en format ObjectId MongoDB OU entier Djongo.
    Retourne : { 'disponible': bool, 'motif': str, 'conflit_type': str|None }
    """
    # ── Résoudre l'ID en ObjectId MongoDB ─────────────────────────────────────
    ressource_oid = None

    # Cas 1 : string ObjectId valide (24 hex chars)
    try:
        ressource_oid = ObjectId(str(ressource_id))
    except Exception:
        pass

    # Cas 2 : entier Djongo — on cherche le document par champ 'id'
    if ressource_oid is None:
        try:
            id_int = int(ressource_id)
            collection = 'bureaux' if ressource_type == 'salle' else 'materiels'
            doc = db[collection].find_one({'id': id_int})
            if doc:
                ressource_oid = doc['_id']
        except Exception:
            pass

    if ressource_oid is None:
        return {
            'disponible':   False,
            'motif':        'ID ressource invalide ou introuvable',
            'conflit_type': None,
        }

    # ── 1) Vérifier les INDISPONIBILITÉS PLANIFIÉES ───────────────────────────
    indispo = db.indisponibilites.find_one({
        'ressource_id':   ressource_oid,
        'ressource_type': ressource_type,
        '$or': [
            {'date_debut': {'$lt': date_fin}, 'date_fin': {'$gt': date_debut}},
        ],
    })
    if indispo:
        type_lib = {
            'maintenance':          'Maintenance programmée',
            'reservation_bloquee':  'Créneau bloqué',
            'fermeture':            'Fermeture exceptionnelle',
        }.get(indispo.get('type_indispo', 'maintenance'), 'Indisponibilité')
        return {
            'disponible':   False,
            'motif':        (
                f"{type_lib} : {indispo.get('titre', 'sans titre')} "
                f"(du {indispo['date_debut'].strftime('%d/%m %H:%M')} "
                f"au {indispo['date_fin'].strftime('%d/%m %H:%M')})"
            ),
            'conflit_type': 'indisponibilite',
            'indispo_id':   str(indispo['_id']),
        }

    # ── 2) Vérifier les RÉSERVATIONS existantes (hors annulées) ───────────────
    query_resa = {
        '$or': [
            {'bureau_id':   ressource_oid},
            {'materiel_id': ressource_oid},
            {'resource_id': ressource_oid},
        ],
        'statut':     {'$in': ['confirmee', 'en_attente']},
        'date_debut': {'$lt': date_fin},
        'date_fin':   {'$gt': date_debut},
    }
    if exclude_resa_id:
        try:
            query_resa['_id'] = {'$ne': ObjectId(exclude_resa_id)}
        except Exception:
            pass

    resa_conflit = db.reservations.find_one(query_resa)
    if resa_conflit:
        return {
            'disponible':   False,
            'motif':        (
                f"Déjà réservé : {resa_conflit.get('titre', 'sans titre')} "
                f"(du {resa_conflit['date_debut'].strftime('%d/%m %H:%M')} "
                f"au {resa_conflit['date_fin'].strftime('%d/%m %H:%M')})"
            ),
            'conflit_type': 'reservation',
        }

    return {'disponible': True, 'motif': '', 'conflit_type': None}
# ====================== AUTHENTIFICATION ======================
def _check_password(password, stored_hash):
    """Vérifie le mot de passe peu importe le format de hash."""
    import bcrypt

    if not stored_hash:
        return False

    # Format bcrypt pur
    if stored_hash.startswith('$2b$') or stored_hash.startswith('$2a$'):
        try:
            return bcrypt.checkpw(
                password.encode('utf-8'),
                stored_hash.encode('utf-8')
            )
        except Exception:
            return False

    # Format Django (pbkdf2, argon2, etc.)
    try:
        from django.contrib.auth.hashers import check_password as django_check
        return django_check(password, stored_hash)
    except Exception:
        return False


def login_view(request):
    next_url = request.GET.get('next', '')

    if request.session.get('user_id'):
        if request.session.get('is_staff'):
            return redirect('dashboard')
        return redirect('employe_espace')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        next_url = request.POST.get('next', next_url)

        user_doc = None
        user_id  = None

        
        # ── 2. Nouveaux users (utilisateurs, ObjectId) ───────────────────
        if not user_doc:
            new_user = db['utilisateurs'].find_one(
                {'username': username, 'is_active': True}
            )
            if new_user and _check_password(password, new_user.get('password', '')):
                user_doc = new_user
                user_id  = str(new_user['_id'])  # ObjectId string

        # ── 3. Connexion ─────────────────────────────────────────────────
        if user_doc and user_id:
            request.session['user_id']      = user_id
            request.session['username']     = user_doc.get('username', '')
            request.session['is_staff']     = user_doc.get('is_staff', False)
            request.session['is_superuser'] = user_doc.get('is_superuser', False)
            request.session['prenom']       = user_doc.get('first_name', '')
            request.session['nom']          = user_doc.get('last_name', '')
            request.session['email']        = user_doc.get('email', '')

            if next_url and next_url.startswith('/'):
                return redirect(next_url)
            if user_doc.get('is_staff'):
                return redirect('dashboard')
            return redirect('employe_espace')
        else:
            messages.error(request, "Nom d'utilisateur ou mot de passe incorrect.")

    return render(request, 'dashboard/login.html', {'next': next_url})

def logout_view(request):
    try:
        session_key = request.session.session_key
        now         = timezone.now()
        user_id     = request.session.get('user_id', '')
 
        if session_key:
            db['dashboard_usersession'].update_one(
                {'session_key': session_key},
                {'$set': {'is_active': False, 'logout_time': now}}
            )
            db['dashboard_sessionlog'].insert_one({
                'user_id':     user_id,
                'action':      'logout',
                'ip_address':  request.META.get('REMOTE_ADDR', ''),
                'user_agent':  request.META.get('HTTP_USER_AGENT', '')[:500],
                'session_key': session_key,
                'timestamp':   now,
            })
    except Exception as e:
        logger.warning(f"logout_view session cleanup error: {e}")
 
    request.session.flush()
    messages.success(request, "Vous avez été déconnecté avec succès.")
    return redirect('login')
 
 


from config.settings import MONGO_DB as db
from datetime import datetime
import bcrypt  # pip install bcrypt --break-system-packages
def register_employe(request):
    import bcrypt
 
    if request.session.get('user_id'):
        if request.session.get('is_staff'):
            return redirect('dashboard')
        return redirect('employe_espace')
 
    if request.method == 'POST':
        username  = request.POST.get('username', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        badge_id  = request.POST.get('badge_id', '').strip()
        nom       = request.POST.get('nom', '').strip()
        prenom    = request.POST.get('prenom', '').strip()
        email     = request.POST.get('email', '').strip()
 
        erreurs = []
 
        if not username:
            erreurs.append("Le nom d'utilisateur est requis.")
        if len(password1) < 6:
            erreurs.append("Le mot de passe doit contenir au moins 6 caractères.")
        if password1 != password2:
            erreurs.append("Les mots de passe ne correspondent pas.")
        if not badge_id:
            erreurs.append("Le numéro de badge est requis.")
        if not nom or not prenom:
            erreurs.append("Le nom et le prénom sont requis.")
 
        employe_mongo = db.employees.find_one({'badge_id': badge_id})
        if not employe_mongo:
            erreurs.append(f"Badge '{badge_id}' non reconnu. Contactez votre administrateur.")
 
        if db.utilisateurs.find_one({'username': username}):
            erreurs.append(f"Le nom d'utilisateur '{username}' est déjà pris.")
 
        if erreurs:
            for e in erreurs:
                messages.error(request, e)
            return render(request, 'dashboard/register_employe.html', {'form_data': request.POST})
 
        hashed_pw = bcrypt.hashpw(password1.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
 
        result = db.utilisateurs.insert_one({
            'username':    username,
            'password':    hashed_pw,
            'email':       email,
            'first_name':  prenom,
            'last_name':   nom,
            'is_staff':    False,
            'is_superuser': False,
            'is_active':   True,
            'date_joined': datetime.now(),
        })
 
        user_id = str(result.inserted_id)
 
        db.employees.update_one(
            {'badge_id': badge_id},
            {'$set': {
                'django_user_id':  user_id,
                'django_username': username,
                'email':           email,
                'compte_cree_le':  datetime.now(),
                'nom':             nom,
                'prenom':          prenom,
            }}
        )
 
        request.session['user_id']      = user_id
        request.session['username']     = username
        request.session['is_staff']     = False
        request.session['is_superuser'] = False
        request.session['prenom']       = prenom
        request.session['nom']          = nom
        request.session['email']        = email
 
        messages.success(request, f"Compte créé avec succès ! Bienvenue {prenom}.")
        return redirect('employe_espace')
 
    return render(request, 'dashboard/register_employe.html', {'form_data': {}})
# ====================== ESPACE EMPLOYÉ ======================
@session_required
def employe_espace(request):
    """Tableau de bord employé amélioré"""
    if request.session.get('is_staff', False) or request.session.get('is_superuser', False):
        return redirect('dashboard')

    from datetime import datetime, timedelta

    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})

    if not employe:
        messages.error(request, "Profil employé introuvable. Contactez l'administrateur.")
        logout(request)
        return redirect('login')

    employe['id']      = str(employe['_id'])
    utilisateur_id     = employe['_id']           # ObjectId — identique à employe_mon_historique
    utilisateur_id_str = str(utilisateur_id)

    # ── Statistiques globales ────────────────────────────────────────────────
    total_acces     = db.acces_logs.count_documents({'utilisateur_id': utilisateur_id})
    acces_autorises = db.acces_logs.count_documents({'utilisateur_id': utilisateur_id, 'resultat': 'AUTORISE'})
    acces_refuses   = total_acces - acces_autorises
    taux_succes     = round(min(100, acces_autorises / total_acces * 100) if total_acces > 0 else 0, 1)

    # ── Statistiques du mois ─────────────────────────────────────────────────
    start_month      = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    total_acces_mois = db.acces_logs.count_documents({
        'utilisateur_id': utilisateur_id,
        'timestamp':      {'$gte': start_month}
    })

    # ── Jours actifs ─────────────────────────────────────────────────────────
    try:
        pipeline_jours = [
            {'$match': {'utilisateur_id': utilisateur_id}},
            {'$group': {
                '_id': {
                    'year':  {'$year':       '$timestamp'},
                    'month': {'$month':      '$timestamp'},
                    'day':   {'$dayOfMonth': '$timestamp'},
                }
            }},
            {'$count': 'total_days'}
        ]
        result             = list(db.acces_logs.aggregate(pipeline_jours))
        jours_actifs_count = result[0]['total_days'] if result else 0
    except Exception:
        jours_actifs_count = 0

    # ── Heures totales (approx. 30 min par accès) ────────────────────────────
    heures_totales = round(total_acces * 0.5, 1)

    # ── Accès récents ────────────────────────────────────────────────────────
    acces = list(db.acces_logs.find({'utilisateur_id': utilisateur_id}).sort('timestamp', -1).limit(10))
    for a in acces:
        bureau        = db.bureaux.find_one({'_id': a.get('bureau_id')})
        a['bureau_nom'] = bureau['nom'] if bureau else 'Zone inconnue'
        if not a.get('type_acces'):
            a['type_acces'] = 'RFID'

    # ── Réservations ─────────────────────────────────────────────────────────
    reservations = list(
        db.reservations.find({'employe_id': utilisateur_id_str}).sort('date_debut', -1)
    )
    now                  = datetime.now()
    a_venir              = 0
    reservations_a_venir = []
    prochaine_resa       = None

    for r in reservations:
        r['id']        = str(r['_id'])
        bureau         = db.bureaux.find_one({'_id': r.get('bureau_id')})
        r['bureau_nom'] = bureau['nom'] if bureau else 'Salle inconnue'

        if r.get('statut') == 'confirmee' and r.get('date_debut') and r['date_debut'] > now:
            a_venir += 1
            reservations_a_venir.append(r)
            if not prochaine_resa:
                prochaine_resa = r

    # ── Suggestions personnalisées ───────────────────────────────────────────

    # Jour le plus fréquent
    frequent_day = "mercredi"
    try:
        day_pipeline = [
            {'$match': {'utilisateur_id': utilisateur_id}},
            {'$group': {
                '_id':   {'$dayOfWeek': '$timestamp'},
                'count': {'$sum': 1}
            }},
            {'$sort': {'count': -1}},
            {'$limit': 1}
        ]
        day_result = list(db.acces_logs.aggregate(day_pipeline))
        if day_result:
            # MongoDB $dayOfWeek : 1=dimanche, 2=lundi … 7=samedi
            days_map     = {1: 'dimanche', 2: 'lundi', 3: 'mardi', 4: 'mercredi',
                            5: 'jeudi', 6: 'vendredi', 7: 'samedi'}
            frequent_day = days_map.get(day_result[0]['_id'], 'mercredi')
    except Exception:
        pass

    # Salle recommandée
    recommended_room = "Salle de réunion A"
    try:
        room_pipeline = [
            {'$match': {'employe_id': utilisateur_id_str, 'statut': 'confirmee'}},
            {'$group': {'_id': '$bureau_id', 'count': {'$sum': 1}}},
            {'$sort': {'count': -1}},
            {'$limit': 1}
        ]
        room_result = list(db.reservations.aggregate(room_pipeline))
        if room_result:
            bureau = db.bureaux.find_one({'_id': room_result[0]['_id']})
            if bureau:
                recommended_room = bureau.get('nom', 'Salle de réunion')
    except Exception:
        pass

    # Meilleur créneau horaire
    best_time = "09h00-10h00"
    try:
        hour_pipeline = [
            {'$match': {'utilisateur_id': utilisateur_id}},
            {'$group': {
                '_id':   {'$hour': '$timestamp'},
                'count': {'$sum': 1}
            }},
            {'$sort': {'count': -1}},
            {'$limit': 1}
        ]
        hour_result = list(db.acces_logs.aggregate(hour_pipeline))
        if hour_result:
            peak_hour = hour_result[0]['_id']
            best_time = f"{peak_hour:02d}h00-{(peak_hour + 1):02d}h00"
    except Exception:
        pass

    # Taux d'occupation actuel
    occupancy_rate = 15
    try:
        one_hour_ago     = datetime.now() - timedelta(hours=1)
        total_occupation = db.acces_logs.count_documents({'timestamp': {'$gte': one_hour_ago}})
        occupancy_rate   = min(100, round(total_occupation / 10 * 100)) if total_occupation > 0 else 15
    except Exception:
        pass

    # ── Notifications non lues ───────────────────────────────────────────────
    notifs_count = 0
    try:
        notifs_count = db.notifications.count_documents({
            'employe_id': utilisateur_id_str,   # string, pas ObjectId
            'status':     {'$ne': 'lu'}          # tout ce qui n'est pas 'lu'
    })
    except Exception:
        notifs_count = 0

    # ── Bureaux (formulaire réservation) ─────────────────────────────────────
    bureaux = list(db.bureaux.find())
    for b in bureaux:
        b['id']           = str(b['_id'])
        b['capacite_max'] = b.get('capacite_max', 10)

    return render(request, 'dashboard/employe_espace.html', {
        'employe':              employe,
        'acces':                acces,
        'reservations':         reservations,
        'reservations_a_venir': reservations_a_venir[:5],
        'prochaine_resa':       prochaine_resa,
        'bureaux':              bureaux,
        # Statistiques
        'total_acces':          total_acces,
        'total_acces_mois':     total_acces_mois,
        'acces_autorises':      acces_autorises,
        'acces_refuses':        acces_refuses,
        'taux_succes':          taux_succes,
        'jours_actifs':         jours_actifs_count,
        'heures_totales':       heures_totales,
        'a_venir':              a_venir,
        # Suggestions
        'frequent_day':         frequent_day,
        'recommended_room':     recommended_room,
        'best_time':            best_time,
        'occupancy_rate':       occupancy_rate,
        # Notifications
        'notifs_count':         notifs_count,
        # Divers
        'now':                  datetime.now(),
    })
# dashboard/views.py - Modifiez la fonction employe_mes_reservations
@session_required
def api_employe_notif_unread_count(request):
    from django.http import JsonResponse
    if request.session.get('is_staff', False) or request.session.get('is_superuser', False):
        return JsonResponse({'count': 0})
    
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})
    if not employe:
        return JsonResponse({'count': 0})
    
    count = db.notifications.count_documents({
        'employe_id': str(employe['_id']),
        'status':     {'$ne': 'lu'}
    })
    return JsonResponse({'count': count})
@session_required
def employe_mes_reservations(request):
    if request.session.get('is_staff', False):
        return redirect('dashboard')

    from datetime import datetime
    from bson import ObjectId
    import json

    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')}) \
              or db.employees.find_one({'django_username': request.session.get('username', '')})

    if not employe:
        messages.error(request, "Profil employé introuvable.")
        return redirect('login')

    employe['id'] = str(employe['_id'])

    # ── Paramètres de réservation (toujours chargés, GET et POST) ──────────
    params_resa  = db.parametres.find_one({'cle': 'reservation'}) or {}
    salle_min    = params_resa.get('salle_min_minutes',      30)
    salle_max    = params_resa.get('salle_max_minutes',    1440)
    materiel_min = params_resa.get('materiel_min_minutes',   60)
    materiel_max = params_resa.get('materiel_max_minutes', 525600)

    # ============ TRAITEMENT POST (création de réservation) ============
    if request.method == 'POST':
        try:
            titre         = (request.POST.get('titre') or '').strip()
            description   = (request.POST.get('description') or '').strip()
            resource_id   = request.POST.get('resource_id') or request.POST.get('bureau_id')
            resource_type = request.POST.get('resource_type', 'salle')
            date_debut_s  = request.POST.get('date_debut')
            date_fin_s    = request.POST.get('date_fin')
            nb_part       = int(request.POST.get('nb_participants', 1) or 1)

            # --- Validations de base ---
            if not titre:
                messages.error(request, "Le titre est obligatoire.")
                return redirect('employe_mes_reservations')

            if not resource_id:
                messages.error(request, "Veuillez choisir une salle ou une ressource.")
                return redirect('employe_mes_reservations')

            if not date_debut_s or not date_fin_s:
                messages.error(request, "Les dates de début et de fin sont obligatoires.")
                return redirect('employe_mes_reservations')

            try:
                date_debut = datetime.strptime(date_debut_s, '%Y-%m-%dT%H:%M')
                date_fin   = datetime.strptime(date_fin_s,   '%Y-%m-%dT%H:%M')
            except ValueError:
                messages.error(request, "Format de date invalide.")
                return redirect('employe_mes_reservations')

            if date_fin <= date_debut:
                messages.error(request, "La date de fin doit être après la date de début.")
                return redirect('employe_mes_reservations')

            # --- Validation durée selon type de ressource ---
            duree_minutes = int((date_fin - date_debut).total_seconds() / 60)

            if resource_type == 'salle':
                if duree_minutes < salle_min:
                    messages.error(request, f"Une salle ne peut pas être réservée moins de {salle_min} minute(s).")
                    return redirect('employe_mes_reservations')
                if duree_minutes > salle_max:
                    h = salle_max // 60
                    messages.error(request, f"Une salle ne peut pas être réservée plus de {salle_max} minutes ({h}h maximum).")
                    return redirect('employe_mes_reservations')
            else:
                if duree_minutes < materiel_min:
                    messages.error(request, f"Une ressource ne peut pas être réservée moins de {materiel_min} minute(s).")
                    return redirect('employe_mes_reservations')
                if duree_minutes > materiel_max:
                    j = materiel_max // 1440
                    messages.error(request, f"Une ressource ne peut pas être réservée plus de {materiel_max} minutes ({j} jour(s) maximum).")
                    return redirect('employe_mes_reservations')

            if date_debut < datetime.now():
                messages.error(request, "Impossible de réserver dans le passé.")
                return redirect('employe_mes_reservations')

            # --- Vérification disponibilité (réservations + indisponibilités) ---
            check = check_ressource_disponibilite(
                resource_id, resource_type, date_debut, date_fin
            )
            if not check['disponible']:
                messages.error(request, f"❌ {check['motif']}")
                return redirect('employe_mes_reservations')

            # --- Vérification capacité (salles uniquement) ---
            try:
                resource_oid = ObjectId(resource_id)
            except Exception:
                messages.error(request, "Identifiant de ressource invalide.")
                return redirect('employe_mes_reservations')

            if resource_type == 'salle':
                bureau = db.bureaux.find_one({'_id': resource_oid})
                if not bureau:
                    messages.error(request, "Salle introuvable.")
                    return redirect('employe_mes_reservations')
                cap_max = bureau.get('capacite_max')
                if cap_max and nb_part > cap_max:
                    messages.error(request,
                        f"⚠️ Capacité dépassée : la salle « {bureau['nom']} » "
                        f"accepte {cap_max} personnes maximum.")
                    return redirect('employe_mes_reservations')
                ressource_nom   = bureau['nom']
                ressource_label = f"🚪 Salle: {ressource_nom}"
            else:
                materiel = db.materiels.find_one({'_id': resource_oid})
                if not materiel:
                    messages.error(request, "Matériel introuvable.")
                    return redirect('employe_mes_reservations')
                ressource_nom   = materiel['nom']
                ressource_label = f"📦 Matériel: {ressource_nom}"

            # --- Construction de la réservation ---
            reservation_data = {
                'titre':           titre,
                'description':     description,
                'employe_id':      str(employe['_id']),
                'employe_nom':     f"{employe.get('nom', '')} {employe.get('prenom', '')}".strip(),
                'date_debut':      date_debut,
                'date_fin':        date_fin,
                'nb_participants': nb_part,
                'statut':          'en_attente',
                'qr_code':         None,
                'resource_type':   resource_type,
                'resource_id':     resource_oid,
                'bureau_nom':      ressource_nom,
                'created_at':      datetime.now(),
                'created_by':      request.session.get('username', ''),
            }

            if resource_type == 'salle':
                reservation_data['bureau_id'] = resource_oid
            else:
                reservation_data['materiel_id']  = resource_oid
                reservation_data['materiel_nom'] = ressource_nom

            result = db.reservations.insert_one(reservation_data)
            reservation_id = str(result.inserted_id)

            # --- Notification employé ---
            db.notifications.insert_one({
                'employe_id':     str(employe['_id']),
                'titre':          '📝 Réservation créée',
                'message':        f"Votre réservation '{titre}' a été créée et est en attente de validation.",
                'categorie':      'reservation',
                'icon':           '📝',
                'status':         'non_lu',
                'action_url':     '/employe/reservations/',
                'reservation_id': reservation_id,
                'created_at':     datetime.now(),
            })

            # --- Notifications admins (PyMongo direct) ---
            admins = list(db['utilisateurs'].find(
                {'is_staff': True, 'is_active': True}, {'_id': 1, 'email': 1}
            ))

            admin_message = (
                f"🆕 NOUVELLE RÉSERVATION EN ATTENTE\n\n"
                f"👤 Employé: {employe.get('prenom', '')} {employe.get('nom', '')}\n"
                f"📋 Titre: {titre}\n"
                f"{ressource_label}\n"
                f"📅 Date: {date_debut.strftime('%d/%m/%Y')}\n"
                f"⏰ Horaire: {date_debut.strftime('%H:%M')} → {date_fin.strftime('%H:%M')}\n"
                f"👥 Participants: {nb_part}\n\n"
                f"🔗 Cliquez pour traiter cette réservation: /reservations/"
            )

            for admin in admins:
                db.admin_notifications.insert_one({
                    'admin_id':       admin.get('_id'),
                    'titre':          '🆕 Nouvelle réservation en attente',
                    'message':        (
                        f"{employe.get('prenom', '')} {employe.get('nom', '')} "
                        f"a demandé une réservation pour '{titre}' "
                        f"({ressource_label}) le {date_debut.strftime('%d/%m/%Y à %H:%M')}."
                    ),
                    'categorie':      'reservation',
                    'icon':           '🆕',
                    'status':         'non_lu',
                    'action_url':     f'/reservations/{reservation_id}/',
                    'reservation_id': reservation_id,
                    'created_at':     datetime.now(),
                })
                if admin.get('email'):
                    try:
                        from dashboard.utils_email import envoyer_email
                        envoyer_email(
                            admin['email'],
                            f"🆕 Nouvelle réservation — {titre}",
                            admin_message,
                        )
                    except Exception as _ee:
                        logger.warning(f"Email admin échoué: {_ee}")

            messages.success(request, "✅ Réservation créée avec succès ! En attente de validation.")
            return redirect('employe_mes_reservations')

        except Exception as e:
            logger.exception("Erreur création réservation employé")
            messages.error(request, f"Erreur: {str(e)}")
            return redirect('employe_mes_reservations')

    # ============ AFFICHAGE GET ============
    reservations = list(
        db.reservations.find({'employe_id': str(employe['_id'])}).sort('date_debut', -1)
    )

    for r in reservations:
        r['id'] = str(r['_id'])
        rtype = r.get('resource_type') or ('materiel' if r.get('materiel_id') else 'salle')

        if rtype == 'materiel':
            mat_id = r.get('materiel_id') or r.get('resource_id')
            mat = None
            if mat_id is not None:
                try:
                    mat = db.materiels.find_one({'_id': mat_id if isinstance(mat_id, ObjectId) else ObjectId(str(mat_id))})
                except Exception:
                    mat = None
            r['bureau_nom']     = (mat['nom'] if mat else r.get('materiel_nom') or r.get('bureau_nom') or 'Matériel inconnu')
            r['ressource_icon'] = '📦'
        else:
            bur_id = r.get('bureau_id') or r.get('resource_id')
            bureau = None
            if bur_id is not None:
                try:
                    bureau = db.bureaux.find_one({'_id': bur_id if isinstance(bur_id, ObjectId) else ObjectId(str(bur_id))})
                except Exception:
                    bureau = None
            r['bureau_nom']     = (bureau['nom'] if bureau else r.get('bureau_nom') or 'Salle inconnue')
            r['ressource_icon'] = '🚪'

        r.setdefault('qr_code', None)

    now = datetime.now()
    employe_id_str = str(employe['_id'])

    total_confirmees = db.reservations.count_documents({
        'employe_id': employe_id_str,
        'statut':     'confirmee'
    })
    a_venir = db.reservations.count_documents({
        'employe_id': employe_id_str,
        'statut':     'confirmee',
        'date_debut': {'$gt': now}
    })
    actives = db.reservations.count_documents({
        'employe_id': employe_id_str,
        'statut':     'confirmee',
        'date_debut': {'$lte': now},
        'date_fin':   {'$gte': now}
    })
    en_attente = db.reservations.count_documents({
        'employe_id': employe_id_str,
        'statut':     'en_attente'
    })

    bureaux = list(db.bureaux.find())
    for b in bureaux:
        b['id']           = str(b['_id'])
        b['capacite_max'] = b.get('capacite_max', 10)

    materiels = list(db.materiels.find()) if 'materiels' in db.list_collection_names() else []
    for m in materiels:
        m['id']        = str(m['_id'])
        m['type_icon'] = get_materiel_icon(m.get('categorie', 'autre'))

    reservations_list = [
        {
            'id':         r['id'],
            'titre':      r.get('titre', ''),
            'bureau_nom': r.get('bureau_nom', ''),
            'statut':     r.get('statut', ''),
            'date_debut': r['date_debut'].isoformat() if r.get('date_debut') else None,
            'date_fin':   r['date_fin'].isoformat()   if r.get('date_fin')   else None,
        }
        for r in reservations if r.get('date_debut')
    ]
    reservations_json = json.dumps(reservations_list, default=str)

    return render(request, 'dashboard/employe_mes_reservations.html', {
        'employe':           employe,
        'reservations':      reservations,
        'bureaux':           bureaux,
        'materiels':         materiels,
        'total':             len(reservations),
        'actives':           actives,
        'a_venir':           a_venir,
        'en_attente':        en_attente,
        'reservations_json': reservations_json,
        'total_confirmees':  total_confirmees,
        'params_resa':       params_resa,
    })
@session_required
def employe_annuler_reservation(request, reservation_id):
    if request.session.get('is_staff', False):
        return redirect('dashboard')
    
    from datetime import datetime
    from bson import ObjectId
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})
    
    if not employe:
        return redirect('login')
    
    if request.method == 'POST':
        try:
            resa = db.reservations.find_one({
                '_id': ObjectId(reservation_id),
                'employe_id': str(employe['_id'])
            })
            
            if resa:
                # Récupérer la salle
                bureau = db.bureaux.find_one({'_id': resa.get('bureau_id')})
                bureau_nom = bureau['nom'] if bureau else 'Salle inconnue'
                
                db.reservations.update_one(
                    {'_id': ObjectId(reservation_id)},
                    {'$set': {
                        'statut': 'annulee', 
                        'cancelled_at': datetime.now(), 
                        'cancelled_by': request.session.get('username', '')
                    }}
                )
                
                # === NOTIFICATION À L'EMPLOYÉ ===
                notification_employe = {
                    'employe_id': str(employe['_id']),
                    'titre': '🗑️ Réservation annulée',
                    'message': f"Votre réservation '{resa.get('titre', 'Sans titre')}' a été annulée.",
                    'categorie': 'annulation',
                    'icon': '🗑️',
                    'status': 'non_lu',
                    'action_url': '/employe/reservations/',
                    'reservation_id': reservation_id,
                    'created_at': datetime.now()
                }
                db.notifications.insert_one(notification_employe)
                
                # === NOTIFICATION AUX ADMINISTRATEURS (PyMongo direct) ===
                try:
                    admins = list(db['utilisateurs'].find(
                        {'is_staff': True, 'is_active': True}, {'_id': 1}
                    ))
                    for admin in admins:
                        db.admin_notifications.insert_one({
                            'admin_id':       admin.get('_id'),
                            'titre':          '🗑️ Réservation annulée',
                            'message':        f"{employe.get('prenom', '')} {employe.get('nom', '')} a annulé sa réservation '{resa.get('titre', 'Sans titre')}' pour la salle {bureau_nom}.",
                            'categorie':      'reservation',
                            'icon':           '🗑️',
                            'status':         'non_lu',
                            'action_url':     f'/reservations/{reservation_id}/',
                            'reservation_id': reservation_id,
                            'created_at':     datetime.now(),
                        })
                except Exception as _e:
                    logger.warning(f"Notifications annulation échouées: {_e}")
                
                messages.success(request, "Réservation annulée avec succès.")
            else:
                messages.error(request, "Réservation introuvable ou non autorisée.")
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    return redirect('employe_mes_reservations')
# ═══════════════════════════════════════════════════════════════
#  EXPORT CSV — RÉSERVATIONS
# ═══════════════════════════════════════════════════════════════
import csv
from django.http import HttpResponse
from datetime import datetime

@session_required
def reservations_export_csv(request):
    if request.session.get('is_staff', False):
        return redirect('dashboard')

    # ── Même logique que employe_mes_reservations ──────────────
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')}) \
              or db.employees.find_one({'django_username': request.session.get('username', '')})

    if not employe:
        return HttpResponse("Profil employé introuvable.", status=403)

    reservations = list(
        db.reservations.find({'employe_id': str(employe['_id'])}).sort('date_debut', -1)
    )

    now_str  = datetime.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="Reservations_{now_str}.csv"'
    response.write('\ufeff')  # BOM UTF-8 pour Excel

    writer = csv.writer(response)
    writer.writerow([
        'Titre', 'Salle / Ressource', 'Type',
        'Début', 'Fin', 'Participants', 'Statut', 'Description', 'Créé le'
    ])

    for r in reservations:
        debut    = r.get('date_debut')
        fin      = r.get('date_fin')
        created  = r.get('created_at')
        writer.writerow([
            r.get('titre', ''),
            r.get('bureau_nom', r.get('materiel_nom', '—')),
            r.get('resource_type', 'salle'),
            debut.strftime('%d/%m/%Y %H:%M')   if hasattr(debut,   'strftime') else '—',
            fin.strftime('%d/%m/%Y %H:%M')     if hasattr(fin,     'strftime') else '—',
            r.get('nb_participants', 1),
            r.get('statut', ''),
            r.get('description', ''),
            created.strftime('%d/%m/%Y %H:%M') if hasattr(created, 'strftime') else '—',
        ])

    return response


# ═══════════════════════════════════════════════════════════════
#  EXPORT PDF — RÉSERVATIONS
# ═══════════════════════════════════════════════════════════════
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether
)
from reportlab.platypus.flowables import Flowable
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.pdfgen import canvas as rl_canvas
import io
from datetime import datetime

C_BLUE_DARK  = colors.HexColor('#0f172a')
C_BLUE_MAIN  = colors.HexColor('#1d4ed8')
C_BLUE_LIGHT = colors.HexColor('#dbeafe')
C_BLUE_MID   = colors.HexColor('#3b82f6')
C_PURPLE     = colors.HexColor('#7c3aed')
C_GREEN      = colors.HexColor('#059669')
C_AMBER      = colors.HexColor('#d97706')
C_RED        = colors.HexColor('#dc2626')
C_GREY_LIGHT = colors.HexColor('#f8fafc')
C_GREY_MID   = colors.HexColor('#e2e8f0')
C_GREY_TEXT  = colors.HexColor('#64748b')
C_WHITE      = colors.white
C_BLACK      = colors.HexColor('#0f172a')


class _NumberedCanvas(rl_canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        n = len(self._saved_page_states)
        for i, state in enumerate(self._saved_page_states):
            self.__dict__.update(state)
            self._draw_footer(i + 1, n)
            rl_canvas.Canvas.showPage(self)
        rl_canvas.Canvas.save(self)

    def _draw_footer(self, page_num, total):
        w, _ = A4
        self.setFillColor(C_GREY_TEXT)
        self.setFont('Helvetica', 8)
        self.drawCentredString(w / 2, 1.2 * cm,
            f"Page {page_num} / {total}  —  SIGR-CA — Document confidentiel")
        self.setStrokeColor(C_GREY_MID)
        self.setLineWidth(0.5)
        self.line(2*cm, 1.6*cm, w - 2*cm, 1.6*cm)


class _ColorBand(Flowable):
    def __init__(self, text, width, height=0.9*cm,
                 bg=C_BLUE_DARK, fg=C_WHITE, font_size=11):
        super().__init__()
        self.text = text
        self.band_width = width
        self.band_height = height
        self.bg = bg; self.fg = fg; self.font_size = font_size

    def wrap(self, *args):
        return self.band_width, self.band_height

    def draw(self):
        c = self.canv
        c.setFillColor(self.bg)
        c.rect(0, 0, self.band_width, self.band_height, fill=1, stroke=0)
        c.setFillColor(self.fg)
        c.setFont('Helvetica-Bold', self.font_size)
        c.drawString(0.4*cm, 0.25*cm, self.text)


def _fmt(val):
    if val is None:
        return '—'
    if hasattr(val, 'strftime'):
        return val.strftime('%d/%m/%Y %H:%M')
    return str(val)

@session_required
def reservations_export_pdf(request):
    """Export PDF réservations — même structure exacte que api_employes_export_pdf."""
    import io, os, traceback
    from datetime import datetime
    from django.http import HttpResponse
    from django.contrib.staticfiles.finders import find as static_find
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, HRFlowable, KeepTogether,
    )
    from reportlab.platypus.flowables import Flowable
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.pdfgen import canvas as rl_canvas

    try:
        # ── Couleurs ──────────────────────────────────────────────
        C_BLUE_DARK  = colors.HexColor('#0f172a')
        C_BLUE_MAIN  = colors.HexColor('#1d4ed8')
        C_BLUE_LIGHT = colors.HexColor('#dbeafe')
        C_BLUE_MID   = colors.HexColor('#3b82f6')
        C_PURPLE     = colors.HexColor('#7c3aed')
        C_GREEN      = colors.HexColor('#059669')
        C_AMBER      = colors.HexColor('#d97706')
        C_RED        = colors.HexColor('#dc2626')
        C_GREY_LIGHT = colors.HexColor('#f8fafc')
        C_GREY_MID   = colors.HexColor('#e2e8f0')
        C_GREY_TEXT  = colors.HexColor('#64748b')
        C_WHITE      = colors.white
        C_BLACK      = colors.HexColor('#0f172a')

        # ── Numérotation des pages ────────────────────────────────
        class NumberedCanvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved_page_states = []
            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                n = len(self._saved_page_states)
                for i, state in enumerate(self._saved_page_states):
                    self.__dict__.update(state)
                    self._draw_footer(i + 1, n)
                    rl_canvas.Canvas.showPage(self)
                rl_canvas.Canvas.save(self)
            def _draw_footer(self, page_num, total_pages):
                w, _ = A4
                self.setFillColor(colors.HexColor('#64748b'))
                self.setFont('Helvetica', 8)
                self.drawCentredString(w / 2, 1.2 * cm,
                    f"Page {page_num} / {total_pages}  —  SIGR-CA — Document confidentiel")
                self.setStrokeColor(colors.HexColor('#e2e8f0'))
                self.setLineWidth(0.5)
                self.line(2 * cm, 1.6 * cm, w - 2 * cm, 1.6 * cm)

        # ── Bande colorée ─────────────────────────────────────────
        class ColorBand(Flowable):
            def __init__(self, text, width, height=0.9*cm,
                         bg=None, fg=None, font_size=11):
                super().__init__()
                self.text       = text
                self.band_width  = width
                self.band_height = height
                self.bg         = bg or colors.HexColor('#0f172a')
                self.fg         = fg or colors.white
                self.font_size  = font_size
            def wrap(self, *args):
                return self.band_width, self.band_height
            def draw(self):
                c = self.canv
                c.setFillColor(self.bg)
                c.rect(0, 0, self.band_width, self.band_height, fill=1, stroke=0)
                c.setFillColor(self.fg)
                c.setFont('Helvetica-Bold', self.font_size)
                c.drawString(0.4 * cm, 0.25 * cm, self.text)

        # ── hex_color helper (identique à export employés) ────────
        def hex_color(c):
            try:
                return f'{int(c.red*255):02x}{int(c.green*255):02x}{int(c.blue*255):02x}'
            except Exception:
                return '0f172a'

        # ── Données ───────────────────────────────────────────────
        employe = (
            db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
            or db.employees.find_one({'django_username': request.session.get('username', '')})
        )
        if not employe:
            return HttpResponse("Profil employé introuvable.", status=403)

        emp_nom = f"{employe.get('prenom', '')} {employe.get('nom', '')}".strip()

        reservations = list(
            db.reservations.find({'employe_id': str(employe['_id'])})
            .sort('date_debut', -1)
        )

        now_str   = datetime.now().strftime('%d/%m/%Y à %H:%M')
        date_file = datetime.now().strftime('%Y%m%d_%H%M')

        # ── Document ──────────────────────────────────────────────
        buffer    = io.BytesIO()
        PAGE_W, _ = A4
        CONTENT_W = PAGE_W - 4 * cm   # 17 cm exactement

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2.2*cm, bottomMargin=2.2*cm,
            title=f"Réservations — {emp_nom}",
            author="SIGR-CA Système",
        )

        # ── Styles ────────────────────────────────────────────────
        _sty = {}
        def ps(name, **kw):
            base = kw.pop('parent', 'Normal')
            sheet = getSampleStyleSheet()
            parent = _sty.get(base) or sheet.get(base, sheet['Normal'])
            p = ParagraphStyle(name, parent=parent, **kw)
            _sty[name] = p
            return p

        th       = ps('TH',  fontName='Helvetica-Bold', fontSize=9,
                       textColor=C_WHITE, alignment=TA_CENTER, leading=12)
        th_left  = ps('THL', parent='TH', alignment=TA_LEFT)
        td       = ps('TD',  fontSize=9, alignment=TA_CENTER, leading=12, textColor=C_BLACK)
        td_left  = ps('TDL', parent='TD', alignment=TA_LEFT)
        td_mono  = ps('TDM', parent='TD', fontSize=8, fontName='Courier')
        sec_sty  = ps('SEC', fontName='Helvetica-Bold', fontSize=12,
                       textColor=C_PURPLE, spaceBefore=20, spaceAfter=8, leading=16)
        foot_sty = ps('FOT', fontSize=7.5, textColor=C_GREY_TEXT,
                       alignment=TA_CENTER, leading=11)

        _NO_PAD = TableStyle([
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
            ('TOPPADDING',    (0,0),(-1,-1), 1),
            ('BOTTOMPADDING', (0,0),(-1,-1), 1),
        ])

        elements = []

        # ════════════════════════════════════════════
        #  EN-TÊTE  (identique à export employés)
        # ════════════════════════════════════════════
        LOGO_PATH = static_find('img/logo.png')

        title_tbl = Table([
            [Paragraph('<font color="#1d4ed8"><b>SIGR-CA</b></font>',
                       ps('LT', fontSize=22, leading=26, alignment=TA_LEFT))],
            [Paragraph("Système Intégré de Gestion des Ressources<br/>"
                       "<font color='#64748b'>et de Contrôle d'Accès</font>",
                       ps('LS', fontSize=9, leading=13,
                          textColor=C_GREY_TEXT, alignment=TA_LEFT))],
        ], colWidths=[11*cm])
        title_tbl.setStyle(_NO_PAD)

        if LOGO_PATH and os.path.exists(LOGO_PATH):
            logo_cell = Image(LOGO_PATH, width=4*cm, height=2.6*cm)
        else:
            logo_cell = Paragraph('<font color="#1d4ed8"><b>SIGR</b></font>',
                                  ps('FL', fontSize=18, alignment=TA_RIGHT))

        meta_tbl = Table([
            [Paragraph(f"<b>Employé :</b> {emp_nom}",
                       ps('M0', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph(f"<b>Date :</b> {now_str}",
                       ps('M1', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph(f"<b>Réservations :</b> {len(reservations)}",
                       ps('M2', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph("<b>Confidentiel</b>",
                       ps('M3', fontSize=8, textColor=C_RED, alignment=TA_RIGHT))],
        ], colWidths=[4.5*cm])
        meta_tbl.setStyle(_NO_PAD)

        right_col = Table([[logo_cell], [meta_tbl]], colWidths=[4.5*cm])
        right_col.setStyle(TableStyle([
            ('ALIGN',         (0,0),(-1,-1), 'RIGHT'),
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
            ('TOPPADDING',    (0,0),(-1,-1), 2),
            ('BOTTOMPADDING', (0,0),(-1,-1), 2),
        ]))

        # colWidths : 12.5 + 4.5 = 17cm = CONTENT_W ✓
        header_tbl = Table([[title_tbl, right_col]],
                           colWidths=[12.5*cm, 4.5*cm])
        header_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0),(-1,-1), 'TOP'),
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
            ('TOPPADDING',    (0,0),(-1,-1), 0),
            ('BOTTOMPADDING', (0,0),(-1,-1), 6),
        ]))

        elements.append(header_tbl)
        elements.append(HRFlowable(width='100%', thickness=2,
                                   color=C_BLUE_MAIN, spaceAfter=4))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(ColorBand(
            f"  MES RÉSERVATIONS — {emp_nom.upper()}",
            CONTENT_W, height=1.1*cm,
            bg=C_BLUE_DARK, fg=C_WHITE, font_size=12,
        ))
        elements.append(Spacer(1, 0.6*cm))

        # ════════════════════════════════════════════
        #  TABLEAU PRINCIPAL
        # ════════════════════════════════════════════
        elements.append(KeepTogether([Paragraph('Liste des réservations', sec_sty)]))

        STATUT_MAP = {
            'confirmee':  ('#059669', '✔ Confirmée'),
            'en_attente': ('#d97706', '⏳ En attente'),
            'annulee':    ('#dc2626', '✖ Annulée'),
            'terminee':   ('#64748b', '■ Terminée'),
        }

        # colWidths : 4.3+3.2+2.7+2.7+1.3+2.8 = 17cm = CONTENT_W ✓
        col_widths = [4.3*cm, 3.2*cm, 2.7*cm, 2.7*cm, 1.3*cm, 2.8*cm]
        data = [[
            Paragraph('Titre',  th_left),
            Paragraph('Salle',  th),
            Paragraph('Début',  th),
            Paragraph('Fin',    th),
            Paragraph('Part.',  th),
            Paragraph('Statut', th),
        ]]

        for r in reservations:
            statut = r.get('statut', '')
            color, label = STATUT_MAP.get(statut, ('#64748b', statut))
            bureau = r.get('bureau_nom') or r.get('materiel_nom') or '—'
            debut  = r.get('date_debut')
            fin    = r.get('date_fin')
            data.append([
                Paragraph(f"<b>{r.get('titre') or '—'}</b>", td_left),
                Paragraph(str(bureau), td),
                Paragraph(debut.strftime('%d/%m/%Y %H:%M') if hasattr(debut,'strftime') else '—', td_mono),
                Paragraph(fin.strftime('%d/%m/%Y %H:%M')   if hasattr(fin,  'strftime') else '—', td_mono),
                Paragraph(str(r.get('nb_participants', 1)), td),
                Paragraph(f'<font color="{color}"><b>{label}</b></font>', td),
            ])

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',     (0,0), (-1,0),   C_BLUE_MAIN),
            ('TEXTCOLOR',      (0,0), (-1,0),   C_WHITE),
            ('ROWBACKGROUNDS', (0,1), (-1,-1),  [C_WHITE, C_BLUE_LIGHT]),
            ('GRID',           (0,0), (-1,-1),  0.4, C_GREY_MID),
            ('LINEBELOW',      (0,0), (-1,0),   1.5, C_BLUE_MAIN),
            ('TOPPADDING',     (0,0), (-1,0),   10),
            ('BOTTOMPADDING',  (0,0), (-1,0),   10),
            ('TOPPADDING',     (0,1), (-1,-1),  7),
            ('BOTTOMPADDING',  (0,1), (-1,-1),  7),
            ('LEFTPADDING',    (0,0), (-1,-1),  7),
            ('VALIGN',         (0,0), (-1,-1),  'MIDDLE'),
            ('BOX',            (0,0), (-1,-1),  1, C_BLUE_MID),
        ]))
        elements.append(tbl)

        # ════════════════════════════════════════════
        #  RÉSUMÉ STATISTIQUE — tableau plat, SANS Spacer en cellule
        # ════════════════════════════════════════════
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph('Résumé statistique', sec_sty))

        total      = len(reservations)
        confirmees = sum(1 for r in reservations if r.get('statut') == 'confirmee')
        en_attente = sum(1 for r in reservations if r.get('statut') == 'en_attente')
        annulees   = sum(1 for r in reservations if r.get('statut') == 'annulee')
        terminees  = sum(1 for r in reservations if r.get('statut') == 'terminee')
        salles     = sum(1 for r in reservations if r.get('resource_type','salle') == 'salle')
        materiels  = sum(1 for r in reservations if r.get('resource_type') == 'materiel')

        def stat_row(label, value, value_color=C_BLACK):
            hv = hex_color(value_color)
            return [
                Paragraph(label, ps(f'SL{label[:4]}', fontSize=9,
                                    textColor=C_GREY_TEXT, alignment=TA_LEFT)),
                Paragraph(f'<font color="#{hv}"><b>{value}</b></font>',
                          ps(f'SV{label[:4]}', fontSize=10,
                             alignment=TA_RIGHT, fontName='Helvetica-Bold')),
            ]

        # Un seul tableau plat — colWidths : 10 + 7 = 17cm = CONTENT_W ✓
        stats_data = [
            stat_row('Total réservations',  str(total)),
            stat_row('Confirmées',          str(confirmees),  C_GREEN),
            stat_row('En attente',          str(en_attente),  C_AMBER),
            stat_row('Annulées',            str(annulees),    C_RED),
            stat_row('Terminées',           str(terminees),   C_GREY_TEXT),
            stat_row('Salles réservées',    str(salles)),
            stat_row('Matériels réservés',  str(materiels)),
        ]

        stats_tbl = Table(stats_data, colWidths=[10*cm, 7*cm])
        stats_tbl.setStyle(TableStyle([
            ('ROWBACKGROUNDS', (0,0),(-1,-1), [C_WHITE, C_GREY_LIGHT]),
            ('GRID',           (0,0),(-1,-1), 0.4, C_GREY_MID),
            ('TOPPADDING',     (0,0),(-1,-1), 7),
            ('BOTTOMPADDING',  (0,0),(-1,-1), 7),
            ('LEFTPADDING',    (0,0),(-1,-1), 10),
            ('RIGHTPADDING',   (0,0),(-1,-1), 10),
            ('VALIGN',         (0,0),(-1,-1), 'MIDDLE'),
            ('BOX',            (0,0),(-1,-1), 1, C_GREY_MID),
        ]))
        elements.append(stats_tbl)

        # ════════════════════════════════════════════
        #  PIED DE PAGE DOCUMENT
        # ════════════════════════════════════════════
        elements.append(Spacer(1, 1.2*cm))
        elements.append(HRFlowable(width='100%', thickness=1,
                                   color=C_GREY_MID, spaceAfter=6))
        elements.append(Paragraph(
            f"Document généré automatiquement le <b>{now_str}</b> par le système SIGR-CA.",
            foot_sty))
        elements.append(Paragraph(
            "Ce document est <b>confidentiel</b> et destiné à un usage interne uniquement.",
            ps('CF2', fontSize=7, textColor=C_GREY_TEXT, alignment=TA_CENTER, leading=10)))

        doc.build(elements, canvasmaker=NumberedCanvas)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'inline; filename="Reservations_{emp_nom.replace(" ","_")}_{date_file}.pdf"')
        return response

    except Exception as e:
        return HttpResponse(
            f"Erreur PDF : {str(e)}\n\n{traceback.format_exc()}",
            content_type='text/plain', status=500)
@session_required
def employe_mon_historique(request):
    if request.session.get('is_staff', False):
        return redirect('dashboard')
    
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})
    
    if not employe:
        return redirect('login')
    
    employe['id'] = str(employe['_id'])
    acces = list(db.acces_logs.find({'utilisateur_id': employe['_id']}).sort('timestamp', -1).limit(200))
    
    for a in acces:
        bureau = db.bureaux.find_one({'_id': a.get('bureau_id')})
        a['bureau_nom'] = bureau['nom'] if bureau else 'Zone inconnue'
    
    total_acces = len(acces)
    acces_autorises = sum(1 for a in acces if a.get('resultat') == 'AUTORISE')
    acces_refuses = total_acces - acces_autorises
    taux_succes = round(acces_autorises / total_acces * 100) if total_acces else 0
    
    return render(request, 'dashboard/employe_mon_historique.html', {
        'employe': employe,
        'acces': acces,
        'total_acces': total_acces,
        'acces_autorises': acces_autorises,
        'acces_refuses': acces_refuses,
        'taux_succes': taux_succes,
    })


# ====================== API RÉSERVATIONS ======================

# dashboard/views.py - Version simplifiée de l'API

@session_required
def api_reservation_details(request, reservation_id):
    """API pour récupérer les détails d'une réservation (version simplifiée)"""
    try:
        reservation = db.reservations.find_one({'_id': ObjectId(reservation_id)})
        
        if not reservation:
            return JsonResponse({'error': 'Réservation non trouvée'}, status=404)
        
        # Données basiques
        response_data = {
            'titre': reservation.get('titre', 'Sans titre'),
            'description': reservation.get('description', ''),
            'bureau_nom': str(reservation.get('bureau_id', 'Salle inconnue')),
            'employe_nom': str(reservation.get('employe_id', 'Inconnu')),
            'date_debut': reservation.get('date_debut'),
            'date_fin': reservation.get('date_fin'),
            'nb_participants': reservation.get('nb_participants', 1),
            'statut': reservation.get('statut', 'en_attente'),
        }
        
        # Essayer d'enrichir avec les vrais noms
        if reservation.get('bureau_id'):
            try:
                bureau = db.bureaux.find_one({'_id': reservation['bureau_id']})
                if bureau:
                    response_data['bureau_nom'] = bureau.get('nom', 'Salle inconnue')
            except:
                pass
        
        if reservation.get('employe_id'):
            try:
                emp = db.employees.find_one({'_id': ObjectId(reservation['employe_id'])})
                if emp:
                    response_data['employe_nom'] = f"{emp.get('nom', '')} {emp.get('prenom', '')}".strip() or 'Employé'
            except:
                pass
        
        # Convertir les dates en string
        if response_data['date_debut']:
            response_data['date_debut'] = response_data['date_debut'].isoformat()
        if response_data['date_fin']:
            response_data['date_fin'] = response_data['date_fin'].isoformat()
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# dashboard/views.py - Remplacer api_reservations_calendrier

@session_required
def api_reservations_calendrier(request):
    """API améliorée pour le calendrier avec filtres"""
    try:
        start_str = request.GET.get('start')
        end_str = request.GET.get('end')
        resource_filter = request.GET.get('resource')
        statut_filter = request.GET.get('statut', 'all')
        employe_filter = request.GET.get('employe', '')
        
        if not start_str or not end_str:
            return JsonResponse({'events': []})
        
        start = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
        
        # Construire la requête
        query = {
            'statut': {'$in': ['confirmee', 'en_attente']},
            'date_debut': {'$lt': end},
            'date_fin': {'$gt': start}
        }
        
        # Filtre par ressource
        if resource_filter and resource_filter != 'all':
            if resource_filter.startswith('bureau_'):
                bureau_id = resource_filter.replace('bureau_', '')
                query['bureau_id'] = ObjectId(bureau_id)
        
        # Filtre par statut
        if statut_filter != 'all':
            query['statut'] = statut_filter
        
        # Filtre par employé
        if employe_filter:
            query['employe_nom'] = {'$regex': employe_filter, '$options': 'i'}
        
        reservations = list(db.reservations.find(query))
        
        colors = {
            'confirmee': '#2dba6f',
            'en_attente': '#e3b341',
            'annulee': '#f85149'
        }
        
        events = []
        for r in reservations:
            if r.get('date_debut') and r.get('date_fin'):
                # Récupérer le nom du bureau
                bureau_nom = 'Salle inconnue'
                if r.get('bureau_id'):
                    bureau = db.bureaux.find_one({'_id': r['bureau_id']})
                    bureau_nom = bureau['nom'] if bureau else 'Salle inconnue'
                
                events.append({
                    'id': str(r['_id']),
                    'title': r.get('titre', 'Réservation'),
                    'start': r['date_debut'].isoformat(),
                    'end': r['date_fin'].isoformat(),
                    'color': colors.get(r.get('statut', 'en_attente'), '#388bfd'),
                    'extendedProps': {
                        'bureau': bureau_nom,
                        'statut': r.get('statut'),
                        'participants': r.get('nb_participants', 1),
                        'description': r.get('description', '')
                    }
                })
        
        return JsonResponse({'events': events})
        
    except Exception as e:
        logger.error(f"Erreur calendrier: {e}")
        return JsonResponse({'events': [], 'error': str(e)})

@session_required
def api_disponibilite_bureau(request, bureau_id):
    """API : disponibilité d'une salle (réservations + indisponibilités)"""
    debut = request.GET.get('debut')
    fin   = request.GET.get('fin')
    if not debut or not fin:
        return JsonResponse({'disponible': False, 'motif': 'Dates manquantes'})
    try:
        date_debut = datetime.fromisoformat(debut)
        date_fin   = datetime.fromisoformat(fin)
    except Exception:
        return JsonResponse({'disponible': False, 'motif': 'Format de date invalide'})
    result = check_ressource_disponibilite(bureau_id, 'salle', date_debut, date_fin)
    return JsonResponse(result)

    


# ====================== DASHBOARD ADMIN ======================
# dashboard/views.py - Version corrigée de dashboard()
from datetime import datetime, timedelta
from django.shortcuts import render, redirect
from bson import ObjectId
import os

@session_required
def dashboard(request):
    if not request.session.get('is_staff', False) and not request.session.get('is_superuser', False):
        return redirect('employe_espace')

    # Statistiques de base
    total_employes = db.employees.count_documents({})
    total_bureaux = db.bureaux.count_documents({})

    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    acces_aujourdhui = db.acces_logs.count_documents({'timestamp': {'$gte': today_start}})
    acces_refuses = db.acces_logs.count_documents({
        'timestamp': {'$gte': today_start}, 
        'resultat': 'REFUSE'
    })
    acces_autorises_today = acces_aujourdhui - acces_refuses

    # Alertes
    alertes = db.alertes.count_documents({'statut': 'NON_TRAITEE'}) if 'alertes' in db.list_collection_names() else 0

    # Derniers logs (10 derniers)
    derniers_logs = list(db.acces_logs.find().sort('timestamp', -1).limit(10))
    for log in derniers_logs:
        # === Employé (champ "employe_id") ===
        emp_id = log.get('employe_id') or log.get('utilisateur_id')
        if emp_id:
            try:
                eid = ObjectId(emp_id) if isinstance(emp_id, str) else emp_id
                employe = db.employees.find_one({'_id': eid})
                if employe:
                    nom_complet = f"{employe.get('prenom','')} {employe.get('nom','')}".strip()
                    log['nom_utilisateur'] = nom_complet or employe.get('email') or 'Utilisateur inconnu'
                else:
                    log['nom_utilisateur'] = 'Utilisateur inconnu'
            except Exception:
                log['nom_utilisateur'] = 'Utilisateur inconnu'
        else:
            log['nom_utilisateur'] = log.get('fait_par', 'Système')

        # === Bureau (peut ne pas exister) ===
        bureau_id = log.get('bureau_id')
        if bureau_id:
            try:
                bid = ObjectId(bureau_id) if isinstance(bureau_id, str) else bureau_id
                bureau = db.bureaux.find_one({'_id': bid})
                log['bureau_nom'] = bureau.get('nom', 'Salle inconnue') if bureau else 'Salle inconnue'
            except Exception:
                log['bureau_nom'] = 'Salle inconnue'
        else:
            # Pas de bureau = action admin (badge_affecte, etc.)
            action = log.get('action', '')
            log['bureau_nom'] = action.replace('_', ' ').capitalize() if action else 'Action système'

        # === Type d'accès (champ "badge_type") ===
        log['type_acces'] = log.get('badge_type') or log.get('type_acces') or 'SYSTEM'

        # === Résultat : si absent, déduit de l'action ===
        if 'resultat' not in log:
            action = log.get('action', '').lower()
            if 'refus' in action or 'denied' in action or 'echec' in action:
                log['resultat'] = 'REFUSE'
            else:
                log['resultat'] = 'AUTORISE'

        # === Heure formatée ===
        if 'timestamp' in log and isinstance(log['timestamp'], datetime):
            log['heure_formatee'] = log['timestamp'].strftime('%H:%M:%S')
        else:
            log['heure_formatee'] = '--:--:--'
    # Stats 7 jours
    seven_days_ago = datetime.now() - timedelta(days=7)
    pipeline = [
        {'$match': {'timestamp': {'$gte': seven_days_ago}}},
        {'$group': {
            '_id': {'$dateToString': {'format': '%Y-%m-%d', 'date': '$timestamp'}},
            'total': {'$sum': 1},
            'autorises': {'$sum': {'$cond': [{'$eq': ['$resultat', 'AUTORISE']}, 1, 0]}}
        }},
        {'$sort': {'_id': 1}}
    ]
    stats_7jours = list(db.acces_logs.aggregate(pipeline))

    # Calcul pourcentages + nom du jour
    max_total = max((s.get('total', 0) for s in stats_7jours), default=1)
    jours_fr = {'Monday':'Lun','Tuesday':'Mar','Wednesday':'Mer','Thursday':'Jeu','Friday':'Ven','Saturday':'Sam','Sunday':'Dim'}

    for s in stats_7jours:
        s['refuses'] = s['total'] - s['autorises']
        s['pct_total'] = int((s['total'] / max_total) * 100) if max_total else 0
        s['pct_autor'] = int((s['autorises'] / max_total) * 100) if max_total else 0
        
        try:
            date_obj = datetime.strptime(s['_id'], '%Y-%m-%d')
            s['jour_court'] = jours_fr.get(date_obj.strftime('%A'), date_obj.strftime('%a')[:3])
        except:
            s['jour_court'] = s['_id'][5:10]

    # IA Models status
    try:
        from dashboard.ai_engine import MODELS_DIR
    except:
        MODELS_DIR = os.path.join(os.path.dirname(__file__), 'models')
        os.makedirs(MODELS_DIR, exist_ok=True)

    models_status = {
        'occupation': os.path.exists(os.path.join(MODELS_DIR, 'occupation_model.pkl')),
        'recommender': os.path.exists(os.path.join(MODELS_DIR, 'reco_model.pkl')),
        'anomaly': os.path.exists(os.path.join(MODELS_DIR, 'anomaly_model.pkl')),
    }

    bureaux_ia = list(db.bureaux.find({}, {'nom': 1}))
    for b in bureaux_ia:
        b['id'] = str(b['_id'])

    context = {
        'total_employes': total_employes,
        'total_bureaux': total_bureaux,
        'acces_aujourdhui': acces_aujourdhui,
        'acces_refuses': acces_refuses,
        'acces_autorises_today': acces_autorises_today,
        'alertes': alertes,
        'derniers_logs': derniers_logs,
        'stats_7jours': stats_7jours,
        'models_status': models_status,
        'bureaux_ia': bureaux_ia,
        'total_resa_ia': db.reservations.count_documents({}),
        'total_emp_ia': total_employes,
    }

    return render(request, 'dashboard/dashboard.html', context)
    
   
# ====================== GESTION DES EMPLOYÉS ======================

# ====================== GESTION DES EMPLOYÉS ======================
# views_employes.py  –  à intégrer dans votre views.py principal
#
# Changements vs version originale :
#  1. employe_list     → statut normalisé : None / '' / absent → 'actif'
#  2. employe_ajouter  → sauvegarde photo (base64 ou fichier)
#  3. employe_modifier → sauvegarde photo + champs cycle de travail + congés
#  4. Helper save_photo() centralisé

import os, base64, uuid
from datetime import datetime
from collections import Counter
from bson import ObjectId

from django.shortcuts import render, redirect
from django.contrib import messages

# ── Votre connexion MongoDB (adaptez selon votre projet) ──
# from .db import db   ← assurez-vous que `db` est importé ici


# ─────────────────────────────────────────────
#  HELPER : enregistre la photo et retourne l'URL
# ─────────────────────────────────────────────
def save_photo(request, employe_id):
    if request.POST.get('photo_data') == '__remove__':
        return False

    # ← CORRECTION ICI : chemin absolu
    from django.conf import settings as _s
    upload_dir = os.path.join(_s.MEDIA_ROOT, 'avatars')
    os.makedirs(upload_dir, exist_ok=True)
    
    filename = f"avatar_{employe_id}_{uuid.uuid4().hex[:8]}.jpg"
    filepath = os.path.join(upload_dir, filename)

    photo_file = request.FILES.get('photo')
    if photo_file:
        with open(filepath, 'wb') as f:
            for chunk in photo_file.chunks():
                f.write(chunk)
        return f"/media/avatars/{filename}"

    photo_data = request.POST.get('photo_data', '').strip()
    if photo_data and photo_data.startswith('data:image'):
        try:
            header, encoded = photo_data.split(',', 1)
            img_bytes = base64.b64decode(encoded)
            with open(filepath, 'wb') as f:
                f.write(img_bytes)
            return f"/media/avatars/{filename}"
        except Exception:
            pass

    return None
# ─────────────────────────────────────────────
#  LISTE DES EMPLOYÉS (VERSION CORRIGÉE)
# ─────────────────────────────────────────────
from django.shortcuts import render
from datetime import datetime
import pymongo

# Connexion MongoDB (à adapter selon votre config)
# db = client.votre_base

@session_required
def employe_list(request):
    try:
        employes_raw = list(db.employees.find({}))
    except Exception:
        employes_raw = []

    employes = []
    total_acces_global = 0
    total_autorises_global = 0

    for e in employes_raw:
        e['id'] = str(e['_id'])
        
        # Valeurs par défaut
        e.setdefault('nom', '')
        e.setdefault('prenom', '')
        e.setdefault('badge_id', '')
        e.setdefault('email', '')
        e.setdefault('telephone', '')
        e.setdefault('departement', '')
        e.setdefault('poste', '')
        e.setdefault('niveau', 'Staff')
        e.setdefault('statut', 'actif')
        e.setdefault('photo_url', None)
        
        # Stats accès
        e['nb_acces'] = db.acces_logs.count_documents({'utilisateur_id': e['_id']})
        total_acces_global += e['nb_acces']
        
        acces_autorises = db.acces_logs.count_documents(
            {'utilisateur_id': e['_id'], 'resultat': 'AUTORISE'}
        )
        total_autorises_global += acces_autorises
        
        e['taux_succes'] = round((acces_autorises / e['nb_acces'] * 100), 1) if e['nb_acces'] > 0 else 0
        
        # Dernier accès
        dernier = db.acces_logs.find_one(
            {'utilisateur_id': e['_id']}, 
            sort=[('timestamp', pymongo.DESCENDING)]
        )
        e['dernier_acces'] = dernier['timestamp'] if dernier else None
        
        employes.append(e)
    
    # Départements uniques
    departements = sorted(set(e.get('departement', '') for e in employes if e.get('departement')))
    
    # Stats par département
    dept_stats = []
    for dept in departements:
        dept_employes = [e for e in employes if e.get('departement') == dept]
        total_acces_dept = sum(e.get('nb_acces', 0) for e in dept_employes)
        total_autorises_dept = sum(
            int(round(e.get('taux_succes', 0) * e.get('nb_acces', 0) / 100)) 
            for e in dept_employes
        )
        dept_stats.append({
            'nom': dept,
            'total': len(dept_employes),
            'total_acces': total_acces_dept,
            'taux': round((total_autorises_dept / total_acces_dept * 100), 1) if total_acces_dept > 0 else 0
        })
    
    # Accès aujourd'hui
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    acces_aujourdhui = db.acces_logs.count_documents({'timestamp': {'$gte': today_start}})
    
    # Taux global
    taux_global = round((total_autorises_global / total_acces_global * 100), 1) if total_acces_global > 0 else 0

    
    return render(request, 'dashboard/employe_list.html', {
        'employes': employes,
        'total_employes': len(employes),
        'actifs': sum(1 for e in employes if e.get('statut') == 'actif'),
        'inactifs': sum(1 for e in employes if e.get('statut') == 'inactif'),
        'total_departements': len(departements),
        'departements': departements,
        'dept_stats': dept_stats,
        'acces_aujourdhui': acces_aujourdhui,
        'taux_global': taux_global,
    })
# ─────────────────────────────────────────────
#  DÉTAIL D'UN EMPLOYÉ
# ─────────────────────────────────────────────
# dashboard/views.py

from pymongo import MongoClient
from bson import ObjectId
from datetime import datetime

mongo_client = MongoClient('localhost', 27017)
mongo_db = mongo_client['general_emballage']

@session_required
def employe_details(request, employe_id):
    employe = mongo_db.employees.find_one({'_id': ObjectId(employe_id)})
    if not employe:
        return render(request, '404.html', {'message': 'Employé non trouvé'})

    employe['id'] = str(employe['_id'])

    # Même requête que employe_list : utilisateur_id en ObjectId
    acces_list = list(mongo_db.acces_logs.find({
        '$or': [
            {'utilisateur_id': employe['_id']},
            {'utilisateur_id': str(employe['_id'])},
            {'employe_id':     str(employe['_id'])},
        ]
    }).sort('timestamp', -1).limit(200))

    for acces in acces_list:
        bid = acces.get('bureau_id')
        if bid:
            try:
                bureau = mongo_db.bureaux.find_one({'_id': ObjectId(str(bid))})
                acces['bureau_nom'] = bureau['nom'] if bureau else 'Salle inconnue'
            except Exception:
                acces['bureau_nom'] = 'Salle inconnue'
        else:
            acces['bureau_nom'] = 'Zone inconnue'

    reservations_list = list(mongo_db.reservations.find({
        '$or': [
            {'employe_id': employe['_id']},
            {'employe_id': str(employe['_id'])},
        ]
    }).sort('date_debut', -1))

    for resa in reservations_list:
        bid = resa.get('bureau_id')
        if bid:
            try:
                bureau = mongo_db.bureaux.find_one({'_id': ObjectId(str(bid))})
                resa['bureau_nom'] = bureau['nom'] if bureau else 'Salle inconnue'
            except Exception:
                resa['bureau_nom'] = 'Salle inconnue'
        else:
            resa['bureau_nom'] = 'Salle inconnue'

    total_acces     = len(acces_list)
    acces_autorises = len([a for a in acces_list if a.get('resultat') == 'AUTORISE'])
    acces_refuses   = total_acces - acces_autorises
    taux_succes = round((acces_autorises / total_acces * 100)) if total_acces > 0 else 0

    return render(request, 'dashboard/employe_details.html', {
        'employe':           employe,
        'acces_list':        acces_list,
        'reservations_list': reservations_list,
        'total_acces':       total_acces,
        'acces_autorises':   acces_autorises,
        'acces_refuses':     acces_refuses,
        'taux_succes':       taux_succes,
        'dernier_acces':     acces_list[0] if acces_list else None,
    })
# ─────────────────────────────────────────────
#  AJOUTER UN EMPLOYÉ  (avec photo)
# ─────────────────────────────────────────────
@session_required
def employe_ajouter(request):
    import re
    JOURS = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']

    if request.method == 'POST':
        # ── Récupération ──
        badge_id = request.POST.get('badge_id', '').strip().upper()
        nom      = request.POST.get('nom', '').strip()
        prenom   = request.POST.get('prenom', '').strip()
        tel      = request.POST.get('telephone', '').strip().replace(' ', '')

        # ── Validation badge obligatoire + format RFID-XXXX ──
        if not badge_id:
            messages.error(request, "Le badge RFID est obligatoire.")
            return render(request, 'dashboard/employe_form.html', {
                'employe': request.POST, 'jours_semaine': JOURS,
            })

        if not re.match(r'^[A-F0-9]{2}:[A-F0-9]{2}:[A-F0-9]{2}:[A-F0-9]{2}$', badge_id):
            messages.error(request, "Le badge doit être au format XX:XX:XX:XX (ex : D7:77:C5:01).")
            return render(request, 'dashboard/employe_form.html', {
            'employe': request.POST, 'jours_semaine': JOURS,
        })

        if db.employees.find_one({'badge_id': badge_id}):
            messages.error(request, f"Le badge {badge_id} est déjà attribué.")
            return render(request, 'dashboard/employe_form.html', {
                'employe': request.POST, 'jours_semaine': JOURS,
            })

        # ── Validation nom (obligatoire, lettres uniquement) ──
        if not re.match(r"^[A-Za-zÀ-ÿ\s\-']+$", nom):
            messages.error(request, "Le nom ne doit contenir que des lettres.")
            return render(request, 'dashboard/employe_form.html', {
                'employe': request.POST, 'jours_semaine': JOURS,
            })

        # ── Validation prénom (optionnel, lettres uniquement si fourni) ──
        if prenom and not re.match(r"^[A-Za-zÀ-ÿ\s\-']+$", prenom):
            messages.error(request, "Le prénom ne doit contenir que des lettres.")
            return render(request, 'dashboard/employe_form.html', {
                'employe': request.POST, 'jours_semaine': JOURS,
            })

        # ── Validation téléphone (optionnel, 10 chiffres algériens) ──
        if tel and not re.match(r'^0[5-7][0-9]{8}$', tel):
            messages.error(request, "Le téléphone doit avoir 10 chiffres et commencer par 05, 06 ou 07.")
            return render(request, 'dashboard/employe_form.html', {
                'employe': request.POST, 'jours_semaine': JOURS,
            })

        # ── Helpers conversion robustes ──
        def to_float(val, default):
            try:
                return float(str(val).replace(',', '.')) if val not in (None, '') else default
            except (ValueError, TypeError):
                return default

        def to_int(val, default):
            try:
                return int(val) if val not in (None, '') else default
            except (ValueError, TypeError):
                return default

        # ── Construction du document ──
        nouvel_employe = {
            'badge_id':         badge_id,
            'nom':              nom.upper(),
            'prenom':           prenom,
            'email':            request.POST.get('email', '').strip().lower(),
            'telephone':        tel,
            'departement':      request.POST.get('departement', ''),
            'poste':            request.POST.get('poste', '').strip(),
            'niveau':           request.POST.get('niveau', 'Staff'),
            'statut':           'actif',
            'type_contrat':     request.POST.get('type_contrat', 'CDI'),
            'manager':          request.POST.get('manager', '').strip(),
            'horaire':          request.POST.get('horaire', '08:00 - 17:00').strip(),
            'heures_hebdo':     to_int(request.POST.get('heures_hebdo'), 35),
            'jours_travailles': request.POST.getlist('jours_travailles') or ['Lun','Mar','Mer','Jeu','Ven'],
            'solde_conges':     to_float(request.POST.get('solde_conges'), 25),
            'solde_rtt':        to_float(request.POST.get('solde_rtt'), 10),
            'solde_maladie':    to_float(request.POST.get('solde_maladie'), 0),
            'photo_url':        None,
            'created_at':       datetime.now(),
        }

        date_str = request.POST.get('date_embauche', '').strip()
        if date_str:
            try:
                nouvel_employe['date_embauche'] = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                pass

        # ── Insertion ──
        try:
            result = db.employees.insert_one(nouvel_employe)
            new_id = str(result.inserted_id)
        except Exception as e:
            import traceback; traceback.print_exc()
            messages.error(request, f"Erreur base de données : {e}")
            return render(request, 'dashboard/employe_form.html', {
                'employe': request.POST, 'jours_semaine': JOURS,
            })

        # ── Photo ──
        try:
            photo_url = save_photo(request, new_id)
            if photo_url:
                db.employees.update_one(
                    {'_id': result.inserted_id},
                    {'$set': {'photo_url': photo_url}}
                )
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"⚠️ Photo non sauvegardée : {e}")

        messages.success(
            request,
            f"Employé {nouvel_employe['nom']} {nouvel_employe['prenom']} créé avec succès !"
        )

        try:
            return redirect('employe_detail', employe_id=new_id)
        except Exception:
            return redirect('employe_list')

    # GET
    return render(request, 'dashboard/employe_form.html', {
        'employe': {}, 'jours_semaine': JOURS,
    })
@session_required
def api_check_badge(request):
    """Vérifie si un badge RFID existe déjà. Retourne {exists: true/false}."""
    from django.http import JsonResponse
    badge = request.GET.get('badge', '').strip()
    exclude_id = request.GET.get('exclude_id', '').strip()  # pour la modification

    if not badge:
        return JsonResponse({'exists': False})

    query = {'badge_id': badge}
    if exclude_id:
        from bson import ObjectId
        try:
            query['_id'] = {'$ne': ObjectId(exclude_id)}
        except Exception:
            pass

    existant = db.employees.find_one(query)
    if existant:
        return JsonResponse({
            'exists': True,
            'nom': f"{existant.get('nom','')} {existant.get('prenom','')}".strip()
        })
    return JsonResponse({'exists': False})
# ─────────────────────────────────────────────
#  MODIFIER UN EMPLOYÉ  (avec photo)
# ─────────────────────────────────────────────
@session_required
def employe_modifier(request, employe_id):
    # Vérification droits
    if not request.session.get('is_staff', False) and not request.session.get('is_superuser', False):
        messages.error(request, "Accès non autorisé.")
        return redirect('employe_list')

    try:
        obj_id = ObjectId(employe_id)
    except Exception:
        messages.error(request, "ID d'employé invalide.")
        return redirect('employe_list')

    employe = db.employees.find_one({'_id': obj_id})
    if not employe:
        messages.error(request, "Employé non trouvé.")
        return redirect('employe_list')

    employe['id'] = str(employe['_id'])
    # Normalisation statut
    if employe.get('statut') not in ('actif', 'inactif'):
        employe['statut'] = 'actif'

    if request.method == 'POST':
        try:
            jours = request.POST.getlist('jours_travailles') or ['Lun','Mar','Mer','Jeu','Ven']

            update_data = {
                'badge_id':     request.POST.get('badge_id', '').strip(),
                'nom':          request.POST.get('nom', '').strip().upper(),
                'prenom':       request.POST.get('prenom', '').strip(),
                'email':        request.POST.get('email', '').strip().lower(),
                'telephone':    request.POST.get('telephone', '').strip(),
                'departement':  request.POST.get('departement', ''),
                'poste':        request.POST.get('poste', '').strip(),
                'niveau':       request.POST.get('niveau', 'Staff'),
                'statut':       request.POST.get('statut', 'actif'),
                'type_contrat': request.POST.get('type_contrat', 'CDI'),
                'manager':      request.POST.get('manager', '').strip(),
                'horaire':      request.POST.get('horaire', '08:00 - 17:00').strip(),
                'heures_hebdo': int(request.POST.get('heures_hebdo', 35) or 35),
                'jours_travailles': jours,
                'solde_conges': float(request.POST.get('solde_conges', 25) or 25),
                'solde_rtt':    float(request.POST.get('solde_rtt', 10) or 10),
                'solde_maladie':float(request.POST.get('solde_maladie', 0) or 0),
                'updated_at':   datetime.now(),
            }

            # Date d'embauche
            date_str = request.POST.get('date_embauche', '').strip()
            if date_str:
                try:
                    update_data['date_embauche'] = datetime.strptime(date_str, '%Y-%m-%d')
                except ValueError:
                    pass

            # Photo
            photo_result = save_photo(request, employe_id)
            if photo_result is False:
                # Suppression demandée
                update_data['photo_url'] = None
            elif photo_result:
                update_data['photo_url'] = photo_result
            # Si None → pas de changement, on ne touche pas photo_url

            result = db.employees.update_one({'_id': obj_id}, {'$set': update_data})

            if result.modified_count > 0:
                messages.success(
                    request,
                    f"Employé '{update_data['nom']} {update_data['prenom']}' mis à jour avec succès !")
            else:
                messages.info(request, "Aucune modification détectée.")

            return redirect('employe_details', employe_id=employe_id)

        except Exception as e:
            import traceback; traceback.print_exc()
            messages.error(request, f"Erreur : {str(e)}")

    return render(request, 'dashboard/employe_form.html', {
        'employe': employe,
        'is_edit': True,
    })


# ─────────────────────────────────────────────
#  ARCHIVER (soft delete)
# ─────────────────────────────────────────────
@session_required
def employe_supprimer(request, employe_id):
    if request.method == 'POST':
        try:
            db.employees.update_one(
                {'_id': ObjectId(employe_id)},
                {'$set': {'statut': 'inactif', 'archived_at': datetime.now()}}
            )
            messages.success(request, "Employé archivé avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
    return redirect('employe_list')

# ─────────────────────────────────────────────
#  EXPORT PDF — RAPPORT EMPLOYÉS AVEC LOGO
from django.http import HttpResponse
from django.conf import settings
from bson import ObjectId
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, HRFlowable, KeepTogether
)
from reportlab.platypus.flowables import Flowable
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.pdfgen import canvas as rl_canvas
import io, os
from datetime import datetime


# ═══════════════════════════════════════════
#  COULEURS PALETTE SIGR-CA
# ═══════════════════════════════════════════
C_BLUE_DARK   = colors.HexColor('#0f172a')   # fond header
C_BLUE_MAIN   = colors.HexColor('#1d4ed8')   # bleu principal
C_BLUE_LIGHT  = colors.HexColor('#dbeafe')   # fond ligne paire
C_BLUE_MID    = colors.HexColor('#3b82f6')   # accents
C_PURPLE      = colors.HexColor('#7c3aed')   # section titles
C_GREEN       = colors.HexColor('#059669')   # taux bon
C_AMBER       = colors.HexColor('#d97706')   # taux moyen
C_RED         = colors.HexColor('#dc2626')   # taux mauvais / archivé
C_GREY_LIGHT  = colors.HexColor('#f8fafc')   # fond ligne impaire
C_GREY_MID    = colors.HexColor('#e2e8f0')   # bordures
C_GREY_TEXT   = colors.HexColor('#64748b')   # texte secondaire
C_WHITE       = colors.white
C_BLACK       = colors.HexColor('#0f172a')


# ═══════════════════════════════════════════
#  NUMÉROTATION DES PAGES (canvas callback)
# ═══════════════════════════════════════════
class NumberedCanvas(rl_canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for i, state in enumerate(self._saved_page_states):
            self.__dict__.update(state)
            self._draw_footer(i + 1, num_pages)
            rl_canvas.Canvas.showPage(self)
        rl_canvas.Canvas.save(self)

    def _draw_footer(self, page_num, total_pages):
        w, _ = A4
        self.setFillColor(colors.HexColor('#64748b'))
        self.setFont('Helvetica', 8)
        self.drawCentredString(
            w / 2, 1.2 * cm,
            f"Page {page_num} / {total_pages}  —  SIGR-CA — Document confidentiel"
        )
        self.setStrokeColor(colors.HexColor('#e2e8f0'))
        self.setLineWidth(0.5)
        self.line(2 * cm, 1.6 * cm, w - 2 * cm, 1.6 * cm)
# ═══════════════════════════════════════════
#  BANDE DE FOND COLORÉE (Flowable custom)
# ═══════════════════════════════════════════
class ColorBand(Flowable):
    """Bande horizontale pleine couleur avec texte centré."""
    def __init__(self, text, width, height=0.9*cm,
                 bg=C_BLUE_DARK, fg=C_WHITE, font_size=11):
        super().__init__()
        self.text = text
        self.band_width = width
        self.band_height = height
        self.bg = bg
        self.fg = fg
        self.font_size = font_size

    def wrap(self, *args):
        return self.band_width, self.band_height

    def draw(self):
        c = self.canv
        c.setFillColor(self.bg)
        c.rect(0, 0, self.band_width, self.band_height, fill=1, stroke=0)
        c.setFillColor(self.fg)
        c.setFont('Helvetica-Bold', self.font_size)
        c.drawString(0.4 * cm, 0.25 * cm, self.text)


# ═══════════════════════════════════════════
#  HELPER : BADGE STATUT coloré inline
# ═══════════════════════════════════════════
def statut_para(statut, style):
    if statut == 'actif':
        color = '#059669'
        label = '● Actif'
    else:
        color = '#dc2626'
        label = '● Archivé'
    return Paragraph(
        f'<font color="{color}"><b>{label}</b></font>', style
    )


# ═══════════════════════════════════════════
#  VUE PRINCIPALE
# ═══════════════════════════════════════════
from django.http import HttpResponse
from django.conf import settings
from bson import ObjectId
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, HRFlowable, KeepTogether
)
from reportlab.platypus.flowables import Flowable
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.pdfgen import canvas as rl_canvas
import io, os, traceback
from datetime import datetime


# ═══════════════════════════════════════════
#  COULEURS PALETTE SIGR-CA
# ═══════════════════════════════════════════
C_BLUE_DARK  = colors.HexColor('#0f172a')
C_BLUE_MAIN  = colors.HexColor('#1d4ed8')
C_BLUE_LIGHT = colors.HexColor('#dbeafe')
C_BLUE_MID   = colors.HexColor('#3b82f6')
C_PURPLE     = colors.HexColor('#7c3aed')
C_GREEN      = colors.HexColor('#059669')
C_AMBER      = colors.HexColor('#d97706')
C_RED        = colors.HexColor('#dc2626')
C_GREY_LIGHT = colors.HexColor('#f8fafc')
C_GREY_MID   = colors.HexColor('#e2e8f0')
C_GREY_TEXT  = colors.HexColor('#64748b')
C_WHITE      = colors.white
C_BLACK      = colors.HexColor('#0f172a')


# ═══════════════════════════════════════════
#  NUMÉROTATION DES PAGES
# ═══════════════════════════════════════════
class NumberedCanvas(rl_canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for i, state in enumerate(self._saved_page_states):
            self.__dict__.update(state)
            self._draw_footer(i + 1, num_pages)
            rl_canvas.Canvas.showPage(self)
        rl_canvas.Canvas.save(self)

    def _draw_footer(self, page_num, total_pages):
        w, _ = A4
        self.setFillColor(colors.HexColor('#64748b'))
        self.setFont('Helvetica', 8)
        self.drawCentredString(
            w / 2, 1.2 * cm,
            f"Page {page_num} / {total_pages}  —  SIGR-CA — Document confidentiel"
        )
        self.setStrokeColor(colors.HexColor('#e2e8f0'))
        self.setLineWidth(0.5)
        self.line(2 * cm, 1.6 * cm, w - 2 * cm, 1.6 * cm)


# ═══════════════════════════════════════════
#  BANDE COLORÉE
# ═══════════════════════════════════════════
class ColorBand(Flowable):
    def __init__(self, text, width, height=0.9*cm,
                 bg=C_BLUE_DARK, fg=C_WHITE, font_size=11):
        super().__init__()
        self.text = text
        self.band_width = width
        self.band_height = height
        self.bg = bg
        self.fg = fg
        self.font_size = font_size

    def wrap(self, *args):
        return self.band_width, self.band_height

    def draw(self):
        c = self.canv
        c.setFillColor(self.bg)
        c.rect(0, 0, self.band_width, self.band_height, fill=1, stroke=0)
        c.setFillColor(self.fg)
        c.setFont('Helvetica-Bold', self.font_size)
        c.drawString(0.4 * cm, 0.25 * cm, self.text)


# ═══════════════════════════════════════════
#  BADGE STATUT
# ═══════════════════════════════════════════
def statut_para(statut, style):
    if statut == 'actif':
        color, label = '#059669', '● Actif'
    else:
        color, label = '#dc2626', '● Archivé'
    return Paragraph(f'<font color="{color}"><b>{label}</b></font>', style)


# ═══════════════════════════════════════════
#  HELPER COULEUR HEX
# ═══════════════════════════════════════════
def hex_color(c):
    """Retourne la valeur hex d'une couleur ReportLab sans le #."""
    try:
        r = int(c.red * 255)
        g = int(c.green * 255)
        b = int(c.blue * 255)
        return f'{r:02x}{g:02x}{b:02x}'
    except Exception:
        return '0f172a'


# ═══════════════════════════════════════════
#  VUE PRINCIPALE
# ═══════════════════════════════════════════
@session_required
def api_employes_export_pdf(request):
    """
    Export PDF professionnel de la liste des employés — SIGR-CA.
    Paramètre GET : ids (optionnel) — liste d'IDs séparés par virgules
    """
    try:
        # ── Récupération des données ──────────────────────────────
        ids_param = request.GET.get('ids', '')
        if ids_param:
            ids = [ObjectId(i) for i in ids_param.split(',') if i]
            employes = list(db.employees.find({'_id': {'$in': ids}}))
        else:
            employes = list(db.employees.find({}))

        now_str   = datetime.now().strftime('%d/%m/%Y à %H:%M')
        date_file = datetime.now().strftime('%Y%m%d_%H%M')

        # ── Mise en page ──────────────────────────────────────────
        buffer   = io.BytesIO()
        PAGE_W, _ = A4
        CONTENT_W = PAGE_W - 4 * cm

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2.2*cm, bottomMargin=2.2*cm,
            title="Rapport Employés — SIGR-CA",
            author="SIGR-CA Système",
            subject="Liste des employés",
        )

        # ── Registre de styles custom ─────────────────────────────
        _custom_styles = {}

        def ps(name, **kw):
            base = kw.pop('parent', 'Normal')
            stylesheet = getSampleStyleSheet()
            parent_style = _custom_styles.get(base) or stylesheet[base]
            p = ParagraphStyle(name, parent=parent_style, **kw)
            _custom_styles[name] = p
            return p

        # Styles définis dans l'ordre (les parents avant les enfants)
        th        = ps('TH',   fontName='Helvetica-Bold', fontSize=9,
                        textColor=C_WHITE, alignment=TA_CENTER,
                        leading=12, spaceAfter=0, spaceBefore=0)
        th_left   = ps('TH_L', parent='TH', alignment=TA_LEFT)
        td        = ps('TD',   fontSize=9, alignment=TA_CENTER,
                        leading=12, textColor=C_BLACK)
        td_left   = ps('TD_L', parent='TD', alignment=TA_LEFT)
        td_mono   = ps('TD_M', parent='TD', fontSize=8, fontName='Courier',
                        alignment=TA_CENTER)
        sec_style = ps('SEC',  fontName='Helvetica-Bold', fontSize=12,
                        textColor=C_PURPLE, spaceBefore=20, spaceAfter=8,
                        leading=16)
        foot_style = ps('FOOT', fontSize=7.5, textColor=C_GREY_TEXT,
                         alignment=TA_CENTER, leading=11)

        elements = []

        # ════════════════════════════════════════════
        #  EN-TÊTE
        # ════════════════════════════════════════════
        from django.contrib.staticfiles.finders import find as static_find

        LOGO_PATH = static_find('img/logo.png')

        title_lines = [
            [Paragraph(
                '<font color="#1d4ed8"><b>SIGR-CA</b></font>',
                ps('LT', fontSize=22, leading=26, alignment=TA_LEFT)
            )],
            [Paragraph(
                "Système Intégré de Gestion des Ressources<br/>"
                "<font color='#64748b'>et de Contrôle d'Accès</font>",
                ps('LS', fontSize=9, leading=13,
                   textColor=C_GREY_TEXT, alignment=TA_LEFT)
            )],
        ]
        title_tbl = Table(title_lines, colWidths=[11*cm])
        title_tbl.setStyle(TableStyle([
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ]))

        if LOGO_PATH and os.path.exists(LOGO_PATH):
            logo_cell = Image(LOGO_PATH, width=4*cm, height=2.6*cm)
        else:
            logo_cell = Paragraph(
                '<font color="#1d4ed8"><b>SIGR</b></font>',
                ps('FL', fontSize=18, alignment=TA_RIGHT)
            )

        meta_lines = [
            [Paragraph(f"<b>Date :</b> {now_str}",
                       ps('M1', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph(f"<b>Employés :</b> {len(employes)}",
                       ps('M2', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph("<b>Confidentiel</b>",
                       ps('M3', fontSize=8, textColor=C_RED, alignment=TA_RIGHT))],
        ]
        meta_tbl = Table(meta_lines, colWidths=[4.5*cm])
        meta_tbl.setStyle(TableStyle([
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))

        right_col = Table([[logo_cell], [meta_tbl]], colWidths=[4.5*cm])
        right_col.setStyle(TableStyle([
            ('ALIGN',         (0,0), (-1,-1), 'RIGHT'),
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))

        header_tbl = Table(
            [[title_tbl, right_col]],
            colWidths=[CONTENT_W - 4.5*cm, 4.5*cm]
        )
        header_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))

        elements.append(header_tbl)
        elements.append(HRFlowable(width='100%', thickness=2,
                                   color=C_BLUE_MAIN, spaceAfter=4))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(ColorBand(
            "  RAPPORT GÉNÉRAL — LISTE DES EMPLOYÉS",
            CONTENT_W, height=1.1*cm,
            bg=C_BLUE_DARK, fg=C_WHITE, font_size=12
        ))
        elements.append(Spacer(1, 0.6*cm))

        # ════════════════════════════════════════════
        #  TABLEAU PRINCIPAL
        # ════════════════════════════════════════════
        elements.append(KeepTogether([
            Paragraph('Liste complete des employes', sec_style),
        ]))

        col_widths = [4.5*cm, 2.8*cm, 2.8*cm, 2*cm, 1.8*cm, 1.8*cm, 2.1*cm]
        headers = [
            Paragraph('Nom & Prenom',  th_left),
            Paragraph('Badge RFID',    th),
            Paragraph('Departement',   th),
            Paragraph('Niveau',        th),
            Paragraph('Acces',         th),
            Paragraph('Succes',        th),
            Paragraph('Statut',        th),
        ]
        data = [headers]

        for emp in employes:
            nom_complet = (
                f"{emp.get('nom','').strip()} {emp.get('prenom','').strip()}".strip()
                or '—'
            )
            badge  = emp.get('badge_id', '—')
            dept   = emp.get('departement', '—')
            niveau = emp.get('niveau', 'Staff')

            nb_acces = db.acces_logs.count_documents({'utilisateur_id': emp['_id']})
            acces_ok = db.acces_logs.count_documents(
                {'utilisateur_id': emp['_id'], 'resultat': 'AUTORISE'}
            )
            taux = round(acces_ok / nb_acces * 100, 1) if nb_acces > 0 else 0.0

            if taux >= 80:
                taux_color = '#059669'
            elif taux >= 50:
                taux_color = '#d97706'
            else:
                taux_color = '#dc2626'

            statut = emp.get('statut', 'actif')

            row = [
                Paragraph(f'<b>{nom_complet}</b>', td_left),
                Paragraph(badge, td_mono),
                Paragraph(dept, td),
                Paragraph(niveau, td),
                Paragraph(str(nb_acces), td),
                Paragraph(f'<font color="{taux_color}"><b>{taux}%</b></font>', td),
                statut_para(statut, td),
            ]
            data.append(row)

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0), (-1, 0), C_BLUE_MAIN),
            ('TEXTCOLOR',      (0, 0), (-1, 0), C_WHITE),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
            ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
            ('LINEBELOW',      (0, 0), (-1, 0), 1.5, C_BLUE_MAIN),
            ('TOPPADDING',     (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING',  (0, 0), (-1, 0), 10),
            ('TOPPADDING',     (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING',  (0, 1), (-1, -1), 7),
            ('LEFTPADDING',    (0, 0), (-1, -1), 7),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX',            (0, 0), (-1, -1), 1, C_BLUE_MID),
        ]))
        elements.append(tbl)

        # ════════════════════════════════════════════
        #  RÉSUMÉ STATISTIQUE
        # ════════════════════════════════════════════
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph('Resume statistique', sec_style))

        total    = len(employes)
        actifs   = sum(1 for e in employes if e.get('statut', 'actif') == 'actif')
        archives = total - actifs

        total_acces = 0
        total_ok    = 0
        for emp in employes:
            nb = db.acces_logs.count_documents({'utilisateur_id': emp['_id']})
            ok = db.acces_logs.count_documents(
                {'utilisateur_id': emp['_id'], 'resultat': 'AUTORISE'}
            )
            total_acces += nb
            total_ok    += ok
        taux_global = round(total_ok / total_acces * 100, 1) if total_acces > 0 else 0.0

        niveaux = {}
        for emp in employes:
            n = emp.get('niveau', 'Staff')
            niveaux[n] = niveaux.get(n, 0) + 1

        def stat_row(label, value, value_color=C_BLACK):
            hex_val = hex_color(value_color)
            return [
                Paragraph(label, ps(f'SL_{label[:4]}', fontSize=9,
                                    textColor=C_GREY_TEXT, alignment=TA_LEFT)),
                Paragraph(
                    f'<font color="#{hex_val}"><b>{value}</b></font>',
                    ps(f'SV_{label[:4]}', fontSize=10,
                       alignment=TA_RIGHT, fontName='Helvetica-Bold')
                ),
            ]

        stats_left = [
            stat_row('Total employes',           str(total)),
            stat_row('Employes actifs',          str(actifs),   C_GREEN),
            stat_row('Employes archives',        str(archives), C_RED),
            stat_row('Total acces enregistres',  str(total_acces)),
            stat_row('Taux de succes global',    f'{taux_global}%',
                     C_GREEN if taux_global >= 80 else C_AMBER if taux_global >= 50 else C_RED),
        ]

        niveau_rows = [[
            Paragraph('Repartition par niveau',
                      ps('NT', fontSize=9, fontName='Helvetica-Bold',
                         textColor=C_PURPLE)),
            Paragraph('', ps('NV', fontSize=9)),
        ]]
        for n, cnt in sorted(niveaux.items()):
            pct = round(cnt / total * 100) if total > 0 else 0
            niveau_rows.append([
                Paragraph(n, ps(f'NL_{n[:4]}', fontSize=9, textColor=C_GREY_TEXT)),
                Paragraph(
                    f'<b>{cnt}</b> <font color="#94a3b8">({pct}%)</font>',
                    ps(f'NR_{n[:4]}', fontSize=9, alignment=TA_RIGHT)
                ),
            ])

        tbl_stats_left  = Table(stats_left,  colWidths=[5*cm, 3*cm])
        tbl_stats_right = Table(niveau_rows, colWidths=[4.5*cm, 2.5*cm])

        common_stat_style = TableStyle([
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [C_WHITE, C_GREY_LIGHT]),
            ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
            ('TOPPADDING',     (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING',  (0, 0), (-1, -1), 7),
            ('LEFTPADDING',    (0, 0), (-1, -1), 10),
            ('RIGHTPADDING',   (0, 0), (-1, -1), 10),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX',            (0, 0), (-1, -1), 1, C_GREY_MID),
        ])
        tbl_stats_left.setStyle(common_stat_style)
        tbl_stats_right.setStyle(common_stat_style)

        stats_container = Table(
            [[tbl_stats_left, Spacer(0.5*cm, 1), tbl_stats_right]],
            colWidths=[8*cm, 0.5*cm, 7.3*cm]
        )
        stats_container.setStyle(TableStyle([
            ('VALIGN',       (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING',  (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(stats_container)

        # ════════════════════════════════════════════
        #  PIED DE PAGE DOCUMENT
        # ════════════════════════════════════════════
        elements.append(Spacer(1, 1.2*cm))
        elements.append(HRFlowable(width='100%', thickness=1,
                                   color=C_GREY_MID, spaceAfter=6))
        elements.append(Paragraph(
            f"Document genere automatiquement le <b>{now_str}</b> par le systeme SIGR-CA.",
            foot_style,
        ))
        elements.append(Paragraph(
            "Ce document est <b>confidentiel</b> et destine a un usage interne uniquement. "
            "Toute reproduction ou diffusion non autorisee est interdite.",
            ps('CONF', fontSize=7, textColor=C_GREY_TEXT,
               alignment=TA_CENTER, leading=10),
        ))

        # ── Build PDF ────────────────────────────────────────────
        doc.build(elements, canvasmaker=NumberedCanvas)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'inline; filename="Rapport_Employes_{date_file}.pdf"'
        )
        return response

    except Exception as e:
        return HttpResponse(
            f"Erreur PDF : {str(e)}\n\n{traceback.format_exc()}",
            content_type='text/plain',
            status=500
        )

# ====================== HISTORIQUE ======================

@session_required
def historique(request):
    logs = list(db.acces_logs.find().sort('timestamp', -1).limit(500))
    for log in logs:
        b = db.bureaux.find_one({'_id': log.get('bureau_id')})
        log['bureau_nom'] = b['nom'] if b else 'Inconnu'
        e = db.employees.find_one({'_id': log.get('utilisateur_id')})
        log['nom_utilisateur'] = f"{e.get('nom','')} {e.get('prenom','')}" if e else 'Inconnu'
        log['emp_statut'] = e.get('statut', 'actif') if e else 'inconnu'
        log['badge_id'] = e.get('badge_id') if e else ''
    
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    return render(request, 'dashboard/historique.html', {
        'logs': logs,
        'total_acces': db.acces_logs.count_documents({}),
        'acces_autorises': db.acces_logs.count_documents({'resultat': 'AUTORISE'}),
        'acces_refuses': db.acces_logs.count_documents({'resultat': 'REFUSE'}),
        'acces_aujourdhui': db.acces_logs.count_documents({'timestamp': {'$gte': today_start}}),
    })


# ====================== LIVE / SUPERVISION ======================
@session_required
def live(request):
    """Surveillance live - Dashboard temps réel"""
    from datetime import datetime, timedelta

    one_hour_ago    = datetime.now() - timedelta(hours=1)
    today_start     = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_start = (datetime.now() - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)

    acces_ok_hour = db.acces_logs.count_documents({'resultat': 'AUTORISE', 'timestamp': {'$gte': one_hour_ago}})
    acces_no_hour = db.acces_logs.count_documents({'resultat': 'REFUSE',   'timestamp': {'$gte': one_hour_ago}})
    total_acces_hour = acces_ok_hour + acces_no_hour

    acces_ok_today = db.acces_logs.count_documents({'resultat': 'AUTORISE', 'timestamp': {'$gte': today_start}})
    acces_no_today = db.acces_logs.count_documents({'resultat': 'REFUSE',   'timestamp': {'$gte': today_start}})
    total_today = acces_ok_today + acces_no_today

    if total_acces_hour > 0:
        taux_succes_hour = round(acces_ok_hour / total_acces_hour * 100, 1)
    elif total_today > 0:
        taux_succes_hour = round(acces_ok_today / total_today * 100, 1)
    else:
        taux_succes_hour = 0

    acces_ok_yesterday = db.acces_logs.count_documents({
        'resultat': 'AUTORISE',
        'timestamp': {'$gte': yesterday_start, '$lt': today_start}
    })
    delta_ok = round(
        (acces_ok_today - acces_ok_yesterday) / acces_ok_yesterday * 100, 1
    ) if acces_ok_yesterday > 0 else 0

    alertes = 0
    alertes_list = []
    if 'alertes' in db.list_collection_names():
        alertes = db.alertes.count_documents({'statut': 'NON_TRAITEE'})
        alertes_list = list(db.alertes.find({'statut': 'NON_TRAITEE'}).sort('timestamp', -1).limit(10))
        for a in alertes_list:
            a['id'] = str(a['_id'])

    TYPES_SYSTEME = {'SYSTEM', 'URGENCE', 'ADMIN', 'EMERGENCY', 'LOCK', 'UNLOCK'}

    derniers_logs = list(db.acces_logs.find().sort('timestamp', -1).limit(30))
    for log in derniers_logs:
        b = db.bureaux.find_one({'_id': log.get('bureau_id')})
        log['bureau_nom'] = b['nom'] if b else 'Inconnu'
        type_acces = log.get('type_acces', 'RFID')
        log['type_acces'] = type_acces
        if type_acces in TYPES_SYSTEME or log.get('utilisateur_id') is None:
            log['nom_utilisateur'] = 'Système'
            log['badge_id'] = 'SYS'
        else:
            e = db.employees.find_one({'_id': log.get('utilisateur_id')})
            if e:
                nom = f"{e.get('nom', '')} {e.get('prenom', '')}".strip()
                log['nom_utilisateur'] = nom if nom else 'Inconnu'
                log['badge_id'] = e.get('badge_id', '???')
            else:
                log['nom_utilisateur'] = 'Inconnu'
                log['badge_id'] = '???'
        log['resultat'] = log.get('resultat', 'REFUSE')

    bureaux = list(db.bureaux.find())
    total_employes = db.employees.count_documents({'statut': 'actif'})
    for b in bureaux:
        b['id'] = str(b['_id'])
        b['capacite_max'] = b.get('capacite_max', 10)
        recent_logs = db.acces_logs.count_documents({'bureau_id': b['_id'], 'timestamp': {'$gte': one_hour_ago}})
        b['occupation'] = min(recent_logs, b['capacite_max'])
        b['taux'] = round(b['occupation'] / b['capacite_max'] * 100, 1) if b['capacite_max'] > 0 else 0

    equipements = []
    if 'equipements' in db.list_collection_names():
        equipements = list(db.equipements.find().limit(10))
        for eq in equipements:
            eq['id'] = str(eq['_id'])

    return render(request, 'dashboard/live.html', {
        'acces_ok':      acces_ok_today,
        'acces_no':      acces_no_today,
        'total_bureaux': len(bureaux),
        'alertes':       alertes,
        'alertes_list':  alertes_list,
        'derniers_logs': derniers_logs,
        'bureaux':       bureaux,
        'equipements':   equipements,
        'total_employes': total_employes,
        'taux_succes':   taux_succes_hour,
        'delta_ok':      delta_ok,
        'taux_label':    'Dernière heure' if total_acces_hour > 0 else "Aujourd'hui",
    })
def _creer_alerte(message, zone='Système', niveau='MEDIUM'):
    """Insère une alerte dans la collection alertes."""
    try:
        if 'alertes' not in db.list_collection_names():
            db.create_collection('alertes')
        db.alertes.insert_one({
            'message':   message,
            'zone':      zone,
            'niveau':    niveau,   # LOW / MEDIUM / HIGH / CRITICAL
            'statut':    'NON_TRAITEE',
            'timestamp': datetime.now(),
        })
    except Exception as e:
        logger.error(f"[ALERTE] Erreur création alerte: {e}")
@require_http_methods(["GET"])
def api_debug_time(request):
    from datetime import datetime
    from django.http import JsonResponse
    
    # Heure Django
    local = datetime.now()
    utc   = datetime.utcnow()
    
    # Heure stockée dans MongoDB (prend le dernier log)
    dernier_log = db.acces_logs.find_one(
        {}, sort=[('timestamp', -1)]
    )
    mongo_time = dernier_log.get('timestamp') if dernier_log else None
    
    return JsonResponse({
        'django_local':     local.strftime('%Y-%m-%d %H:%M:%S'),
        'django_utc':       utc.strftime('%Y-%m-%d %H:%M:%S'),
        'mongodb_last_log': str(mongo_time),
        'difference_minutes': round((local - utc).total_seconds() / 60, 1),
    })
import json
import logging
from datetime import datetime, timedelta
from bson import ObjectId
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.conf import settings
 
# Import de ta connexion MongoDB (adapter selon ton projet)
# from .db import db   ← décommente et adapte
 
logger = logging.getLogger(__name__)
 
# Clé secrète partagée avec l'ESP32 (à mettre dans settings.py)
# ESP32_API_KEY = "CHANGE_MOI_SECRET_KEY_123"
 
 
# ─────────────────────────────────────────────────────────────────────
#  HELPER : Vérification clé ESP32
# ─────────────────────────────────────────────────────────────────────
 
def _check_esp32_key(request):
    """Vérifie le header X-ESP32-Key envoyé par l'ESP32."""
    key = request.headers.get('X-ESP32-Key', '')
    expected = getattr(settings, 'ESP32_API_KEY', 'CHANGE_MOI_SECRET_KEY_123')
    return key == expected
 
 
# ─────────────────────────────────────────────────────────────────────
#  ENDPOINT PRINCIPAL : Réception d'un scan RFID depuis l'ESP32
# ─────────────────────────────────────────────────────────────────────
@csrf_exempt
@require_http_methods(["POST"])
def api_rfid_scan(request):
    
    """
    Reçoit un scan RFID depuis l'ESP32 et enregistre l'accès.

    Body JSON attendu :
    {
        "uid":             "D7:77:C5:01",
        "equipement_code": "RDR-001",
        "type_acces":      "RFID"
    }

    Réponse JSON :
    {
        "resultat": "AUTORISE" | "REFUSE",
        "message":  "...",
        "nom":      "Nom Prenom"  (si trouvé)
    }
    """

    # ── Vérification clé sécurité ──
    if not _check_esp32_key(request):
        logger.warning(f"[RFID] Clé API invalide depuis {request.META.get('REMOTE_ADDR')}")
        return JsonResponse({'resultat': 'REFUSE', 'message': 'Clé API invalide'}, status=403)

    # ── Parse JSON ──
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'resultat': 'REFUSE', 'message': 'JSON invalide'}, status=400)

    uid             = data.get('uid', '').strip().upper()
    equipement_code = data.get('equipement_code', '').strip()
    type_acces      = data.get('type_acces', 'RFID')

    if not uid or not equipement_code:
        return JsonResponse({'resultat': 'REFUSE', 'message': 'UID ou code équipement manquant'}, status=400)

    logger.info(f"[RFID] Scan reçu — UID: {uid} | Équipement: {equipement_code}")

    # ── Vérification équipement ──
    equipement = db.equipements.find_one({'code': equipement_code, 'statut': 'actif'})
    if not equipement:
        logger.warning(f"[RFID] Équipement inconnu ou inactif: {equipement_code}")
        return JsonResponse({'resultat': 'REFUSE', 'message': f'Équipement {equipement_code} non trouvé ou inactif'})

    bureau_id  = equipement.get('bureau_id')
    bureau     = db.bureaux.find_one({'_id': bureau_id}) if bureau_id else None
    bureau_nom = bureau['nom'] if bureau else 'Zone inconnue'

    # ── Recherche employé par badge_id (UID RFID) ──
    uid_normalized = uid.replace(':', '').replace(' ', '')

    employe = db.employees.find_one({
        '$or': [
            {'badge_id': uid},
            {'badge_id': uid_normalized},
            {'badge_id': uid.lower()},
        ],
        'statut': 'actif'
    })

    from datetime import timezone as tz
    maintenant    = datetime.now()
    timestamp     = maintenant

    if employe:
        nom_complet = f"{employe.get('nom', '')} {employe.get('prenom', '')}".strip()

        # ── Vérification réservation pour cette salle ──
        if bureau_id:
            employe_oid    = employe['_id']
            from datetime import timedelta as td
            fenetre_retard = td(minutes=15)
            

            # Chercher les réservations d'aujourd'hui pour cette salle
            # ✅ Après — utiliser datetime.now() comme MongoDB
            maintenant    = datetime.now()
            timestamp     = maintenant
            debut_journee = maintenant.replace(hour=0,  minute=0,  second=0,  microsecond=0)
            fin_journee   = maintenant.replace(hour=23, minute=59, second=59, microsecond=0)
            resa_aujourd_hui = db.reservations.find_one(
                {
                    'employe_id': employe_oid,
                    'bureau_id':  bureau_id,
                    'statut':     'confirmee',
                    'date_debut': {'$gte': debut_journee, '$lte': fin_journee},
                },
                sort=[('date_debut', 1)]
            )

            logger.info(f"[RFID] Réservation aujourd'hui — {resa_aujourd_hui}")

            if not resa_aujourd_hui:
                # Aucune réservation aujourd'hui pour cette salle
                resultat = 'REFUSE'
                message  = f"Aucune réservation confirmée pour la salle {bureau_nom} aujourd'hui."

            else:
                debut = resa_aujourd_hui['date_debut']
                fin   = resa_aujourd_hui['date_fin']
                limite_retard = debut + fenetre_retard  # debut + 15 min

                if maintenant < debut:
                    # ── Trop tôt ──
                    early_minutes = int((debut - maintenant).total_seconds() / 60)
                    resultat = 'REFUSE'
                    message  = f"Trop tôt ! Votre réservation commence à {debut.strftime('%H:%M')} (dans {early_minutes} min)."

                elif maintenant <= limite_retard:
                    # ── Dans la fenêtre autorisée : entre debut et debut+15min ──
                    resultat = 'AUTORISE'
                    message  = f"Bienvenue, {nom_complet} !"
                    if maintenant > debut:
                        retard_min = int((maintenant - debut).total_seconds() / 60)
                        message = f"Bienvenue, {nom_complet} ! ({retard_min} min de retard)"

                else:
                    # ── Plus de 15 min de retard → réservation ratée ──
                    resultat = 'REFUSE'
                    message  = (
                        f"Accès refusé : vous avez dépassé les 15 minutes de tolérance. "
                        f"Réservation à {debut.strftime('%H:%M')} — créneau expiré."
                    )

            # ── Log & retour si REFUSÉ ──
            if resultat == 'REFUSE':
                logger.warning(f"[RFID] REFUSÉ (réservation) — {nom_complet} → {bureau_nom} : {message}")

                db.acces_logs.insert_one({
                    'utilisateur_id':  employe['_id'],
                    'badge_id':        uid,
                    'equipement_code': equipement_code,
                    'equipement_id':   equipement['_id'],
                    'bureau_id':       bureau_id,
                    'type_acces':      type_acces,
                    'resultat':        resultat,
                    'message':         message,
                    'timestamp':       timestamp,
                    'ip_source':       request.META.get('REMOTE_ADDR', ''),
                })

                _creer_alerte(
                    message=f"Accès refusé (réservation) — {nom_complet} | {bureau_nom} : {message}",
                    zone=bureau_nom,
                    niveau='MEDIUM'
                )

                return JsonResponse({
                    'resultat': resultat,
                    'message':  message,
                    'nom':      nom_complet,
                    'bureau':   bureau_nom,
                })

        # ── ACCÈS AUTORISÉ ──
        resultat = resultat if 'resultat' in dir() else 'AUTORISE'
        message  = message  if 'message'  in dir() else f'Bienvenue, {nom_complet}'

        logger.info(f"[RFID] AUTORISÉ — {nom_complet} ({uid}) → {bureau_nom}")

        db.equipements.update_one(
            {'_id': equipement['_id']},
            {'$set': {'derniere_connexion': timestamp}}
        )

        db.acces_logs.insert_one({
            'utilisateur_id':  employe['_id'],
            'badge_id':        uid,
            'equipement_code': equipement_code,
            'equipement_id':   equipement['_id'],
            'bureau_id':       bureau_id,
            'type_acces':      type_acces,
            'resultat':        'AUTORISE',
            'message':         message,
            'timestamp':       timestamp,
            'ip_source':       request.META.get('REMOTE_ADDR', ''),
        })

        return JsonResponse({
            'resultat': 'AUTORISE',
            'message':  message,
            'nom':      nom_complet,
            'bureau':   bureau_nom,
        })

    else:
        # ── ACCÈS REFUSÉ (badge inconnu) ──
        resultat = 'REFUSE'
        message  = f'Badge non reconnu ({uid})'

        logger.warning(f"[RFID] REFUSÉ — UID inconnu: {uid} sur {equipement_code}")

        db.acces_logs.insert_one({
            'utilisateur_id':  None,
            'badge_id':        uid,
            'equipement_code': equipement_code,
            'equipement_id':   equipement['_id'],
            'bureau_id':       bureau_id,
            'type_acces':      type_acces,
            'resultat':        resultat,
            'timestamp':       timestamp,
            'ip_source':       request.META.get('REMOTE_ADDR', ''),
        })

        _creer_alerte(
            message=f"Tentative accès non autorisée — Badge: {uid} sur {equipement_code} ({bureau_nom})",
            zone=bureau_nom,
            niveau='MEDIUM'
        )

        return JsonResponse({
            'resultat': resultat,
            'message':  message,
            'nom':      'Inconnu',
            'bureau':   bureau_nom,
        })
# ====================== GESTION DES RESSOURCES (ZONES & MATÉRIEL) ======================

@session_required
def ressources(request):
    """Gestion des ressources - Zones et matériel"""
    from datetime import datetime, timedelta
    import json

    # ── Bureaux / Zones ──────────────────────────────────────────────────────
    bureaux = list(db.bureaux.find())
    for b in bureaux:
        b['id'] = str(b['_id'])
        b['capacite_max'] = b.get('capacite_max', 10)
        one_hour_ago = datetime.now() - timedelta(hours=1)
        recent = db.acces_logs.count_documents({
            'bureau_id': b['_id'],
            'timestamp': {'$gte': one_hour_ago}
        })
        b['occupation'] = min(recent, b['capacite_max'])
        b['taux_occupation'] = round((b['occupation'] / b['capacite_max'] * 100), 1) if b['capacite_max'] > 0 else 0

    total_bureaux   = len(bureaux)
    zones_actives   = sum(1 for b in bureaux if b.get('statut', 'actif') == 'actif')
    capacite_totale = sum(b.get('capacite_max', 0) for b in bureaux)
    total_occ       = sum(min(b.get('occupation', 0), b.get('capacite_max', 10)) for b in bureaux)
    total_cap       = sum(b.get('capacite_max', 10) for b in bureaux)
    occupation_moy  = round((total_occ / total_cap * 100), 1) if total_cap > 0 else 0

    # ── Matériel ─────────────────────────────────────────────────────────────
    if 'materiels' not in db.list_collection_names():
        db.create_collection('materiels')

    materiels = list(db.materiels.find())
    for m in materiels:
        m['id'] = str(m['_id'])

    total_materiel       = len(materiels)
    materiel_disponible  = sum(1 for m in materiels if m.get('statut') == 'disponible')
    materiel_maintenance = sum(1 for m in materiels if m.get('statut') in ['maintenance', 'hors_service'])

    materiels_json = json.dumps([{
        'id':           str(m['_id']),
        'nom':          m.get('nom', ''),
        'categorie':    m.get('categorie', ''),
        'fournisseur':  m.get('fournisseur', ''),
        'annee_inv':    m.get('annee_inv', ''),
        'seq_inv':      m.get('seq_inv', ''),
        'numero_serie': m.get('numero_serie', ''),
        'num_inventaire': m.get('num_inventaire', ''),
        'statut':       m.get('statut', 'disponible'),
        'zone':         m.get('zone', ''),
        'description':  m.get('description', ''),
        'photo':        m.get('photo', ''),
        'marque':       m.get('marque', ''),
        'modele':       m.get('modele', ''),
        'date_achat':   m.get('date_achat', ''),
        'valeur':       m.get('valeur', ''),
        'processeur':   m.get('processeur', ''),
        'ram':          m.get('ram', ''),
        'stockage':     m.get('stockage', ''),
        'os':           m.get('os', ''),
        'ecran':        m.get('ecran', ''),
    } for m in materiels], default=str)
    import json as _json
    bureaux_json = _json.dumps([{
        'id':              str(b['_id']),
        'nom':             b.get('nom', ''),
        'code_bureau':     b.get('code_bureau', ''),
        'etage':           b.get('etage', 0),
        'capacite_max':    b.get('capacite_max', 10),
        'niveau_securite': b.get('niveau_securite', 'standard'),
        'description':     b.get('description', ''),
        'statut':          b.get('statut', 'actif'),
        'taux_occupation': b.get('taux_occupation', 0),
    } for b in bureaux], default=str)
    params_resa = db.parametres.find_one({'cle': 'reservation'}) or {}
    return render(request, 'dashboard/ressources.html', {
        'bureaux':             bureaux,
        'bureaux_json':        bureaux_json,
        'total_bureaux':       total_bureaux,
        'zones_actives':       zones_actives,
        'capacite_totale':     capacite_totale,
        'occupation_moy':      occupation_moy,
        'materiels':           materiels,
        'total_materiel':      total_materiel,
        'materiel_disponible': materiel_disponible,
        'materiel_maintenance':materiel_maintenance,
        'materiels_json':      materiels_json,
        'params_resa': params_resa,
    })


# ─── Génération du numéro d'inventaire ────────────────────────────────────────

def _generer_num_inventaire(categorie, fournisseur='', annee_inv=None, seq_inv=None):
    """
    Génère un numéro d'inventaire :
      [FOURNISSEUR-]CAT-AAAA-NNNNN
    Exemple : DELL-INF-2026-00001
    """
    from datetime import datetime

    prefixes = {
        'informatique': 'INF',
        'mobilier':     'MOB',
        'audiovisuel':  'AUD',
        'imprimante':   'IMP',
        'securite':     'SEC',
        'vehicule':     'VEH',
        'laboratoire':  'LAB',
        'autre':        'MAT',
    }
    prefix = prefixes.get(categorie, 'MAT')
    annee  = int(annee_inv) if annee_inv else datetime.now().year

    # Numéro séquentiel : fourni par l'utilisateur ou auto-incrémenté
    if seq_inv:
        seq = int(seq_inv)
    else:
        # Chercher le dernier numéro pour ce préfixe + année
        fournisseur_clean = (fournisseur or '').strip().upper().replace(' ', '-')[:10]
        pattern_parts = [fournisseur_clean, prefix, str(annee)] if fournisseur_clean else [prefix, str(annee)]
        pattern = '-'.join(pattern_parts) + '-'
        dernier = db.materiels.find_one(
            {'num_inventaire': {'$regex': f'^{pattern}'}},
            sort=[('num_inventaire', -1)]
        )
        if dernier and dernier.get('num_inventaire'):
            try:
                seq = int(dernier['num_inventaire'].split('-')[-1]) + 1
            except Exception:
                seq = 1
        else:
            seq = 1

    fournisseur_clean = (fournisseur or '').strip().upper().replace(' ', '-')[:10]
    if fournisseur_clean:
        return f"{fournisseur_clean}-{prefix}-{annee}-{seq:05d}"
    return f"{prefix}-{annee}-{seq:05d}"


# ─── CRUD Zones ───────────────────────────────────────────────────────────────

@session_required
def bureau_ajouter(request):
    from bson import ObjectId
    from datetime import datetime

    if request.method == 'POST':
        try:
            bureau_id = request.POST.get('bureau_id')
            data = {
                'nom':             request.POST.get('nom'),
                'code_bureau':     request.POST.get('code_bureau', ''),
                'etage':           int(request.POST.get('etage', 0)),
                'capacite_max':    int(request.POST.get('capacite_max', 10)),
                'niveau_securite': request.POST.get('niveau_securite', 'standard'),
                'description':     request.POST.get('description', ''),
                'statut':          request.POST.get('statut', 'actif'),
                'updated_at':      datetime.now(),
            }
            if bureau_id:
                db.bureaux.update_one({'_id': ObjectId(bureau_id)}, {'$set': data})
                messages.success(request, f"Zone '{data['nom']}' modifiée avec succès !")
            else:
                data['created_at'] = datetime.now()
                db.bureaux.insert_one(data)
                messages.success(request, f"Zone '{data['nom']}' ajoutée avec succès !")
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")

    return redirect('ressources')


@session_required
def bureau_supprimer(request, bureau_id):
    from bson import ObjectId

    if request.method == 'POST':
        try:
            result = db.bureaux.delete_one({'_id': ObjectId(bureau_id)})
            if result.deleted_count > 0:
                messages.success(request, "Zone supprimée avec succès !")
            else:
                messages.error(request, "Zone non trouvée")
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
    return redirect('ressources')


# ─── API CRUD Matériel ─────────────────────────────────────────────────────────

@session_required
def api_materiel_list(request):
    materiels = list(db.materiels.find()) if 'materiels' in db.list_collection_names() else []
    result = [{
        'id':             str(m['_id']),
        'nom':            m.get('nom', ''),
        'categorie':      m.get('categorie', ''),
        'num_inventaire': m.get('num_inventaire', ''),
        'fournisseur':    m.get('fournisseur', ''),
        'numero_serie':   m.get('numero_serie', ''),
        'statut':         m.get('statut', 'disponible'),
        'zone':           m.get('zone', ''),
        'marque':         m.get('marque', ''),
        'modele':         m.get('modele', ''),
    } for m in materiels]
    return JsonResponse({'materiels': result})

@session_required
def parametres_reservation_save(request):
    """Sauvegarde les durées max/min de réservation (admin seulement)."""
    if not request.session.get('is_staff'):
        return JsonResponse({'status': 'error', 'message': 'Accès refusé'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            db.parametres.update_one(
                {'cle': 'reservation'},
                {'$set': {
                    'cle': 'reservation',
                    'salle_max_minutes':     int(data.get('salle_max_minutes',     1440)),
                    'salle_min_minutes':     int(data.get('salle_min_minutes',       30)),
                    'materiel_max_minutes':  int(data.get('materiel_max_minutes', 525600)),
                    'materiel_min_minutes':  int(data.get('materiel_min_minutes',     60)),
                    'updated_at': datetime.now(),
                    'updated_by': request.session.get('username', ''),
                }},
                upsert=True
            )
            return JsonResponse({'status': 'success', 'message': 'Paramètres sauvegardés.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode invalide'}, status=405)
@session_required
@require_http_methods(["POST"])
def api_materiel_ajouter(request):
    from bson import ObjectId
    from datetime import datetime

    try:
        data        = json.loads(request.body)
        materiel_id = data.get('id', '').strip()
        categorie   = data.get('categorie', 'autre')
        fournisseur = data.get('fournisseur', '').strip().upper()
        annee_inv   = data.get('annee_inv', '')
        seq_inv     = data.get('seq_inv', '')

        if 'materiels' not in db.list_collection_names():
            db.create_collection('materiels')

        is_edit = materiel_id and len(materiel_id) == 24
        if is_edit:
            existing       = db.materiels.find_one({'_id': ObjectId(materiel_id)})
            num_inventaire = existing.get('num_inventaire', '') if existing else ''
        else:
            num_inventaire = _generer_num_inventaire(categorie, fournisseur, annee_inv, seq_inv)

        materiel_data = {
            'nom':          data.get('nom', '').strip(),
            'categorie':    categorie,
            'fournisseur':  fournisseur,
            'annee_inv':    annee_inv,
            'seq_inv':      seq_inv,
            'numero_serie': data.get('numero_serie', '').strip(),
            'statut':       data.get('statut', 'disponible'),
            'zone':         data.get('zone', ''),
            'description':  data.get('description', '').strip(),
            'photo':        data.get('photo', '').strip(),
            'marque':       data.get('marque', '').strip(),
            'modele':       data.get('modele', '').strip(),
            'date_achat':   data.get('date_achat', '').strip(),
            'valeur':       data.get('valeur', '').strip(),
            # Spécifications techniques machine
            'processeur':   data.get('processeur', '').strip(),
            'ram':          data.get('ram', '').strip(),
            'stockage':     data.get('stockage', '').strip(),
            'os':           data.get('os', '').strip(),
            'ecran':        data.get('ecran', '').strip(),
            'num_inventaire': num_inventaire,
            'updated_at':   datetime.now(),
        }

        if not materiel_data['nom']:
            return JsonResponse({'status': 'error', 'message': 'Le nom est obligatoire'}, status=400)

        if is_edit:
            db.materiels.update_one({'_id': ObjectId(materiel_id)}, {'$set': materiel_data})
            return JsonResponse({'status': 'success', 'message': 'Matériel modifié', 'id': materiel_id, 'num_inventaire': num_inventaire})
        else:
            materiel_data['created_at'] = datetime.now()
            materiel_data['created_by'] = request.session.get('username', '')
            result = db.materiels.insert_one(materiel_data)
            return JsonResponse({'status': 'success', 'message': f'Matériel ajouté', 'id': str(result.inserted_id), 'num_inventaire': num_inventaire})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_materiel_supprimer(request, materiel_id):
    from bson import ObjectId

    if request.method not in ('DELETE', 'POST'):
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    try:
        result = db.materiels.delete_one({'_id': ObjectId(materiel_id)})
        if result.deleted_count > 0:
            return JsonResponse({'status': 'success', 'message': 'Matériel supprimé'})
        return JsonResponse({'status': 'error', 'message': 'Matériel non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# ─── RAPPORT PDF — Un seul matériel ───────────────────────────────────────────


# ─── FICHE D'INVENTAIRE — 1 seul matériel ────────────────────────────────────
@session_required
def api_materiel_pdf(request, materiel_id):
    """Génère la fiche d'inventaire PDF complète pour un matériel — SIGR-CA."""
    try:
        m = db.materiels.find_one({'_id': ObjectId(materiel_id)})
        if not m:
            return HttpResponse('Matériel non trouvé', status=404)

        now_str   = datetime.now().strftime('%d/%m/%Y à %H:%M')
        date_file = datetime.now().strftime('%Y%m%d_%H%M')

        # ── Mise en page ──────────────────────────────────────────
        buffer   = io.BytesIO()
        PAGE_W, _ = A4
        CONTENT_W = PAGE_W - 4 * cm

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2.2*cm, bottomMargin=2.2*cm,
            title=f"Fiche Inventaire — {m.get('nom', 'Matériel')}",
            author="SIGR-CA Système",
            subject="Fiche d'inventaire matériel",
        )

        # ── Registre de styles custom ─────────────────────────────
        _custom_styles = {}

        def ps(name, **kw):
            base = kw.pop('parent', 'Normal')
            stylesheet = getSampleStyleSheet()
            parent_style = _custom_styles.get(base) or stylesheet[base]
            p = ParagraphStyle(name, parent=parent_style, **kw)
            _custom_styles[name] = p
            return p

        # Styles
        th        = ps('TH',   fontName='Helvetica-Bold', fontSize=9,
                        textColor=C_WHITE, alignment=TA_CENTER,
                        leading=12, spaceAfter=0, spaceBefore=0)
        th_left   = ps('TH_L', parent='TH', alignment=TA_LEFT)
        td        = ps('TD',   fontSize=9, alignment=TA_CENTER,
                        leading=12, textColor=C_BLACK)
        td_left   = ps('TD_L', parent='TD', alignment=TA_LEFT)
        td_mono   = ps('TD_M', parent='TD', fontSize=8, fontName='Courier',
                        alignment=TA_CENTER)
        sec_style = ps('SEC',  fontName='Helvetica-Bold', fontSize=12,
                        textColor=C_PURPLE, spaceBefore=20, spaceAfter=8,
                        leading=16)
        foot_style = ps('FOOT', fontSize=7.5, textColor=C_GREY_TEXT,
                         alignment=TA_CENTER, leading=11)
        label_style = ps('LABEL', fontSize=9, textColor=C_GREY_TEXT,
                          fontName='Helvetica-Bold', alignment=TA_LEFT)
        val_style = ps('VAL', fontSize=10, fontName='Helvetica',
                        textColor=C_BLACK, alignment=TA_LEFT)

        # ✅ Fonction info_row DÉFINIE ICI (à l'intérieur de la fonction)
        def info_row(label, value):
            if not value and value != 0:
                return [Paragraph(label, label_style), Paragraph('—', val_style)]
            return [Paragraph(label, label_style), Paragraph(str(value), val_style)]
        
        # ✅ Fonction pour formater les valeurs numériques
        def format_valeur(value):
            """Formate une valeur avec séparateur de milliers"""
            if value is None or value == '':
                return ''
            try:
                if isinstance(value, str):
                    value = value.strip().replace(' ', '').replace('\u202f', '')
                    if value == '':
                        return ''
                    value = float(value) if '.' in value else int(value)
                return f"{int(value):,}"
            except (ValueError, TypeError):
                return str(value)

        elements = []

        # ════════════════════════════════════════════
        #  EN-TÊTE
        # ════════════════════════════════════════════
        from django.contrib.staticfiles.finders import find as static_find

        LOGO_PATH = static_find('img/logo.png')

        title_lines = [
            [Paragraph(
                '<font color="#1d4ed8"><b>SIGR-CA</b></font>',
                ps('LT', fontSize=22, leading=26, alignment=TA_LEFT)
            )],
            [Paragraph(
                "Système Intégré de Gestion des Ressources<br/>"
                "<font color='#64748b'>et de Contrôle d'Accès</font>",
                ps('LS', fontSize=9, leading=13,
                   textColor=C_GREY_TEXT, alignment=TA_LEFT)
            )],
        ]
        title_tbl = Table(title_lines, colWidths=[11*cm])
        title_tbl.setStyle(TableStyle([
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ]))

        if LOGO_PATH and os.path.exists(LOGO_PATH):
            logo_cell = Image(LOGO_PATH, width=4*cm, height=2.6*cm)
        else:
            logo_cell = Paragraph(
                '<font color="#1d4ed8"><b>SIGR</b></font>',
                ps('FL', fontSize=18, alignment=TA_RIGHT)
            )

        meta_lines = [
            [Paragraph(f"<b>Date :</b> {now_str}",
                       ps('M1', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph("<b>Fiche d'inventaire</b>",
                       ps('M2', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph("<b>Confidentiel</b>",
                       ps('M3', fontSize=8, textColor=C_RED, alignment=TA_RIGHT))],
        ]
        meta_tbl = Table(meta_lines, colWidths=[4.5*cm])
        meta_tbl.setStyle(TableStyle([
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))

        right_col = Table([[logo_cell], [meta_tbl]], colWidths=[4.5*cm])
        right_col.setStyle(TableStyle([
            ('ALIGN',         (0,0), (-1,-1), 'RIGHT'),
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))

        header_tbl = Table(
            [[title_tbl, right_col]],
            colWidths=[CONTENT_W - 4.5*cm, 4.5*cm]
        )
        header_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))

        elements.append(header_tbl)
        elements.append(HRFlowable(width='100%', thickness=2,
                                   color=C_BLUE_MAIN, spaceAfter=4))
        elements.append(Spacer(1, 0.3*cm))

        # ── Bandeau titre ─────────────────────────────────────────
        inv = m.get('num_inventaire', 'N/A')
        elements.append(ColorBand(
            f"  FICHE D'INVENTAIRE — N° {inv}",
            CONTENT_W, height=1.1*cm,
            bg=C_BLUE_DARK, fg=C_WHITE, font_size=12
        ))
        elements.append(Spacer(1, 0.6*cm))

        # ════════════════════════════════════════════
        #  INFORMATIONS GÉNÉRALES
        # ════════════════════════════════════════════
        elements.append(Paragraph('Informations générales', sec_style))

        # Photo
        photo_path = m.get('photo', '')
        photo_img = None
        if photo_path:
            full_photo = os.path.join(settings.MEDIA_ROOT, photo_path)
            if os.path.exists(full_photo):
                try:
                    from reportlab.platypus import Image as RLImage
                    photo_img = RLImage(full_photo, width=5*cm, height=5*cm)
                except:
                    pass

        # ✅ Données générales (corrigé avec format_valeur)
        gen_rows = [
            info_row('Nom',                    m.get('nom', '')),
            info_row('Catégorie',              m.get('categorie', '').title() if m.get('categorie') else ''),
            info_row('Sous-catégorie',         m.get('sous_categorie', '').title() if m.get('sous_categorie') else ''),
            info_row('Marque',                 m.get('marque', '')),
            info_row('Modèle',                 m.get('modele', '')),
            info_row('N° de série',            m.get('numero_serie', '')),
            info_row('N° inventaire',          inv),
            info_row('Fournisseur',            m.get('fournisseur', '')),
            info_row('Date d\'achat',          m.get('date_achat', '')),
            info_row('Date de mise en service', m.get('date_mise_service', '')),
            info_row('Garantie jusqu\'au',     m.get('garantie', '')),
            info_row('Valeur d\'achat (DA)',   f"{format_valeur(m.get('valeur'))} DA" if m.get('valeur') else ''),
            info_row('Valeur résiduelle (DA)', f"{format_valeur(m.get('valeur_residuelle'))} DA" if m.get('valeur_residuelle') else ''),
            info_row('Statut',                 m.get('statut', '').replace('_', ' ').title() if m.get('statut') else ''),
            info_row('État',                   m.get('etat', '').replace('_', ' ').title() if m.get('etat') else ''),
            info_row('Zone assignée',          m.get('zone', '')),
            info_row('Utilisateur assigné',    m.get('utilisateur_nom', '')),
            info_row('Département',            m.get('departement', '')),
            info_row('Code-barres',            m.get('code_barres', '')),
            info_row('RFID/NFC Tag',           m.get('rfid_tag', '')),
        ]

        # Filtrer les lignes vides (corrigé)
        gen_rows = [r for r in gen_rows if len(r) >= 2 and r[1].text not in ('', '—', None)]

        if photo_img:
            t_gen = Table(gen_rows, colWidths=[5*cm, 7*cm])
            t_gen.setStyle(TableStyle([
                ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
                ('BACKGROUND',     (0, 0), (0, -1), C_GREY_LIGHT),
                ('PADDING',        (0, 0), (-1, -1), 7),
                ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
            ]))

            photo_cell = Table([[photo_img]], colWidths=[5.5*cm])
            photo_cell.setStyle(TableStyle([
                ('ALIGN',  (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOX',    (0, 0), (-1, -1), 1, C_GREY_MID),
                ('PADDING', (0, 0), (-1, -1), 5),
            ]))

            side_table = Table(
                [[t_gen, photo_cell]],
                colWidths=[12*cm, 5.5*cm],
            )
            side_table.setStyle(TableStyle([
                ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING',  (1, 0), (1, 0), 10),
                ('RIGHTPADDING', (0, 0), (0, 0), 0),
            ]))
            elements.append(side_table)
        else:
            t_gen = Table(gen_rows, colWidths=[5*cm, 12.5*cm])
            t_gen.setStyle(TableStyle([
                ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
                ('BACKGROUND',     (0, 0), (0, -1), C_GREY_LIGHT),
                ('PADDING',        (0, 0), (-1, -1), 7),
                ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
            ]))
            elements.append(t_gen)

        # ════════════════════════════════════════════
        #  SPÉCIFICATIONS TECHNIQUES
        # ════════════════════════════════════════════
        tech_rows = [
            info_row('Processeur (CPU)',         m.get('processeur', '')),
            info_row('Mémoire RAM',              m.get('ram', '')),
            info_row('Type de RAM',              m.get('type_ram', '')),
            info_row('Stockage',                 m.get('stockage', '')),
            info_row('Type de stockage',         m.get('type_stockage', '')),
            info_row('Carte graphique',          m.get('carte_graphique', '')),
            info_row('Système d\'exploitation',  m.get('os', '')),
            info_row('Version OS',               m.get('version_os', '')),
            info_row('Adresse IP',               m.get('adresse_ip', '')),
            info_row('Adresse MAC',              m.get('adresse_mac', '')),
            info_row('Nom d\'hôte',              m.get('hostname', '')),
            info_row('Écran / Résolution',       m.get('ecran', '')),
            info_row('Ports / Connectique',      m.get('ports', '')),
            info_row('Alimentation',             m.get('alimentation', '')),
            info_row('Poids (kg)',               m.get('poids', '')),
            info_row('Dimensions',               m.get('dimensions', '')),
        ]
        tech_rows = [r for r in tech_rows if len(r) >= 2 and r[1].text not in ('', '—', None)]

        if tech_rows:
            elements.append(Paragraph('Spécifications techniques', sec_style))
            t_tech = Table(tech_rows, colWidths=[5*cm, 12.5*cm])
            t_tech.setStyle(TableStyle([
                ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
                ('BACKGROUND',     (0, 0), (0, -1), C_BLUE_LIGHT),
                ('PADDING',        (0, 0), (-1, -1), 7),
                ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [C_WHITE, C_GREY_LIGHT]),
            ]))
            elements.append(t_tech)

        # ════════════════════════════════════════════
        #  LOGICIELS & LICENCES
        # ════════════════════════════════════════════
        logiciels = m.get('logiciels', [])
        if logiciels:
            elements.append(Paragraph('Logiciels & Licences', sec_style))
            log_headers = [
                Paragraph('Logiciel', th_left),
                Paragraph('Version', th),
                Paragraph('Licence', th),
                Paragraph('Expiration', th),
            ]
            log_data = [log_headers]
            for log in logiciels:
                if isinstance(log, dict):
                    log_data.append([
                        Paragraph(log.get('nom', '—'), td_left),
                        Paragraph(log.get('version', '—'), td),
                        Paragraph(log.get('licence', '—'), td),
                        Paragraph(log.get('expiration', '—'), td),
                    ])
                else:
                    log_data.append([
                        Paragraph(str(log), td_left),
                        Paragraph('—', td),
                        Paragraph('—', td),
                        Paragraph('—', td),
                    ])

            t_log = Table(log_data, colWidths=[6*cm, 3.5*cm, 4*cm, 4*cm], repeatRows=1)
            t_log.setStyle(TableStyle([
                ('BACKGROUND',     (0, 0), (-1, 0), C_BLUE_MAIN),
                ('TEXTCOLOR',      (0, 0), (-1, 0), C_WHITE),
                ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
                ('PADDING',        (0, 0), (-1, -1), 7),
                ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
            ]))
            elements.append(t_log)

        # ════════════════════════════════════════════
        #  MAINTENANCE & INTERVENTIONS
        # ════════════════════════════════════════════
        maintenances = list(db.maintenances.find(
            {'materiel_id': m['_id']}
        ).sort('date', -1).limit(10)) if 'maintenances' in db.list_collection_names() else []

        if maintenances:
            elements.append(Paragraph('Historique de maintenance', sec_style))
            maint_headers = [
                Paragraph('Date', th_left),
                Paragraph('Type', th),
                Paragraph('Intervenant', th),
                Paragraph('Coût (DA)', th),
                Paragraph('Description', th_left),
            ]
            maint_data = [maint_headers]
            for mt in maintenances:
                cout = mt.get('cout', 0)
                cout_str = f"{int(cout):,}" if cout else '—'
                maint_data.append([
                    Paragraph(str(mt.get('date', '—')), td_left),
                    Paragraph(str(mt.get('type', '—')).replace('_', ' ').title(), td),
                    Paragraph(str(mt.get('intervenant', '—')), td),
                    Paragraph(cout_str, td),
                    Paragraph(str(mt.get('description', '—'))[:80], td_left),
                ])

            t_maint = Table(maint_data, colWidths=[2.5*cm, 2.8*cm, 3.5*cm, 2.5*cm, 6.2*cm], repeatRows=1)
            t_maint.setStyle(TableStyle([
                ('BACKGROUND',     (0, 0), (-1, 0), C_BLUE_MAIN),
                ('TEXTCOLOR',      (0, 0), (-1, 0), C_WHITE),
                ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
                ('PADDING',        (0, 0), (-1, -1), 7),
                ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
            ]))
            elements.append(t_maint)
        else:
            elements.append(Paragraph('Maintenance', sec_style))
            elements.append(Paragraph(
                '<i>Aucune intervention de maintenance enregistrée.</i>',
                ps('NOMT', fontSize=9, textColor=C_GREY_TEXT, alignment=TA_LEFT)
            ))

        # ════════════════════════════════════════════
        #  DESCRIPTION / REMARQUES
        # ════════════════════════════════════════════
        if m.get('description'):
            elements.append(Paragraph('Description / Remarques', sec_style))
            elements.append(Paragraph(m['description'], ps('DESC', fontSize=10, textColor=C_BLACK,
                                   alignment=TA_LEFT, leading=14, spaceAfter=6)))

        # ════════════════════════════════════════════
        #  ZONE SIGNATURE
        # ════════════════════════════════════════════
        elements.append(Spacer(1, 0.8*cm))
        elements.append(Paragraph('Validation & Signature', sec_style))

        sig_data = [
            [
                Paragraph('<b>Vérifié par</b>', label_style),
                Paragraph('', val_style),
                Paragraph('<b>Date de vérification</b>', label_style),
                Paragraph('', val_style),
            ],
            [
                Paragraph('<b>Signature</b>', label_style),
                Paragraph('', val_style),
                Paragraph('<b>Visa responsable</b>', label_style),
                Paragraph('', val_style),
            ],
            [
                Paragraph('<b>Commentaire</b>', label_style),
                Paragraph('', val_style),
                Paragraph('', val_style),
                Paragraph('', val_style),
            ],
        ]
        sig_table = Table(sig_data, colWidths=[3.5*cm, 5.5*cm, 3.5*cm, 5*cm])
        sig_table.setStyle(TableStyle([
            ('BOX',            (0, 0), (-1, -1), 1, C_GREY_MID),
            ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
            ('PADDING',        (0, 0), (-1, -1), 10),
            ('BACKGROUND',     (0, 0), (0, -1), C_GREY_LIGHT),
            ('BACKGROUND',     (2, 0), (2, -1), C_GREY_LIGHT),
            ('MINROWHEIGHTS',  (0, 0), (-1, -1), 1.3*cm),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(sig_table)

        # ════════════════════════════════════════════
        #  PIED DE PAGE
        # ════════════════════════════════════════════
        elements.append(Spacer(1, 1.2*cm))
        elements.append(HRFlowable(width='100%', thickness=1,
                                   color=C_GREY_MID, spaceAfter=6))
        elements.append(Paragraph(
            f"Document généré automatiquement le <b>{now_str}</b> par le système SIGR-CA.",
            foot_style,
        ))
        elements.append(Paragraph(
            f"Fiche d'inventaire N° <b>{inv}</b> — Ce document est <b>confidentiel</b> "
            "et destiné à un usage interne uniquement. Toute reproduction ou diffusion "
            "non autorisée est interdite.",
            ps('CONF', fontSize=7, textColor=C_GREY_TEXT,
               alignment=TA_CENTER, leading=10),
        ))

        # ── Build PDF ────────────────────────────────────────────
        doc.build(elements, canvasmaker=NumberedCanvas)
        buffer.seek(0)

        nom_fichier = (m.get('nom', 'materiel') + '_' + inv).replace(' ', '_')
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="fiche_{nom_fichier}.pdf"'
        return response

    except Exception as e:
        import traceback
        return HttpResponse(
            f"Erreur PDF : {str(e)}\n\n{traceback.format_exc()}",
            content_type='text/plain',
            status=500
        )
# ─── RAPPORT PDF — Zone + tout son matériel ────────────────────────────────────
@session_required
def api_zone_pdf(request, zone_id):
    """Rapport PDF d'une zone avec tout son matériel — logo SIGR-CA."""
    try:
        bureau = db.bureaux.find_one({'_id': ObjectId(zone_id)})
        if not bureau:
            return HttpResponse('Zone non trouvée', status=404)

        zone_nom  = bureau.get('nom', 'Zone')
        materiels = (
            list(db.materiels.find({'zone': zone_nom}))
            if 'materiels' in db.list_collection_names() else []
        )

        now_str   = datetime.now().strftime('%d/%m/%Y à %H:%M')
        date_file = datetime.now().strftime('%Y%m%d_%H%M')

        # ── Mise en page ──────────────────────────────────────────
        buffer   = io.BytesIO()
        PAGE_W, PAGE_H = landscape(A4)
        CONTENT_W = PAGE_W - 3 * cm

        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=2*cm, bottomMargin=2*cm,
            title=f"Rapport Zone — {zone_nom}",
            author="SIGR-CA Système",
            subject=f"Rapport zone {zone_nom}",
        )

        # ── Registre de styles custom ─────────────────────────────
        _custom_styles = {}

        def ps(name, **kw):
            base = kw.pop('parent', 'Normal')
            stylesheet = getSampleStyleSheet()
            parent_style = _custom_styles.get(base) or stylesheet[base]
            p = ParagraphStyle(name, parent=parent_style, **kw)
            _custom_styles[name] = p
            return p

        # Styles
        th        = ps('TH',   fontName='Helvetica-Bold', fontSize=7,
                        textColor=C_WHITE, alignment=TA_CENTER,
                        leading=10, spaceAfter=0, spaceBefore=0)
        th_left   = ps('TH_L', parent='TH', alignment=TA_LEFT)
        td        = ps('TD',   fontSize=7, alignment=TA_CENTER,
                        leading=10, textColor=C_BLACK)
        td_left   = ps('TD_L', parent='TD', alignment=TA_LEFT)
        td_mono   = ps('TD_M', parent='TD', fontSize=7, fontName='Courier',
                        alignment=TA_CENTER)
        sec_style = ps('SEC',  fontName='Helvetica-Bold', fontSize=13,
                        textColor=C_PURPLE, spaceBefore=18, spaceAfter=8,
                        leading=17)
        foot_style = ps('FOOT', fontSize=7.5, textColor=C_GREY_TEXT,
                         alignment=TA_CENTER, leading=11)
        label_style = ps('LABEL', fontSize=9, textColor=C_GREY_TEXT,
                          fontName='Helvetica-Bold', alignment=TA_LEFT)
        val_style = ps('VAL', fontSize=10, fontName='Helvetica',
                        textColor=C_BLACK, alignment=TA_LEFT)

        elements = []

        # ════════════════════════════════════════════
        #  EN-TÊTE
        # ════════════════════════════════════════════
        from django.contrib.staticfiles.finders import find as static_find

        LOGO_PATH = static_find('img/logo.png')

        title_lines = [
            [Paragraph(
                '<font color="#1d4ed8"><b>SIGR-CA</b></font>',
                ps('LT', fontSize=22, leading=26, alignment=TA_LEFT)
            )],
            [Paragraph(
                "Système Intégré de Gestion des Ressources<br/>"
                "<font color='#64748b'>et de Contrôle d'Accès</font>",
                ps('LS', fontSize=9, leading=13,
                   textColor=C_GREY_TEXT, alignment=TA_LEFT)
            )],
        ]
        title_tbl = Table(title_lines, colWidths=[14*cm])
        title_tbl.setStyle(TableStyle([
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ]))

        if LOGO_PATH and os.path.exists(LOGO_PATH):
            logo_cell = Image(LOGO_PATH, width=4*cm, height=4*cm)
        else:
            logo_cell = Paragraph(
                '<font color="#1d4ed8"><b>SIGR</b></font>',
                ps('FL', fontSize=16, alignment=TA_RIGHT)
            )

        meta_lines = [
            [Paragraph(f"<b>Date :</b> {now_str}",
                       ps('M1', fontSize=7, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph(f"<b>Zone :</b> {zone_nom}",
                       ps('M2', fontSize=7, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph(f"<b>Articles :</b> {len(materiels)}",
                       ps('M3', fontSize=7, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph("<b>Confidentiel</b>",
                       ps('M4', fontSize=7, textColor=C_RED, alignment=TA_RIGHT))],
        ]
        meta_tbl = Table(meta_lines, colWidths=[4.5*cm])
        meta_tbl.setStyle(TableStyle([
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))

        right_col = Table([[logo_cell], [meta_tbl]], colWidths=[4.5*cm])
        right_col.setStyle(TableStyle([
            ('ALIGN',         (0,0), (-1,-1), 'RIGHT'),
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))

        header_tbl = Table(
            [[title_tbl, right_col]],
            colWidths=[CONTENT_W - 4.5*cm, 4.5*cm]
        )
        header_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))

        elements.append(header_tbl)
        elements.append(HRFlowable(width='100%', thickness=2,
                                   color=C_BLUE_MAIN, spaceAfter=4))
        elements.append(Spacer(1, 0.3*cm))

        # ── Bandeau titre ─────────────────────────────────────────
        elements.append(ColorBand(
            f"  RAPPORT DE ZONE : {zone_nom.upper()}",
            CONTENT_W, height=1*cm,
            bg=C_BLUE_DARK, fg=C_WHITE, font_size=11
        ))
        elements.append(Spacer(1, 0.6*cm))

        # ════════════════════════════════════════════
        #  INFORMATIONS DE LA ZONE
        # ════════════════════════════════════════════
        elements.append(Paragraph('Informations de la zone', sec_style))

        code_bureau = bureau.get('code_bureau', '—')
        etage = bureau.get('etage', 'RDC')
        capacite = bureau.get('capacite_max', 0)
        securite = bureau.get('niveau_securite', 'standard').title()
        statut_zone = bureau.get('statut', 'actif').title()
        description = bureau.get('description', '')
        responsable = bureau.get('responsable', '')
        telephone = bureau.get('telephone', '')
        email = bureau.get('email', '')

        zone_info = [
            ['Code bureau',      code_bureau],
            ['Étage',            str(etage)],
            ['Capacité max',     f"{capacite} personnes"],
            ['Niveau sécurité',  securite],
            ['Statut',           statut_zone],
        ]
        if responsable:
            zone_info.append(['Responsable', responsable])
        if telephone:
            zone_info.append(['Téléphone', telephone])
        if email:
            zone_info.append(['Email', email])

        t_zone = Table(
            [[Paragraph(r[0], label_style), Paragraph(str(r[1]), val_style)]
             for r in zone_info],
            colWidths=[4.5*cm, 18.5*cm],
        )
        t_zone.setStyle(TableStyle([
            ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
            ('BACKGROUND',     (0, 0), (0, -1), C_BLUE_LIGHT),
            ('PADDING',        (0, 0), (-1, -1), 7),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [C_WHITE, C_GREY_LIGHT]),
        ]))
        elements.append(t_zone)

        if description:
            elements.append(Spacer(1, 0.3*cm))
            elements.append(Paragraph(
                f"<b>Description :</b> {description}",
                ps('DESC', fontSize=9, textColor=C_GREY_TEXT, leading=13, spaceAfter=6),
            ))

        # ── Statistiques rapides ──
        dispo_zone = sum(1 for m in materiels if m.get('statut') == 'disponible')
        utilise_zone = sum(1 for m in materiels if m.get('statut') == 'utilise')
        maint_zone = sum(1 for m in materiels if m.get('statut') in ['maintenance', 'hors_service'])
        valeur_zone = sum(
            float(m.get('valeur', 0)) for m in materiels
            if m.get('valeur') and str(m.get('valeur')).replace('.', '').isdigit()
        )

        elements.append(Spacer(1, 0.4*cm))
        elements.append(Paragraph(
            f"📊 <b>{len(materiels)} articles</b> — "
            f"✅ {dispo_zone} disponibles  |  "
            f"🔄 {utilise_zone} utilisés  |  "
            f"⚠️ {maint_zone} en maintenance/HS  |  "
            f"💰 Valeur totale : <b>{valeur_zone:,.0f} DA</b>",
            ps('STATSZ', fontSize=9, textColor=C_BLUE_MAIN, leading=13, spaceAfter=6),
        ))

        # ════════════════════════════════════════════
        #  MATÉRIEL INVENTORIÉ
        # ════════════════════════════════════════════
        elements.append(Paragraph(
            f"Matériel inventorié ({len(materiels)} article(s))", sec_style,
        ))

        if materiels:
            headers = [
                Paragraph('N° Inventaire', th_left),
                Paragraph('Nom', th_left),
                Paragraph('Catégorie', th),
                Paragraph('Marque/Modèle', th_left),
                Paragraph('N° Série', th),
                Paragraph('Spécifications', th_left),
                Paragraph('Date achat', th),
                Paragraph('Valeur (DA)', th),
                Paragraph('Statut', th),
            ]

            statut_map = {
                'disponible':    '✅ Dispo',
                'utilise':       '🔄 Utilisé',
                'maintenance':   '⚠️ Maint.',
                'hors_service':  '❌ H.S.',
                'reforme':       '🗑️ Réformé',
            }

            rows = [headers]
            for mat in materiels:
                specs_parts = []
                if mat.get('processeur'): specs_parts.append(f"CPU: {mat['processeur']}")
                if mat.get('ram'):        specs_parts.append(f"RAM: {mat['ram']}")
                if mat.get('stockage'):   specs_parts.append(f"DD: {mat['stockage']}")
                if mat.get('os'):         specs_parts.append(f"OS: {mat['os']}")
                specs = ' | '.join(specs_parts) if specs_parts else '—'

                statut = mat.get('statut', '')
                statut_display = statut_map.get(statut, statut.replace('_', ' ').title())

                valeur = mat.get('valeur', '')
                valeur_display = f"{float(valeur):,.0f}" if valeur and str(valeur).replace('.', '').isdigit() else str(valeur) if valeur else '—'

                rows.append([
                    Paragraph(mat.get('num_inventaire', '—'), td_mono),
                    Paragraph(mat.get('nom', '—'), td_left),
                    Paragraph(mat.get('categorie', '—').title(), td),
                    Paragraph(
                        f"{mat.get('marque', '')} {mat.get('modele', '')}".strip() or '—',
                        td_left
                    ),
                    Paragraph(mat.get('numero_serie', '—'), td_mono),
                    Paragraph(specs, td_left),
                    Paragraph(mat.get('date_achat', '—'), td),
                    Paragraph(valeur_display, td),
                    Paragraph(statut_display, td),
                ])

            col_widths = [
                2.8*cm, 3*cm, 2.2*cm, 3.5*cm, 2.5*cm,
                5*cm, 2.2*cm, 2.2*cm, 2.2*cm,
            ]
            t_mat = Table(rows, colWidths=col_widths, repeatRows=1)
            t_mat.setStyle(TableStyle([
                ('BACKGROUND',     (0, 0), (-1, 0), C_BLUE_MAIN),
                ('TEXTCOLOR',      (0, 0), (-1, 0), C_WHITE),
                ('GRID',           (0, 0), (-1, -1), 0.3, C_GREY_MID),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
                ('PADDING',        (0, 0), (-1, -1), 4),
                ('TOPPADDING',     (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING',  (0, 0), (-1, 0), 8),
                ('TOPPADDING',     (0, 1), (-1, -1), 5),
                ('BOTTOMPADDING',  (0, 1), (-1, -1), 5),
                ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                ('BOX',            (0, 0), (-1, -1), 1, C_BLUE_MID),
            ]))
            elements.append(t_mat)

            # ── Répartition par catégorie dans la zone ──
            categories_zone = {}
            for m in materiels:
                cat = m.get('categorie', 'Non classé').title()
                categories_zone[cat] = categories_zone.get(cat, 0) + 1

            if len(categories_zone) > 1:
                elements.append(Spacer(1, 0.5*cm))
                elements.append(Paragraph('Répartition par catégorie', sec_style))
                cat_data = [[
                    Paragraph('<b>Catégorie</b>', th_left),
                    Paragraph('<b>Nombre</b>', th),
                    Paragraph('<b>%</b>', th),
                ]]
                for cat, cnt in sorted(categories_zone.items(), key=lambda x: -x[1]):
                    pct = round(cnt / len(materiels) * 100, 1) if len(materiels) > 0 else 0
                    cat_data.append([
                        Paragraph(cat, td_left),
                        Paragraph(str(cnt), td),
                        Paragraph(f'{pct}%', td),
                    ])

                t_cat_zone = Table(cat_data, colWidths=[10*cm, 3*cm, 3*cm])
                t_cat_zone.setStyle(TableStyle([
                    ('BACKGROUND',     (0, 0), (-1, 0), C_BLUE_MAIN),
                    ('TEXTCOLOR',      (0, 0), (-1, 0), C_WHITE),
                    ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
                    ('PADDING',        (0, 0), (-1, -1), 5),
                    ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
                ]))
                elements.append(t_cat_zone)
        else:
            elements.append(Paragraph(
                '<i>Aucun matériel assigné à cette zone.</i>',
                ps('NOMAT', fontSize=9, textColor=C_GREY_TEXT, alignment=TA_LEFT,
                   spaceBefore=4, spaceAfter=4),
            ))

        # ════════════════════════════════════════════
        #  PIED DE PAGE
        # ════════════════════════════════════════════
        elements.append(Spacer(1, 1*cm))
        elements.append(HRFlowable(width='100%', thickness=1,
                                   color=C_GREY_MID, spaceAfter=6))
        elements.append(Paragraph(
            f"Document généré automatiquement le <b>{now_str}</b> par le système SIGR-CA.",
            foot_style,
        ))
        elements.append(Paragraph(
            f"Rapport zone <b>{zone_nom}</b> — <b>{len(materiels)} articles</b> — "
            "Ce document est <b>confidentiel</b> et destiné à un usage interne uniquement. "
            "Toute reproduction ou diffusion non autorisée est interdite.",
            ps('CONF', fontSize=7, textColor=C_GREY_TEXT,
               alignment=TA_CENTER, leading=10),
        ))

        # ── Build PDF ────────────────────────────────────────────
        doc.build(elements, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        nom_fichier = f"Rapport_Zone_{zone_nom.replace(' ', '_')}_{date_file}"
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{nom_fichier}.pdf"'
        return response

    except Exception as e:
        return HttpResponse(
            f"Erreur PDF : {str(e)}\n\n{traceback.format_exc()}",
            content_type='text/plain',
            status=500
        )

# ─── RAPPORT PDF — Toutes les zones ───────────────────────────────────────────
@session_required
def api_zones_rapport_pdf(request):
    """Rapport PDF global toutes zones — avec logo SIGR-CA."""
    try:
        bureaux       = list(db.bureaux.find())
        materiels_all = (
            list(db.materiels.find()) if 'materiels' in db.list_collection_names() else []
        )

        now_str   = datetime.now().strftime('%d/%m/%Y à %H:%M')
        date_file = datetime.now().strftime('%Y%m%d_%H%M')

        # ── Mise en page ──────────────────────────────────────────
        buffer   = io.BytesIO()
        PAGE_W, PAGE_H = landscape(A4)
        CONTENT_W = PAGE_W - 3 * cm

        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=2*cm, bottomMargin=2*cm,
            title="Rapport Global des Ressources — SIGR-CA",
            author="SIGR-CA Système",
            subject="Rapport toutes zones",
        )

        # ── Registre de styles custom ─────────────────────────────
        _custom_styles = {}

        def ps(name, **kw):
            base = kw.pop('parent', 'Normal')
            stylesheet = getSampleStyleSheet()
            parent_style = _custom_styles.get(base) or stylesheet[base]
            p = ParagraphStyle(name, parent=parent_style, **kw)
            _custom_styles[name] = p
            return p

        # Styles
        th        = ps('TH',   fontName='Helvetica-Bold', fontSize=7,
                        textColor=C_WHITE, alignment=TA_CENTER,
                        leading=10, spaceAfter=0, spaceBefore=0)
        th_left   = ps('TH_L', parent='TH', alignment=TA_LEFT)
        td        = ps('TD',   fontSize=7, alignment=TA_CENTER,
                        leading=10, textColor=C_BLACK)
        td_left   = ps('TD_L', parent='TD', alignment=TA_LEFT)
        td_mono   = ps('TD_M', parent='TD', fontSize=7, fontName='Courier',
                        alignment=TA_CENTER)
        sec_style = ps('SEC',  fontName='Helvetica-Bold', fontSize=13,
                        textColor=C_PURPLE, spaceBefore=18, spaceAfter=8,
                        leading=17)
        foot_style = ps('FOOT', fontSize=7.5, textColor=C_GREY_TEXT,
                         alignment=TA_CENTER, leading=11)
        label_style = ps('LABEL', fontSize=8, textColor=C_GREY_TEXT,
                          fontName='Helvetica-Bold', alignment=TA_LEFT)
        val_style = ps('VAL', fontSize=12, fontName='Helvetica-Bold',
                        textColor=C_BLUE_MAIN, alignment=TA_LEFT)
        zone_title_style = ps('ZTITLE', fontName='Helvetica-Bold', fontSize=12,
                               textColor=C_BLUE_MAIN, spaceBefore=10, spaceAfter=4,
                               leading=16)

        elements = []

        # ════════════════════════════════════════════
        #  EN-TÊTE
        # ════════════════════════════════════════════
        from django.contrib.staticfiles.finders import find as static_find

        LOGO_PATH = static_find('img/logo.png')

        title_lines = [
            [Paragraph(
                '<font color="#1d4ed8"><b>SIGR-CA</b></font>',
                ps('LT', fontSize=22, leading=26, alignment=TA_LEFT)
            )],
            [Paragraph(
                "Système Intégré de Gestion des Ressources<br/>"
                "<font color='#64748b'>et de Contrôle d'Accès</font>",
                ps('LS', fontSize=9, leading=13,
                   textColor=C_GREY_TEXT, alignment=TA_LEFT)
            )],
        ]
        title_tbl = Table(title_lines, colWidths=[14*cm])
        title_tbl.setStyle(TableStyle([
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ]))

        if LOGO_PATH and os.path.exists(LOGO_PATH):
            logo_cell = Image(LOGO_PATH, width=4*cm, height=4*cm)
        else:
            logo_cell = Paragraph(
                '<font color="#1d4ed8"><b>SIGR</b></font>',
                ps('FL', fontSize=16, alignment=TA_RIGHT)
            )

        meta_lines = [
            [Paragraph(f"<b>Date :</b> {now_str}",
                       ps('M1', fontSize=7, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph(f"<b>Zones :</b> {len(bureaux)}",
                       ps('M2', fontSize=7, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph(f"<b>Articles :</b> {len(materiels_all)}",
                       ps('M3', fontSize=7, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph("<b>Confidentiel</b>",
                       ps('M4', fontSize=7, textColor=C_RED, alignment=TA_RIGHT))],
        ]
        meta_tbl = Table(meta_lines, colWidths=[4.5*cm])
        meta_tbl.setStyle(TableStyle([
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))

        right_col = Table([[logo_cell], [meta_tbl]], colWidths=[4.5*cm])
        right_col.setStyle(TableStyle([
            ('ALIGN',         (0,0), (-1,-1), 'RIGHT'),
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('TOPPADDING',    (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))

        header_tbl = Table(
            [[title_tbl, right_col]],
            colWidths=[CONTENT_W - 4.5*cm, 4.5*cm]
        )
        header_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))

        elements.append(header_tbl)
        elements.append(HRFlowable(width='100%', thickness=2,
                                   color=C_BLUE_MAIN, spaceAfter=4))
        elements.append(Spacer(1, 0.3*cm))

        # ── Bandeau titre ─────────────────────────────────────────
        elements.append(ColorBand(
            "  RAPPORT GLOBAL DES RESSOURCES — TOUTES ZONES",
            CONTENT_W, height=1*cm,
            bg=C_BLUE_DARK, fg=C_WHITE, font_size=11
        ))
        elements.append(Spacer(1, 0.5*cm))

        # ════════════════════════════════════════════
        #  RÉSUMÉ GLOBAL
        # ════════════════════════════════════════════
        elements.append(Paragraph('Résumé global', sec_style))

        total_mat = len(materiels_all)
        dispo     = sum(1 for m in materiels_all if m.get('statut') == 'disponible')
        utilise   = sum(1 for m in materiels_all if m.get('statut') == 'utilise')
        maint     = sum(1 for m in materiels_all if m.get('statut') in ['maintenance', 'hors_service'])
        reforme   = sum(1 for m in materiels_all if m.get('statut') == 'reforme')

        # Valeur totale
        valeur_totale = sum(
            float(m.get('valeur', 0)) for m in materiels_all
            if m.get('valeur') and str(m.get('valeur')).replace('.', '').isdigit()
        )

        def stat_card(label, value, color=C_BLUE_MAIN):
            return [
                Paragraph(label, ps(f'SCL_{label[:6]}', fontSize=8,
                                    textColor=C_GREY_TEXT, alignment=TA_LEFT)),
                Paragraph(
                    f'<font color="#{hex_color(color)}"><b>{value}</b></font>',
                    ps(f'SCV_{label[:6]}', fontSize=14,
                       fontName='Helvetica-Bold', alignment=TA_LEFT)
                ),
            ]

        resume_data = [
            stat_card('Total zones',          str(len(bureaux))),
            stat_card('Total articles',       str(total_mat)),
            stat_card('Disponibles',          str(dispo),       C_GREEN),
            stat_card('Utilisés',             str(utilise),     C_BLUE_MAIN),
            stat_card('En maintenance / HS',  str(maint),       C_AMBER),
            stat_card('Réformés',             str(reforme),     C_RED),
            stat_card('Valeur totale (DA)',   f'{valeur_totale:,.0f} DA', C_PURPLE),
        ]

        t_resume = Table(resume_data, colWidths=[3.5*cm, 3.5*cm]*3 + [4*cm])
        t_resume.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0), (-1, -1), C_BLUE_LIGHT),
            ('BOX',            (0, 0), (-1, -1), 1, C_BLUE_MAIN),
            ('GRID',           (0, 0), (-1, -1), 0.4, C_BLUE_MID),
            ('PADDING',        (0, 0), (-1, -1), 8),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
        ]))
        elements.append(t_resume)

        # ── Répartition par catégorie ──
        categories = {}
        for m in materiels_all:
            cat = m.get('categorie', 'Non classé').title()
            categories[cat] = categories.get(cat, 0) + 1

        if categories:
            elements.append(Spacer(1, 0.4*cm))
            cat_data = [[
                Paragraph('<b>Répartition par catégorie</b>',
                          ps('CATT', fontSize=8, textColor=C_GREY_TEXT)),
                Paragraph('<b>Nb</b>',
                          ps('CATN', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_CENTER)),
                Paragraph('<b>%</b>',
                          ps('CATP', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_CENTER)),
            ]]
            for cat, cnt in sorted(categories.items(), key=lambda x: -x[1]):
                pct = round(cnt / total_mat * 100, 1) if total_mat > 0 else 0
                cat_data.append([
                    Paragraph(cat, ps(f'CN_{cat[:8]}', fontSize=8)),
                    Paragraph(str(cnt), ps(f'CC_{cat[:8]}', fontSize=8, alignment=TA_CENTER)),
                    Paragraph(f'{pct}%', ps(f'CP_{cat[:8]}', fontSize=8, alignment=TA_CENTER)),
                ])

            t_cat = Table(cat_data, colWidths=[10*cm, 3*cm, 3*cm])
            t_cat.setStyle(TableStyle([
                ('GRID',           (0, 0), (-1, -1), 0.4, C_GREY_MID),
                ('BACKGROUND',     (0, 0), (-1, 0), C_GREY_LIGHT),
                ('PADDING',        (0, 0), (-1, -1), 5),
                ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
            ]))
            elements.append(t_cat)

        elements.append(Spacer(1, 0.6*cm))

        # ════════════════════════════════════════════
        #  UNE SECTION PAR ZONE
        # ════════════════════════════════════════════
        for idx, bureau in enumerate(bureaux):
            zone_nom  = bureau.get('nom', 'Zone sans nom')
            mats_zone = [m for m in materiels_all if m.get('zone') == zone_nom]

            if idx > 0:
                elements.append(Spacer(1, 0.3*cm))

            elements.append(HRFlowable(width='100%', thickness=1.5,
                                       color=C_BLUE_MID, spaceAfter=6))

            elements.append(Paragraph(
                f"📍  {zone_nom}  —  {len(mats_zone)} article(s)",
                zone_title_style,
            ))

            code_bureau = bureau.get('code_bureau', '—')
            etage = bureau.get('etage', 'RDC')
            capacite = bureau.get('capacite_max', 0)
            securite = bureau.get('niveau_securite', 'standard').title()
            statut_zone = bureau.get('statut', 'actif').title()
            description = bureau.get('description', '')

            zone_infos = (
                f"<b>Code :</b> {code_bureau}  |  "
                f"<b>Étage :</b> {etage}  |  "
                f"<b>Capacité :</b> {capacite} pers  |  "
                f"<b>Sécurité :</b> {securite}  |  "
                f"<b>Statut :</b> {statut_zone}"
            )
            elements.append(Paragraph(
                zone_infos,
                ps('ZINFO', fontSize=8, textColor=C_GREY_TEXT, spaceAfter=4, leading=12),
            ))

            if description:
                elements.append(Paragraph(
                    f"<i>{description[:200]}</i>",
                    ps('ZDESC', fontSize=7, textColor=C_GREY_TEXT, spaceAfter=6, leading=10),
                ))

            dispo_zone = sum(1 for m in mats_zone if m.get('statut') == 'disponible')
            maint_zone = sum(1 for m in mats_zone if m.get('statut') in ['maintenance', 'hors_service'])
            valeur_zone = sum(
                float(m.get('valeur', 0)) for m in mats_zone
                if m.get('valeur') and str(m.get('valeur')).replace('.', '').isdigit()
            )

            elements.append(Paragraph(
                f"✅ {dispo_zone} dispo  |  ⚠️ {maint_zone} en maintenance/HS  |  "
                f"💰 Valeur zone : {valeur_zone:,.0f} DA",
                ps('ZSTATS', fontSize=7.5, textColor=C_BLUE_MAIN, spaceAfter=8, leading=11),
            ))

            if mats_zone:
                headers = [
                    Paragraph('N° Inventaire', th_left),
                    Paragraph('Nom', th_left),
                    Paragraph('Catégorie', th),
                    Paragraph('Marque/Modèle', th_left),
                    Paragraph('N° Série', th),
                    Paragraph('Spécifications', th_left),
                    Paragraph('Date achat', th),
                    Paragraph('Valeur (DA)', th),
                    Paragraph('Statut', th),
                ]

                statut_map = {
                    'disponible':    '✅ Dispo',
                    'utilise':       '🔄 Utilisé',
                    'maintenance':   '⚠️ Maint.',
                    'hors_service':  '❌ H.S.',
                    'reforme':       '🗑️ Réformé',
                }

                rows = [headers]
                for mat in mats_zone:
                    specs_parts = []
                    if mat.get('processeur'): specs_parts.append(f"CPU: {mat['processeur']}")
                    if mat.get('ram'):        specs_parts.append(f"RAM: {mat['ram']}")
                    if mat.get('stockage'):   specs_parts.append(f"DD: {mat['stockage']}")
                    if mat.get('os'):         specs_parts.append(f"OS: {mat['os']}")
                    specs = ' | '.join(specs_parts) if specs_parts else '—'

                    statut = mat.get('statut', '')
                    statut_display = statut_map.get(statut, statut.replace('_', ' ').title())

                    valeur = mat.get('valeur', '')
                    valeur_display = f"{float(valeur):,.0f}" if valeur and str(valeur).replace('.', '').isdigit() else str(valeur) if valeur else '—'

                    rows.append([
                        Paragraph(mat.get('num_inventaire', '—'), td_mono),
                        Paragraph(mat.get('nom', '—'), td_left),
                        Paragraph(mat.get('categorie', '—').title(), td),
                        Paragraph(
                            f"{mat.get('marque', '')} {mat.get('modele', '')}".strip() or '—',
                            td_left
                        ),
                        Paragraph(mat.get('numero_serie', '—'), td_mono),
                        Paragraph(specs, td_left),
                        Paragraph(mat.get('date_achat', '—'), td),
                        Paragraph(valeur_display, td),
                        Paragraph(statut_display, td),
                    ])

                col_widths = [
                    2.8*cm, 3*cm, 2.2*cm, 3.5*cm, 2.5*cm,
                    5*cm, 2.2*cm, 2.2*cm, 2.2*cm,
                ]
                t = Table(rows, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND',     (0, 0), (-1, 0), C_BLUE_MAIN),
                    ('TEXTCOLOR',      (0, 0), (-1, 0), C_WHITE),
                    ('GRID',           (0, 0), (-1, -1), 0.3, C_GREY_MID),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [C_WHITE, C_BLUE_LIGHT]),
                    ('PADDING',        (0, 0), (-1, -1), 4),
                    ('TOPPADDING',     (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING',  (0, 0), (-1, 0), 8),
                    ('TOPPADDING',     (0, 1), (-1, -1), 5),
                    ('BOTTOMPADDING',  (0, 1), (-1, -1), 5),
                    ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                    ('BOX',            (0, 0), (-1, -1), 1, C_BLUE_MID),
                ]))
                elements.append(t)
            else:
                elements.append(Paragraph(
                    '<i>Aucun matériel assigné à cette zone.</i>',
                    ps('NOMAT', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_LEFT,
                       spaceBefore=4, spaceAfter=4),
                ))

        # ════════════════════════════════════════════
        #  PIED DE PAGE
        # ════════════════════════════════════════════
        elements.append(Spacer(1, 1*cm))
        elements.append(HRFlowable(width='100%', thickness=1,
                                   color=C_GREY_MID, spaceAfter=6))
        elements.append(Paragraph(
            f"Document généré automatiquement le <b>{now_str}</b> par le système SIGR-CA.",
            foot_style,
        ))
        elements.append(Paragraph(
            f"Rapport global — <b>{len(bureaux)} zones</b>, <b>{total_mat} articles</b> — "
            "Ce document est <b>confidentiel</b> et destiné à un usage interne uniquement. "
            "Toute reproduction ou diffusion non autorisée est interdite.",
            ps('CONF', fontSize=7, textColor=C_GREY_TEXT,
               alignment=TA_CENTER, leading=10),
        ))

        # ── Build PDF ────────────────────────────────────────────
        doc.build(elements, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        filename = f"Rapport_Zones_{date_file}.pdf"
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    except Exception as e:
        return HttpResponse(
            f"Erreur PDF : {str(e)}\n\n{traceback.format_exc()}",
            content_type='text/plain',
            status=500
        )


# ─── Export CSV ───────────────────────────────────────────────────────────────

@session_required
def api_export_ressources_csv(request):
    import csv
    from django.http import HttpResponse
    from datetime import datetime

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="ressources_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    response.write('\ufeff')

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Type','Nom','N° Inventaire','Fournisseur','Code/Série',
                     'Catégorie','Marque','Modèle','Processeur','RAM','Stockage','OS',
                     'Capacité','Étage','Niveau sécurité','Statut','Zone','Date achat','Valeur DA'])

    for b in db.bureaux.find():
        writer.writerow(['Zone', b.get('nom',''), '', '', b.get('code_bureau',''),
                         '','','','','','','',
                         b.get('capacite_max',0), b.get('etage',0),
                         b.get('niveau_securite','standard'), b.get('statut','actif'), '', '', ''])

    materiels = list(db.materiels.find()) if 'materiels' in db.list_collection_names() else []
    for m in materiels:
        writer.writerow(['Matériel', m.get('nom',''), m.get('num_inventaire',''),
                         m.get('fournisseur',''), m.get('numero_serie',''),
                         m.get('categorie',''), m.get('marque',''), m.get('modele',''),
                         m.get('processeur',''), m.get('ram',''), m.get('stockage',''), m.get('os',''),
                         '', '', '', m.get('statut',''), m.get('zone',''),
                         m.get('date_achat',''), m.get('valeur','')])

    return response


# ─── API bureau stats ─────────────────────────────────────────────────────────

@session_required
def api_bureau_stats(request, bureau_id):
    from bson import ObjectId
    from datetime import datetime, timedelta

    try:
        bureau = db.bureaux.find_one({'_id': ObjectId(bureau_id)})
        if not bureau:
            return JsonResponse({'error': 'Bureau non trouvé'}, status=404)

        dates, acces_par_jour = [], []
        for i in range(6, -1, -1):
            day_start = (datetime.now() - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
            day_end   = day_start + timedelta(days=1)
            count = db.acces_logs.count_documents({
                'bureau_id': ObjectId(bureau_id),
                'timestamp': {'$gte': day_start, '$lt': day_end}
            })
            dates.append(day_start.strftime('%a'))
            acces_par_jour.append(count)

        return JsonResponse({
            'dates':         dates,
            'acces_par_jour':acces_par_jour,
            'nom':           bureau.get('nom'),
            'capacite':      bureau.get('capacite_max', 0),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



# ====================== CALENDRIER ET RÈGLES ======================

# ====================== CALENDRIER ET RÈGLES ======================

@session_required
def calendrier(request):
    """Page du calendrier des règles d'accès"""
    from bson import ObjectId
    
    employes = list(db.employees.find({'statut': 'actif'}))
    for e in employes:
        e['id'] = str(e['_id'])
    
    bureaux = list(db.bureaux.find())
    for b in bureaux:
        b['id'] = str(b['_id'])
    
    return render(request, 'dashboard/calendrier.html', {
        'employes': employes,
        'bureaux': bureaux,
    })


@session_required
def api_get_employee_rules(request, employe_id):
    """API pour récupérer les règles d'un employé"""
    from bson import ObjectId
    
    try:
        rules_cursor = db.access_rules.find({'employe_id': employe_id})
        formatted_rules = {}
        for rule in rules_cursor:
            jour, mois, annee = rule.get('jour'), rule.get('mois'), rule.get('annee')
            if not (jour and mois and annee):
                continue
            key = f"{annee}-{mois}-{jour}"
            if key not in formatted_rules:
                formatted_rules[key] = {}
            formatted_rules[key][rule['zone_nom']] = {
                'heure_debut': rule.get('heure_debut', '08:00'),
                'heure_fin': rule.get('heure_fin', '18:00'),
                'acces_autorise': rule.get('acces_autorise', True)
            }
        return JsonResponse({'rules': formatted_rules, 'status': 'success'})
    except Exception as e:
        return JsonResponse({'rules': {}, 'status': 'error', 'message': str(e)})


@session_required
def api_save_day_rules(request):
    """API pour sauvegarder les règles d'un jour"""
    from datetime import datetime
    import json
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        rules = data.get('rules', [])
        employe_id = data.get('employe_id')
        
        if not employe_id and rules:
            employe_id = rules[0].get('employe_id')
        if not employe_id:
            return JsonResponse({'error': 'Employé ID manquant'}, status=400)
        
        saved_count = 0
        for rule in rules:
            jour, mois, annee = rule.get('jour'), rule.get('mois'), rule.get('annee')
            zone_nom = rule.get('zone_nom', '')
            
            if zone_nom == '__DELETE__':
                db.access_rules.delete_many({
                    'employe_id': employe_id,
                    'jour': jour,
                    'mois': mois,
                    'annee': annee
                })
            else:
                db.access_rules.delete_one({
                    'employe_id': employe_id,
                    'zone_nom': zone_nom,
                    'jour': jour,
                    'mois': mois,
                    'annee': annee
                })
                db.access_rules.insert_one({
                    'employe_id': employe_id,
                    'zone_nom': zone_nom,
                    'jour': jour,
                    'mois': mois,
                    'annee': annee,
                    'heure_debut': rule.get('heure_debut', '08:00'),
                    'heure_fin': rule.get('heure_fin', '18:00'),
                    'acces_autorise': rule.get('acces_autorise', True),
                    'created_at': datetime.now(),
                    'updated_at': datetime.now()
                })
                saved_count += 1
        
        return JsonResponse({'status': 'success', 'message': f'{saved_count} règle(s) sauvegardée(s)'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_save_all_rules(request):
    """API pour sauvegarder toutes les règles d'un employé"""
    from datetime import datetime
    import json
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        employe_id = data.get('employe_id')
        rules = data.get('rules', [])
        
        if not employe_id:
            return JsonResponse({'error': 'Employé ID manquant'}, status=400)
        
        # Supprimer toutes les règles existantes
        db.access_rules.delete_many({'employe_id': employe_id})
        
        if rules:
            rules_to_insert = []
            for r in rules:
                rules_to_insert.append({
                    'employe_id': employe_id,
                    'zone_nom': r.get('zone_nom', ''),
                    'jour': r.get('jour'),
                    'mois': r.get('mois'),
                    'annee': r.get('annee'),
                    'heure_debut': r.get('heure_debut', '08:00'),
                    'heure_fin': r.get('heure_fin', '18:00'),
                    'acces_autorise': r.get('acces_autorise', True),
                    'created_at': datetime.now(),
                    'updated_at': datetime.now()
                })
            db.access_rules.insert_many(rules_to_insert)
        
        return JsonResponse({'status': 'success', 'message': f'{len(rules)} règle(s) enregistrée(s)'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_bureaux(request):
    """API pour récupérer la liste des bureaux"""
    bureaux = list(db.bureaux.find())
    result = [{
        'id': str(b['_id']),
        'nom': b.get('nom', ''),
        'niveau': b.get('niveau_securite', 'standard'),
        'capacite': b.get('capacite_max', 0),
        'etage': b.get('etage', 0)
    } for b in bureaux]
    
    # Si aucun bureau n'existe, retourner des données par défaut
    if not result:
        result = [
            {'id': '1', 'nom': 'Direction Générale', 'niveau': 'critique', 'capacite': 5, 'etage': 1},
            {'id': '2', 'nom': 'Atelier Production', 'niveau': 'standard', 'capacite': 20, 'etage': 0},
            {'id': '3', 'nom': 'Salle Serveur', 'niveau': 'critique', 'capacite': 2, 'etage': 0},
            {'id': '4', 'nom': 'Archives', 'niveau': 'restreint', 'capacite': 3, 'etage': 0},
            {'id': '5', 'nom': 'Bureau RH', 'niveau': 'standard', 'capacite': 4, 'etage': 1},
            {'id': '6', 'nom': 'Laboratoire', 'niveau': 'restreint', 'capacite': 6, 'etage': 1},
            {'id': '7', 'nom': 'Entrée Principale', 'niveau': 'public', 'capacite': 50, 'etage': 0},
        ]
    
    return JsonResponse({'bureaux': result})


# ====================== STATISTIQUES ======================
@session_required
def statistiques(request):
    from datetime import datetime, timedelta
    import json
    
    now = datetime.now()
    start_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # ===== KPIs de base =====
    total_mois = db.acces_logs.count_documents({'timestamp': {'$gte': start_month}})
    total_all = db.acces_logs.count_documents({})
    autorises = db.acces_logs.count_documents({'resultat': 'AUTORISE'})
    taux_succes = round(autorises / total_all * 100) if total_all else 0
    
    # ===== Top employés =====
    top_employes = []
    for t in db.acces_logs.aggregate([
        {'$match': {'timestamp': {'$gte': start_month}}},
        {'$group': {'_id': '$utilisateur_id', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}},
        {'$limit': 5}
    ]):
        emp = db.employees.find_one({'_id': t['_id']})
        if emp:
            auto_emp = db.acces_logs.count_documents({'utilisateur_id': t['_id'], 'resultat': 'AUTORISE'})
            top_employes.append({
                'nom': emp.get('nom', ''),
                'prenom': emp.get('prenom', ''),
                'departement': emp.get('departement', ''),
                'nb_acces': t['count'],
                'taux_succes': round(auto_emp / t['count'] * 100) if t['count'] else 0,
                'dernier_acces': None,
            })
    
    # ===== Zones stats =====
    zones_stats = []
    for z in db.acces_logs.aggregate([
        {'$match': {'timestamp': {'$gte': start_month}}},
        {'$group': {'_id': '$bureau_id', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}},
        {'$limit': 5}
    ]):
        b = db.bureaux.find_one({'_id': z['_id']})
        if b:
            zones_stats.append({
                'nom': b.get('nom', 'Inconnu'), 
                'count': z['count']
            })
    
    # Calculer les pourcentages
    total_zones = sum(z['count'] for z in zones_stats)
    for z in zones_stats:
        z['pct'] = round(z['count'] / total_zones * 100) if total_zones else 0
    
    # ===== Données pour le graphique (30 jours) =====
    labels = []
    autorises_list = []
    refuses_list = []
    
    for i in range(29, -1, -1):
        day_start = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        a = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'AUTORISE'})
        r = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'REFUSE'})
        labels.append(day_start.strftime('%d/%m'))
        autorises_list.append(a)
        refuses_list.append(r)
    
    # Prédictions
    prediction_list = []
    for i in range(len(autorises_list)):
        window = autorises_list[max(0, i-6):i+1]
        avg = sum(window) / len(window) if window else 0
        prediction_list.append(round(avg * 1.05, 1))
    
    # Calcul de la prédiction globale
    last_7 = sum(autorises_list[-7:]) if len(autorises_list) >= 7 else sum(autorises_list)
    prev_7 = sum(autorises_list[-14:-7]) if len(autorises_list) >= 14 else last_7
    prediction_pct = round(((last_7 - prev_7) / prev_7 * 100) if prev_7 else 0, 1)
    
    # ===== Données occupation salles =====
    total_salles = db.bureaux.count_documents({})
    reservations_mois = db.reservations.count_documents({
        'date_debut': {'$gte': start_month},
        'statut': 'confirmee'
    })
    heures_possibles = total_salles * 240 if total_salles > 0 else 1
    heures_occupees = reservations_mois * 2
    occupation_moy = min(100, round((heures_occupees / heures_possibles) * 100, 1)) if heures_possibles > 0 else 0
    
    total_reservations = reservations_mois
    
    # Salles disponibles maintenant
    salles_reservees = db.reservations.distinct('bureau_id', {
        'date_debut': {'$lte': now},
        'date_fin': {'$gte': now},
        'statut': 'confirmee'
    })
    salles_disponibles = total_salles - len(salles_reservees)
    
    # Graphique occupation des salles
    occupation_labels = []
    occupation_values = []
    for bureau in db.bureaux.find().limit(8):
        res_count = db.reservations.count_documents({
            'bureau_id': bureau['_id'],
            'date_debut': {'$gte': start_month},
            'statut': 'confirmee'
        })
        taux = min(100, round((res_count * 2 / 240) * 100, 1)) if res_count > 0 else 0
        occupation_labels.append(bureau.get('nom', 'Salle')[:15])
        occupation_values.append(taux)
    
    # Top ressources
    top_ressources_list = []
    pipeline = [
        {'$match': {'date_debut': {'$gte': start_month}, 'statut': 'confirmee'}},
        {'$group': {'_id': '$bureau_id', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}},
        {'$limit': 5}
    ]
    results = list(db.reservations.aggregate(pipeline))
    total_res = sum(r['count'] for r in results) if results else 1
    for r in results:
        bureau = db.bureaux.find_one({'_id': r['_id']})
        if bureau:
            top_ressources_list.append({
                'nom': bureau.get('nom', 'Salle')[:20],
                'reservations': r['count'],
                'taux': round(r['count'] / total_res * 100, 1)
            })
    
    # ===== Construction du contexte =====
    context = {
        # KPIs
        'total_mois': total_mois,
        'taux_succes': taux_succes,
        'taux_refus': 100 - taux_succes,
        'pic_heure': '08h30',
        'zone_active': zones_stats[0]['nom'] if zones_stats else 'N/A',
        'top_employes': top_employes,
        'prediction': prediction_pct,
        
        # KPIs ressources
        'occupation_moy': occupation_moy,
        'total_reservations': total_reservations,
        'salles_disponibles': salles_disponibles,
        'total_salles': total_salles,
        
        # Données JSON (stringifiées correctement)
        'chart_labels': json.dumps(labels),
        'chart_autorises': json.dumps(autorises_list),
        'chart_refuses': json.dumps(refuses_list),
        'chart_prediction': json.dumps(prediction_list),
        'zones_stats': json.dumps(zones_stats),
        'top_ressources': json.dumps(top_ressources_list),
        'occupation_labels': json.dumps(occupation_labels),
        'occupation_values': json.dumps(occupation_values),
    }
    
    return render(request, 'dashboard/statistiques.html', context)
    # ====================== STATISTIQUES AVANCÉES ======================

@session_required
def api_stats_export_csv(request):
    """Export des statistiques en CSV"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    import csv
    from datetime import datetime, timedelta
    from django.http import HttpResponse
    
    # Récupérer la période
    days = int(request.GET.get('days', 30))
    start_date = datetime.now() - timedelta(days=days)
    
    # Récupérer les données
    stats = []
    for i in range(days, -1, -1):
        day_start = (datetime.now() - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        a = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'AUTORISE'})
        r = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'REFUSE'})
        
        stats.append({
            'date': day_start.strftime('%d/%m/%Y'),
            'autorises': a,
            'refuses': r,
            'total': a + r,
            'taux_succes': round(a / (a + r) * 100, 1) if (a + r) > 0 else 0
        })
    
    # Créer la réponse CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="statistiques_acces_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Accès autorisés', 'Accès refusés', 'Total', 'Taux de succès (%)'])
    
    for s in stats:
        writer.writerow([s['date'], s['autorises'], s['refuses'], s['total'], s['taux_succes']])
    
    return response


@session_required
def api_stats_export_pdf(request):
    """Export des statistiques en PDF"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    from datetime import datetime, timedelta
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    import io
    
    # Récupérer la période
    days = int(request.GET.get('days', 30))
    start_date = datetime.now() - timedelta(days=days)
    
    # Récupérer les données
    stats = []
    for i in range(days, -1, -1):
        day_start = (datetime.now() - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        a = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'AUTORISE'})
        r = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'REFUSE'})
        
        stats.append([day_start.strftime('%d/%m/%Y'), str(a), str(r), str(a + r), f"{round(a / (a + r) * 100, 1) if (a + r) > 0 else 0}%"])
    
    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Style
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, alignment=1)
    
    # Titre
    elements.append(Paragraph(f"Rapport des statistiques d'accès - {datetime.now().strftime('%d/%m/%Y')}", title_style))
    elements.append(Spacer(1, 0.5 * cm))
    
    # Tableau
    data = [['Date', 'Autorisés', 'Refusés', 'Total', 'Taux succès']] + stats
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="statistiques_acces_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


@session_required
def api_stats_departement(request):
    """Statistiques par département"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    from datetime import datetime, timedelta
    
    days = int(request.GET.get('days', 30))
    start_date = datetime.now() - timedelta(days=days)
    
    # Récupérer tous les employés par département
    pipeline = [
        {'$match': {'statut': 'actif'}},
        {'$group': {'_id': '$departement', 'count': {'$sum': 1}}}
    ]
    dept_counts = list(db.employees.aggregate(pipeline))
    
    # Statistiques par département
    dept_stats = []
    for dept in dept_counts:
        dept_name = dept['_id'] or 'Non défini'
        
        # Récupérer les employés de ce département
        employees = list(db.employees.find({'departement': dept_name}))
        employee_ids = [e['_id'] for e in employees]
        
        # Compter les accès
        total_acces = db.acces_logs.count_documents({
            'utilisateur_id': {'$in': employee_ids},
            'timestamp': {'$gte': start_date}
        })
        
        autorises = db.acces_logs.count_documents({
            'utilisateur_id': {'$in': employee_ids},
            'timestamp': {'$gte': start_date},
            'resultat': 'AUTORISE'
        })
        
        dept_stats.append({
            'nom': dept_name,
            'employes': dept['count'],
            'acces': total_acces,
            'taux_succes': round(autorises / total_acces * 100, 1) if total_acces > 0 else 0
        })
    
    return JsonResponse({'departements': dept_stats})


@session_required
def api_stats_period_custom(request):
    """Statistiques pour une période personnalisée"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    from datetime import datetime
    
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    if not date_debut_str or not date_fin_str:
        return JsonResponse({'error': 'Dates manquantes'}, status=400)
    
    try:
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d')
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d')
        date_fin = date_fin.replace(hour=23, minute=59, second=59)
    except ValueError:
        return JsonResponse({'error': 'Format de date invalide'}, status=400)
    
    # Statistiques globales
    total_acces = db.acces_logs.count_documents({'timestamp': {'$gte': date_debut, '$lte': date_fin}})
    autorises = db.acces_logs.count_documents({'timestamp': {'$gte': date_debut, '$lte': date_fin}, 'resultat': 'AUTORISE'})
    
    # Données quotidiennes
    stats = []
    current = date_debut
    while current <= date_fin:
        day_end = current.replace(hour=23, minute=59, second=59)
        a = db.acces_logs.count_documents({'timestamp': {'$gte': current, '$lte': day_end}, 'resultat': 'AUTORISE'})
        r = db.acces_logs.count_documents({'timestamp': {'$gte': current, '$lte': day_end}, 'resultat': 'REFUSE'})
        
        stats.append({
            'date': current.strftime('%d/%m'),
            'autorises': a,
            'refuses': r
        })
        current += timedelta(days=1)
    
    return JsonResponse({
        'total_acces': total_acces,
        'taux_succes': round(autorises / total_acces * 100, 1) if total_acces > 0 else 0,
        'stats': stats
    })


@session_required
def api_stats_trend_cache(request):
    """Version avec cache des statistiques de tendance"""
    from django.core.cache import cache
    from datetime import datetime, timedelta
    
    days = int(request.GET.get('days', 30))
    cache_key = f"stats_trend_{days}"
    
    # Vérifier le cache
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data)
    
    # Calculer les données
    now = datetime.now()
    labels, autorises_list, refuses_list, prediction_list = [], [], [], []
    
    for i in range(days - 1, -1, -1):
        day_start = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        a = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'AUTORISE'})
        r = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'REFUSE'})
        labels.append(day_start.strftime('%d/%m'))
        autorises_list.append(a)
        refuses_list.append(r)
    
    # Prédictions
    for i in range(len(autorises_list)):
        window = autorises_list[max(0, i-6):i+1]
        avg = sum(window) / len(window) if window else 0
        prediction_list.append(round(avg * 1.05, 1))
    
    data = {
        'labels': labels,
        'autorises': autorises_list,
        'refuses': refuses_list,
        'prediction': prediction_list
    }
    
    # Mettre en cache pour 5 minutes
    cache.set(cache_key, data, 300)
    
    return JsonResponse(data)

# ================================================================
# PARAMETRES VIEWS — version finale (MongoDB, pas Django ORM)
# Collez ce code dans votre views.py
# ================================================================

import json
from datetime import datetime
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST
from django.http import JsonResponse

CONFIG_COLLECTION = 'config_systeme'
CONFIG_DOC_ID     = 'global'

DEFAULTS = {
    's_2fa_global'    : False,
    's_block5'        : True,
    's_alert3'        : True,
    's_keep'          : True,
    's_refresh'       : True,
    'session_timeout' : '30',
    'max_failures'    : '5',
    'dark_default'    : True,
    'default_language': 'fr',
    'log_retention'   : '365',
    'n_email_resa'    : True,
    'n_reminder'      : True,
    'n_weekly'        : False,
    'n_unauth'        : True,
    'n_maintenance'   : True,
    'n_sms'           : False,
    'notification_email': '',
    'backup_auto'     : True,
    'backup_ftp'      : False,
    'api_external'    : False,
    'ldap_auth'       : False,
    'webhook_alerts'  : False,
    'api_key'         : '',
    'webhook_url'     : '',
    'last_backup_date': 'Aucune',
}


def _load_config():
    doc = db[CONFIG_COLLECTION].find_one({'_id': CONFIG_DOC_ID}) or {}
    config = dict(DEFAULTS)
    config.update({k: v for k, v in doc.items() if k != '_id'})
    return config


def _save_config(data: dict):
    db[CONFIG_COLLECTION].update_one(
        {'_id': CONFIG_DOC_ID},
        {'$set': data},
        upsert=True,
    )


@session_required
def parametres(request):
    """Affiche la page de paramètres."""
    if not request.session.get('is_staff', False) and not request.session.get('is_superuser', False):
        return redirect('employe_espace')

    config = _load_config()

    user_id = str(request.session.get('user_id', ''))
    admin_profile_doc = db['admin_profiles'].find_one({'user_id': user_id}) or {}
    admin_profile = {
        'phone': admin_profile_doc.get('phone', ''),
        'role' : admin_profile_doc.get('role', 'Administrateur Système'),
    }

    context = {
        'config'       : config,
        'user'         : get_session_user(request),
        'admin_profile': admin_profile,
    }
    return render(request, 'dashboard/parametres.html', context)


@session_required
@require_POST
def api_parametres_save(request):
    """Reçoit JSON, sauvegarde config + profil admin dans MongoDB."""
    if not request.session.get('is_staff', False) and not request.session.get('is_superuser', False):
        return JsonResponse({'status': 'error', 'message': 'Accès refusé'}, status=403)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError) as e:
        return JsonResponse({'status': 'error', 'message': f'JSON invalide : {e}'}, status=400)

    errors = []

    # ── 1. Mise à jour du profil admin ────────────────────────────────────────
    profile_data = body.pop('profile_update', None)
    if profile_data:
        from django.contrib.auth.hashers import check_password, make_password
        from bson import ObjectId

        user_id  = request.session.get('user_id', '')
        username = request.session.get('username', '')

        # Chercher l'utilisateur dans MongoDB
        u_doc = None
        if user_id:
            try:
                u_doc = db['utilisateurs'].find_one({'_id': ObjectId(str(user_id))})
            except Exception:
                pass
        if not u_doc:
            u_doc = db['utilisateurs'].find_one({'username': username})

        if not u_doc:
            return JsonResponse({'status': 'error', 'message': 'Utilisateur introuvable en base.'})

        update_fields = {}

        # Email
        new_email = (profile_data.get('email') or '').strip()
        current_email = (u_doc.get('email') or '').strip()
        if new_email and new_email != current_email:
            if db['utilisateurs'].find_one({'email': new_email, '_id': {'$ne': u_doc['_id']}}):
                errors.append('Cet email est déjà utilisé.')
            else:
                update_fields['email'] = new_email

        # Prénom / Nom
        new_firstname = (profile_data.get('first_name') or '').strip()
        new_lastname  = (profile_data.get('last_name')  or '').strip()
        if new_firstname:
            update_fields['first_name'] = new_firstname
        if new_lastname:
            update_fields['last_name'] = new_lastname

        # Mot de passe
        current_pwd = profile_data.get('current_password', '')
        new_pwd     = profile_data.get('new_password', '')
        confirm_pwd = profile_data.get('confirm_password', '')

        if new_pwd:
            stored_hash = u_doc.get('password', '')
            if not check_password(current_pwd, stored_hash):
                errors.append('Mot de passe actuel incorrect.')
            elif new_pwd != confirm_pwd:
                errors.append('Les mots de passe ne correspondent pas.')
            elif len(new_pwd) < 8:
                errors.append('Le mot de passe doit contenir au moins 8 caractères.')
            else:
                update_fields['password'] = make_password(new_pwd)

        if errors:
            return JsonResponse({'status': 'error', 'message': ' | '.join(errors)})

        # Sauvegarder dans utilisateurs
        if update_fields:
            db['utilisateurs'].update_one(
                {'_id': u_doc['_id']},
                {'$set': update_fields}
            )
            # Mettre à jour la session
            if 'email' in update_fields:
                request.session['email'] = update_fields['email']
            if 'first_name' in update_fields:
                request.session['prenom'] = update_fields['first_name']
            if 'last_name' in update_fields:
                request.session['nom'] = update_fields['last_name']
            request.session.modified = True

        # Profil étendu (téléphone, fonction)
        user_id_str = str(u_doc['_id'])
        db['admin_profiles'].update_one(
            {'user_id': user_id_str},
            {'$set': {
                'phone'     : (profile_data.get('phone') or '').strip(),
                'role'      : (profile_data.get('role')  or 'Administrateur Système').strip(),
                'updated_at': datetime.now(),
            }},
            upsert=True,
        )

    # ── 2. Sauvegarde configuration système ───────────────────────────────────
    allowed_keys = set(DEFAULTS.keys())
    clean_data = {k: v for k, v in body.items() if k in allowed_keys}
    clean_data['updated_at'] = datetime.now()
    clean_data['updated_by'] = request.session.get('username', '')

    _save_config(clean_data)

    return JsonResponse({
        'status' : 'success',
        'message': 'Configuration sauvegardée avec succès',
    })


@session_required
@require_POST
def api_parametres_reset(request):
    """Remet la configuration aux valeurs par défaut."""
    if not request.session.get('is_superuser', False):
        return JsonResponse({'status': 'error', 'message': 'Réservé au superadmin'}, status=403)
    _save_config(dict(DEFAULTS))
    return JsonResponse({'status': 'success', 'message': 'Configuration réinitialisée'})


@session_required
def api_system_info(request):
    """Retourne des infos système dynamiques."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'status': 'error'}, status=403)
    try:
        db.command('ping')
        db_status = 'Connecté'
    except Exception:
        db_status = 'Erreur'

    import django, sys
    return JsonResponse({
        'status'   : 'success',
        'django'   : django.get_version(),
        'python'   : sys.version.split()[0],
        'db_status': db_status,
        'timestamp': datetime.now().isoformat(),
    })

# ====================== API OCCUPATION ======================

@session_required
def api_occupation(request):
    bureaux = list(db.bureaux.find())
    result = []
    one_hour_ago = datetime.now() - timedelta(hours=1)
    for b in bureaux:
        recent = db.acces_logs.count_documents({'bureau_id': b['_id'], 'timestamp': {'$gte': one_hour_ago}})
        cap = b.get('capacite_max', 10)
        occ = min(recent * 3, cap)
        taux = round(occ / cap * 100) if cap else 0
        result.append({'id': str(b['_id']), 'nom': b['nom'], 'occupation': occ, 'capacite': cap, 'taux': taux})
    return JsonResponse({'bureaux': result})


@session_required
def api_bureau_stats(request, bureau_id):
    dates = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    acces = [random.randint(20, 90) for _ in range(7)]
    return JsonResponse({'dates': dates, 'acces_par_jour': acces})


# ====================== API LIVE FEED ======================
@session_required
def api_live_feed(request):
    """API JSON — flux live pour le refresh automatique"""
    import json
    from datetime import datetime, timedelta

    one_hour_ago = datetime.now() - timedelta(hours=1)
    today_start  = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    TYPES_SYSTEME = {'SYSTEM', 'URGENCE', 'ADMIN', 'EMERGENCY', 'LOCK', 'UNLOCK'}

    # ── Logs ─────────────────────────────────────────────
    raw_logs = list(db.acces_logs.find().sort('timestamp', -1).limit(20))
    logs_out = []
    for log in raw_logs:
        b = db.bureaux.find_one({'_id': log.get('bureau_id')})
        zone = b['nom'] if b else 'Zone inconnue'

        type_acces = log.get('type_acces', 'RFID')

        if type_acces in TYPES_SYSTEME or log.get('utilisateur_id') is None:
            nom = 'Système'
        else:
            e = db.employees.find_one({'_id': log.get('utilisateur_id')})
            if e:
                nom = f"{e.get('nom', '')} {e.get('prenom', '')}".strip() or 'Inconnu'
            else:
                nom = 'Inconnu'

        ts = log.get('timestamp')
        time_str = ts.strftime('%H:%M:%S') if ts else '--:--:--'

        logs_out.append({
            'nom':      nom,
            'zone':     zone,
            'method':   type_acces,
            'resultat': log.get('resultat', 'REFUSE'),
            'time':     time_str,
        })

    # ── Stats ─────────────────────────────────────────────
    acces_ok_hour = db.acces_logs.count_documents({'resultat': 'AUTORISE', 'timestamp': {'$gte': one_hour_ago}})
    acces_no_hour = db.acces_logs.count_documents({'resultat': 'REFUSE',   'timestamp': {'$gte': one_hour_ago}})
    total_hour = acces_ok_hour + acces_no_hour

    acces_ok_today = db.acces_logs.count_documents({'resultat': 'AUTORISE', 'timestamp': {'$gte': today_start}})
    acces_no_today = db.acces_logs.count_documents({'resultat': 'REFUSE',   'timestamp': {'$gte': today_start}})
    total_today = acces_ok_today + acces_no_today

    if total_hour > 0:
        taux = round(acces_ok_hour / total_hour * 100, 1)
    elif total_today > 0:
        taux = round(acces_ok_today / total_today * 100, 1)
    else:
        taux = 0

    alertes = 0
    if 'alertes' in db.list_collection_names():
        alertes = db.alertes.count_documents({'statut': 'NON_TRAITEE'})

    return JsonResponse({
        'logs': logs_out,
        'stats': {
            'acces_ok':    acces_ok_today,
            'acces_no':    acces_no_today,
            'taux_succes': taux,
            'alertes':     alertes,
        }
    })

# ====================== API DÉVERROUILLAGE D'URGENCE ======================

@session_required
@csrf_exempt
@require_http_methods(["POST"])
def api_emergency_unlock(request):
    """Journalise et déclenche un déverrouillage d'urgence."""
    try:
        db.acces_logs.insert_one({
            'type_acces':  'URGENCE',
            'resultat':    'AUTORISE',
            'message':     'Déverrouillage d\'urgence déclenché',
            'utilisateur': request.session.get('username', ''),
            'timestamp':   datetime.now(),
        })
        _creer_alerte(
            message=f"⚠️ DÉVERROUILLAGE D'URGENCE déclenché par {request.session.get('username', '')}",
            zone='SYSTÈME',
            niveau='CRITICAL'
        )
        logger.warning(f"[URGENCE] Déverrouillage déclenché par {request.session.get('username', '')}")
        return JsonResponse({'status': 'success', 'message': 'Déverrouillage d\'urgence activé et journalisé'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
 
 

# ====================== API STATISTIQUES TENDANCE ======================

@session_required
def api_stats_trend(request):
    days = min(int(request.GET.get('days', 30)), 365)
    now = datetime.now()
    labels, autorises_list, refuses_list, prediction_list = [], [], [], []
    for i in range(days - 1, -1, -1):
        day_start = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        a = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'AUTORISE'})
        r = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'REFUSE'})
        labels.append(day_start.strftime('%d/%m'))
        autorises_list.append(a)
        refuses_list.append(r)
    for i in range(len(autorises_list)):
        window = autorises_list[max(0, i - 6):i + 1]
        avg = sum(w * (j + 1) for j, w in enumerate(window)) / sum(range(1, len(window) + 1)) if window else 0
        prediction_list.append(round(avg * 1.05, 1))
    return JsonResponse({'labels': labels, 'autorises': autorises_list, 'refuses': refuses_list, 'prediction': prediction_list})


# ====================== API RÉSERVATIONS ACTIVES ======================

@session_required
def api_reservations_active(request):
    now = datetime.now()
    reservations_actives = list(db.reservations.find({
        'statut': 'confirmee',
        'date_debut': {'$lte': now},
        'date_fin': {'$gte': now},
    }).limit(20))
    result = []
    for r in reservations_actives:
        bureau = db.bureaux.find_one({'_id': r.get('bureau_id')})
        emp = db.employees.find_one({'_id': r.get('employe_id')})
        result.append({
            'id': str(r['_id']),
            'titre': r.get('titre', 'Réservation'),
            'bureau': bureau['nom'] if bureau else 'Inconnu',
            'employe': f"{emp.get('nom','?')} {emp.get('prenom','')}" if emp else 'Inconnu',
            'debut': r['date_debut'].isoformat() if r.get('date_debut') else '',
            'fin': r['date_fin'].isoformat() if r.get('date_fin') else '',
        })
    return JsonResponse({'reservations': result, 'total': len(result)})


# ====================== API TOP RESSOURCES ======================

@session_required
def api_resources_top(request):
    pipeline = [{'$group': {'_id': '$bureau_id', 'reservations': {'$sum': 1}}},
                {'$sort': {'reservations': -1}},
                {'$limit': 5}]
    top = list(db.reservations.aggregate(pipeline))
    total_reservations = db.reservations.count_documents({})
    result = []
    for t in top:
        b = db.bureaux.find_one({'_id': t['_id']})
        if b:
            result.append({
                'nom': b.get('nom', 'Inconnu'),
                'reservations': t['reservations'],
                'taux': round(t['reservations'] / total_reservations * 100) if total_reservations else 0,
            })
    return JsonResponse({'resources': result})


# ====================== GESTION DES ÉQUIPEMENTS ======================

@session_required
def equipement_list(request):
    equipements = list(db.equipements.find().sort('type', 1))
    nb_rfid = sum(1 for e in equipements if e.get('type') == 'RFID')
    nb_qr = sum(1 for e in equipements if e.get('type') == 'QR')
    nb_actifs = sum(1 for e in equipements if e.get('statut') == 'actif')
    nb_inactifs = len(equipements) - nb_actifs
    
    for equip in equipements:
        equip['id'] = str(equip['_id'])
        if equip.get('bureau_id'):
            bureau = db.bureaux.find_one({'_id': equip['bureau_id']})
            equip['bureau_nom'] = bureau['nom'] if bureau else 'Non assigné'
        else:
            equip['bureau_nom'] = 'Non assigné'
    
    yesterday = datetime.now() - timedelta(days=1)
    for equip in equipements:
        equip['logs_24h'] = db.acces_logs.count_documents({
            'equipement_code': equip.get('code'),
            'timestamp': {'$gte': yesterday}
        })
    
    return render(request, 'dashboard/equipement_list.html', {
        'equipements': equipements,
        'nb_total': len(equipements),
        'nb_rfid': nb_rfid,
        'nb_qr': nb_qr,
        'nb_actifs': nb_actifs,
        'nb_inactifs': nb_inactifs,
    })

@session_required
def equipement_detail(request, equipement_id):
    try:
        equipement = db.equipements.find_one({'_id': ObjectId(equipement_id)})
        if not equipement:
            messages.error(request, "Équipement non trouvé")
            return redirect('equipement_list')

        equipement['id'] = str(equipement['_id'])
        if equipement.get('bureau_id'):
            equipement['bureau'] = db.bureaux.find_one({'_id': equipement['bureau_id']})

        logs = list(db.acces_logs.find(
            {'equipement_code': equipement.get('code')}
        ).sort('timestamp', -1).limit(100))

        for log in logs:
            employe = db.employees.find_one({'_id': log.get('utilisateur_id')})
            log['nom_utilisateur'] = (
                f"{employe.get('nom', '')} {employe.get('prenom', '')}"
                if employe else 'Inconnu'
            )

        yesterday = datetime.now() - timedelta(days=1)
        week_ago  = datetime.now() - timedelta(days=7)

        logs_24h  = db.acces_logs.count_documents({
            'equipement_code': equipement.get('code'),
            'timestamp': {'$gte': yesterday}
        })
        logs_7j   = db.acces_logs.count_documents({
            'equipement_code': equipement.get('code'),
            'timestamp': {'$gte': week_ago}
        })
        autorises = db.acces_logs.count_documents({
            'equipement_code': equipement.get('code'),
            'resultat': 'AUTORISE'
        })
        refuses   = db.acces_logs.count_documents({
            'equipement_code': equipement.get('code'),
            'resultat': 'REFUSE'
        })

        total = autorises + refuses

        # Taux de succès / refus
        taux_succes = round(autorises / total * 100) if total > 0 else 0
        taux_refus  = round(refuses  / total * 100) if total > 0 else 0

        # Barre activité 24h : % par rapport à 150% de la moyenne journalière 7j
        moyenne_jour = logs_7j / 7 if logs_7j > 0 else 0
        if moyenne_jour > 0:
            activite_24h_pct = min(round(logs_24h / (moyenne_jour * 1.5) * 100), 100)
        else:
            activite_24h_pct = 100 if logs_24h > 0 else 0

        # Tendance 7j : part des logs récents sur le total historique
        logs_total  = db.acces_logs.count_documents({'equipement_code': equipement.get('code')})
        tendance_7j = min(round(logs_7j / logs_total * 100), 100) if logs_total > 0 else 0

        stats = {
            'logs_24h':         logs_24h,
            'logs_7j':          logs_7j,
            'autorises':        autorises,
            'refuses':          refuses,
            'taux_succes':      taux_succes,
            'taux_refus':       taux_refus,
            'activite_24h_pct': activite_24h_pct,
            'tendance_7j':      tendance_7j,
        }

        return render(request, 'dashboard/equipement_detail.html', {
            'equipement': equipement,
            'logs': logs,
            'stats': stats,
        })

    except Exception as e:
        messages.error(request, f"Erreur: {str(e)}")
        return redirect('equipement_list')
@session_required
def equipement_ajouter(request):
    bureaux = list(db.bureaux.find())
    for b in bureaux:
        b['id'] = str(b['_id'])
    if request.method == 'POST':
        try:
            equip_type = request.POST.get('type')
            prefix = 'RDR' if equip_type == 'RFID' else 'QR'
            count = db.equipements.count_documents({'type': equip_type}) + 1
            code = f"{prefix}-{str(count).zfill(3)}"
            equipement = {
                'nom': request.POST.get('nom'),
                'type': equip_type,
                'code': code,
                'emplacement': request.POST.get('emplacement'),
                'bureau_id': ObjectId(request.POST.get('bureau_id')) if request.POST.get('bureau_id') else None,
                'ip_address': request.POST.get('ip_address'),
                'port': int(request.POST.get('port', 5000)),
                'statut': request.POST.get('statut', 'actif'),
                'description': request.POST.get('description', ''),
                'created_at': datetime.now()
            }
            db.equipements.insert_one(equipement)
            messages.success(request, f"Équipement ajouté avec succès!")
            return redirect('equipement_list')
        except Exception as e:
            messages.error(request, f"Erreur lors de l'ajout: {str(e)}")
    return render(request, 'dashboard/equipement_form.html', {'bureaux': bureaux, 'equipement': {}, 'is_edit': False})


@session_required
def equipement_modifier(request, equipement_id):
    try:
        equipement = db.equipements.find_one({'_id': ObjectId(equipement_id)})
        if not equipement:
            messages.error(request, "Équipement non trouvé")
            return redirect('equipement_list')
        bureaux = list(db.bureaux.find())
        for b in bureaux:
            b['id'] = str(b['_id'])
        if request.method == 'POST':
            update_data = {
                'nom': request.POST.get('nom'),
                'emplacement': request.POST.get('emplacement'),
                'bureau_id': ObjectId(request.POST.get('bureau_id')) if request.POST.get('bureau_id') else None,
                'ip_address': request.POST.get('ip_address'),
                'port': int(request.POST.get('port', 5000)),
                'statut': request.POST.get('statut', 'actif'),
                'description': request.POST.get('description', ''),
                'updated_at': datetime.now()
            }
            db.equipements.update_one({'_id': ObjectId(equipement_id)}, {'$set': update_data})
            messages.success(request, "Équipement modifié avec succès!")
            return redirect('equipement_detail', equipement_id=equipement_id)
        equipement['id'] = str(equipement['_id'])
        return render(request, 'dashboard/equipement_form.html', {'equipement': equipement, 'bureaux': bureaux, 'is_edit': True})
    except Exception as e:
        messages.error(request, f"Erreur: {str(e)}")
        return redirect('equipement_list')


@session_required
def equipement_supprimer(request, equipement_id):
    if request.method == 'POST':
        try:
            db.equipements.update_one({'_id': ObjectId(equipement_id)},
                                      {'$set': {'statut': 'inactif', 'deleted_at': datetime.now()}})
            messages.success(request, "Équipement désactivé avec succès!")
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    return redirect('equipement_list')


@session_required
def equipement_tester(request, equipement_id):
    try:
        equipement = db.equipements.find_one({'_id': ObjectId(equipement_id)})
        if not equipement:
            return JsonResponse({'status': 'error', 'message': 'Équipement non trouvé'})
        response_time = random.randint(10, 50)
        db.equipements.update_one({'_id': ObjectId(equipement_id)},
                                  {'$set': {'derniere_connexion': datetime.now(), 'statut': 'actif'}})
        return JsonResponse({'status': 'success', 'message': 'Connexion réussie', 'response_time': response_time})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@session_required
def api_equipements(request):
    equipements = list(db.equipements.find({'statut': 'actif'}))
    resultats = []
    for eq in equipements:
        bureau = db.bureaux.find_one({'_id': eq.get('bureau_id')})
        resultats.append({
            'id': str(eq['_id']),
            'nom': eq.get('nom', ''),
            'type': eq.get('type', ''),
            'code': eq.get('code', ''),
            'emplacement': eq.get('emplacement', ''),
            'bureau_nom': bureau['nom'] if bureau else 'Non assigné',
            'ip_address': eq.get('ip_address', ''),
            'port': eq.get('port', 0),
            'statut': eq.get('statut', 'actif'),
            'derniere_connexion': eq.get('derniere_connexion'),
        })
    return JsonResponse({'equipements': resultats}, encoder=JSONEncoder)


@session_required
def api_equipement_logs(request, equipement_id):
    try:
        equipement = db.equipements.find_one({'_id': ObjectId(equipement_id)})
        if not equipement:
            return JsonResponse({'error': 'Équipement non trouvé'}, status=404)
        logs = list(db.acces_logs.find({'equipement_code': equipement.get('code')}).sort('timestamp', -1).limit(50))
        resultats = []
        for log in logs:
            employe = db.employees.find_one({'_id': log.get('utilisateur_id')})
            resultats.append({
                'id': str(log['_id']),
                'timestamp': log['timestamp'],
                'nom_utilisateur': f"{employe.get('nom', '')} {employe.get('prenom', '')}" if employe else 'Inconnu',
                'resultat': log.get('resultat', ''),
                'type_acces': log.get('type_acces', ''),
            })
        return JsonResponse({'logs': resultats}, encoder=JSONEncoder)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@session_required
def api_equipement_commande(request, equipement_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    try:
        data = json.loads(request.body)
        commande = data.get('commande')
        equipement = db.equipements.find_one({'_id': ObjectId(equipement_id)})
        if not equipement:
            return JsonResponse({'error': 'Équipement non trouvé'}, status=404)
        if 'commandes' not in db.list_collection_names():
            db.create_collection('commandes')
        db.commandes.insert_one({
            'equipement_id': ObjectId(equipement_id),
            'equipement_nom': equipement['nom'],
            'commande': commande,
            'statut': 'envoyee',
            'envoyee_par': request.session.get('username', ''),
            'timestamp': datetime.now()
        })
        return JsonResponse({'status': 'success', 'message': f'Commande "{commande}" envoyée à {equipement["nom"]}'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# ====================== RÉSERVATIONS ADMIN ======================
@session_required
def reservation_list(request):
    """Liste des réservations avec vue calendrier et tableau"""
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')
    
    from datetime import datetime, timedelta
    import json
    from bson import ObjectId
    
    reservations = list(db.reservations.find().sort('date_debut', -1))
    
    # Enrichir les données
    for r in reservations:
        r['id'] = str(r['_id'])
        
        # Employé
        employe_id = r.get('employe_id')
        if employe_id:
            try:
                if isinstance(employe_id, str):
                    emp = db.employees.find_one({'_id': ObjectId(employe_id)})
                else:
                    emp = db.employees.find_one({'_id': employe_id})
                if emp:
                    r['employe_nom'] = f"{emp.get('nom', '')} {emp.get('prenom', '')}".strip() or 'Inconnu'
                    r['employe_badge'] = emp.get('badge_id', '—')
                else:
                    r['employe_nom'] = 'Inconnu'
                    r['employe_badge'] = '—'
            except:
                r['employe_nom'] = 'Inconnu'
                r['employe_badge'] = '—'
        else:
            r['employe_nom'] = 'Inconnu'
            r['employe_badge'] = '—'
        
        # Bureau
        bureau_id = r.get('bureau_id')
        if bureau_id:
            try:
                if isinstance(bureau_id, str):
                    bureau = db.bureaux.find_one({'_id': ObjectId(bureau_id)})
                else:
                    bureau = db.bureaux.find_one({'_id': bureau_id})
                r['bureau_nom'] = bureau['nom'] if bureau else 'Salle inconnue'
            except:
                r['bureau_nom'] = 'Salle inconnue'
        else:
            r['bureau_nom'] = 'Salle inconnue'
        
        # QR code
        if 'qr_code' not in r:
            r['qr_code'] = None
    
    now = datetime.now()
    confirmees = sum(1 for r in reservations if r.get('statut') == 'confirmee')
    en_attente = sum(1 for r in reservations if r.get('statut') == 'en_attente')
    annulees = sum(1 for r in reservations if r.get('statut') == 'annulee')
    
    a_venir = sum(1 for r in reservations if r.get('statut') == 'confirmee' 
                  and r.get('date_debut') and r['date_debut'] > now)
    
    # Taux d'occupation
    total_bureaux = db.bureaux.count_documents({})
    if total_bureaux > 0:
        occupied_bureaux = set()
        for r in reservations:
            if r.get('statut') == 'confirmee' and r.get('date_debut') and r.get('date_fin'):
                if r['date_debut'] <= now <= r['date_fin']:
                    occupied_bureaux.add(str(r.get('bureau_id')))
        taux_occupation = round((len(occupied_bureaux) / total_bureaux) * 100)
    else:
        taux_occupation = 0
    
    # Convertir les données en JSON sécurisé
    reservations_list = []
    for r in reservations:
        if r.get('date_debut'):
            reservations_list.append({
                'id': str(r['_id']),
                'titre': r.get('titre', ''),
                'bureau_id': str(r.get('bureau_id')) if r.get('bureau_id') else None,
                'bureau_nom': r.get('bureau_nom', ''),
                'employe_nom': r.get('employe_nom', ''),
                'statut': r.get('statut', ''),
                'date_debut': r['date_debut'].isoformat() if r.get('date_debut') else None,
                'date_fin': r['date_fin'].isoformat() if r.get('date_fin') else None,
            })
    
    reservations_json = json.dumps(reservations_list, default=str)
    
    # Liste des bureaux
    bureaux_list = []
    for b in db.bureaux.find():
        bureaux_list.append({
            'id': str(b['_id']),
            'nom': b.get('nom', ''),
        })
    
    return render(request, 'dashboard/reservation_list.html', {
        'reservations': reservations,
        'total': len(reservations),
        'confirmees': confirmees,
        'en_attente': en_attente,
        'annulees': annulees,
        'a_venir': a_venir,
        'taux_occupation': taux_occupation,
        'reservations_json': reservations_json,
        'bureaux': bureaux_list,
    })
@session_required
def reservation_ajouter(request):
    """Ajouter une nouvelle réservation (ressources et matériel) avec workflow d'approbation et file d'attente"""
    from bson import ObjectId
    from datetime import datetime

    # Récupérer les bureaux/zones
    bureaux = list(db.bureaux.find())
    for b in bureaux:
        b['id'] = str(b['_id'])
        b['type'] = 'salle'
        b['type_icon'] = '🚪'

    # Récupérer le matériel
    materiels = list(db.materiels.find()) if 'materiels' in db.list_collection_names() else []
    for m in materiels:
        m['id'] = str(m['_id'])
        m['type'] = 'materiel'
        m['type_icon'] = get_materiel_icon(m.get('categorie', 'autre'))
        m['capacite_max'] = 1
        m['nom_affichage'] = f"{m['nom']} ({m.get('categorie', 'Matériel')})"

    # Récupérer les employés
    employes = list(db.employees.find({'statut': 'actif'}))
    for e in employes:
        e['id'] = str(e['_id'])

    # ── Paramètres de réservation (toujours chargés) ─────────────────────────
    params_resa  = db.parametres.find_one({'cle': 'reservation'}) or {}
    salle_min    = params_resa.get('salle_min_minutes',      30)
    salle_max    = params_resa.get('salle_max_minutes',    1440)
    materiel_min = params_resa.get('materiel_min_minutes',   60)
    materiel_max = params_resa.get('materiel_max_minutes', 525600)

    if request.method == 'POST':
        try:
            date_debut     = datetime.strptime(request.POST.get('date_debut'), '%Y-%m-%dT%H:%M')
            date_fin       = datetime.strptime(request.POST.get('date_fin'),   '%Y-%m-%dT%H:%M')
            resource_id    = request.POST.get('resource_id')
            resource_type  = request.POST.get('resource_type', 'salle')
            employe_id_str = request.POST.get('employe_id')
            join_queue     = request.POST.get('join_queue') == '1'
            flexible_minutes = int(request.POST.get('flexible_minutes', 120))

            form_ctx = {
                'bureaux': bureaux, 'materiels': materiels, 'employes': employes,
                'ressources': bureaux + materiels, 'reservation': request.POST, 'is_edit': False,
                'params_resa': params_resa,
            }

            # ── Validations de base ───────────────────────────────────────────
            if date_fin <= date_debut:
                messages.error(request, "La date de fin doit être après la date de début.")
                return render(request, 'dashboard/reservation_form.html', form_ctx)

            duree_minutes = int((date_fin - date_debut).total_seconds() / 60)

            if resource_type == 'salle':
                if duree_minutes < salle_min:
                    messages.error(request, f"Une salle ne peut pas être réservée moins de {salle_min} minute(s).")
                    return render(request, 'dashboard/reservation_form.html', form_ctx)
                if duree_minutes > salle_max:
                    h = salle_max // 60
                    messages.error(request, f"Une salle ne peut pas être réservée plus de {salle_max} minutes ({h}h maximum).")
                    return render(request, 'dashboard/reservation_form.html', form_ctx)
            else:
                if duree_minutes < materiel_min:
                    messages.error(request, f"Une ressource ne peut pas être réservée moins de {materiel_min} minute(s).")
                    return render(request, 'dashboard/reservation_form.html', form_ctx)
                if duree_minutes > materiel_max:
                    j = materiel_max // 1440
                    messages.error(request, f"Une ressource ne peut pas être réservée plus de {materiel_max} minutes ({j} jour(s) maximum).")
                    return render(request, 'dashboard/reservation_form.html', form_ctx)

            # ── Vérifier les conflits ─────────────────────────────────────────
            conflit = None
            try:
                if resource_type == 'salle':
                    conflit = db.reservations.find_one({
                        'bureau_id': ObjectId(resource_id),
                        'statut':    {'$in': ['confirmee', 'en_attente']},
                        'date_debut': {'$lt': date_fin},
                        'date_fin':   {'$gt': date_debut},
                    })
                else:
                    conflit = db.reservations.find_one({
                        'materiel_id': resource_id,
                        'statut':      {'$in': ['confirmee', 'en_attente']},
                        'date_debut':  {'$lt': date_fin},
                        'date_fin':    {'$gt': date_debut},
                    })
            except Exception as e:
                logger.warning(f"Vérification conflit échouée: {e}")

            if conflit:
                alternatives = []
                if join_queue:
                    try:
                        from dashboard.queue_service import QueueService
                        QueueService.ajouter(
                            user=get_session_user(request),
                            resource_type=resource_type,
                            resource_id=resource_id,
                            date_debut=date_debut,
                            date_fin=date_fin,
                            flexible_minutes=flexible_minutes,
                            employe_id=employe_id_str,
                            titre=request.POST.get('titre', '').strip(),
                        )
                        messages.success(request,
                            "Vous avez rejoint la file d'attente. Vous serez notifié dès qu'un créneau se libère.")
                        return redirect('reservation_list')
                    except Exception as e:
                        logger.warning(f"File d'attente échouée: {e}")
                        messages.error(request, f"Erreur file d'attente: {str(e)}")

                try:
                    from dashboard.queue_service import QueueService
                    alternatives = QueueService.proposer_alternatives(
                        resource_type=resource_type,
                        resource_id=resource_id,
                        date_debut=date_debut,
                        date_fin=date_fin,
                        flexible_minutes=flexible_minutes,
                    )
                except Exception:
                    alternatives = []

                messages.error(request, "Cette ressource est déjà réservée sur ce créneau.")
                return render(request, 'dashboard/reservation_form.html', {
                    **form_ctx,
                    'conflict':       True,
                    'alternatives':   alternatives,
                    'can_join_queue': True,
                })

            # ── Vérifier les indisponibilités (maintenance) ───────────────────
            try:
                if 'indisponibilites' in db.list_collection_names():
                    indispo = db.indisponibilites.find_one({
                        'ressource_type': resource_type,
                        'ressource_id':   ObjectId(resource_id),
                        'date_debut':     {'$lt': date_fin},
                        'date_fin':       {'$gt': date_debut},
                    })
                    if indispo:
                        messages.error(request,
                            f"Cette ressource est en maintenance du "
                            f"{indispo['date_debut'].strftime('%d/%m/%Y %H:%M')} au "
                            f"{indispo['date_fin'].strftime('%d/%m/%Y %H:%M')} "
                            f"({indispo.get('titre', 'maintenance planifiée')}).")
                        return render(request, 'dashboard/reservation_form.html', form_ctx)
            except Exception as e:
                logger.warning(f"Vérification indisponibilités échouée: {e}")

            # ── Préparer les données ──────────────────────────────────────────
            reservation_data = {
                'titre':           request.POST.get('titre', '').strip(),
                'description':     request.POST.get('description', '').strip(),
                'resource_type':   resource_type,
                'nb_participants': int(request.POST.get('nb_participants', 1)),
                'statut':          'en_attente',
                'created_at':      datetime.now(),
                'created_by':      request.session.get('username', ''),
                'date_debut':      date_debut,
                'date_fin':        date_fin,
            }

            employe = None
            if employe_id_str:
                try:
                    employe = db.employees.find_one({'_id': ObjectId(employe_id_str)})
                except Exception:
                    pass

            if employe:
                reservation_data['employe_nom'] = f"{employe.get('nom', '')} {employe.get('prenom', '')}".strip()

            if resource_type == 'salle':
                reservation_data['bureau_id']  = ObjectId(resource_id)
                reservation_data['employe_id'] = ObjectId(employe_id_str) if employe_id_str else None
                bureau = db.bureaux.find_one({'_id': ObjectId(resource_id)})
                reservation_data['bureau_nom'] = bureau['nom'] if bureau else 'Salle inconnue'
            else:
                reservation_data['materiel_id'] = resource_id
                reservation_data['employe_id']  = ObjectId(employe_id_str) if employe_id_str else None
                materiel = db.materiels.find_one({'_id': ObjectId(resource_id)})
                reservation_data['materiel_nom'] = materiel['nom'] if materiel else 'Matériel inconnu'

            # ── Insérer la réservation ────────────────────────────────────────
            result = db.reservations.insert_one(reservation_data)
            reservation_id = str(result.inserted_id)

            # ── Workflow d'approbation (facultatif, ne bloque pas) ────────────
            auto_approved = False
            try:
                from dashboard.approval_service import ApprovalService
                auto_approved = ApprovalService.creer_workflow(
                    reservation_id=reservation_id,
                    reservation_data=reservation_data,
                    user=get_session_user(request),
                )
            except Exception as e:
                import traceback
                logger.warning(
                    f"[WARN] approval workflow: {type(e).__name__}: {e}\n"
                    f"{traceback.format_exc()}"
                )

            if auto_approved:
                db.reservations.update_one(
                    {'_id': result.inserted_id},
                    {'$set': {
                        'statut':        'confirmee',
                        'auto_approved': True,
                        'approved_at':   datetime.now(),
                    }}
                )
                reservation_data['statut'] = 'confirmee'

            # ── Notification employé (PyMongo direct) ─────────────────────────
            try:
                if employe_id_str:
                    db.notifications.insert_one({
                        'employe_id':     employe_id_str,
                        'titre':          '📝 Réservation créée',
                        'message':        (
                            f"Votre réservation '{reservation_data['titre']}' a été créée "
                            f"et est en attente de validation."
                        ),
                        'categorie':      'reservation',
                        'icon':           '📝',
                        'status':         'non_lu',
                        'action_url':     '/employe/reservations/',
                        'reservation_id': reservation_id,
                        'created_at':     datetime.now(),
                    })
            except Exception as e:
                logger.warning(f"Notification employé échouée: {e}")

            # ── Notifications admins (PyMongo direct, sans ORM) ───────────────
            try:
                ressource_label = (
                    f"🚪 Salle: {reservation_data.get('bureau_nom', '')}"
                    if resource_type == 'salle'
                    else f"📦 Matériel: {reservation_data.get('materiel_nom', '')}"
                )
                emp_prenom = employe.get('prenom', '') if employe else ''
                emp_nom    = employe.get('nom', '')    if employe else ''

                admins = list(db['utilisateurs'].find(
                    {'is_staff': True, 'is_active': True},
                    {'_id': 1, 'email': 1, 'username': 1}
                ))

                for admin in admins:
                    db.admin_notifications.insert_one({
                        'admin_id':       admin.get('_id'),
                        'titre':          '🆕 Nouvelle réservation en attente',
                        'message':        (
                            f"{emp_prenom} {emp_nom} a demandé une réservation "
                            f"pour '{reservation_data['titre']}' "
                            f"({ressource_label}) le {date_debut.strftime('%d/%m/%Y à %H:%M')}."
                        ),
                        'categorie':      'reservation',
                        'icon':           '🆕',
                        'status':         'non_lu',
                        'action_url':     f'/reservations/{reservation_id}/',
                        'reservation_id': reservation_id,
                        'created_at':     datetime.now(),
                    })
                    if admin.get('email'):
                        try:
                            from dashboard.utils_email import envoyer_email
                            envoyer_email(
                                admin['email'],
                                f"🆕 Nouvelle réservation — {reservation_data['titre']}",
                                (
                                    f"Nouvelle réservation de {emp_prenom} {emp_nom}\n"
                                    f"{ressource_label}\n"
                                    f"Le {date_debut.strftime('%d/%m/%Y à %H:%M')}\n"
                                    f"Lien: /reservations/{reservation_id}/"
                                ),
                            )
                        except Exception as _ee:
                            logger.warning(f"Email admin échoué: {_ee}")
            except Exception as e:
                logger.warning(f"Notifications admins échouées: {e}")

            if auto_approved:
                messages.success(request, "Réservation créée et approuvée automatiquement ✅")
            else:
                messages.success(request,
                    "Réservation créée. Elle est en attente d'approbation par le responsable de zone.")
            return redirect('reservation_list')

        except Exception as e:
            import traceback
            logger.exception("Erreur création réservation admin")
            messages.error(request, f"Erreur: {str(e)}")

    return render(request, 'dashboard/reservation_form.html', {
        'bureaux':    bureaux,
        'materiels':  materiels,
        'employes':   employes,
        'ressources': bureaux + materiels,
        'reservation': {},
        'is_edit':    False,
        'params_resa': params_resa,
    })

def get_materiel_icon(categorie):
    """Retourne l'icône correspondant à la catégorie du matériel"""
    icons = {
        'informatique': '💻',
        'mobilier': '🪑',
        'audiovisuel': '📽️',
        'imprimante': '🖨️',
        'securite': '🔒',
        'vehicule': '🚗',
        'outillage': '🔧',
        'autre': '📦'
    }
    return icons.get(categorie, '📦')

@session_required
def reservation_modifier(request, reservation_id):
    try:
        reservation = db.reservations.find_one({'_id': ObjectId(reservation_id)})
        if not reservation:
            messages.error(request, "Réservation non trouvée")
            return redirect('reservation_list')
        reservation['id'] = str(reservation['_id'])
        if reservation.get('date_debut'):
            reservation['date_debut_str'] = reservation['date_debut'].strftime('%Y-%m-%dT%H:%M')
        if reservation.get('date_fin'):
            reservation['date_fin_str'] = reservation['date_fin'].strftime('%Y-%m-%dT%H:%M')
        
        bureaux = list(db.bureaux.find())
        for b in bureaux:
            b['id'] = str(b['_id'])
        employes = list(db.employees.find({'statut': 'actif'}))
        for e in employes:
            e['id'] = str(e['_id'])
        
        if request.method == 'POST':
            date_debut = datetime.strptime(request.POST.get('date_debut'), '%Y-%m-%dT%H:%M')
            date_fin = datetime.strptime(request.POST.get('date_fin'), '%Y-%m-%dT%H:%M')
            db.reservations.update_one(
                {'_id': ObjectId(reservation_id)},
                {'$set': {
                    'titre': request.POST.get('titre', '').strip(),
                    'description': request.POST.get('description', '').strip(),
                    'bureau_id': ObjectId(request.POST.get('bureau_id')),
                    'employe_id': ObjectId(request.POST.get('employe_id')),
                    'date_debut': date_debut,
                    'date_fin': date_fin,
                    'nb_participants': int(request.POST.get('nb_participants', 1)),
                    'statut': request.POST.get('statut', 'confirmee'),
                    'updated_at': datetime.now(),
                }}
            )
            messages.success(request, "Réservation modifiée!")
            return redirect('reservation_list')
        return render(request, 'dashboard/reservation_form.html',
                      {'reservation': reservation, 'bureaux': bureaux, 'employes': employes, 'is_edit': True})
    except Exception as e:
        messages.error(request, f"Erreur: {str(e)}")
        return redirect('reservation_list')

@session_required
def reservation_annuler(request, reservation_id):
    """Annuler une réservation et notifier la file d'attente"""
    from bson import ObjectId
    from datetime import datetime
    from dashboard.queue_service import QueueService

    if request.method == 'POST':
        # Récupérer la réservation AVANT annulation (pour connaître la ressource libérée)
        reservation = db.reservations.find_one({'_id': ObjectId(reservation_id)})

        # Annuler
        db.reservations.update_one(
            {'_id': ObjectId(reservation_id)},
            {'$set': {
                'statut': 'annulee',
                'cancelled_at': datetime.now(),
                'cancelled_by': request.session.get('username', ''),
            }}
        )

        # ── NOUVEAU: notifier la file d'attente ──────────────────────────────
        if reservation:
            try:
                resource_type = reservation.get('resource_type', 'salle')
                if resource_type == 'salle':
                    resource_id = str(reservation.get('bureau_id'))
                else:
                    resource_id = str(reservation.get('materiel_id'))

                notified = QueueService.notifier_prochain(
                    resource_type=resource_type,
                    resource_id=resource_id,
                    date_debut=reservation.get('date_debut'),
                    date_fin=reservation.get('date_fin'),
                )
                if notified:
                    messages.info(request,
                        f"{notified} personne(s) en file d'attente ont été notifiées.")
            except Exception as e:
                print(f"[WARN] queue notify: {e}")

        messages.success(request, "Réservation annulée.")
    return redirect('reservation_list')

# ═══════════════════════════════════════════════════════════════
#  ADMIN — EXPORT CSV RÉSERVATIONS
# ═══════════════════════════════════════════════════════════════
import csv
from django.http import HttpResponse
from datetime import datetime

@session_required
def admin_reservations_export_csv(request):
    if not request.session.get('is_staff', False) and not request.session.get('is_superuser', False):
        return HttpResponse("Accès non autorisé.", status=403)

    reservations = list(db.reservations.find().sort('date_debut', -1))

    # Enrichir avec le nom de l'employé
    emp_cache = {}
    for r in reservations:
        eid = str(r.get('employe_id', ''))
        if eid and eid not in emp_cache:
            try:
                emp = db.employees.find_one({'_id': ObjectId(eid)})
                emp_cache[eid] = f"{emp.get('nom','')} {emp.get('prenom','')}" if emp else '—'
            except Exception:
                emp_cache[eid] = eid
        r['_emp_nom'] = emp_cache.get(eid, '—')

    now_str  = datetime.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="Admin_Reservations_{now_str}.csv"'
    response.write('\ufeff')  # BOM UTF-8 pour Excel

    writer = csv.writer(response)
    writer.writerow([
        'Titre', 'Employé', 'Salle / Ressource', 'Type',
        'Début', 'Fin', 'Participants', 'Statut', 'Description', 'Créé le'
    ])

    for r in reservations:
        debut   = r.get('date_debut')
        fin     = r.get('date_fin')
        created = r.get('created_at')
        writer.writerow([
            r.get('titre', ''),
            r.get('_emp_nom', '—'),
            r.get('bureau_nom', r.get('materiel_nom', '—')),
            r.get('resource_type', 'salle'),
            debut.strftime('%d/%m/%Y %H:%M')   if hasattr(debut,   'strftime') else '—',
            fin.strftime('%d/%m/%Y %H:%M')     if hasattr(fin,     'strftime') else '—',
            r.get('nb_participants', 1),
            r.get('statut', ''),
            r.get('description', ''),
            created.strftime('%d/%m/%Y %H:%M') if hasattr(created, 'strftime') else '—',
        ])

    return response


# ═══════════════════════════════════════════════════════════════
#  ADMIN — EXPORT PDF RÉSERVATIONS
# ═══════════════════════════════════════════════════════════════
@session_required
def admin_reservations_export_pdf(request):
    if not request.session.get('is_staff', False) and not request.session.get('is_superuser', False):
        return HttpResponse("Accès non autorisé.", status=403)

    import io, os, traceback
    from datetime import datetime
    from django.contrib.staticfiles.finders import find as static_find
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, HRFlowable, KeepTogether,
    )
    from reportlab.platypus.flowables import Flowable
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.pdfgen import canvas as rl_canvas

    try:
        C_BLUE_DARK  = colors.HexColor('#0f172a')
        C_BLUE_MAIN  = colors.HexColor('#1d4ed8')
        C_BLUE_LIGHT = colors.HexColor('#dbeafe')
        C_BLUE_MID   = colors.HexColor('#3b82f6')
        C_PURPLE     = colors.HexColor('#7c3aed')
        C_GREEN      = colors.HexColor('#059669')
        C_AMBER      = colors.HexColor('#d97706')
        C_RED        = colors.HexColor('#dc2626')
        C_GREY_LIGHT = colors.HexColor('#f8fafc')
        C_GREY_MID   = colors.HexColor('#e2e8f0')
        C_GREY_TEXT  = colors.HexColor('#64748b')
        C_WHITE      = colors.white
        C_BLACK      = colors.HexColor('#0f172a')

        class NumberedCanvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved_page_states = []
            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                n = len(self._saved_page_states)
                for i, state in enumerate(self._saved_page_states):
                    self.__dict__.update(state)
                    self._draw_footer(i + 1, n)
                    rl_canvas.Canvas.showPage(self)
                rl_canvas.Canvas.save(self)
            def _draw_footer(self, page_num, total_pages):
                w, _ = A4
                self.setFillColor(colors.HexColor('#64748b'))
                self.setFont('Helvetica', 8)
                self.drawCentredString(w / 2, 1.2 * cm,
                    f"Page {page_num} / {total_pages}  —  SIGR-CA — Document confidentiel")
                self.setStrokeColor(colors.HexColor('#e2e8f0'))
                self.setLineWidth(0.5)
                self.line(2*cm, 1.6*cm, w - 2*cm, 1.6*cm)

        class ColorBand(Flowable):
            def __init__(self, text, width, height=0.9*cm, bg=None, fg=None, font_size=11):
                super().__init__()
                self.text = text; self.band_width = width; self.band_height = height
                self.bg = bg or colors.HexColor('#0f172a')
                self.fg = fg or colors.white; self.font_size = font_size
            def wrap(self, *args):
                return self.band_width, self.band_height
            def draw(self):
                c = self.canv
                c.setFillColor(self.bg)
                c.rect(0, 0, self.band_width, self.band_height, fill=1, stroke=0)
                c.setFillColor(self.fg)
                c.setFont('Helvetica-Bold', self.font_size)
                c.drawString(0.4*cm, 0.25*cm, self.text)

        def hex_color(c):
            try:
                return f'{int(c.red*255):02x}{int(c.green*255):02x}{int(c.blue*255):02x}'
            except Exception:
                return '0f172a'

        # ── Données : TOUTES les réservations ────────────────────
        reservations = list(db.reservations.find().sort('date_debut', -1))

        emp_cache = {}
        for r in reservations:
            eid = str(r.get('employe_id', ''))
            if eid and eid not in emp_cache:
                try:
                    emp = db.employees.find_one({'_id': ObjectId(eid)})
                    emp_cache[eid] = f"{emp.get('nom','')} {emp.get('prenom','')}" if emp else '—'
                except Exception:
                    emp_cache[eid] = '—'
            r['_emp_nom'] = emp_cache.get(eid, '—')

        now_str   = datetime.now().strftime('%d/%m/%Y à %H:%M')
        date_file = datetime.now().strftime('%Y%m%d_%H%M')
        admin_nom = request.session.get('full_name', request.session.get('username', 'Admin'))

        buffer   = io.BytesIO()
        PAGE_W, _ = A4
        CONTENT_W = PAGE_W - 4*cm

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2.2*cm, bottomMargin=2.2*cm,
            title="Rapport Réservations — Admin SIGR-CA",
            author="SIGR-CA Système",
        )

        _sty = {}
        def ps(name, **kw):
            base  = kw.pop('parent', 'Normal')
            sheet = getSampleStyleSheet()
            parent = _sty.get(base) or sheet.get(base, sheet['Normal'])
            p = ParagraphStyle(name, parent=parent, **kw)
            _sty[name] = p
            return p

        th      = ps('TH',  fontName='Helvetica-Bold', fontSize=8, textColor=C_WHITE, alignment=TA_CENTER, leading=11)
        th_left = ps('THL', parent='TH', alignment=TA_LEFT)
        td      = ps('TD',  fontSize=8, alignment=TA_CENTER, leading=11, textColor=C_BLACK)
        td_left = ps('TDL', parent='TD', alignment=TA_LEFT)
        td_mono = ps('TDM', parent='TD', fontSize=7, fontName='Courier')
        sec_sty = ps('SEC', fontName='Helvetica-Bold', fontSize=12,
                     textColor=C_PURPLE, spaceBefore=20, spaceAfter=8, leading=16)
        foot_sty = ps('FOT', fontSize=7.5, textColor=C_GREY_TEXT, alignment=TA_CENTER, leading=11)

        _NO_PAD = TableStyle([
            ('LEFTPADDING',  (0,0),(-1,-1), 0),
            ('RIGHTPADDING', (0,0),(-1,-1), 0),
            ('TOPPADDING',   (0,0),(-1,-1), 1),
            ('BOTTOMPADDING',(0,0),(-1,-1), 1),
        ])

        elements = []

        # ── En-tête ───────────────────────────────────────────────
        LOGO_PATH = static_find('img/logo.png')

        title_tbl = Table([
            [Paragraph('<font color="#1d4ed8"><b>SIGR-CA</b></font>',
                       ps('LT', fontSize=22, leading=26, alignment=TA_LEFT))],
            [Paragraph("Système Intégré de Gestion des Ressources<br/>"
                       "<font color='#64748b'>et de Contrôle d'Accès</font>",
                       ps('LS', fontSize=9, leading=13, textColor=C_GREY_TEXT, alignment=TA_LEFT))],
        ], colWidths=[11*cm])
        title_tbl.setStyle(_NO_PAD)

        logo_cell = (Image(LOGO_PATH, width=4*cm, height=2.6*cm)
                     if LOGO_PATH and os.path.exists(LOGO_PATH)
                     else Paragraph('<font color="#1d4ed8"><b>SIGR</b></font>',
                                    ps('FL', fontSize=18, alignment=TA_RIGHT)))

        meta_tbl = Table([
            [Paragraph(f"<b>Généré par :</b> {admin_nom}",
                       ps('M0', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph(f"<b>Date :</b> {now_str}",
                       ps('M1', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph(f"<b>Total réservations :</b> {len(reservations)}",
                       ps('M2', fontSize=8, textColor=C_GREY_TEXT, alignment=TA_RIGHT))],
            [Paragraph("<b>CONFIDENTIEL</b>",
                       ps('M3', fontSize=8, textColor=C_RED, alignment=TA_RIGHT))],
        ], colWidths=[4.5*cm])
        meta_tbl.setStyle(_NO_PAD)

        right_col = Table([[logo_cell], [meta_tbl]], colWidths=[4.5*cm])
        right_col.setStyle(TableStyle([
            ('ALIGN',        (0,0),(-1,-1), 'RIGHT'),
            ('LEFTPADDING',  (0,0),(-1,-1), 0),
            ('RIGHTPADDING', (0,0),(-1,-1), 0),
            ('TOPPADDING',   (0,0),(-1,-1), 2),
            ('BOTTOMPADDING',(0,0),(-1,-1), 2),
        ]))

        header_tbl = Table([[title_tbl, right_col]], colWidths=[12.5*cm, 4.5*cm])
        header_tbl.setStyle(TableStyle([
            ('VALIGN',       (0,0),(-1,-1), 'TOP'),
            ('LEFTPADDING',  (0,0),(-1,-1), 0),
            ('RIGHTPADDING', (0,0),(-1,-1), 0),
            ('TOPPADDING',   (0,0),(-1,-1), 0),
            ('BOTTOMPADDING',(0,0),(-1,-1), 6),
        ]))

        elements.append(header_tbl)
        elements.append(HRFlowable(width='100%', thickness=2, color=C_BLUE_MAIN, spaceAfter=4))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(ColorBand(
            "  RAPPORT RÉSERVATIONS — ADMINISTRATION",
            CONTENT_W, height=1.1*cm,
            bg=C_PURPLE, fg=C_WHITE, font_size=12,
        ))
        elements.append(Spacer(1, 0.6*cm))

        # ── Tableau principal ─────────────────────────────────────
        elements.append(KeepTogether([Paragraph('Liste complète des réservations', sec_sty)]))

        STATUT_MAP = {
            'confirmee':  ('#059669', '✔ Confirmée'),
            'en_attente': ('#d97706', '⏳ En attente'),
            'annulee':    ('#dc2626', '✖ Annulée'),
            'terminee':   ('#64748b', '■ Terminée'),
        }

        # colWidths : 3.6+3.0+2.8+2.6+2.6+1.2+1.2 = 17cm ✓
        col_widths = [3.6*cm, 3.0*cm, 2.8*cm, 2.6*cm, 2.6*cm, 1.2*cm, 1.2*cm]
        data = [[
            Paragraph('Titre',     th_left),
            Paragraph('Employé',   th_left),
            Paragraph('Salle',     th),
            Paragraph('Début',     th),
            Paragraph('Fin',       th),
            Paragraph('Part.',     th),
            Paragraph('Statut',    th),
        ]]

        for r in reservations:
            statut = r.get('statut', '')
            color, label = STATUT_MAP.get(statut, ('#64748b', statut))
            bureau = r.get('bureau_nom') or r.get('materiel_nom') or '—'
            debut  = r.get('date_debut')
            fin    = r.get('date_fin')
            data.append([
                Paragraph(f"<b>{r.get('titre') or '—'}</b>", td_left),
                Paragraph(str(r.get('_emp_nom', '—')), td_left),
                Paragraph(str(bureau), td),
                Paragraph(debut.strftime('%d/%m %H:%M') if hasattr(debut,'strftime') else '—', td_mono),
                Paragraph(fin.strftime('%d/%m %H:%M')   if hasattr(fin,  'strftime') else '—', td_mono),
                Paragraph(str(r.get('nb_participants', 1)), td),
                Paragraph(f'<font color="{color}"><b>{label.split()[0]}</b></font>', td),
            ])

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  C_PURPLE),
            ('TEXTCOLOR',     (0,0), (-1,0),  C_WHITE),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [C_WHITE, C_GREY_LIGHT]),
            ('GRID',          (0,0), (-1,-1), 0.4, C_GREY_MID),
            ('LINEBELOW',     (0,0), (-1,0),  1.5, C_PURPLE),
            ('TOPPADDING',    (0,0), (-1,0),  9),
            ('BOTTOMPADDING', (0,0), (-1,0),  9),
            ('TOPPADDING',    (0,1), (-1,-1), 6),
            ('BOTTOMPADDING', (0,1), (-1,-1), 6),
            ('LEFTPADDING',   (0,0), (-1,-1), 6),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('BOX',           (0,0), (-1,-1), 1, C_PURPLE),
        ]))
        elements.append(tbl)

        # ── Résumé statistique ────────────────────────────────────
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph('Résumé statistique', sec_sty))

        total      = len(reservations)
        confirmees = sum(1 for r in reservations if r.get('statut') == 'confirmee')
        en_attente = sum(1 for r in reservations if r.get('statut') == 'en_attente')
        annulees   = sum(1 for r in reservations if r.get('statut') == 'annulee')
        terminees  = sum(1 for r in reservations if r.get('statut') == 'terminee')
        salles     = sum(1 for r in reservations if r.get('resource_type', 'salle') == 'salle')
        mats       = sum(1 for r in reservations if r.get('resource_type') == 'materiel')
        employes_u = len(set(str(r.get('employe_id','')) for r in reservations if r.get('employe_id')))

        def stat_row(label, value, value_color=C_BLACK):
            hv = hex_color(value_color)
            return [
                Paragraph(label, ps(f'SL{label[:5]}', fontSize=9, textColor=C_GREY_TEXT, alignment=TA_LEFT)),
                Paragraph(f'<font color="#{hv}"><b>{value}</b></font>',
                          ps(f'SV{label[:5]}', fontSize=10, alignment=TA_RIGHT, fontName='Helvetica-Bold')),
            ]

        stats_data = [
            stat_row('Total réservations',      str(total)),
            stat_row('Confirmées',              str(confirmees),  C_GREEN),
            stat_row('En attente de validation',str(en_attente),  C_AMBER),
            stat_row('Annulées',                str(annulees),    C_RED),
            stat_row('Terminées',               str(terminees),   C_GREY_TEXT),
            stat_row('Réservations de salles',  str(salles)),
            stat_row('Réservations de matériels', str(mats)),
            stat_row('Employés ayant réservé',  str(employes_u),  C_PURPLE),
        ]

        stats_tbl = Table(stats_data, colWidths=[10*cm, 7*cm])
        stats_tbl.setStyle(TableStyle([
            ('ROWBACKGROUNDS',(0,0),(-1,-1), [C_WHITE, C_GREY_LIGHT]),
            ('GRID',          (0,0),(-1,-1), 0.4, C_GREY_MID),
            ('TOPPADDING',    (0,0),(-1,-1), 7),
            ('BOTTOMPADDING', (0,0),(-1,-1), 7),
            ('LEFTPADDING',   (0,0),(-1,-1), 10),
            ('RIGHTPADDING',  (0,0),(-1,-1), 10),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ('BOX',           (0,0),(-1,-1), 1, C_GREY_MID),
        ]))
        elements.append(stats_tbl)

        # ── Pied de page document ─────────────────────────────────
        elements.append(Spacer(1, 1.2*cm))
        elements.append(HRFlowable(width='100%', thickness=1, color=C_GREY_MID, spaceAfter=6))
        elements.append(Paragraph(
            f"Rapport généré le <b>{now_str}</b> par <b>{admin_nom}</b> via le système SIGR-CA.",
            foot_sty))
        elements.append(Paragraph(
            "Ce document est <b>confidentiel</b> et destiné à un usage interne uniquement.",
            ps('CF2', fontSize=7, textColor=C_GREY_TEXT, alignment=TA_CENTER, leading=10)))

        doc.build(elements, canvasmaker=NumberedCanvas)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Admin_Reservations_{date_file}.pdf"'
        return response

    except Exception as e:
        return HttpResponse(
            f"Erreur PDF : {str(e)}\n\n{traceback.format_exc()}",
            content_type='text/plain', status=500)
    


# ====================== API HISTORIQUE EMPLOYÉ ======================
@session_required
def api_employee_history(request, employe_id):
    """API pour récupérer l'historique d'un employé"""
    from bson import ObjectId
    from datetime import datetime
    
    try:
        # Essayer de convertir l'ID en ObjectId
        try:
            emp_id = ObjectId(employe_id)
        except:
            # Si ce n'est pas un ObjectId valide, chercher par string
            emp_id = employe_id
        
        # Chercher l'employé
        employe = db.employees.find_one({'_id': emp_id})
        if not employe and isinstance(emp_id, str):
            employe = db.employees.find_one({'_id': ObjectId(emp_id)})
        
        if not employe:
            return JsonResponse({'error': 'Employé non trouvé', 'total_acces': 0, 'logs': []}, status=404)
        
        # Récupérer les logs
        logs = list(db.acces_logs.find({'utilisateur_id': employe['_id']}).sort('timestamp', -1).limit(100))
        
        logs_data = []
        for log in logs:
            # Récupérer le nom du bureau
            bureau_nom = 'Inconnu'
            if log.get('bureau_id'):
                try:
                    bureau = db.bureaux.find_one({'_id': log['bureau_id']})
                    if bureau:
                        bureau_nom = bureau.get('nom', 'Inconnu')
                except:
                    pass
            
            logs_data.append({
                'date': log['timestamp'].strftime('%d/%m/%Y %H:%M:%S') if log.get('timestamp') else '',
                'zone': bureau_nom,
                'resultat': log.get('resultat', ''),
            })
        
        return JsonResponse({
            'total_acces': len(logs),
            'logs': logs_data
        })
        
    except Exception as e:
        logger.error(f"Erreur dans api_employee_history: {str(e)}")
        return JsonResponse({
            'error': str(e),
            'total_acces': 0,
            'logs': []
        }, status=500)
# ====================== API PROFIL ADMIN ======================

@session_required
def update_admin_profile(request):
    if not request.session.get('is_staff', False) and not request.session.get('is_superuser', False):
        return JsonResponse({'status': 'error', 'message': 'Accès non autorisé'}, status=403)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user = get_session_user(request)
            user.username = data.get('username', user.username)
            user.first_name = data.get('first_name', user.first_name)
            user.last_name = data.get('last_name', user.last_name)
            user.email = data.get('email', user.email)
            
            current_password = data.get('current_password')
            new_password = data.get('new_password')
            if new_password:
                if not current_password:
                    return JsonResponse({'status': 'error', 'message': 'Mot de passe actuel requis'})
                if not check_password(current_password, user.password):
                    return JsonResponse({'status': 'error', 'message': 'Mot de passe actuel incorrect'})
                if len(new_password) < 6:
                    return JsonResponse({'status': 'error', 'message': 'Le mot de passe doit contenir au moins 6 caractères'})
                user.password = make_password(new_password)
            
            phone = data.get('phone')
            if phone:
                db.admin_profiles.update_one({'user_id': user.id},
                                             {'$set': {'phone': phone, 'updated_at': datetime.now()}}, upsert=True)
            user.save()
            db.system_logs.insert_one({
                'user_id': user.id,
                'username': user.username,
                'action': 'PROFILE_UPDATE',
                'timestamp': datetime.now(),
                'ip': request.META.get('REMOTE_ADDR')
            })
            return JsonResponse({'status': 'success', 'message': 'Profil mis à jour avec succès'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée'}, status=405)


@session_required
def admin_login_history(request):
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    logs = list(db.system_logs.find({'user_id': request.session.get('user_id', '')}).sort('timestamp', -1).limit(50))
    history = [{'timestamp': log['timestamp'].isoformat() if log.get('timestamp') else '',
                'ip_address': log.get('ip', '—'),
                'user_agent': log.get('user_agent', '—'),
                'success': True} for log in logs]
    return JsonResponse({'history': history}, encoder=JSONEncoder)


@session_required
def update_admin_avatar(request):
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            avatar_base64 = data.get('avatar')
            if avatar_base64:
                db.admin_profiles.update_one({'user_id': request.session.get('user_id', '')},
                                             {'$set': {'avatar': avatar_base64, 'updated_at': datetime.now()}}, upsert=True)
                return JsonResponse({'status': 'success', 'message': 'Avatar mis à jour'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée'}, status=405)


# ====================== GESTION DES RESSOURCES ======================



# ====================== GESTION DES RESSOURCES (SUITE) ======================

@session_required
@require_http_methods(["POST"])
def api_materiel_upload_photo(request):
    import base64, uuid
    from django.conf import settings
    try:
        data = json.loads(request.body)
        photo_data = data.get('photo_data', '')
        if not photo_data:
            return JsonResponse({'status': 'error', 'message': 'Pas de photo'}, status=400)
        
        if ',' in photo_data:
            photo_data = photo_data.split(',')[1]
        
        img_bytes = base64.b64decode(photo_data)
        filename = f"materiels/{uuid.uuid4().hex}.jpg"
        filepath = os.path.join(settings.MEDIA_ROOT, filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'wb') as f:
            f.write(img_bytes)
        
        return JsonResponse({'status': 'success', 'path': filename})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
@session_required
def bureau_detail(request, bureau_id):
    """Détail d'un bureau/zone avec affichage de la hiérarchie"""
    from bson import ObjectId
    from datetime import datetime, timedelta
    
    try:
        bureau = db.bureaux.find_one({'_id': ObjectId(bureau_id)})
        if not bureau:
            messages.error(request, "Zone non trouvée")
            return redirect('ressources')
        
        bureau['id'] = str(bureau['_id'])
        
        # Récupérer la hiérarchie complète
        hierarchie = []
        if bureau.get('etage_id'):
            etage = db.etages.find_one({'_id': bureau['etage_id']})
            if etage:
                hierarchie.append(('Étage', etage.get('nom')))
                if etage.get('batiment_id'):
                    batiment = db.batiments.find_one({'_id': etage['batiment_id']})
                    if batiment:
                        hierarchie.append(('Bâtiment', batiment.get('nom')))
                        if batiment.get('site_id'):
                            site = db.sites.find_one({'_id': batiment['site_id']})
                            if site:
                                hierarchie.append(('Site', site.get('nom')))
                                if site.get('domaine_id'):
                                    domaine = db.domainesp.find_one({'_id': site['domaine_id']})
                                    if domaine:
                                        hierarchie.append(('Domaine', domaine.get('nom')))
        
        # Statistiques d'occupation
        one_hour_ago = datetime.now() - timedelta(hours=1)
        occupation_recente = db.acces_logs.count_documents({
            'bureau_id': ObjectId(bureau_id),
            'timestamp': {'$gte': one_hour_ago}
        })
        
        capacite = bureau.get('capacite_max', 10)
        taux_occupation = min(100, round((occupation_recente * 3 / capacite) * 100)) if capacite > 0 else 0
        
        # Historique des 7 derniers jours
        historique = []
        for i in range(6, -1, -1):
            day_start = (datetime.now() - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            count = db.acces_logs.count_documents({
                'bureau_id': ObjectId(bureau_id),
                'timestamp': {'$gte': day_start, '$lt': day_end}
            })
            historique.append({
                'date': day_start.strftime('%d/%m'),
                'acces': count
            })
        
        return render(request, 'dashboard/bureau_detail.html', {
            'bureau': bureau,
            'hierarchie': hierarchie,
            'taux_occupation': taux_occupation,
            'historique': historique,
            'capacite': capacite,
            'occupation_recente': occupation_recente
        })
        
    except Exception as e:
        messages.error(request, f"Erreur: {str(e)}")
        return redirect('ressources')


@session_required
def api_bureau_reservations(request, bureau_id):
    """API pour les réservations d'un bureau (simplifié)"""
    try:
        reservations = list(db.reservations.find({
            'bureau_id': ObjectId(bureau_id),
            'statut': 'confirmee',
            'date_debut': {'$gte': datetime.now()}
        }).sort('date_debut', 1).limit(5))
        
        resultats = []
        for r in reservations:
            employe = db.employees.find_one({'_id': r.get('employe_id')})
            resultats.append({
                'id': str(r['_id']),
                'debut': r['date_debut'].isoformat(),
                'fin': r['date_fin'].isoformat(),
                'employe': f"{employe.get('nom', '')} {employe.get('prenom', '')}".strip() if employe else 'Inconnu'
            })
        
        return JsonResponse({'reservations': resultats})
    except Exception as e:
        return JsonResponse({'reservations': [], 'error': str(e)})


@session_required
def api_materiel_list(request):
    """API pour la liste du matériel"""
    materiels = list(db.materiels.find()) if 'materiels' in db.list_collection_names() else []
    for m in materiels:
        m['id'] = str(m['_id'])
    return JsonResponse({'materiels': materiels})


@session_required
def api_materiel_supprimer(request, materiel_id):
    """API pour supprimer du matériel"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        result = db.materiels.delete_one({'_id': ObjectId(materiel_id)})
        if result.deleted_count > 0:
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Matériel non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# ====================== API RESSOURCES ======================

@session_required
def api_resources(request):
    ressources = list(db.resources.find({'statut': {'$ne': 'hors_service'}}))
    result = []
    for r in ressources:
        result.append({
            'id': str(r['_id']),
            'nom': r.get('nom', ''),
            'categorie': r.get('categorie', ''),
            'description': r.get('description', ''),
            'photo': r.get('photo', ''),
            'localisation': r.get('localisation', ''),
            'capacite': r.get('capacite', 1),
            'statut': r.get('statut', 'disponible'),
            'disponible': r.get('statut') == 'disponible',
        })
    return JsonResponse({'resources': result})


# ====================== GESTION DES RÉSERVATIONS AVANCÉES ======================

@session_required
def reservation_ajouter_avance(request):
    bureaux = list(db.bureaux.find())
    for b in bureaux:
        b['id'] = str(b['_id'])
    
    resources = list(db.resources.find({'statut': 'disponible'}))
    for r in resources:
        r['id'] = str(r['_id'])
    
    employes = list(db.employees.find({'statut': 'actif'}))
    for e in employes:
        e['id'] = str(e['_id'])
    
    if request.method == 'POST':
        try:
            date_debut = datetime.strptime(request.POST.get('date_debut'), '%Y-%m-%dT%H:%M')
            date_fin = datetime.strptime(request.POST.get('date_fin'), '%Y-%m-%dT%H:%M')
            resource_id = request.POST.get('resource_id')
            resource_type = request.POST.get('resource_type', 'salle')
            employe_id = request.POST.get('employe_id')
            recurrence = request.POST.get('recurrence', 'none')
            recurrence_end = request.POST.get('recurrence_end')
            
            if date_fin <= date_debut:
                messages.error(request, "La date de fin doit être après la date de début")
                return render(request, 'dashboard/reservation_form_avance.html', {
                    'bureaux': bureaux, 'resources': resources, 'employes': employes,
                })
            
            conflit = db.reservations.find_one({
                'resource_id': resource_id,
                'statut': {'$in': ['confirmee', 'en_attente']},
                'date_debut': {'$lt': date_fin},
                'date_fin': {'$gt': date_debut},
            })
            
            if conflit:
                suggestions = suggest_alternative_slots(resource_id, date_debut, date_fin)
                messages.warning(request, "Conflit détecté ! Suggestions disponibles.")
                return render(request, 'dashboard/reservation_form_avance.html', {
                    'bureaux': bureaux, 'resources': resources, 'employes': employes,
                    'suggestions': suggestions, 'form_data': request.POST,
                })
            
            employe = db.employees.find_one({'_id': ObjectId(employe_id)})
            reservation_data = {
                'titre': request.POST.get('titre', 'Réservation'),
                'description': request.POST.get('description', ''),
                'resource_id': resource_id,
                'resource_type': resource_type,
                'bureau_id': request.POST.get('bureau_id'),
                'employe_id': employe_id,
                'employe_nom': f"{employe.get('nom', '')} {employe.get('prenom', '')}" if employe else '',
                'date_debut': date_debut,
                'date_fin': date_fin,
                'nb_participants': int(request.POST.get('nb_participants', 1)),
                'statut': 'confirmee',
                'recurrence': recurrence if recurrence != 'none' else '',
                'created_by': request.session.get('username', ''),
                'created_at': datetime.now(),
            }
            
            if recurrence_end:
                reservation_data['recurrence_end'] = datetime.strptime(recurrence_end, '%Y-%m-%d')
            
            result = db.reservations.insert_one(reservation_data)
            send_reservation_notification(employe_id, reservation_data)
            messages.success(request, "Réservation créée avec succès !")
            return redirect('reservation_list')
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    return render(request, 'dashboard/reservation_form_avance.html', {
        'bureaux': bureaux,
        'resources': resources,
        'employes': employes,
    })

# dashboard/views.py - Remplacer l'ancienne fonction

from dashboard.ai_suggestions import get_suggestion_engine

def suggest_alternative_slots_advanced(resource_id, date_debut, date_fin, employe_id=None):
    """Version avancée avec IA"""
    engine = get_suggestion_engine(db)
    return engine.suggest_alternative_slots_advanced(resource_id, date_debut, date_fin, employe_id)


def send_reservation_notification(employe_id, reservation_data):
    employe = db.employees.find_one({'_id': ObjectId(employe_id)})
    if not employe:
        return
    
    message = f"""
    Bonjour {employe.get('prenom', '')} {employe.get('nom', '')},
    
    Votre réservation a été confirmée :
    - Titre: {reservation_data.get('titre')}
    - Date: {reservation_data['date_debut'].strftime('%d/%m/%Y %H:%M')} → {reservation_data['date_fin'].strftime('%H:%M')}
    
    Merci d'utiliser vos accès avec votre badge RFID.
    
    SIGR-CA
    """
    
    db.notifications.insert_one({
        'destinataire': employe.get('email', ''),
        'type_notification': 'email',
        'categorie': 'confirmation',
        'sujet': f"Confirmation de réservation - {reservation_data.get('titre')}",
        'message': message,
        'statut': 'envoyee',
        'reservation_id': str(reservation_data.get('_id')),
        'created_at': datetime.now(),
    })
    
    if employe.get('email'):
        try:
            send_mail(
                f"Confirmation de réservation - {reservation_data.get('titre')}",
                message,
                settings.DEFAULT_FROM_EMAIL,
                [employe['email']],
                fail_silently=True,
            )
        except:
            pass


# ====================== CONTRÔLE D'ACCÈS PHYSIQUE ======================

@csrf_exempt
def api_verify_access(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        badge_id = data.get('badge_id')
        zone_code = data.get('zone_code')
        access_method = data.get('method', 'RFID')
        
        employe = db.employees.find_one({'badge_id': badge_id, 'statut': 'actif'})
        
        if not employe:
            log_access(None, zone_code, 'REFUSE', 'Badge inconnu', access_method)
            
            # Notifier les admins d'une tentative avec badge inconnu
            from dashboard.views import notify_admins_security_alert
            notify_admins_security_alert(zone_code, badge_id, "Tentative d'accès avec badge non reconnu")
            
            return JsonResponse({'autorise': False, 'message': 'Badge non reconnu'})
        
        now = datetime.now()
        reservation_valide = db.reservations.find_one({
            'employe_id': str(employe['_id']),
            'statut': 'confirmee',
            'date_debut': {'$lte': now},
            'date_fin': {'$gte': now},
        })
        
        zone = db.bureaux.find_one({'code': zone_code})
        if not zone:
            zone = db.bureaux.find_one({'nom': zone_code})
        
        access_rule = db.access_rules.find_one({
            'employe_id': str(employe['_id']),
            'zone_nom': zone.get('nom', zone_code) if zone else zone_code,
            'jour': now.day,
            'mois': now.month,
            'annee': now.year,
        })
        
        current_hour = now.strftime('%H:%M')
        acces_autorise = False
        motif_refus = ""
        
        if access_rule:
            if access_rule.get('acces_autorise', True):
                if access_rule.get('heure_debut', '00:00') <= current_hour <= access_rule.get('heure_fin', '23:59'):
                    acces_autorise = True
                else:
                    motif_refus = "Horaire non autorisé"
            else:
                motif_refus = "Règle d'accès restreinte"
        elif reservation_valide:
            acces_autorise = True
        else:
            motif_refus = "Aucune réservation active"
        
        emergency = db.system_config.find_one({'type': 'emergency'})
        if emergency and emergency.get('active', False):
            acces_autorise = True
        
        # Log l'accès
        log_access(employe['_id'], zone_code, 'AUTORISE' if acces_autorise else 'REFUSE',
                  'Accès ' + ('autorisé' if acces_autorise else 'refusé'), access_method)
        
        # Si accès refusé, notifier les admins (après 3 refus dans la même heure)
        if not acces_autorise:
            # Compter les refus récents
            one_hour_ago = now - timedelta(hours=1)
            recent_refus = db.acces_logs.count_documents({
                'utilisateur_id': employe['_id'],
                'resultat': 'REFUSE',
                'timestamp': {'$gte': one_hour_ago}
            })
            
            if recent_refus >= 3:
                from dashboard.views import notify_admins_security_alert
                zone_nom = zone.get('nom', zone_code) if zone else zone_code
                notify_admins_security_alert(
                    zone_nom, 
                    badge_id, 
                    f"Tentatives d'accès multiples refusées ({recent_refus} fois en 1h) - {motif_refus}"
                )
        
        return JsonResponse({
            'autorise': acces_autorise,
            'message': 'Accès autorisé' if acces_autorise else f'Accès refusé: {motif_refus}',
            'employe_nom': f"{employe.get('nom', '')} {employe.get('prenom', '')}",
        })
        
    except Exception as e:
        return JsonResponse({'autorise': False, 'error': str(e)})


def log_access(utilisateur_id, zone_code, resultat, message, method):
    db.acces_logs.insert_one({
        'utilisateur_id': utilisateur_id,
        'bureau_code': zone_code,
        'resultat': resultat,
        'message': message,
        'type_acces': method,
        'timestamp': datetime.now(),
    })


# ====================== NOTIFICATIONS ET ALERTES ======================

@session_required
def api_send_notification(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        destinataire = data.get('destinataire')
        type_notif = data.get('type', 'email')
        categorie = data.get('categorie', 'info')
        sujet = data.get('sujet', 'Notification SIGR-CA')
        message = data.get('message', '')
        
        notification = {
            'destinataire': destinataire,
            'type_notification': type_notif,
            'categorie': categorie,
            'sujet': sujet,
            'message': message,
            'statut': 'envoyee',
            'created_at': datetime.now(),
        }
        
        db.notifications.insert_one(notification)
        
        if type_notif == 'email' and destinataire:
            try:
                send_mail(sujet, message, settings.DEFAULT_FROM_EMAIL, [destinataire])
            except:
                notification['statut'] = 'echouee'
                db.notifications.update_one({'_id': notification['_id']}, {'$set': {'statut': 'echouee'}})
        
        return JsonResponse({'status': 'success', 'message': 'Notification envoyée'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_alerts(request):
    alerts = list(db.alertes.find({'statut': 'NON_TRAITEE'}).sort('timestamp', -1).limit(50))
    result = []
    for a in alerts:
        result.append({
            'id': str(a['_id']),
            'type': a.get('type', 'ALERT'),
            'message': a.get('message', ''),
            'timestamp': a.get('timestamp'),
            'zone': a.get('zone', ''),
        })
    return JsonResponse({'alerts': result})


# ====================== STATISTIQUES AVANCÉES ======================

@session_required
def api_stats_predictions(request):
    today = datetime.now()
    predictions = []
    for i in range(1, 8):
        pred_date = today + timedelta(days=i)
        history = db.acces_logs.count_documents({
            'timestamp': {'$gte': pred_date - timedelta(days=7), '$lt': pred_date}
        })
        predicted = int(history / 7 * (0.9 + (i * 0.02)))
        predictions.append({
            'date': pred_date.strftime('%Y-%m-%d'),
            'predicted_access': predicted,
            'confidence': min(95, 70 + i * 3),
        })
    
    zones_stats = list(db.acces_logs.aggregate([
        {'$group': {'_id': '$bureau_code', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}},
        {'$limit': 5}
    ]))
    
    conflict_zones = []
    for z in zones_stats:
        conflict_zones.append({
            'zone': z['_id'],
            'activity': z['count'],
            'risk': 'high' if z['count'] > 100 else 'medium' if z['count'] > 50 else 'low',
        })
    
    return JsonResponse({
        'predictions': predictions,
        'conflict_zones': conflict_zones,
        'recommendations': [
            "Les zones Atelier et Direction sont très sollicitées en début de semaine",
            "Optimisez les créneaux du mercredi après-midi (moins d'affluence)",
            "Prévoyez des ressources supplémentaires pour la zone Production",
        ]
    })

# ====================== HELPERS PARTAGÉS ======================
from django.utils.timezone import make_aware, is_naive
import pytz

def _get_db_algiers():
    from django.conf import settings
    db      = settings.MONGO_DB
    algiers = pytz.timezone(settings.TIME_ZONE)
    return db, algiers

def _make_aware_dt(dt, algiers):
    if dt is None:
        return None
    return make_aware(dt, algiers) if is_naive(dt) else dt

def _normalize_uid(raw_uid):
    """Retourne uid_key — int si possible, sinon str."""
    if not raw_uid:
        return 0, True
    try:
        return int(raw_uid), True
    except (ValueError, TypeError):
        return str(raw_uid), False

def _get_users_map(db, user_ids):
    """
    Construit un dict  id → doc utilisateur.
    Supporte les anciens user_id entiers ET les nouveaux ObjectId string.
    """
    if not user_ids:
        return {}

    int_ids = []
    str_ids = []
    for uid in user_ids:
        if not uid:
            continue
        try:
            int_ids.append(int(uid))
        except (ValueError, TypeError):
            str_ids.append(str(uid))

    users_map = {}

    if int_ids:
        for u in db['utilisateurs'].find({'id': {'$in': int_ids}}):
            users_map[u['id']] = u

    if str_ids:
        from bson import ObjectId
        oids = []
        for s in str_ids:
            try:
                oids.append(ObjectId(s))
            except Exception:
                pass
        if oids:
            for u in db['utilisateurs'].find({'_id': {'$in': oids}}):
                users_map[str(u['_id'])] = u

    return users_map

def _get_employee_photo(db, uid):
    emp = db['employees'].find_one({'django_user_id': uid}, {'avatar': 1, 'photo': 1})
    if not emp:
        return None
    return emp.get('avatar') or emp.get('photo')


def _dedup_sessions(sessions_data):
    """
    Supprime les doublons par session_key (le middleware peut insérer
    plusieurs documents pour la même session Django).
    Garde le document avec la last_activity la plus récente.
    """
    best = {}   # session_key → doc
    for s in sessions_data:
        key = s.get('session_key') or str(s.get('_id', ''))
        if not key:
            continue
        existing = best.get(key)
        if existing is None:
            best[key] = s
        else:
            # Garde le plus récent
            cur_ts  = s.get('last_activity')
            prev_ts = existing.get('last_activity')
            if cur_ts and (prev_ts is None or cur_ts > prev_ts):
                best[key] = s
    return list(best.values())


# ─────────────────────────────────────────────────────────────
@session_required
def active_sessions(request):
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')

    from types import SimpleNamespace
    db, algiers = _get_db_algiers()

    raw_sessions = list(
        db['dashboard_usersession'].find({
            'is_active': True,
            '$or': [
                {'logout_time': None},
                {'logout_time': {'$exists': False}},
            ]
        }).sort('last_activity', -1)
    )

    # ── Dédupliquer par session_key ──────────────────────────
    sessions_data = _dedup_sessions(raw_sessions)

    user_ids  = [s.get('user_id') for s in sessions_data if s.get('user_id')]
    users_map = _get_users_map(db, user_ids)

    active_sessions_list = []
    for s in sessions_data:
        uid, _     = _normalize_uid(s.get('user_id'))
        user_data  = users_map.get(uid, {})
        login_time_aware    = _make_aware_dt(s.get('login_time'), algiers)
        last_activity_aware = _make_aware_dt(s.get('last_activity'), algiers)
        fn       = user_data.get('first_name', '')
        ln       = user_data.get('last_name', '')
        fullname = f"{fn} {ln}".strip() or user_data.get('username', '?')

        user = SimpleNamespace(
            id=str(uid),
            username=user_data.get('username', '?'),
            first_name=fn,
            last_name=ln,
            email=user_data.get('email', ''),
            is_staff=bool(user_data.get('is_staff', False)),
            get_full_name=lambda n=fullname: n,
        )
        session = SimpleNamespace(
            id=str(s.get('_id', '')),
            user=user,
            session_key=s.get('session_key', ''),
            ip_address=s.get('ip_address', ''),
            device_type=s.get('device_type', 'desktop'),
            location=s.get('location', ''),
            user_agent=s.get('user_agent', ''),
            login_time=login_time_aware,
            last_activity=last_activity_aware,
            logout_time=_make_aware_dt(s.get('logout_time'), algiers),
            is_active=s.get('is_active', True),
            get_duration=lambda lt=login_time_aware: str(timezone.now() - lt).split('.')[0] if lt else '—',
        )
        active_sessions_list.append(session)

    now                = timezone.now()
    inactive_threshold = now - timedelta(minutes=30)
    total_connected    = len(active_sessions_list)
    admin_sessions     = sum(1 for s in active_sessions_list if s.user.is_staff)
    employee_sessions  = total_connected - admin_sessions
    inactive_sessions  = sum(1 for s in active_sessions_list if s.last_activity and s.last_activity < inactive_threshold)
    desktop_sessions   = sum(1 for s in active_sessions_list if s.device_type == 'desktop')
    mobile_sessions    = sum(1 for s in active_sessions_list if s.device_type == 'mobile')
    tablet_sessions    = sum(1 for s in active_sessions_list if s.device_type == 'tablet')
    total_users        = db['utilisateurs'].count_documents({'is_active': True})

    # ── Historique 24h ────────────────────────────────────────
    last_24h    = now - timedelta(hours=24)
    raw_history = list(
        db['dashboard_sessionlog'].find({'timestamp': {'$gte': last_24h}})
        .sort('timestamp', -1).limit(100)
    )

    log_int_ids = []
    log_str_ids = []
    for l in raw_history:
        uid = l.get('user_id')
        if not uid:
            continue
        try:
            log_int_ids.append(int(uid))
        except (ValueError, TypeError):
            log_str_ids.append(str(uid))

    log_users_map = {}
    if log_int_ids:
        for u in db['utilisateurs'].find({'id': {'$in': log_int_ids}}):
            log_users_map[u['id']] = u
    if log_str_ids:
        from bson import ObjectId
        oids = []
        for s in log_str_ids:
            try:
                oids.append(ObjectId(s))
            except Exception:
                pass
        if oids:
            for u in db['utilisateurs'].find({'_id': {'$in': oids}}):
                log_users_map[str(u['_id'])] = u

    from types import SimpleNamespace
    recent_history = []
    for log in raw_history:
        raw_uid  = log.get('user_id')
        uid, _   = _normalize_uid(raw_uid)
        u_data   = log_users_map.get(uid, {})
        fn       = u_data.get('first_name', '')
        ln       = u_data.get('last_name', '')
        fullname = f"{fn} {ln}".strip() or u_data.get('username', '?')
        user_ns  = SimpleNamespace(
            username=u_data.get('username', '?'),
            get_full_name=lambda n=fullname: n,
        )
        recent_history.append(SimpleNamespace(
            user=user_ns,
            action=log.get('action', ''),
            ip_address=log.get('ip_address', ''),
            timestamp=_make_aware_dt(log.get('timestamp'), algiers),
        ))

    return render(request, 'dashboard/active_sessions.html', {
        'active_sessions':   active_sessions_list,
        'total_connected':   total_connected,
        'total_users':       total_users,
        'admin_sessions':    admin_sessions,
        'employee_sessions': employee_sessions,
        'inactive_sessions': inactive_sessions,
        'desktop_sessions':  desktop_sessions,
        'mobile_sessions':   mobile_sessions,
        'tablet_sessions':   tablet_sessions,
        'recent_history':    recent_history,
        'now':               now,
    })


# ─────────────────────────────────────────────────────────────
@session_required
def api_connected_users(request):
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    db, algiers = _get_db_algiers()

    raw_sessions = list(
        db['dashboard_usersession'].find({
            'is_active': True,
            '$or': [
                {'logout_time': None},
                {'logout_time': {'$exists': False}},
            ]
        }).sort('last_activity', -1)
    )

    # ── Dédupliquer par session_key avant tout traitement ────
    sessions_data = _dedup_sessions(raw_sessions)

    user_ids  = [s.get('user_id') for s in sessions_data if s.get('user_id')]
    users_map = _get_users_map(db, user_ids)

    now                = timezone.now()
    inactive_threshold = now - timedelta(minutes=30)
    result             = []

    for s in sessions_data:
        uid, _     = _normalize_uid(s.get('user_id'))
        user_data  = users_map.get(uid, {})
        login_time    = _make_aware_dt(s.get('login_time'), algiers)
        last_activity = _make_aware_dt(s.get('last_activity'), algiers)

        if login_time:
            total_s  = int((now - login_time).total_seconds())
            duration = f"{total_s // 3600}h {(total_s % 3600) // 60}m"
        else:
            duration = '—'

        if last_activity:
            is_inactive = last_activity < inactive_threshold
            minutes_ago = int((now - last_activity).total_seconds() / 60)
        else:
            is_inactive = True
            minutes_ago = None

        full_name = f"{user_data.get('first_name', '')} {user_data.get('last_name', '')}".strip()
        avatar    = _get_employee_photo(db, uid)

        result.append({
            'id':            str(s['_id']),
            'user_id':       str(uid),
            'username':      user_data.get('username', '?'),
            'full_name':     full_name or user_data.get('username', '?'),
            'is_staff':      bool(user_data.get('is_staff', False)),
            'is_superuser':  bool(user_data.get('is_superuser', False)),
            'avatar':        avatar,
            'session_key':   s.get('session_key', ''),
            'ip_address':    s.get('ip_address', ''),
            'device_type':   s.get('device_type', 'desktop'),
            'location':      s.get('location', ''),
            'user_agent':    s.get('user_agent', ''),
            'login_time':    login_time.isoformat()    if login_time    else None,
            'last_activity': last_activity.isoformat() if last_activity else None,
            'duration':      duration,
            'is_inactive':   is_inactive,
            'minutes_ago':   minutes_ago,
            'is_current':    s.get('session_key') == request.session.session_key,
        })

    stats = {
        'total':     len(result),
        'admins':    sum(1 for s in result if s['is_staff']),
        'employees': sum(1 for s in result if not s['is_staff']),
        'inactive':  sum(1 for s in result if s['is_inactive']),
        'desktop':   sum(1 for s in result if s['device_type'] == 'desktop'),
        'mobile':    sum(1 for s in result if s['device_type'] == 'mobile'),
        'tablet':    sum(1 for s in result if s['device_type'] == 'tablet'),
    }

    return JsonResponse({'sessions': result, 'stats': stats, 'now': now.isoformat()})


# ─────────────────────────────────────────────────────────────
@session_required
def terminate_session(request, session_id):
    if request.method != 'POST':
        return redirect('active_sessions' if request.session.get('is_staff', False) else 'employe_profil')

    db, algiers = _get_db_algiers()

    try:
        from bson import ObjectId
        oid     = ObjectId(session_id)
        session = db['dashboard_usersession'].find_one({'_id': oid})

        if not session:
            messages.error(request, "Session non trouvée.")
            return redirect('active_sessions' if request.session.get('is_staff', False) else 'employe_profil')

        session_user_id = session.get('user_id')
        if not request.session.get('is_staff', False):
            if str(session_user_id) != str(request.session.get('user_id', '')):
                messages.error(request, "Vous ne pouvez pas terminer la session d'un autre utilisateur.")
                return redirect('employe_profil')

        now = timezone.now()

        try:
            uid = int(session_user_id) if session_user_id is not None else 0
            db['dashboard_sessionlog'].insert_one({
                'user_id':       uid,
                'action':        'terminated',
                'terminated_by': request.session.get('user_id', ''),
                'ip_address':    session.get('ip_address', ''),
                'session_key':   session.get('session_key', ''),
                'timestamp':     now,
            })
        except Exception:
            pass

        # Désactiver TOUS les documents portant ce session_key (nettoie les doublons)
        sk = session.get('session_key', '')
        if sk:
            db['dashboard_usersession'].update_many(
                {'session_key': sk},
                {'$set': {'is_active': False, 'logout_time': now}}
            )
            try:
                db['django_session'].delete_one({'session_key': sk})
            except Exception:
                pass
        else:
            db['dashboard_usersession'].update_one(
                {'_id': oid},
                {'$set': {'is_active': False, 'logout_time': now}}
            )

        messages.success(request, "Session terminée avec succès.")

    except Exception as e:
        messages.error(request, f"Erreur : {e}")

    return redirect('active_sessions' if request.session.get('is_staff', False) else 'employe_profil')


# ─────────────────────────────────────────────────────────────
@session_required
def terminate_all_sessions(request):
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    if request.method != 'POST':
        return redirect('active_sessions')

    db, algiers = _get_db_algiers()
    current_key = request.session.session_key
    now         = timezone.now()

    other_sessions = list(
        db['dashboard_usersession'].find({
            'is_active': True,
            '$or': [
                {'logout_time': None},
                {'logout_time': {'$exists': False}},
            ],
            'session_key': {'$ne': current_key},
        })
    )

    # Collecter les session_keys uniques à terminer
    keys_to_terminate = set()
    for session in other_sessions:
        uid, _ = _normalize_uid(session.get('user_id'))
        db['dashboard_sessionlog'].insert_one({
            'user_id':     uid,
            'action':      'terminated',
            'ip_address':  session.get('ip_address', ''),
            'session_key': session.get('session_key', ''),
            'timestamp':   now,
        })
        sk = session.get('session_key', '')
        if sk:
            keys_to_terminate.add(sk)
        else:
            db['dashboard_usersession'].update_one(
                {'_id': session['_id']},
                {'$set': {'is_active': False, 'logout_time': now}}
            )

    # Désactiver tous les documents de ces session_keys (y compris doublons)
    if keys_to_terminate:
        db['dashboard_usersession'].update_many(
            {'session_key': {'$in': list(keys_to_terminate)}},
            {'$set': {'is_active': False, 'logout_time': now}}
        )
        for sk in keys_to_terminate:
            try:
                db['django_session'].delete_one({'session_key': sk})
            except Exception:
                pass

    messages.success(request, f"{len(keys_to_terminate)} session(s) terminée(s)")
    return redirect('active_sessions')


# ─────────────────────────────────────────────────────────────
@session_required
def api_session_stats(request):
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    db, algiers    = _get_db_algiers()
    now            = timezone.now()
    today_start    = now.replace(hour=0, minute=0, second=0, microsecond=0)
    seven_days_ago = now - timedelta(days=7)

    today_logins = db['dashboard_sessionlog'].count_documents({
        'action':    'login',
        'timestamp': {'$gte': today_start},
    })
    week_logins = db['dashboard_sessionlog'].count_documents({
        'action':    'login',
        'timestamp': {'$gte': seven_days_ago},
    })
    avg_logins = round(week_logins / 7, 1)

    device_stats = {
        d: db['dashboard_usersession'].count_documents({
            'is_active':   True,
            'logout_time': None,
            'device_type': d,
        })
        for d in ('desktop', 'mobile', 'tablet')
    }

    return JsonResponse({
        'today_logins': today_logins,
        'avg_logins':   avg_logins,
        'device_stats': device_stats,
    })


# ─────────────────────────────────────────────────────────────
@session_required
def clear_session_history(request):
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

    db, algiers = _get_db_algiers()
    days        = int(request.POST.get('days', 30))
    cutoff      = timezone.now() - timedelta(days=days)

    result_logs     = db['dashboard_sessionlog'].delete_many({'timestamp': {'$lt': cutoff}})
    result_sessions = db['dashboard_usersession'].delete_many({
        'is_active':   False,
        'logout_time': {'$lt': cutoff},
    })

    total = result_logs.deleted_count + result_sessions.deleted_count
    messages.success(request, f"{total} entrée(s) supprimée(s) (antérieures à {days} jours)")
    return redirect('active_sessions')

# ====================== ESPACE EMPLOYÉ - PROFIL ======================
# ====================== ESPACE EMPLOYÉ - PROFIL ======================
# ── Helpers profil employé (utilisent _check_password existant ligne ~125) ──

def _get_user_hash(user):
    """
    Récupère le hash du mot de passe pour un MongoUser.
    Cherche dans dashboard_utilisateur puis utilisateurs, par username.
    """
    # Essai 1 : attribut direct sur MongoUser
    try:
        if hasattr(user, 'password') and user.password:
            return user.password
    except Exception:
        pass
    # Essai 2 : dashboard_utilisateur (anciens comptes Django)
    try:
        doc = db['utilisateurs'].find_one({'username': user.username})
        if doc and doc.get('password'):
            return doc['password']
    except Exception:
        pass
    # Essai 3 : utilisateurs (nouveaux comptes)
    try:
        doc = db['utilisateurs'].find_one({'username': user.username})
        if doc and doc.get('password'):
            return doc['password']
    except Exception:
        pass
    return None


def _save_user_password(user, new_password, request=None):
    """
    Sauvegarde le nouveau mot de passe en bcrypt (même format que l'existant).
    Met à jour dashboard_utilisateur ET utilisateurs par username.
    Si request fourni, maintient la session active.
    Retourne True si succès.
    """
    import bcrypt as _bcrypt
    from django.contrib.auth import update_session_auth_hash

    new_hash = _bcrypt.hashpw(
        new_password.encode('utf-8'),
        _bcrypt.gensalt(12)
    ).decode('utf-8')

    updated = 0
    try:
        updated += db['utilisateurs'].update_one(
            {'username': user.username},
            {'$set': {'password': new_hash}}
        ).modified_count
    except Exception as _e:
        logger.warning(f"_save_user_password dashboard_utilisateur échoué: {_e}")

    if not updated:
        try:
            db['utilisateurs'].update_one(
                {'username': user.username},
                {'$set': {'password': new_hash}}
            )
        except Exception as _e:
            logger.warning(f"_save_user_password utilisateurs échoué: {_e}")

    if request is not None:
        try:
            update_session_auth_hash(request, user)
        except Exception:
            pass
    return True


def _update_django_user(username, **fields):
    """
    Met à jour les champs du user Django via PyMongo (par username).
    Remplace user.save() partout dans le projet.
    Exemple: _update_django_user(request.session.get('username', ''), first_name='Jean')
    """
    if not fields:
        return
    for collection in ('dashboard_utilisateur', 'utilisateurs'):
        try:
            db[collection].update_one({'username': username}, {'$set': fields})
        except Exception as _e:
            logger.warning(f"_update_django_user [{collection}] échoué: {_e}")
@session_required
def employe_profil(request):
    """Modification du profil employé"""
    if request.session.get('is_staff', False):
        return redirect('dashboard')

    from datetime import datetime

    # Récupérer l'employé
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})
    if not employe:
        messages.error(request, "Profil employé introuvable.")
        return redirect('login')

    employe['id'] = str(employe['_id'])
    utilisateur_id = employe['_id']

    # === STATISTIQUES ===
    total_acces     = db.acces_logs.count_documents({'utilisateur_id': utilisateur_id})
    acces_autorises = db.acces_logs.count_documents({'utilisateur_id': utilisateur_id, 'resultat': 'AUTORISE'})
    acces_refuses   = total_acces - acces_autorises
    taux_succes     = round(min(100, (acces_autorises / total_acces * 100)) if total_acces > 0 else 0, 1)
    reservations_count = db.reservations.count_documents({'employe_id': str(employe['_id'])})

    start_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    acces_mois  = db.acces_logs.count_documents({
        'utilisateur_id': utilisateur_id,
        'timestamp': {'$gte': start_month}
    })

    dernier_acces_doc = db.acces_logs.find_one(
        {'utilisateur_id': utilisateur_id}, sort=[('timestamp', -1)]
    )
    dernier_acces = dernier_acces_doc['timestamp'] if dernier_acces_doc else None

    try:
        pipeline = [
            {'$match': {'utilisateur_id': utilisateur_id}},
            {'$group': {'_id': {
                'year':  {'$year': '$timestamp'},
                'month': {'$month': '$timestamp'},
                'day':   {'$dayOfMonth': '$timestamp'}
            }}},
            {'$count': 'total_days'}
        ]
        result = list(db.acces_logs.aggregate(pipeline))
        jours_actifs = result[0]['total_days'] if result else 0
    except Exception:
        jours_actifs = 0

    preferences = employe.get('preferences_notifications', {}) or {'email': True, 'rappel': True}

    # Sessions actives
    active_sessions = []
    try:
        raw_sessions = list(db['dashboard_usersession'].find({
            'user_id': request.session.get('user_id', ''),
            'is_active': True,
            '$or': [{'logout_time': None}, {'logout_time': {'$exists': False}}],
        }).sort('last_activity', -1))
        for s in raw_sessions:
            def _fmt(dt):
                if dt is None: return '—'
                try: return dt.strftime('%d/%m/%Y %H:%M:%S')
                except Exception: return str(dt)
            active_sessions.append({
                'id':            str(s.get('_id', '')),
                'device_type':   s.get('device_type') or 'desktop',
                'ip_address':    s.get('ip_address') or '—',
                'login_time':    _fmt(s.get('login_time')),
                'last_activity': _fmt(s.get('last_activity')),
                'is_current':    s.get('session_key') == request.session.session_key,
            })
    except Exception:
        active_sessions = []

    # ============================================================
    # TRAITEMENT POST
    # ============================================================
    if request.method == 'POST':

        # ── Changement de mot de passe ───────────────────────────
        if 'change_password' in request.POST:
            old_password  = request.POST.get('old_password', '')
            new_password1 = request.POST.get('new_password1', '')
            new_password2 = request.POST.get('new_password2', '')

            stored = _get_user_hash(get_session_user(request))
            if not _check_password(old_password, stored):
                messages.error(request, "L'ancien mot de passe est incorrect.")
            elif len(new_password1) < 6:
                messages.error(request, "Le nouveau mot de passe doit contenir au moins 6 caractères.")
            elif new_password1 != new_password2:
                messages.error(request, "Les mots de passe ne correspondent pas.")
            elif new_password1 == old_password:
                messages.error(request, "Le nouveau mot de passe doit être différent de l'ancien.")
            else:
                if _save_user_password(get_session_user(request), new_password1, request=request):
                    messages.success(request, "Mot de passe changé avec succès.")
                else:
                    messages.error(request, "Erreur lors du changement de mot de passe.")
            return redirect('employe_profil')

        # ── Mise à jour des préférences ──────────────────────────
        elif 'update_preferences' in request.POST:
            notif_email  = request.POST.get('notif_email')  == 'on'
            notif_rappel = request.POST.get('notif_rappel') == 'on'
            db.employees.update_one(
                {'_id': employe['_id']},
                {'$set': {
                    'preferences_notifications': {'email': notif_email, 'rappel': notif_rappel},
                    'updated_at': datetime.now()
                }}
            )
            messages.success(request, "Préférences mises à jour.")
            return redirect('employe_profil')

        # ── Mise à jour du profil (infos + photo) ────────────────
        else:
            try:
                prenom      = request.POST.get('prenom', '').strip()
                nom         = request.POST.get('nom', '').strip()
                email       = request.POST.get('email', '').strip()
                telephone   = request.POST.get('telephone', '').strip()
                poste       = request.POST.get('poste', '').strip()
                departement = request.POST.get('departement', '').strip()

                if not prenom or not nom:
                    messages.error(request, "Le nom et le prénom sont requis.")
                    return redirect('employe_profil')

                update_data = {
                    'nom': nom, 'prenom': prenom, 'email': email,
                    'telephone': telephone, 'poste': poste,
                    'departement': departement, 'updated_at': datetime.now()
                }

                if 'photo' in request.FILES:
                    import base64 as b64mod
                    photo_file = request.FILES['photo']
                    if photo_file.size > 2 * 1024 * 1024:
                        messages.error(request, "La photo ne doit pas dépasser 2 Mo.")
                        return redirect('employe_profil')
                    if photo_file.content_type not in ['image/jpeg', 'image/png', 'image/webp', 'image/gif']:
                        messages.error(request, "Format accepté : JPG, PNG, WEBP.")
                        return redirect('employe_profil')
                    photo_b64 = b64mod.b64encode(photo_file.read()).decode('utf-8')
                    update_data['photo'] = f"data:{photo_file.content_type};base64,{photo_b64}"

                db.employees.update_one({'_id': employe['_id']}, {'$set': update_data})
                _update_django_user(request.session.get('username', ''), first_name=prenom, last_name=nom, email=email)
                messages.success(request, "Profil mis à jour avec succès.")

            except Exception as e:
                messages.error(request, f"Erreur: {str(e)}")

            return redirect('employe_profil')

    # ============================================================
    # AFFICHAGE GET
    # ============================================================
    return render(request, 'dashboard/employe_profil.html', {
        'employe':            employe,
        'user':               get_session_user(request),
        'total_acces':        total_acces,
        'acces_autorises':    acces_autorises,
        'acces_refuses':      acces_refuses,
        'taux_succes':        taux_succes,
        'reservations_count': reservations_count,
        'acces_mois':         acces_mois,
        'jours_actifs':       jours_actifs,
        'dernier_acces':      dernier_acces,
        'preferences':        preferences,
        'active_sessions':    active_sessions,
    })
@session_required
def employe_change_password(request):
    """Changer le mot de passe de l'employé"""
    if request.session.get('is_staff', False):
        return redirect('dashboard')

    if request.method == 'POST':
        old_password  = request.POST.get('old_password', '')
        new_password1 = request.POST.get('new_password1', '')
        new_password2 = request.POST.get('new_password2', '')

        stored = _get_user_hash(get_session_user(request))
        if not _check_password(old_password, stored):
            messages.error(request, "L'ancien mot de passe est incorrect.")
            return redirect('employe_profil')

        if len(new_password1) < 6:
            messages.error(request, "Le nouveau mot de passe doit contenir au moins 6 caractères.")
            return redirect('employe_profil')

        if new_password1 != new_password2:
            messages.error(request, "Les mots de passe ne correspondent pas.")
            return redirect('employe_profil')

        if new_password1 == old_password:
            messages.error(request, "Le nouveau mot de passe doit être différent de l'ancien.")
            return redirect('employe_profil')

        if _save_user_password(get_session_user(request), new_password1, request=request):
            try:
                db.system_logs.insert_one({
                    'user_id':   request.session.get('user_id', ''),
                    'username':  request.session.get('username', ''),
                    'action':    'PASSWORD_CHANGE',
                    'timestamp': datetime.now(),
                    'ip':        request.META.get('REMOTE_ADDR'),
                })
            except Exception:
                pass
            messages.success(request, "Votre mot de passe a été changé avec succès.")
        else:
            messages.error(request, "Erreur lors du changement de mot de passe.")

    return redirect('employe_profil')
@session_required
def api_save_preferences(request):
    """API pour sauvegarder les préférences utilisateur"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
        if not employe:
            employe = db.employees.find_one({'django_username': request.session.get('username', '')})
        
        if not employe:
            return JsonResponse({'error': 'Employé non trouvé'}, status=404)
        
        # Récupérer les préférences existantes
        preferences = employe.get('preferences_notifications', {})
        
        # Mettre à jour les préférences
        if 'email' in data:
            preferences['email'] = data['email']
        if 'rappel' in data:
            preferences['rappel'] = data['rappel']
        if 'theme' in data:
            db.employees.update_one(
                {'_id': employe['_id']},
                {'$set': {'theme': data['theme']}}
            )
        
        db.employees.update_one(
            {'_id': employe['_id']},
            {'$set': {'preferences_notifications': preferences, 'updated_at': datetime.now()}}
        )
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
@session_required
def employe_update_profil(request):
    """API pour mettre à jour le profil employé (AJAX)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)

        employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
        if not employe:
            employe = db.employees.find_one({'django_username': request.session.get('username', '')})
        if not employe:
            return JsonResponse({'error': 'Employé non trouvé'}, status=404)

        update_data = {}
        for field in ('prenom', 'nom', 'email', 'telephone', 'poste', 'departement'):
            if field in data:
                update_data[field] = str(data[field]).strip()
        update_data['updated_at'] = datetime.now()

        if update_data:
            db.employees.update_one({'_id': employe['_id']}, {'$set': update_data})
            django_fields = {k: v for k, v in {
                'first_name': update_data.get('prenom'),
                'last_name':  update_data.get('nom'),
                'email':      update_data.get('email'),
            }.items() if v is not None}
            if django_fields:
                _update_django_user(request.session.get('username', ''), **django_fields)

        return JsonResponse({'status': 'success', 'message': 'Profil mis à jour'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
@session_required
def api_employee_stats(request):
    """API pour les statistiques de l'employé (graphiques)"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        period = request.GET.get('period', 'month')
        
        employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
        if not employe:
            employe = db.employees.find_one({'django_username': request.session.get('username', '')})
        
        if not employe:
            return JsonResponse({'error': 'Employé non trouvé'}, status=404)
        
        now = datetime.now()
        labels = []
        values = []
        
        if period == 'week':
            # 7 derniers jours
            for i in range(6, -1, -1):
                day_start = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
                day_end = day_start + timedelta(days=1)
                count = db.acces_logs.count_documents({
                    'utilisateur_id': employe['_id'],
                    'timestamp': {'$gte': day_start, '$lt': day_end}
                })
                labels.append(day_start.strftime('%a'))
                values.append(count)
        
        elif period == 'month':
            # 30 derniers jours
            for i in range(29, -1, -1):
                day_start = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
                day_end = day_start + timedelta(days=1)
                count = db.acces_logs.count_documents({
                    'utilisateur_id': employe['_id'],
                    'timestamp': {'$gte': day_start, '$lt': day_end}
                })
                labels.append(day_start.strftime('%d/%m'))
                values.append(count)
        
        else:
            # 12 derniers mois
            for i in range(11, -1, -1):
                month_start = (now.replace(day=1) - timedelta(days=30 * i)).replace(day=1, hour=0, minute=0, second=0)
                if i == 0:
                    month_end = now
                else:
                    month_end = (month_start + timedelta(days=32)).replace(day=1)
                count = db.acces_logs.count_documents({
                    'utilisateur_id': employe['_id'],
                    'timestamp': {'$gte': month_start, '$lt': month_end}
                })
                labels.append(month_start.strftime('%b'))
                values.append(count)
        
        return JsonResponse({
            'labels': labels,
            'values': values,
            'period': period
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
        # dashboard/views.py - Ajoutez ces fonctions

@session_required
def reservation_list(request):
    """Liste des réservations avec vue calendrier et tableau"""
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')
    
    from datetime import datetime, timedelta
    import json
    from bson import ObjectId
    
    reservations = list(db.reservations.find().sort('date_debut', -1))
    
    print(f"🔍 Nombre de réservations trouvées: {len(reservations)}")  # Debug
    
    # Enrichir les données
    for r in reservations:
        r['id'] = str(r['_id'])
        
        # Employé
        employe_id = r.get('employe_id')
        if employe_id:
            try:
                if isinstance(employe_id, str):
                    emp = db.employees.find_one({'_id': ObjectId(employe_id)})
                else:
                    emp = db.employees.find_one({'_id': employe_id})
                if emp:
                    r['employe_nom'] = f"{emp.get('nom', '')} {emp.get('prenom', '')}".strip() or 'Inconnu'
                    r['employe_badge'] = emp.get('badge_id', '—')
                else:
                    r['employe_nom'] = 'Inconnu'
                    r['employe_badge'] = '—'
            except Exception as e:
                print(f"Erreur employé: {e}")
                r['employe_nom'] = 'Inconnu'
                r['employe_badge'] = '—'
        else:
            r['employe_nom'] = 'Inconnu'
            r['employe_badge'] = '—'
        
        # Bureau
        bureau_id = r.get('bureau_id')
        if bureau_id:
            try:
                if isinstance(bureau_id, str):
                    bureau = db.bureaux.find_one({'_id': ObjectId(bureau_id)})
                else:
                    bureau = db.bureaux.find_one({'_id': bureau_id})
                r['bureau_nom'] = bureau['nom'] if bureau else 'Salle inconnue'
            except Exception as e:
                print(f"Erreur bureau: {e}")
                r['bureau_nom'] = 'Salle inconnue'
        else:
            r['bureau_nom'] = 'Salle inconnue'
        
        # QR code
        if 'qr_code' not in r:
            r['qr_code'] = None
    
    now = datetime.now()
    confirmees = sum(1 for r in reservations if r.get('statut') == 'confirmee')
    en_attente = sum(1 for r in reservations if r.get('statut') == 'en_attente')
    annulees = sum(1 for r in reservations if r.get('statut') == 'annulee')
    
    a_venir = sum(1 for r in reservations if r.get('statut') == 'confirmee' 
                  and r.get('date_debut') and r['date_debut'] > now)
    
    # Taux d'occupation
    total_bureaux = db.bureaux.count_documents({})
    if total_bureaux > 0:
        occupied_bureaux = set()
        for r in reservations:
            if r.get('statut') == 'confirmee' and r.get('date_debut') and r.get('date_fin'):
                if r['date_debut'] <= now <= r['date_fin']:
                    occupied_bureaux.add(str(r.get('bureau_id')))
        taux_occupation = round((len(occupied_bureaux) / total_bureaux) * 100)
    else:
        taux_occupation = 0
    
    # Convertir les données en JSON sécurisé
    reservations_list = []
    for r in reservations:
        if r.get('date_debut'):
            reservations_list.append({
                'id': str(r['_id']),
                'titre': r.get('titre', ''),
                'bureau_id': str(r.get('bureau_id')) if r.get('bureau_id') else None,
                'bureau_nom': r.get('bureau_nom', ''),
                'employe_nom': r.get('employe_nom', ''),
                'statut': r.get('statut', ''),
                'date_debut': r['date_debut'].isoformat() if r.get('date_debut') else None,
                'date_fin': r['date_fin'].isoformat() if r.get('date_fin') else None,
            })
    
    reservations_json = json.dumps(reservations_list, default=str)
    print(f"📊 JSON généré avec {len(reservations_list)} réservations")  # Debug
    
    # Liste des bureaux
    bureaux_list = []
    for b in db.bureaux.find():
        bureaux_list.append({
            'id': str(b['_id']),
            'nom': b.get('nom', ''),
        })
    
    return render(request, 'dashboard/reservation_list.html', {
        'reservations': reservations,
        'total': len(reservations),
        'confirmees': confirmees,
        'en_attente': en_attente,
        'annulees': annulees,
        'a_venir': a_venir,
        'taux_occupation': taux_occupation,
        'reservations_json': reservations_json,
        'bureaux': bureaux_list,
    })
# ================================================================
# REMPLACEZ CES DEUX FONCTIONS dans votre views.py
# 1. reservation_confirmer  (cherchez @session_required + def reservation_confirmer)
# 2. reservation_annuler    (cherchez @session_required + def reservation_annuler)
# ================================================================


@session_required
def reservation_confirmer(request, reservation_id):
    """Confirmer une réservation et générer un QR code"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    try:
        from bson import ObjectId
        from datetime import datetime, timedelta
        import qrcode
        from io import BytesIO
        import base64

        reservation = db.reservations.find_one({'_id': ObjectId(reservation_id)})
        if not reservation:
            messages.error(request, "Réservation non trouvée")
            return redirect('reservation_list')

        if reservation.get('statut') == 'confirmee':
            messages.warning(request, "Cette réservation est déjà confirmée")
            return redirect('reservation_detail', reservation_id=reservation_id)

        if request.method == 'POST':
            # ── QR code ─────────────────────────────────────────────
            qr_data = (
                f"RESA-{reservation_id}"
                f"-{reservation.get('employe_id')}"
                f"-{reservation.get('date_debut').strftime('%Y%m%d%H%M')}"
            )
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_data)
            qr.make(fit=True)
            img    = qr.make_image(fill_color="black", back_color="white")
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            qr_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

            # ── Mettre à jour la réservation ─────────────────────────
            db.reservations.update_one(
                {'_id': ObjectId(reservation_id)},
                {'$set': {
                    'statut':       'confirmee',
                    'qr_code':      qr_base64,
                    'confirmed_at': datetime.now(),
                    'confirmed_by': request.session.get('username', ''),
                }}
            )

            # ── Récupérer l'employé ──────────────────────────────────
            employe_id = reservation.get('employe_id')
            employe    = None
            try:
                if isinstance(employe_id, str) and len(employe_id) == 24:
                    employe = db.employees.find_one({'_id': ObjectId(employe_id)})
                else:
                    employe = db.employees.find_one({'django_user_id': employe_id})
            except Exception:
                pass

            # ── Récupérer la salle / matériel ────────────────────────
            bureau    = None
            bureau_id = reservation.get('bureau_id')
            if bureau_id:
                try:
                    bureau = db.bureaux.find_one({'_id': ObjectId(bureau_id)})
                except Exception:
                    pass
            bureau_nom = bureau['nom'] if bureau else reservation.get('materiel_nom') or 'Ressource'

            # ── Notification MongoDB (une seule fois) ────────────────
            if employe:
                existing_notif = db.notifications.find_one({
                    'employe_id':     str(employe['_id']),
                    'reservation_id': str(reservation['_id']),
                    'categorie':      'confirmation',
                })

                if not existing_notif:
                    db.notifications.insert_one({
                        'employe_id':     str(employe['_id']),
                        'titre':          '✅ Réservation confirmée',
                        'message': (
                            f"Votre réservation '{reservation.get('titre', 'Sans titre')}' "
                            f"a été confirmée pour le "
                            f"{reservation['date_debut'].strftime('%d/%m/%Y à %H:%M')} "
                            f"— {bureau_nom}."
                        ),
                        'categorie':      'confirmation',
                        'icon':           '✅',
                        'status':         'non_lu',
                        'action_url':     '/employe/reservations/',
                        'reservation_id': str(reservation['_id']),
                        'created_at':     datetime.now(),
                    })

                    # ── Email (résolution email avec fallback utilisateurs) ──
                    email_dest = (employe.get('email') or '').strip()

                    if not email_dest:
                        uid   = employe.get('django_user_id') or employe.get('user_id', '')
                        uname = employe.get('django_username') or employe.get('username', '')
                        or_q  = []
                        if uid:
                            try:
                                or_q.append({'_id': ObjectId(str(uid))})
                            except Exception:
                                pass
                            or_q.append({'id': uid})
                        if uname:
                            or_q.append({'username': uname})
                        if or_q:
                            u_doc = db['utilisateurs'].find_one({'$or': or_q}, {'email': 1})
                            if u_doc:
                                email_dest = (u_doc.get('email') or '').strip()

                    if email_dest:
                        try:
                            from django.core.mail import send_mail
                            prenom     = employe.get('prenom', employe.get('first_name', ''))
                            nom_emp    = employe.get('nom',    employe.get('last_name',  ''))
                            date_str   = reservation['date_debut'].strftime('%d/%m/%Y à %H:%M')
                            heure_fin  = reservation['date_fin'].strftime('%H:%M') if reservation.get('date_fin') else '—'
                            titre_resa = reservation.get('titre', 'Sans titre')
                            corps = (
                                f"Bonjour {prenom} {nom_emp},\n\n"
                                f"Votre réservation a été CONFIRMÉE.\n\n"
                                f"Réservation : {titre_resa}\n"
                                f"Ressource   : {bureau_nom}\n"
                                f"Date        : {date_str} → {heure_fin}\n\n"
                                f"QR code disponible dans votre espace :\n"
                                f"→ /employe/reservations/\n\n"
                                f"Cordialement,\nL'équipe SIGR-CA"
                            )
                            send_mail(
                                subject=f"✅ Réservation confirmée — {titre_resa}",
                                message=corps,
                                from_email=settings.DEFAULT_FROM_EMAIL,
                                recipient_list=[email_dest],
                                fail_silently=False,
                            )
                            logger.info(f"Email confirmation envoyé → {email_dest}")
                        except Exception as email_err:
                            logger.error(f"Échec email confirmation → {email_dest} : {email_err}")

            # ── Rappel J-1 (matériel uniquement) ─────────────────────
            resource_type = reservation.get('resource_type', 'salle')
            if resource_type != 'salle' and employe:
                existing_rappel = db.rappels_email.find_one({
                    'reservation_id': str(reservation['_id']),
                    'type':           'retour_ressource',
                })
                if not existing_rappel:
                    rappel_date = reservation['date_fin'] - timedelta(days=1)
                    db.rappels_email.insert_one({
                        'reservation_id': str(reservation['_id']),
                        'employe_id':     str(employe['_id']),
                        'employe_email':  employe.get('email', ''),
                        'employe_prenom': employe.get('prenom', ''),
                        'ressource_nom':  bureau_nom,
                        'date_fin_resa':  reservation['date_fin'],
                        'type':           'retour_ressource',
                        'a_envoyer_le':   rappel_date,
                        'envoye':         False,
                        'created_at':     datetime.now(),
                    })

            messages.success(request, f"Réservation '{reservation.get('titre')}' confirmée avec QR code généré.")

            if request.POST.get('redirect_to') == 'list':
                return redirect('reservation_list')
            return redirect('reservation_detail', reservation_id=reservation_id)

        # ── GET : affichage ──────────────────────────────────────────
        employe    = None
        employe_id = reservation.get('employe_id')
        if employe_id:
            try:
                if isinstance(employe_id, str) and len(employe_id) == 24:
                    employe = db.employees.find_one({'_id': ObjectId(employe_id)})
                else:
                    employe = db.employees.find_one({'django_user_id': employe_id})
            except Exception:
                pass

        bureau    = None
        bureau_id = reservation.get('bureau_id')
        if bureau_id:
            try:
                bureau = db.bureaux.find_one({'_id': ObjectId(bureau_id)})
            except Exception:
                pass

        return render(request, 'dashboard/reservation_confirmer.html', {
            'reservation': reservation,
            'employe':     employe,
            'bureau':      bureau,
        })

    except Exception as e:
        messages.error(request, f"Erreur: {str(e)}")
        return redirect('reservation_list')


@session_required
def reservation_refuser(request, reservation_id):
    """Refuser une réservation (admin)"""
    from bson import ObjectId
    from datetime import datetime

    if request.method == 'POST':
        try:
            reservation = db.reservations.find_one({'_id': ObjectId(reservation_id)})
            if not reservation:
                messages.error(request, "Réservation non trouvée")
                return redirect('reservation_list')

            db.reservations.update_one(
                {'_id': ObjectId(reservation_id)},
                {'$set': {
                    'statut':       'annulee',
                    'cancelled_at': datetime.now(),
                    'cancelled_by': request.session.get('username', ''),
                }}
            )

            # ── Notification MongoDB à l'employé ─────────────────────
            employe_id = reservation.get('employe_id')
            employe    = None
            try:
                if isinstance(employe_id, str) and len(employe_id) == 24:
                    employe = db.employees.find_one({'_id': ObjectId(employe_id)})
                else:
                    employe = db.employees.find_one({'django_user_id': employe_id})
            except Exception:
                pass

            if employe:
                existing_notif = db.notifications.find_one({
                    'employe_id':     str(employe['_id']),
                    'reservation_id': str(reservation['_id']),
                    'categorie':      'annulation',
                })
                if not existing_notif:
                    db.notifications.insert_one({
                        'employe_id':     str(employe['_id']),
                        'titre':          '🗑️ Réservation annulée',
                        'message': (
                            f"Votre réservation '{reservation.get('titre', 'Sans titre')}' "
                            f"a été annulée par l'administrateur."
                        ),
                        'categorie':      'annulation',
                        'icon':           '🗑️',
                        'status':         'non_lu',
                        'action_url':     '/employe/reservations/',
                        'reservation_id': str(reservation['_id']),
                        'created_at':     datetime.now(),
                    })

            messages.success(request, "Réservation annulée.")

        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")

    return redirect('reservation_list')

@session_required
def reservation_detail(request, reservation_id):
    """Voir les détails d'une réservation (avec QR code si confirmée)"""
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')
    
    from bson import ObjectId
    
    try:
        reservation = db.reservations.find_one({'_id': ObjectId(reservation_id)})
        if not reservation:
            messages.error(request, "Réservation non trouvée")
            return redirect('reservation_list')
        
        reservation['id'] = str(reservation['_id'])
        
        # Récupérer l'employé
        employe = None
        employe_id = reservation.get('employe_id')
        if employe_id:
            try:
                if isinstance(employe_id, str) and len(employe_id) == 24:
                    employe = db.employees.find_one({'_id': ObjectId(employe_id)})
                else:
                    employe = db.employees.find_one({'django_user_id': employe_id})
            except:
                pass
        
        # Récupérer la salle
        bureau = None
        bureau_id = reservation.get('bureau_id')
        if bureau_id:
            try:
                bureau = db.bureaux.find_one({'_id': ObjectId(bureau_id)})
            except:
                pass
        
        return render(request, 'dashboard/reservation_detail.html', {
            'reservation': reservation,
            'employe': employe,
            'bureau': bureau,
        })
        
    except Exception as e:
        messages.error(request, f"Erreur: {str(e)}")
        return redirect('reservation_list')


def send_reservation_confirmation_email(employe, reservation, qr_base64):
    """Envoie un email de confirmation avec QR code"""
    message = f"""
    Bonjour {employe.get('prenom', '')} {employe.get('nom', '')},
    
    ✅ Votre réservation a été CONFIRMÉE par l'administrateur !
    
    Détails de la réservation:
    - Titre: {reservation.get('titre')}
    - Date: {reservation['date_debut'].strftime('%d/%m/%Y %H:%M')} → {reservation['date_fin'].strftime('%H:%M')}
    - Participants: {reservation.get('nb_participants', 1)}
    
    🔐 QR Code d'accès:
    Présentez ce QR code au lecteur à l'entrée de la salle.
    
    (Le QR code est également disponible dans votre espace employé)
    
    Merci d'utiliser SIGR-CA.
    """
    
    db.notifications.insert_one({
        'destinataire': employe.get('email'),
        'type_notification': 'email',
        'categorie': 'confirmation',
        'sujet': f"Réservation confirmée - {reservation.get('titre')}",
        'message': message,
        'statut': 'envoyee',
        'reservation_id': str(reservation['_id']),
        'created_at': datetime.now(),
    })
    
    try:
        from django.core.mail import send_mail
        send_mail(
            f"Réservation confirmée - {reservation.get('titre')}",
            message,
            settings.DEFAULT_FROM_EMAIL,
            [employe.get('email')],
            fail_silently=True,
        )
    except:
        pass


def send_reservation_refusal_email(employe, reservation, motif):
    """Envoie un email de refus de réservation"""
    message = f"""
    Bonjour {employe.get('prenom', '')} {employe.get('nom', '')},
    
    ❌ Votre réservation a été REFUSÉE par l'administrateur.
    
    Détails de la réservation:
    - Titre: {reservation.get('titre')}
    - Date: {reservation['date_debut'].strftime('%d/%m/%Y %H:%M')} → {reservation['date_fin'].strftime('%H:%M')}
    
    Motif du refus: {motif}
    
    Vous pouvez effectuer une nouvelle demande de réservation depuis votre espace employé.
    
    Cordialement,
    SIGR-CA
    """
    
    db.notifications.insert_one({
        'destinataire': employe.get('email'),
        'type_notification': 'email',
        'categorie': 'annulation',
        'sujet': f"Réservation refusée - {reservation.get('titre')}",
        'message': message,
        'statut': 'envoyee',
        'reservation_id': str(reservation['_id']),
        'created_at': datetime.now(),
    })
    
    try:
        from django.core.mail import send_mail
        send_mail(
            f"Réservation refusée - {reservation.get('titre')}",
            message,
            settings.DEFAULT_FROM_EMAIL,
            [employe.get('email')],
            fail_silently=True,
        )
    except:
        pass
        # dashboard/views.py - Ajoutez cette API

#@session_required
#def api_reservation_qr(request, reservation_id):
   # """API pour récupérer le QR code d'une réservation"""
   # try:
        # Vérifier que l'utilisateur a le droit d'accéder au QR code
      #  if request.session.get('is_staff', False):
        #    reservation = db.reservations.find_one({'_id': ObjectId(reservation_id)})
       # else:
            # Pour les employés, vérifier que c'est bien leur réservation
         #   employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
          #  if not employe:
          #      return JsonResponse({'error': 'Employé non trouvé'}, status=404)
          #  reservation = db.reservations.find_one({
          #      '_id': ObjectId(reservation_id),
           #     'employe_id': str(employe['_id'])
          #  })
        
       # if not reservation:
        #    return JsonResponse({'error': 'Réservation non trouvée'}, status=404)
        
       # return JsonResponse({
          #  'qr_code': reservation.get('qr_code'),
          #  'date_debut': reservation.get('date_debut'),
          #  'date_fin': reservation.get('date_fin'),
          #  'titre': reservation.get('titre'),
          #  'statut': reservation.get('statut'),
       # })
    #except Exception as e:
     #   return JsonResponse({'error': str(e)}, status=500)
        # dashboard/views.py


# ====================== NOTIFICATIONS EMPLOYÉ ======================

# dashboard/views.py - Remplacez la fonction employe_notifications par celle-ci

@session_required
def employe_notifications(request):
    """Centre de notifications de l'employé - Version MongoDB"""
    if request.session.get('is_staff', False):
        return redirect('dashboard')
    
    from bson import ObjectId
    from datetime import datetime
    
    # Récupérer l'employé
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})
    
    if not employe:
        return redirect('login')
    
    employe['id'] = str(employe['_id'])
    
    # Récupérer les notifications depuis MongoDB
    notifications = list(db.notifications.find({
        'employe_id': str(employe['_id'])
    }).sort('created_at', -1))
    
    for n in notifications:
        n['id'] = str(n['_id'])
        if 'status' not in n:
            n['status'] = 'non_lu'
        if 'categorie' not in n:
            n['categorie'] = n.get('type', 'info')
        if 'icon' not in n:
            n['icon'] = '🔔'
    
    # Compter les non lues
    unread_count = sum(1 for n in notifications if n.get('status') == 'non_lu')
    
    # Traitement POST pour marquer comme lu
    if request.method == 'POST':
        if 'mark_read' in request.POST:
            notification_id = request.POST.get('notification_id')
            if notification_id:
                db.notifications.update_one(
                    {'_id': ObjectId(notification_id), 'employe_id': str(employe['_id'])},
                    {'$set': {'status': 'lu', 'read_at': datetime.now()}}
                )
            else:
                # Marquer toutes comme lues
                db.notifications.update_many(
                    {'employe_id': str(employe['_id']), 'status': 'non_lu'},
                    {'$set': {'status': 'lu', 'read_at': datetime.now()}}
                )
            return redirect('employe_notifications')
        
        elif 'delete_all' in request.POST:
            db.notifications.delete_many({'employe_id': str(employe['_id'])})
            return redirect('employe_notifications')
    
    return render(request, 'dashboard/employe_notifications.html', {
        'employe': employe,
        'notifications': notifications,
        'unread_count': unread_count,
    })


@session_required
def api_mark_notification_read(request):
    """API pour marquer une notification comme lue (AJAX) - Version MongoDB"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        from bson import ObjectId
        from datetime import datetime
        
        data = json.loads(request.body)
        notification_id = data.get('notification_id')
        
        employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
        if not employe:
            employe = db.employees.find_one({'django_username': request.session.get('username', '')})
        
        if notification_id:
            db.notifications.update_one(
                {'_id': ObjectId(notification_id), 'employe_id': str(employe['_id'])},
                {'$set': {'status': 'lu', 'read_at': datetime.now()}}
            )
        else:
            # Marquer toutes comme lues
            db.notifications.update_many(
                {'employe_id': str(employe['_id']), 'status': 'non_lu'},
                {'$set': {'status': 'lu', 'read_at': datetime.now()}}
            )
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_delete_notification(request):
    """API pour supprimer une notification (AJAX) - Version MongoDB"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        from bson import ObjectId
        
        data = json.loads(request.body)
        notification_id = data.get('notification_id')
        
        employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
        if not employe:
            employe = db.employees.find_one({'django_username': request.session.get('username', '')})
        
        if notification_id:
            db.notifications.delete_one({
                '_id': ObjectId(notification_id), 
                'employe_id': str(employe['_id'])
            })
        else:
            db.notifications.delete_many({'employe_id': str(employe['_id'])})
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_delete_all_notifications(request):
    """API pour supprimer toutes les notifications - Version MongoDB"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
        if not employe:
            employe = db.employees.find_one({'django_username': request.session.get('username', '')})
        
        db.notifications.delete_many({'employe_id': str(employe['_id'])})
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_send_test_notification(request):
    """API pour envoyer une notification de test - Version MongoDB"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    from datetime import datetime
    
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})
    
    notification = {
        'employe_id': str(employe['_id']),
        'titre': '🔔 Notification de test',
        'message': 'Ceci est une notification de test pour vérifier le bon fonctionnement du centre de notifications.',
        'categorie': 'info',
        'icon': '🔔',
        'status': 'non_lu',
        'action_url': '/employe/notifications/',
        'created_at': datetime.now()
    }
    
    db.notifications.insert_one(notification)
    
    return JsonResponse({'status': 'success', 'message': 'Notification envoyée'})


def send_reservation_notification(employe_id, reservation_data, action='created'):
    """Fonction utilitaire pour envoyer une notification de réservation"""
    from datetime import datetime
    
    notifications_data = {
        'created': {
            'titre': '📝 Réservation créée',
            'message': f"Votre réservation '{reservation_data.get('titre')}' a été créée et est en attente de validation.",
            'categorie': 'reservation',
            'icon': '📝'
        },
        'confirmed': {
            'titre': '✅ Réservation confirmée',
            'message': f"Votre réservation '{reservation_data.get('titre')}' a été confirmée.",
            'categorie': 'confirmation',
            'icon': '✅'
        },
        'refused': {
            'titre': '❌ Réservation refusée',
            'message': f"Votre réservation '{reservation_data.get('titre')}' a été refusée.",
            'categorie': 'annulation',
            'icon': '❌'
        },
        'reminder': {
            'titre': '⏰ Rappel de réservation',
            'message': f"Rappel: Votre réservation '{reservation_data.get('titre')}' commence dans 30 minutes.",
            'categorie': 'rappel',
            'icon': '⏰'
        },
        'cancelled': {
            'titre': '🗑️ Réservation annulée',
            'message': f"Votre réservation '{reservation_data.get('titre')}' a été annulée.",
            'categorie': 'annulation',
            'icon': '🗑️'
        }
    }
    
    data = notifications_data.get(action, notifications_data['created'])
    
    db.notifications.insert_one({
        'employe_id': str(employe_id),
        'titre': data['titre'],
        'message': data['message'],
        'categorie': data['categorie'],
        'icon': data['icon'],
        'status': 'non_lu',
        'action_url': '/employe/reservations/',
        'reservation_id': reservation_data.get('id'),
        'created_at': datetime.now()
    })
    
@session_required
def api_notifications_unread_count(request):
    """API pour récupérer le nombre de notifications non lues"""
    try:
        employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
        if not employe:
            employe = db.employees.find_one({'django_username': request.session.get('username', '')})
        
        if not employe:
            return JsonResponse({'count': 0})
        
        count = db.notifications.count_documents({
            'employe_id': str(employe['_id']),
            'status': 'non_lu'
        })
        
        return JsonResponse({'count': count})
        
    except Exception as e:
        return JsonResponse({'count': 0, 'error': str(e)}, status=500)
        # ====================== STATISTIQUES GESTION DES RESSOURCES ======================
# Ajoutez ces fonctions à la fin de votre fichier views.py

@session_required
def get_ressource_stats(request):
    """Statistiques de gestion des ressources pour le dashboard"""
    from datetime import datetime, timedelta
    
    now = datetime.now()
    start_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Taux d'occupation des salles
    total_salles = db.bureaux.count_documents({})
    
    # Compter les réservations actives du mois
    reservations_mois = db.reservations.count_documents({
        'date_debut': {'$gte': start_month},
        'statut': 'confirmee'
    })
    
    # Calcul du taux d'occupation (basé sur 8h/jour * 30 jours = 240h par salle)
    heures_possibles = total_salles * 240 if total_salles > 0 else 1
    heures_occupees = reservations_mois * 2  # moyenne 2h par réservation
    taux_occupation = min(100, round((heures_occupees / heures_possibles) * 100, 1)) if heures_possibles > 0 else 0
    
    # Réservations totales du mois
    total_reservations = db.reservations.count_documents({
        'date_debut': {'$gte': start_month},
        'statut': 'confirmee'
    })
    
    # Salles disponibles actuellement
    salles_reservees = db.reservations.distinct('bureau_id', {
        'date_debut': {'$lte': now},
        'date_fin': {'$gte': now},
        'statut': 'confirmee'
    })
    salles_disponibles = total_salles - len(salles_reservees)
    
    # Taux d'annulation (30 derniers jours)
    thirty_days_ago = now - timedelta(days=30)
    total_commandes = db.reservations.count_documents({
        'date_debut': {'$gte': thirty_days_ago}
    })
    annulations = db.reservations.count_documents({
        'date_debut': {'$gte': thirty_days_ago},
        'statut': 'annulee'
    })
    taux_annulation = round((annulations / total_commandes) * 100, 1) if total_commandes > 0 else 0
    
    return {
        'taux_occupation': taux_occupation,
        'total_reservations': total_reservations,
        'salles_disponibles': salles_disponibles,
        'total_salles': total_salles,
        'taux_annulation': taux_annulation,
    }


@session_required
def get_occupation_stats(request):
    """Statistiques d'occupation par salle"""
    from datetime import datetime, timedelta
    
    now = datetime.now()
    thirty_days_ago = now - timedelta(days=30)
    
    bureaux = list(db.bureaux.find())
    occupation_data = []
    
    for bureau in bureaux:
        # Compter les réservations des 30 derniers jours
        reservations_count = db.reservations.count_documents({
            'bureau_id': bureau['_id'],
            'date_debut': {'$gte': thirty_days_ago},
            'statut': 'confirmee'
        })
        
        # Calculer le taux d'occupation (max 30 jours * 8h par jour = 240h)
        # Chaque réservation dure en moyenne 2h
        heures_occupees = reservations_count * 2
        heures_possibles = 240  # 30 jours * 8h
        taux = min(100, round((heures_occupees / heures_possibles) * 100, 1)) if heures_possibles > 0 else 0
        
        occupation_data.append({
            'nom': bureau.get('nom', 'Salle inconnue'),
            'taux': taux,
            'reservations': reservations_count
        })
    
    # Trier par taux d'occupation décroissant
    occupation_data.sort(key=lambda x: x['taux'], reverse=True)
    
    return {
        'labels': [o['nom'] for o in occupation_data[:10]],
        'values': [o['taux'] for o in occupation_data[:10]]
    }


@session_required
def get_top_ressources(request):
    """Top ressources les plus réservées"""
    from datetime import datetime, timedelta
    
    now = datetime.now()
    thirty_days_ago = now - timedelta(days=30)
    
    pipeline = [
        {'$match': {
            'date_debut': {'$gte': thirty_days_ago},
            'statut': 'confirmee'
        }},
        {'$group': {
            '_id': '$bureau_id',
            'count': {'$sum': 1}
        }},
        {'$sort': {'count': -1}},
        {'$limit': 5}
    ]
    
    results = list(db.reservations.aggregate(pipeline))
    total_reservations = sum(r['count'] for r in results)
    
    top_ressources = []
    for r in results:
        bureau = db.bureaux.find_one({'_id': r['_id']})
        if bureau:
            top_ressources.append({
                'nom': bureau.get('nom', 'Salle inconnue'),
                'reservations': r['count'],
                'pct': round((r['count'] / total_reservations) * 100, 1) if total_reservations > 0 else 0
            })
    
    return top_ressources


@session_required
def get_weekly_schedule(request):
    """Planning des réservations pour les 7 prochains jours"""
    from datetime import datetime, timedelta
    
    now = datetime.now()
    week_later = now + timedelta(days=7)
    
    reservations = list(db.reservations.find({
        'date_debut': {'$gte': now, '$lte': week_later},
        'statut': 'confirmee'
    }).sort('date_debut', 1).limit(10))
    
    schedule = []
    for r in reservations:
        bureau = db.bureaux.find_one({'_id': r.get('bureau_id')})
        schedule.append({
            'date': r['date_debut'].strftime('%d/%m'),
            'heure': r['date_debut'].strftime('%H:%M'),
            'titre': r.get('titre', 'Sans titre'),
            'salle': bureau.get('nom', 'Salle inconnue') if bureau else 'Salle inconnue',
            'participants': r.get('nb_participants', 1)
        })
    
    return schedule


@session_required
def get_hour_stats(request):
    """Statistiques horaires des accès"""
    from datetime import datetime, timedelta
    
    now = datetime.now()
    thirty_days_ago = now - timedelta(days=30)
    
    # Créer les tranches horaires
    hours = [f"{h:02d}h-{h+1:02d}h" for h in range(0, 24)]
    hour_counts = [0] * 24
    
    # Compter les accès par heure
    logs = db.acces_logs.find({
        'timestamp': {'$gte': thirty_days_ago}
    })
    
    for log in logs:
        if log.get('timestamp'):
            hour = log['timestamp'].hour
            hour_counts[hour] += 1
    
    return {
        'labels': hours,
        'values': hour_counts
    }


@session_required
def get_zone_stats_data(request):
    """Statistiques par zone pour les graphiques"""
    from datetime import datetime, timedelta
    
    now = datetime.now()
    thirty_days_ago = now - timedelta(days=30)
    
    pipeline = [
        {'$match': {'timestamp': {'$gte': thirty_days_ago}}},
        {'$group': {
            '_id': '$bureau_id',
            'count': {'$sum': 1}
        }},
        {'$sort': {'count': -1}},
        {'$limit': 5}
    ]
    
    results = list(db.acces_logs.aggregate(pipeline))
    total_acces = sum(r['count'] for r in results)
    
    labels = []
    values = []
    details = []
    
    for r in results:
        bureau = db.bureaux.find_one({'_id': r['_id']})
        if bureau:
            nom = bureau.get('nom', 'Zone inconnue')
            pct = round((r['count'] / total_acces) * 100, 1) if total_acces > 0 else 0
            labels.append(nom)
            values.append(r['count'])
            details.append({
                'nom': nom,
                'count': r['count'],
                'pct': pct
            })
    
    return {
        'labels': labels,
        'values': values,
        'details': details,
        'total': total_acces
    }


# ====================== API ENDPOINTS POUR LES STATISTIQUES ======================

@session_required
def api_stats_overview(request):
    """API endpoint pour les statistiques globales (pour AJAX)"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        ressource_stats = get_ressource_stats(request)
        occupation_stats = get_occupation_stats(request)
        top_ressources = get_top_ressources(request)
        weekly_schedule = get_weekly_schedule(request)
        hour_stats = get_hour_stats(request)
        zone_stats = get_zone_stats_data(request)
        
        return JsonResponse({
            'status': 'success',
            'ressource_stats': ressource_stats,
            'occupation_stats': occupation_stats,
            'top_ressources': top_ressources,
            'weekly_schedule': weekly_schedule,
            'hour_stats': hour_stats,
            'zone_stats': zone_stats,
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@session_required
def api_occupation(request):
    """Retourne l'occupation en temps réel de chaque bureau."""
    one_hour_ago = datetime.now() - timedelta(hours=1)
    bureaux_raw  = list(db.bureaux.find())
    bureaux      = []
 
    for b in bureaux_raw:
        capacite = b.get('capacite_max', 10)
        recent   = db.acces_logs.count_documents({
            'bureau_id':  b['_id'],
            'resultat':   'AUTORISE',
            'timestamp':  {'$gte': one_hour_ago}
        })
        occupation = min(recent, capacite)
        taux       = round(occupation / capacite * 100, 1) if capacite > 0 else 0
 
        bureaux.append({
            'id':        str(b['_id']),
            'nom':       b.get('nom', ''),
            'occupation': occupation,
            'capacite':  capacite,
            'taux':      taux,
        })
 
    return JsonResponse({'bureaux': bureaux})
 

@session_required
def api_top_ressources(request):
    """API pour le top des ressources"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        top = get_top_ressources(request)
        return JsonResponse({
            'status': 'success',
            'data': top
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_weekly_schedule(request):
    """API pour le planning hebdomadaire"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        schedule = get_weekly_schedule(request)
        return JsonResponse({
            'status': 'success',
            'data': schedule
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_hour_stats(request):
    """API pour les statistiques horaires"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        stats = get_hour_stats(request)
        return JsonResponse({
            'status': 'success',
            'labels': stats['labels'],
            'values': stats['values']
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        # ====================== PLAN DES ZONES ======================
@session_required
def employe_plan_zones(request):
    """Plan interactif des zones et matériels — côté employé"""
    if request.session.get('is_staff', False):
        return redirect('dashboard')

    from datetime import datetime, timedelta

    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})

    now = datetime.now()

    # ═══════════════════════════════════════════════════════════════
    # ZONES / BUREAUX
    # ═══════════════════════════════════════════════════════════════
    bureaux = list(db.bureaux.find())
    for b in bureaux:
        b['id'] = str(b['_id'])

        # Activité sur la dernière heure
        one_hour_ago = now - timedelta(hours=1)
        recent_acces = db.acces_logs.count_documents({
            'bureau_id': b['_id'],
            'timestamp': {'$gte': one_hour_ago}
        })
        b['acces_recents'] = recent_acces

        # Niveau d'activité
        if recent_acces == 0:
            b['niveau'] = 'low'
        elif recent_acces <= 5:
            b['niveau'] = 'mid'
        else:
            b['niveau'] = 'high'

        # Dernière activité
        last_log = db.acces_logs.find_one(
            {'bureau_id': b['_id']},
            sort=[('timestamp', -1)]
        )
        b['derniere_activite'] = last_log['timestamp'].strftime('%d/%m %H:%M') if last_log else None

        # Réservations actives en ce moment
        b['reservations_actives'] = db.reservations.count_documents({
            'bureau_id': b['_id'],
            'statut':    'confirmee',
            'date_debut': {'$lte': now},
            'date_fin':   {'$gte': now},
        })

        # Prochaine réservation
        next_resa = db.reservations.find_one(
            {'bureau_id': b['_id'], 'statut': 'confirmee', 'date_debut': {'$gt': now}},
            sort=[('date_debut', 1)]
        )
        b['prochaine_resa'] = next_resa['date_debut'].strftime('%d/%m à %H:%M') if next_resa else None

        # Accès du jour (depuis minuit)
        start_today = now.replace(hour=0, minute=0, second=0, microsecond=0)
        b['acces_jour'] = db.acces_logs.count_documents({
            'bureau_id': b['_id'],
            'timestamp': {'$gte': start_today}
        })

        b.setdefault('type_zone',       'Salle')
        b.setdefault('niveau_securite', 'Standard')
        b.setdefault('responsable',     None)
        b.setdefault('etage',           0)
        b.setdefault('equipements',     [])
        b.setdefault('description',     '')
        b.setdefault('code_bureau',     '')

    etages = sorted(set(b.get('etage', 0) for b in bureaux))

    # ═══════════════════════════════════════════════════════════════
    # MATÉRIELS
    # ═══════════════════════════════════════════════════════════════
    if 'materiels' not in db.list_collection_names():
        db.create_collection('materiels')

    materiels = list(db.materiels.find())
    for m in materiels:
        m['id'] = str(m['_id'])

        # ── Normalisation de la photo ────────────────────────────
        photo = m.get('photo') or ''
        if photo and not photo.startswith('data:'):
            # Détection du type MIME via les premiers octets base64
            if photo.startswith('/9j') or photo.startswith('FFD8') or photo.startswith('ffd8'):
                mime = 'image/jpeg'
            elif photo.startswith('iVBOR'):
                mime = 'image/png'
            elif photo.startswith('R0lG'):
                mime = 'image/gif'
            elif photo.startswith('UklG') or photo.startswith('AAAA'):
                mime = 'image/webp'
            else:
                mime = 'image/jpeg'  # fallback
            photo = f'data:{mime};base64,{photo}'
        m['photo'] = photo
        # ─────────────────────────────────────────────────────────

        # Réservation active en ce moment sur ce matériel
        m['reservation_active'] = db.reservations.count_documents({
            'materiel_id': m['_id'],
            'statut':      'confirmee',
            'date_debut':  {'$lte': now},
            'date_fin':    {'$gte': now},
        })

        # Prochaine réservation du matériel
        next_m = db.reservations.find_one(
            {'materiel_id': m['_id'], 'statut': 'confirmee', 'date_debut': {'$gt': now}},
            sort=[('date_debut', 1)]
        )
        m['prochaine_resa'] = next_m['date_debut'].strftime('%d/%m à %H:%M') if next_m else None

        m.setdefault('statut',         'disponible')
        m.setdefault('categorie',      'autre')
        m.setdefault('description',    '')
        m.setdefault('zone',           '')
        m.setdefault('marque',         '')
        m.setdefault('modele',         '')
        m.setdefault('processeur',     '')
        m.setdefault('ram',            '')
        m.setdefault('stockage',       '')
        m.setdefault('os',             '')
        m.setdefault('ecran',          '')
        m.setdefault('num_inventaire', '')

    categories_mat = sorted(set(m.get('categorie', 'autre') for m in materiels))

    # ═══════════════════════════════════════════════════════════════
    # STATISTIQUES GLOBALES
    # ═══════════════════════════════════════════════════════════════
    total_zones   = len(bureaux)
    zones_libres  = sum(1 for b in bureaux if b['niveau'] == 'low')
    zones_moderes = sum(1 for b in bureaux if b['niveau'] == 'mid')
    zones_actives = sum(1 for b in bureaux if b['niveau'] == 'high')
    total_acces_h = sum(b['acces_recents'] for b in bureaux)

    total_materiels  = len(materiels)
    mat_disponibles  = sum(1 for m in materiels if m['statut'] == 'disponible' and m['reservation_active'] == 0)
    mat_utilises     = sum(1 for m in materiels if m['reservation_active'] > 0)
    mat_maintenance  = sum(1 for m in materiels if m['statut'] in ('maintenance', 'hors_service'))

    return render(request, 'dashboard/employe_plan_zones.html', {
        'employe':         employe,
        'bureaux':         bureaux,
        'etages':          etages,
        'materiels':       materiels,
        'categories_mat':  categories_mat,
        'user':            get_session_user(request),
        'now':             now,
        # Stats zones
        'total_zones':     total_zones,
        'zones_libres':    zones_libres,
        'zones_moderes':   zones_moderes,
        'zones_actives':   zones_actives,
        'total_acces_h':   total_acces_h,
        # Stats matériels
        'total_materiels': total_materiels,
        'mat_disponibles': mat_disponibles,
        'mat_utilises':    mat_utilises,
        'mat_maintenance': mat_maintenance,
    })
    #============= BADGE VIRTUEL ======================#
@session_required
def employe_badge_virtuel(request):
    """Badge virtuel avec QR code, stats d'accès et horaires"""
    if request.session.get('is_staff', False):
        return redirect('dashboard')

    from datetime import datetime, timedelta

    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})
    if not employe:
        return redirect('login')

    employe['id'] = str(employe['_id'])

    # ── Photo employé ─────────────────────────────────────────────
    photo = employe.get('photo') or ''
    if photo and not photo.startswith('data:'):
        if photo.startswith('/9j') or photo.startswith('ffd8') or photo.startswith('FFD8'):
            mime = 'image/jpeg'
        elif photo.startswith('iVBOR'):
            mime = 'image/png'
        elif photo.startswith('R0lG'):
            mime = 'image/gif'
        else:
            mime = 'image/jpeg'
        photo = f'data:{mime};base64,{photo}'
    employe['photo'] = photo

    now = datetime.now()
    emp_id = employe['_id']

    # ── Statistiques d'accès ───────────────────────────────────────
    total_acces  = db.acces_logs.count_documents({'utilisateur_id': emp_id})
    total_succes = db.acces_logs.count_documents({'utilisateur_id': emp_id, 'resultat': 'succes'})
    taux_succes  = round((total_succes / total_acces * 100)) if total_acces > 0 else 0

    start_today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    start_week  = now - timedelta(days=7)

    acces_aujourd_hui = db.acces_logs.count_documents({
        'utilisateur_id': emp_id,
        'timestamp': {'$gte': start_today},
    })
    acces_semaine = db.acces_logs.count_documents({
        'utilisateur_id': emp_id,
        'timestamp': {'$gte': start_week},
    })

    # ── Derniers accès (10 derniers) ───────────────────────────────
    derniers_acces_raw = list(db.acces_logs.find(
        {'utilisateur_id': emp_id},
        sort=[('timestamp', -1)],
        limit=10,
    ))
    derniers_acces = []
    for log in derniers_acces_raw:
        bureau = db.bureaux.find_one({'_id': log.get('bureau_id')})
        derniers_acces.append({
            'zone':     bureau['nom'] if bureau else 'Zone inconnue',
            'resultat': log.get('resultat', 'succes'),
            'heure':    log['timestamp'].strftime('%H:%M'),
            'date':     log['timestamp'].strftime('%d/%m'),
        })

    # ── Zones accessibles ──────────────────────────────────────────
    zones_accessibles = list(db.bureaux.find())
    for z in zones_accessibles:
        z['id'] = str(z['_id'])
        z.setdefault('niveau_securite', 'standard')
        z.setdefault('description', 'Zone de travail')
        z.setdefault('etage', 0)

    # ── Validité du badge ──────────────────────────────────────────
    expiration_year    = 2026
    badge_expiration   = f'31/12/{expiration_year}'
    badge_validite_pct = 65
    date_creation = employe.get('date_creation') or employe.get('date_embauche')
    if date_creation:
        try:
            if hasattr(date_creation, 'year'):
                total_days = (datetime(expiration_year, 12, 31) - datetime(date_creation.year, 1, 1)).days
                elapsed    = (now - datetime(date_creation.year, 1, 1)).days
                badge_validite_pct = min(100, round(elapsed / total_days * 100)) if total_days > 0 else 65
        except Exception:
            pass

    # ── Horaires d'accès ───────────────────────────────────────────
    horaires_acces = {
        'Lundi':    {'debut': '08:00', 'fin': '18:00', 'ouvert': True},
        'Mardi':    {'debut': '08:00', 'fin': '18:00', 'ouvert': True},
        'Mercredi': {'debut': '08:00', 'fin': '18:00', 'ouvert': True},
        'Jeudi':    {'debut': '08:00', 'fin': '18:00', 'ouvert': True},
        'Vendredi': {'debut': '08:00', 'fin': '17:00', 'ouvert': True},
        'Samedi':   {'debut': '09:00', 'fin': '13:00', 'ouvert': True},
        'Dimanche': {'debut': '—',     'fin': '—',     'ouvert': False},
    }
    jours_fr  = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    jour_actuel = jours_fr[now.weekday()]

    return render(request, 'dashboard/employe_badge_virtuel.html', {
        'employe':            employe,
        'zones_accessibles':  zones_accessibles,
        'horaires_acces':     horaires_acces,
        'jour_actuel':        jour_actuel,
        'derniers_acces':     derniers_acces,
        'total_acces':        total_acces,
        'taux_succes':        taux_succes,
        'acces_aujourd_hui':  acces_aujourd_hui,
        'acces_semaine':      acces_semaine,
        'badge_expiration':   badge_expiration,
        'badge_validite_pct': badge_validite_pct,
        'now':                now,
    })

    # ====================== CENTRE D'AIDE ======================

@session_required
def employe_aide(request):
    """Centre d'information et d'aide"""
    if request.session.get('is_staff', False):
        return redirect('dashboard')
    
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})
    
    # FAQ
    faqs = [
        {'question': 'Comment réserver une salle ?', 
         'reponse': 'Rendez-vous dans "Mes réservations" puis cliquez sur "Nouvelle réservation". Sélectionnez la salle, la date et l\'heure.'},
        {'question': 'Comment utiliser mon badge virtuel ?', 
         'reponse': 'Le QR code dans "Badge virtuel" peut être scanné par les lecteurs. Vous pouvez aussi le télécharger et l\'imprimer.'},
        {'question': 'Que faire en cas de refus d\'accès ?', 
         'reponse': 'Vérifiez vos horaires d\'accès dans "Badge virtuel". Si le problème persiste, contactez votre administrateur.'},
        {'question': 'Comment annuler une réservation ?', 
         'reponse': 'Allez dans "Mes réservations", trouvez la réservation concernée et cliquez sur "Annuler".'},
        {'question': 'Comment modifier mon profil ?', 
         'reponse': 'Rendez-vous dans "Mon profil" pour modifier vos informations personnelles et préférences.'},
        {'question': 'Où voir mon historique d\'accès ?', 
         'reponse': 'La section "Mon historique" vous montre tous vos accès avec filtres et export CSV.'},
    ]
    
    # Contacts support
    contacts = {
        'email': 'support@sigr-ca.com',
        'telephone': '+213 00 00 00 00',
        'horaires': 'Lun-Ven: 08:00 - 18:00',
    }
    
    return render(request, 'dashboard/employe_aide.html', {
        'employe': employe,
        'faqs': faqs,
        'contacts': contacts,
    })


# Assurez-vous que ces fonctions sont au bon niveau d'indentation (sans espace avant def)
@session_required
def api_reservation_qr(request, reservation_id):
    """API pour récupérer le QR code d'une réservation"""
    from bson import ObjectId
    
    try:
        reservation = db.reservations.find_one({'_id': ObjectId(reservation_id)})
        if not reservation:
            return JsonResponse({'error': 'Réservation non trouvée'}, status=404)
        
        # Vérifier que l'utilisateur a le droit d'accéder à ce QR code
        employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
        if not employe:
            employe = db.employees.find_one({'django_username': request.session.get('username', '')})
        
        if employe and str(employe['_id']) != reservation.get('employe_id'):
            if not request.session.get('is_staff', False):
                return JsonResponse({'error': 'Non autorisé'}, status=403)
        
        return JsonResponse({
            'qr_code': reservation.get('qr_code'),
            'titre': reservation.get('titre', 'Sans titre'),
            'bureau_nom': get_bureau_name(reservation.get('bureau_id')),
            'date_debut': reservation.get('date_debut').isoformat() if reservation.get('date_debut') else None,
            'date_fin': reservation.get('date_fin').isoformat() if reservation.get('date_fin') else None,
            'statut': reservation.get('statut'),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def get_bureau_name(bureau_id):
    """Récupère le nom du bureau à partir de son ID"""
    from bson import ObjectId
    if not bureau_id:
        return 'Salle inconnue'
    try:
        bureau = db.bureaux.find_one({'_id': ObjectId(bureau_id)})
        return bureau['nom'] if bureau else 'Salle inconnue'
    except:
        return 'Salle inconnue'


@session_required
def api_reservation_duplicate(request, reservation_id):
    """API pour dupliquer une réservation"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        original = db.reservations.find_one({'_id': ObjectId(reservation_id)})
        if not original:
            return JsonResponse({'error': 'Réservation non trouvée'}, status=404)
        
        # Créer une copie
        del original['_id']
        original['titre'] = f"Copie de {original.get('titre', 'Réservation')}"
        original['statut'] = 'en_attente'
        original['created_at'] = datetime.now()
        original['qr_code'] = None
        
        result = db.reservations.insert_one(original)
        
        return JsonResponse({'status': 'success', 'id': str(result.inserted_id)})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_bureau_schedule(request, bureau_id):
    """API pour récupérer les créneaux d'une salle"""
    date = request.GET.get('date')
    if not date:
        return JsonResponse({'creneaux': []})
    
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        date_end = date_obj + timedelta(days=1)
        
        reservations = list(db.reservations.find({
            'bureau_id': ObjectId(bureau_id),
            'statut': {'$in': ['confirmee', 'en_attente']},
            'date_debut': {'$gte': date_obj, '$lt': date_end}
        }).sort('date_debut', 1))
        
        creneaux = []
        for r in reservations:
            creneaux.append({
                'debut': r['date_debut'].strftime('%H:%M'),
                'fin': r['date_fin'].strftime('%H:%M'),
                'titre': r.get('titre', 'Sans titre'),
                'employe': r.get('employe_nom', 'Inconnu'),
            })
        
        return JsonResponse({'creneaux': creneaux})
    except Exception as e:
        return JsonResponse({'creneaux': [], 'error': str(e)})


@session_required
def api_bureau_suggestions(request, bureau_id):
    """API pour suggérer des créneaux disponibles"""
    suggestions = [
        {'date': 'Aujourd\'hui', 'debut': '14:00', 'fin': '15:00', 'taux': 25, 'disponibilite': 'Libre'},
        {'date': 'Aujourd\'hui', 'debut': '15:00', 'fin': '16:00', 'taux': 30, 'disponibilite': 'Libre'},
        {'date': 'Demain', 'debut': '09:00', 'fin': '10:00', 'taux': 15, 'disponibilite': 'Très disponible'},
        {'date': 'Demain', 'debut': '10:00', 'fin': '11:00', 'taux': 20, 'disponibilite': 'Disponible'},
        {'date': 'Jeudi', 'debut': '14:00', 'fin': '15:00', 'taux': 10, 'disponibilite': 'Peu fréquenté'},
    ]
    return JsonResponse({'suggestions': suggestions})

# ====================== Chatbot IA ======================
import json, re
from datetime import datetime, timedelta
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import ChatbotConversation, ChatbotMessage


@session_required
def api_chatbot_message(request):
    """API pour le chatbot employé — PyMongo direct (MongoUser sans ForeignKey)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data            = json.loads(request.body)
        user_message    = data.get('message', '').strip()
        conversation_id = data.get('conversation_id', '')

        if not user_message:
            return JsonResponse({'error': 'Message vide'}, status=400)

        user_key = str(request.session.get('user_id', request.session.get('username', '')))
        now      = datetime.now()

        # ── Récupérer ou créer la conversation ───────────────────
        conversation = None
        if conversation_id:
            try:
                conversation = db['chatbot_conversations'].find_one({
                    '_id':      ObjectId(conversation_id),
                    'user_key': user_key,
                })
            except Exception:
                pass
        if not conversation:
            result       = db['chatbot_conversations'].insert_one({
                'user_key':  user_key,
                'username':  request.session.get('username', ''),
                'is_active': True,
                'created_at': now,
                'updated_at': now,
            })
            conversation = db['chatbot_conversations'].find_one({'_id': result.inserted_id})

        conv_id = conversation['_id']

        # ── Sauvegarder le message utilisateur ───────────────────
        db['chatbot_messages'].insert_one({
            'conversation_id': conv_id,
            'role':            'user',
            'content':         user_message,
            'created_at':      now,
        })

        # ── Générer la réponse ────────────────────────────────────
        response_data = process_chatbot_message(get_session_user(request), user_message, conversation)

        # ── Sauvegarder la réponse assistante ────────────────────
        db['chatbot_messages'].insert_one({
            'conversation_id': conv_id,
            'role':            'assistant',
            'content':         response_data['message'],
            'intent':          response_data.get('intent', ''),
            'entities':        response_data.get('entities', {}),
            'created_at':      datetime.now(),
        })

        # Mettre à jour updated_at
        db['chatbot_conversations'].update_one(
            {'_id': conv_id},
            {'$set': {'updated_at': datetime.now()}}
        )

        return JsonResponse({
            'status':          'success',
            'message':         response_data['message'],
            'intent':          response_data.get('intent', ''),
            'data':            response_data.get('data', {}),
            'conversation_id': str(conv_id),
            'suggestions':     response_data.get('suggestions', []),
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@session_required
def api_chatbot_conversations(request):
    """Récupérer l'historique des conversations — PyMongo direct"""
    user_key = str(request.session.get('user_id', request.session.get('username', '')))
    raw = list(
        db['chatbot_conversations']
        .find({'user_key': user_key, 'is_active': True})
        .sort('updated_at', -1)
        .limit(10)
    )
    data = []
    for conv in raw:
        conv_id = conv['_id']
        last_msg = db['chatbot_messages'].find_one(
            {'conversation_id': conv_id, 'role': 'assistant'},
            sort=[('created_at', -1)]
        )
        msg_count = db['chatbot_messages'].count_documents({'conversation_id': conv_id})
        created   = conv.get('created_at')
        data.append({
            'id':            str(conv_id),
            'created_at':    created.strftime('%d/%m/%Y %H:%M') if created else '',
            'last_message':  (last_msg['content'][:100] if last_msg else ''),
            'message_count': msg_count,
        })
    return JsonResponse({'conversations': data})


@session_required
def api_chatbot_conversation_detail(request, conversation_id):
    """Détail d'une conversation — PyMongo direct"""
    user_key = str(request.session.get('user_id', request.session.get('username', '')))
    try:
        conv = db['chatbot_conversations'].find_one({
            '_id':      ObjectId(conversation_id),
            'user_key': user_key,
        })
        if not conv:
            return JsonResponse({'error': 'Conversation non trouvée'}, status=404)
        msgs_raw = list(
            db['chatbot_messages']
            .find({'conversation_id': conv['_id']})
            .sort('created_at', 1)
        )
        msgs = []
        for m in msgs_raw:
            ts = m.get('created_at')
            msgs.append({
                'role':       m.get('role', ''),
                'content':    m.get('content', ''),
                'created_at': ts.strftime('%H:%M') if ts else '',
            })
        return JsonResponse({'messages': msgs, 'conversation_id': str(conv['_id'])})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)


def get_available_rooms():
    """Liste des salles avec leurs vrais IDs MongoDB."""
    rooms = []
    try:
        for b in db.bureaux.find():
            rooms.append({
                'id':       str(b['_id']),
                'nom':      b.get('nom', 'Salle'),
                'capacite': b.get('capacite_max', b.get('capacite', 10)),
                'niveau':   b.get('niveau', 'standard'),
            })
    except Exception as e:
        print(f"get_available_rooms error: {e}")
    return rooms


def process_chatbot_message(user, message, conversation):
    """Tente Gemini, fallback mots-clés."""
    try:
        return _ai_response(user, message, conversation)
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"IA indisponible, fallback mots-cles : {e}")
        return _keyword_response(user, message, conversation)


# ─────────────────────────────────────────────────────────────────────────────
# Réponse IA (Gemini)
# ─────────────────────────────────────────────────────────────────────────────

def _ai_response(user, message, conversation):
    """Gemini + détection bloc d'action pour créer une réservation réelle."""
    import os
    from google import genai
    from google.genai import types

    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY non configurée")

    client = genai.Client(api_key=api_key)

    # Contexte salles
    salles = get_available_rooms()
    if salles:
        salles_txt = "\n".join(
            [f"- {s['nom']} (id={s['id']}, capacite {s['capacite']})" for s in salles[:10]]
        )
    else:
        salles_txt = "Aucune salle"

    # Contexte réservations utilisateur
    mes_resa_txt = "Aucune"
    employe = None
    try:
        from bson import ObjectId
        employe = db.employees.find_one({'django_user_id': user.id})
        if not employe:
            employe = db.employees.find_one({'django_username': user.username})
        if employe:
            resa = list(
                db.reservations
                .find({'employe_id': str(employe['_id'])})
                .sort('date_debut', -1)
                .limit(5)
            )
            if resa:
                lines = []
                for r in resa:
                    bid    = r.get('bureau_id')
                    bureau = db.bureaux.find_one({'_id': ObjectId(bid) if isinstance(bid, str) else bid}) if bid else None
                    nom_b  = bureau['nom'] if bureau else 'Salle'
                    date_s = r['date_debut'].strftime('%d/%m %H:%M') if r.get('date_debut') else '?'
                    lines.append(
                        f"- {r.get('titre','Sans titre')} | {nom_b} | {date_s} | {r.get('statut','?')}"
                    )
                mes_resa_txt = "\n".join(lines)
    except Exception:
        pass

    prenom      = user.first_name or user.username
    today_iso   = datetime.now().strftime('%Y-%m-%d')
    today_label = datetime.now().strftime('%A %d %B %Y')

    system_prompt = (
        "Tu es l'assistant intelligent de SIGR-CA (gestion d'accès et réservation de salles).\n\n"
        f"CONTEXTE :\n- Utilisateur : {prenom}\n- Date du jour : {today_label} (ISO: {today_iso})\n\n"
        f"SALLES DISPONIBLES (utilise EXACTEMENT ces noms et ids) :\n{salles_txt}\n\n"
        f"RÉSERVATIONS DE L'UTILISATEUR :\n{mes_resa_txt}\n\n"
        "RÈGLES :\n"
        "1. Réponds en français, tutoie l'utilisateur, sois concis (3-6 lignes), emojis avec parcimonie.\n"
        "2. Pour réserver, collecte progressivement : salle, date, heure debut, heure fin, nombre de participants.\n"
        "3. N'invente JAMAIS une salle absente de la liste ci-dessus.\n"
        "4. Pour annuler, dis d'aller dans Mes réservations.\n"
        "5. IMPORTANT : les réservations créées sont EN ATTENTE de validation par un administrateur.\n"
        "   Informe toujours l'utilisateur que sa réservation sera confirmée après validation admin.\n\n"
        "IMPORTANT - CRÉATION DE RÉSERVATION :\n"
        "Dès que tu as les 5 champs (salle, date, heure_debut, heure_fin, participants),\n"
        "tu DOIS terminer ta réponse par un bloc d'action exactement dans ce format :\n\n"
        "```action\n"
        '{"type":"create_reservation","bureau_id":"<id_de_la_salle>","date":"YYYY-MM-DD",'
        '"heure_debut":"HH:MM","heure_fin":"HH:MM","participants":N,"titre":"<titre court>"}\n'
        "```\n\n"
        f"Exemple pour aujourd'hui de 16h00 à 16h30 :\n"
        "```action\n"
        '{"type":"create_reservation","bureau_id":"<id_exact_de_la_liste>","date":"'
        + today_iso + '","heure_debut":"16:00","heure_fin":"16:30","participants":3,"titre":"Réunion"}\n'
        "```\n\n"
        "N'inclus le bloc action QUE quand tu as les 5 infos. Sinon continue à demander."
    )

    # Mémoire conversationnelle — PyMongo direct (conversation est un dict MongoDB)
    contents  = []
    try:
        conv_id   = conversation.get('_id') if isinstance(conversation, dict) else None
        if conv_id:
            raw_msgs  = list(
                db['chatbot_messages']
                .find({'conversation_id': conv_id})
                .sort('created_at', -1)
                .limit(12)
            )
            raw_msgs.reverse()
            for m in raw_msgs:
                role = 'user' if m.get('role') == 'user' else 'model'
                contents.append(types.Content(role=role, parts=[types.Part(text=m.get('content', ''))]))
    except Exception as _he:
        logger.warning(f"_ai_response history load échoué: {_he}")
    if not contents or (contents and contents[-1].parts[0].text != message):
        contents.append(types.Content(role='user', parts=[types.Part(text=message)]))

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=contents,
        config=types.GenerateContentConfig(
            system_instruction=system_prompt,
            temperature=0.4,
            max_output_tokens=500,
        ),
    )
    reply = (response.text or "").strip() or "Je n'ai pas pu générer de réponse."

    # ── Détection du bloc action ──────────────────────────────────
    action_match = re.search(r'```action\s*(\{.*?\})\s*```', reply, re.DOTALL)
    if action_match and employe:
        try:
            action_data = json.loads(action_match.group(1))
            if action_data.get('type') == 'create_reservation':
                ok, info = _create_reservation_from_chat(user, employe, action_data)
                # Supprimer le bloc technique de la réponse visible
                clean_reply = re.sub(r'```action\s*\{.*?\}\s*```', '', reply, flags=re.DOTALL).strip()
                if ok:
                    reply = (
                        clean_reply +
                        f"\n\n✅ Demande de réservation envoyée ! (Référence : {info})\n"
                        "⏳ Elle est actuellement **en attente de validation** par un administrateur. "
                        "Tu recevras une notification dès qu'elle sera confirmée."
                    )
                else:
                    reply = clean_reply + f"\n\n❌ Impossible d'enregistrer la réservation : {info}"
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"Erreur parsing action : {e}")

    msg_lower = message.lower()
    if any(k in msg_lower for k in ['réserver', 'reserver', 'salle']):
        suggestions = ["Mes réservations", "Voir disponibilités", "Aide"]
    elif 'mes' in msg_lower or 'réservation' in msg_lower or 'reservation' in msg_lower:
        suggestions = ["Nouvelle réservation", "Annuler", "Aide"]
    else:
        suggestions = ["Réserver", "Mes réservations", "Aide"]

    return {
        'intent':    'ai_gemini',
        'message':   reply,
        'suggestions': suggestions,
        'entities':  {'model': 'gemini-2.5-flash'},
    }


# ─────────────────────────────────────────────────────────────────────────────
# Création de réservation depuis le chat
# STATUT = 'en_attente'  →  l'admin doit valider
# ─────────────────────────────────────────────────────────────────────────────

def _create_reservation_from_chat(user, employe, data):
    """
    Crée la réservation en base avec statut 'en_attente'.
    Envoie les notifications à l'employé ET aux admins.
    Retourne (ok: bool, info: str).
    """
    from bson import ObjectId

    try:
        bureau_id_str = str(data.get('bureau_id', '')).strip()
        date_str      = str(data.get('date', '')).strip()
        heure_debut   = str(data.get('heure_debut', '')).strip()
        heure_fin     = str(data.get('heure_fin', '')).strip()
        participants  = int(data.get('participants', 1))
        titre         = str(data.get('titre', 'Réservation')).strip() or 'Réservation'

        if not (bureau_id_str and date_str and heure_debut and heure_fin):
            return False, "informations incomplètes"

        # ── Vérification salle ────────────────────────────────────
        try:
            bureau_oid = ObjectId(bureau_id_str)
        except Exception:
            return False, f"identifiant salle invalide ({bureau_id_str})"

        bureau = db.bureaux.find_one({'_id': bureau_oid})
        if not bureau:
            return False, "salle introuvable"

        # ── Parsing dates ─────────────────────────────────────────
        try:
            date_debut = datetime.strptime(f"{date_str} {heure_debut}", '%Y-%m-%d %H:%M')
            date_fin   = datetime.strptime(f"{date_str} {heure_fin}",   '%Y-%m-%d %H:%M')
        except ValueError:
            return False, "format de date/heure invalide (attendu YYYY-MM-DD et HH:MM)"

        if date_fin <= date_debut:
            return False, "l'heure de fin doit être après l'heure de début"

        if date_debut < datetime.now():
            return False, "impossible de réserver dans le passé"

        # ── Durée minimale 30 min ─────────────────────────────────
        duree_minutes = int((date_fin - date_debut).total_seconds() / 60)
        if duree_minutes < 30:
            return False, "durée minimale de 30 minutes requise"

        # ── Capacité ──────────────────────────────────────────────
        cap = bureau.get('capacite_max', bureau.get('capacite', 999))
        if participants > cap:
            return False, f"trop de participants ({participants} > capacité {cap})"

        # ── Conflit de créneaux (confirmée OU en_attente) ─────────
        conflit = db.reservations.find_one({
            'bureau_id':  bureau_oid,
            'statut':     {'$in': ['confirmee', 'en_attente']},
            'date_debut': {'$lt': date_fin},
            'date_fin':   {'$gt': date_debut},
        })
        if conflit:
            return False, f"créneau déjà occupé sur « {bureau.get('nom')} »"

        # ── Création de la réservation — statut EN ATTENTE ────────
        employe_nom = f"{employe.get('prenom', '')} {employe.get('nom', '')}".strip()

        reservation_data = {
            'titre':           titre,
            'description':     'Réservation créée via chatbot IA',
            'employe_id':      str(employe['_id']),
            'employe_nom':     employe_nom,
            'bureau_id':       bureau_oid,
            'bureau_nom':      bureau.get('nom', ''),
            'resource_type':   'salle',
            'resource_id':     bureau_oid,
            'date_debut':      date_debut,
            'date_fin':        date_fin,
            'nb_participants': participants,
            'statut':          'en_attente',        # ← TOUJOURS en_attente
            'qr_code':         None,
            'cree_par':        'chatbot',
            'created_at':      datetime.now(),
            'created_by':      user.username,
        }

        result       = db.reservations.insert_one(reservation_data)
        reservation_id = str(result.inserted_id)

        # ── Notification à l'employé ──────────────────────────────
        db.notifications.insert_one({
            'employe_id':     str(employe['_id']),
            'titre':          '📝 Réservation créée',
            'message':        (
                f"Votre réservation « {titre} » a été créée via le chatbot "
                f"et est en attente de validation par un administrateur."
            ),
            'categorie':      'reservation',
            'icon':           '📝',
            'status':         'non_lu',
            'action_url':     '/employe/reservations/',
            'reservation_id': reservation_id,
            'created_at':     datetime.now(),
        })

        # ── Notifications aux admins ──────────────────────────────
        ressource_label = f"🚪 Salle : {bureau.get('nom', '')}"
        admin_message_email = (
            f"🆕 NOUVELLE RÉSERVATION EN ATTENTE (via chatbot)\n\n"
            f"👤 Employé  : {employe_nom}\n"
            f"📋 Titre    : {titre}\n"
            f"{ressource_label}\n"
            f"📅 Date     : {date_debut.strftime('%d/%m/%Y')}\n"
            f"⏰ Horaire  : {date_debut.strftime('%H:%M')} → {date_fin.strftime('%H:%M')}\n"
            f"👥 Participants : {participants}\n\n"
            f"🔗 Traiter : /reservations/"
        )

        try:
            admins = list(db['utilisateurs'].find(
                {'is_staff': True, 'is_active': True}, {'_id': 1, 'email': 1}
            ))
            for admin in admins:
                db.admin_notifications.insert_one({
                    'admin_id':       admin.get('_id'),
                    'titre':          '🆕 Nouvelle réservation en attente (chatbot)',
                    'message':        (
                        f"{employe_nom} a demandé une réservation pour « {titre} » "
                        f"({ressource_label}) le {date_debut.strftime('%d/%m/%Y à %H:%M')}."
                    ),
                    'categorie':      'reservation',
                    'icon':           '🆕',
                    'status':         'non_lu',
                    'action_url':     f'/reservations/{reservation_id}/',
                    'reservation_id': reservation_id,
                    'created_at':     datetime.now(),
                })
                if admin.get('email'):
                    try:
                        from dashboard.utils_email import envoyer_email
                        envoyer_email(admin['email'], f"🆕 Nouvelle réservation (chatbot) — {titre}", admin_message_email)
                    except Exception as _ee:
                        logger.warning(f"Email admin chatbot échoué: {_ee}")
        except Exception as _e:
            logger.warning(f"Notifications admins chatbot échouées: {_e}")

        return True, reservation_id

    except Exception as e:
        import traceback; traceback.print_exc()
        return False, str(e)


# ─────────────────────────────────────────────────────────────────────────────
# Fallback mots-clés (si Gemini est indisponible)
# ─────────────────────────────────────────────────────────────────────────────

def _keyword_response(user, message, conversation):
    """Fallback mots-clés si Gemini est indisponible."""
    msg    = message.lower()
    prenom = user.first_name or user.username

    if any(k in msg for k in ['bonjour', 'salut', 'hello', 'coucou', 'hey']):
        return {
            'intent':      'bonjour',
            'message':     f"Bonjour {prenom} ! Je suis ton assistant SIGR-CA. Comment puis-je t'aider ?",
            'suggestions': ["Réserver une salle", "Mes réservations", "Aide"],
        }
    if any(k in msg for k in ['merci', 'thanks']):
        return {
            'intent':      'merci',
            'message':     "Avec plaisir ! 😊",
            'suggestions': ["Réserver", "Mes réservations"],
        }
    if any(k in msg for k in ['réserver', 'reserver', 'réservation', 'reservation', 'salle']):
        salles = get_available_rooms()
        if salles:
            txt = "\n".join([f"- {s['nom']} (cap. {s['capacite']})" for s in salles[:5]])
            return {
                'intent':  'reserver',
                'message': (
                    f"Voici les salles disponibles :\n\n{txt}\n\n"
                    "Laquelle t'intéresse ? Indique aussi la date, les horaires et le nombre de participants.\n\n"
                    "⚠️ Note : ta réservation sera en **attente de validation** par un administrateur."
                ),
                'suggestions': ["Mes réservations", "Aide"],
            }
        return {'intent': 'reserver', 'message': "Aucune salle disponible.", 'suggestions': ["Aide"]}
    if any(k in msg for k in ['aide', 'help']):
        return {
            'intent':  'aide',
            'message': (
                "Je peux t'aider à :\n"
                "- 📅 **Réserver une salle** (les réservations sont validées par un admin)\n"
                "- 📋 **Consulter tes réservations**\n"
                "- 🔑 **Informations sur tes accès**\n\n"
                "Que souhaites-tu faire ?"
            ),
            'suggestions': ["Réserver", "Mes réservations"],
        }
    return {
        'intent':      'general',
        'message':     "Je n'ai pas bien compris. Peux-tu reformuler ?",
        'suggestions': ["Réserver une salle", "Mes réservations", "Aide"],
    }


@session_required
def api_chatbot_conversation_detail(request, conversation_id):
    """Détail d'une conversation"""
    try:
        conversation = ChatbotConversation.objects.get(id=conversation_id, user=get_session_user(request))
        messages = []
        for msg in conversation.messages.all():
            messages.append({
                'role': msg.role,
                'content': msg.content,
                'created_at': msg.created_at.strftime('%H:%M')
            })
        
        return JsonResponse({'messages': messages, 'conversation_id': conversation.id})
    except ChatbotConversation.DoesNotExist:
        return JsonResponse({'error': 'Conversation non trouvée'}, status=404)
        # dashboard/views.py - Ajoutez ces fonctions

# ====================== NOTIFICATIONS ADMINISTRATEUR ======================
def _normalize_notif(n, source, algiers):
    """
    Normalise un document notification quelle que soit sa collection.
    Ajoute n['_source'] pour que les APIs sachent où écrire.
    """
    n['id']      = str(n['_id'])
    n['source'] = source

    if 'status' not in n:
        n['status'] = 'non_lu'
    if 'titre' not in n:
        n['titre'] = n.get('sujet', 'Notification')
    if 'categorie' not in n:
        n['categorie'] = n.get('type', n.get('type_notification', 'info'))
    if 'icon' not in n:
        n['icon'] = '🔔'
    if 'message' not in n:
        n['message'] = ''
    if 'action_url' not in n and n.get('reservation_id'):
        n['action_url'] = f"/employe/reservations/"

    raw_dt = n.get('created_at')
    if raw_dt is not None:
        n['created_at'] = _make_aware_dt(raw_dt, algiers)

    return n


def _try_update_notif(db, notification_id, update, now):
    """
    Tente de mettre à jour une notification dans 'notifications'
    puis dans 'admin_notifications'. Retourne True si trouvée.
    """
    from bson import ObjectId
    try:
        oid = ObjectId(notification_id)
    except Exception:
        return False

    for col in ('notifications', 'admin_notifications'):
        res = db[col].update_one({'_id': oid}, update)
        if res.matched_count:
            return True
    return False


def _try_delete_notif(db, notification_id):
    """Supprime par _id dans les deux collections."""
    from bson import ObjectId
    try:
        oid = ObjectId(notification_id)
    except Exception:
        return False
    for col in ('notifications', 'admin_notifications'):
        res = db[col].delete_one({'_id': oid})
        if res.deleted_count:
            return True
    return False


# ── Helper pour résoudre l'admin_id entier ───────────────────────────────────
def _resolve_admin_id(request, db):
    """Retourne toujours l'id entier depuis dashboard_utilisateur."""
    raw_id = request.session.get('user_id', '')
    try:
        return int(raw_id)  # '1' → 1 ✅
    except (ValueError, TypeError):
        pass
    # ObjectId string → chercher par username
    username = getattr(get_session_user(request), 'username', None)
    if username:
        u = db['utilisateurs'].find_one({'username': username}, {'_id': 1})
        if u and u.get('id'):
            return int(u['id'])
    return raw_id


@session_required
# ====================== NOTIFICATIONS ADMIN - VUES CORRIGÉES ======================

def _get_employe_from_session(request, db):
    """Récupère l'employé depuis la session (cohérent avec le côté employé)"""
    user_id  = request.session.get('user_id', '')
    username = request.session.get('username', '')
    employe  = db['employees'].find_one({'django_user_id': user_id})
    if not employe:
        employe = db['employees'].find_one({'django_username': username})
    return employe


@session_required
def admin_notifications(request):
    """Centre de notifications pour les administrateurs."""
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')

    from django.utils import timezone

    db, algiers = _get_db_algiers()
    now         = timezone.now()

    # ── 1. Toutes les notifications employés (ce que l'admin surveille) ──
    #    Collection : 'notifications'  →  champ : employe_id
    notifs_employes = list(
        db['notifications']
        .find({})
        .sort('created_at', -1)
        .limit(300)
    )

    # ── 2. Notifications admin-spécifiques (alertes système, etc.) ──
    #    Collection : 'admin_notifications'
    notifs_admin = list(
        db['admin_notifications']
        .find({})
        .sort('created_at', -1)
        .limit(100)
    )

    # ── Fusion + normalisation ────────────────────────────────────────
    notifications = (
        [_normalize_notif(n, 'notifications',       algiers) for n in notifs_employes] +
        [_normalize_notif(n, 'admin_notifications', algiers) for n in notifs_admin]
    )

    # Trier par date décroissante (None en dernier)
    notifications.sort(
        key=lambda x: x.get('created_at') or now.replace(year=2000),
        reverse=True,
    )

    # Enrichir avec le nom de l'employé quand disponible
    emp_ids = list({n.get('employe_id') for n in notifications if n.get('employe_id')})
    emp_map = {}
    if emp_ids:
        from bson import ObjectId
        oids = []
        for eid in emp_ids:
            try:
                oids.append(ObjectId(eid))
            except Exception:
                pass
        if oids:
            for emp in db['employees'].find({'_id': {'$in': oids}}, {'prenom': 1, 'nom': 1}):
                emp_map[str(emp['_id'])] = f"{emp.get('prenom', '')} {emp.get('nom', '')}".strip()

    for n in notifications:
        eid = n.get('employe_id')
        n['employe_nom'] = emp_map.get(str(eid), '') if eid else ''

    unread_count = sum(1 for n in notifications if n.get('status') == 'non_lu')

    return render(request, 'dashboard/admin_notifications.html', {
        'notifications': notifications,
        'unread_count':  unread_count,
    })


@session_required
def api_admin_mark_notification_read(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    from django.utils import timezone
    from bson import ObjectId

    db, _           = _get_db_algiers()
    now             = timezone.now()
    update_op       = {'$set': {'status': 'lu', 'read_at': now}}

    try:
        data            = json.loads(request.body)
        notification_id = data.get('notification_id')

        if notification_id:
            # Marquer UNE notification (essaie les deux collections)
            _try_update_notif(db, notification_id, update_op, now)
        else:
            # Marquer TOUTES dans les deux collections
            db['notifications'].update_many(
                {'status': 'non_lu'},
                update_op,
            )
            db['admin_notifications'].update_many(
                {'status': 'non_lu'},
                update_op,
            )

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@session_required
def api_admin_delete_notification(request):
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    db, _ = _get_db_algiers()

    try:
        data            = json.loads(request.body)
        notification_id = data.get('notification_id')

        if notification_id:
            _try_delete_notif(db, notification_id)
        else:
            db['notifications'].delete_many({})
            db['admin_notifications'].delete_many({})

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@session_required
def api_admin_send_test_notification(request):
    """API - envoyer une notification de test admin"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    from django.utils import timezone

    db, _    = _get_db_algiers()
    admin_id = _resolve_admin_id(request, db)

    db['admin_notifications'].insert_one({
        'admin_id':   admin_id,
        'titre':      '🔔 Notification de test',
        'message':    "Ceci est une notification de test pour le centre d'administration.",
        'categorie':  'info',
        'icon':       '🔔',
        'status':     'non_lu',
        'action_url': '/admin/notifications/',
        'created_at': timezone.now(),
    })
    return JsonResponse({'status': 'success', 'message': 'Notification envoyée'})

# ====================== FONCTIONS D'ENVOI DE NOTIFICATIONS ADMIN ======================

def send_admin_notification(admin_id, titre, message, categorie='info', icon='🔔', action_url=None, reservation_id=None):
    """Envoie une notification à un administrateur"""
    from datetime import datetime
    
    notification = {
        'admin_id': admin_id,
        'titre': titre,
        'message': message,
        'categorie': categorie,
        'icon': icon,
        'status': 'non_lu',
        'action_url': action_url,
        'reservation_id': reservation_id,
        'created_at': datetime.now()
    }
    
    db.admin_notifications.insert_one(notification)


def send_notification_to_all_admins(titre, message, categorie='info', icon='🔔', action_url=None, reservation_id=None):
    """Envoie une notification à tous les administrateurs (PyMongo direct)"""
    from datetime import datetime
    try:
        admins = list(db['utilisateurs'].find(
            {'is_staff': True, 'is_active': True}, {'_id': 1}
        ))
        if not admins:
            return
        notifications = [{
            'admin_id':       admin.get('_id'),
            'titre':          titre,
            'message':        message,
            'categorie':      categorie,
            'icon':           icon,
            'status':         'non_lu',
            'action_url':     action_url,
            'reservation_id': reservation_id,
            'created_at':     datetime.now(),
        } for admin in admins]
        if notifications:
            db.admin_notifications.insert_many(notifications)
    except Exception as _e:
        import logging; logging.getLogger(__name__).warning(f"send_notification_to_all_admins échoué: {_e}")


# ====================== ÉVÉNEMENTS DÉCLENCHANT DES NOTIFICATIONS ADMIN ======================

# 1. Nouvelle réservation créée par un employé
def notify_admins_new_reservation(employe, reservation_data):
    """Notifie tous les admins d'une nouvelle réservation"""
    titre = f"🆕 Nouvelle réservation en attente"
    message = f"{employe.get('prenom', '')} {employe.get('nom', '')} a demandé une réservation pour '{reservation_data.get('titre')}' le {reservation_data['date_debut'].strftime('%d/%m/%Y à %H:%M')}."
    send_notification_to_all_admins(
        titre=titre,
        message=message,
        categorie='reservation',
        icon='🆕',
        action_url=f'/reservations/{reservation_data.get("id")}/',
        reservation_id=reservation_data.get('id')
    )


# 2. Alerte de sécurité (tentative d'accès non autorisée)
def notify_admins_security_alert(zone, badge_id, message):
    """Notifie les admins d'une alerte de sécurité"""
    titre = f"⚠️ ALERTE SÉCURITÉ"
    message_complet = f"Tentative d'accès non autorisée détectée.\nZone: {zone}\nBadge: {badge_id}\nDétails: {message}"
    send_notification_to_all_admins(
        titre=titre,
        message=message_complet,
        categorie='alerte',
        icon='⚠️'
    )


# 3. Équipement hors ligne / maintenance
def notify_admins_equipment_offline(equipement_nom, equipement_id):
    """Notifie les admins qu'un équipement est hors ligne"""
    titre = f"🔧 Équipement hors ligne"
    message = f"L'équipement '{equipement_nom}' est actuellement hors ligne. Une intervention est nécessaire."
    send_notification_to_all_admins(
        titre=titre,
        message=message,
        categorie='maintenance',
        icon='🔧',
        action_url=f'/equipements/{equipement_id}/'
    )


# 4. Réservation modifiée/annulée par un employé
def notify_admins_reservation_cancelled(employe, reservation):
    """Notifie les admins qu'une réservation a été annulée"""
    titre = f"🗑️ Réservation annulée"
    message = f"{employe.get('prenom', '')} {employe.get('nom', '')} a annulé sa réservation '{reservation.get('titre')}'."
    send_notification_to_all_admins(
        titre=titre,
        message=message,
        categorie='reservation',
        icon='🗑️',
        action_url=f'/reservations/{reservation.get("id")}/'
    )


# 5. Réservation bientôt pleine (alerte occupation)
def notify_admins_high_occupation(zone_nom, occupation_rate):
    """Notifie les admins qu'une zone a un taux d'occupation élevé"""
    if occupation_rate >= 80:
        titre = f"📊 Taux d'occupation critique"
        message = f"La zone '{zone_nom}' a atteint {occupation_rate}% d'occupation. Une attention particulière est recommandée."
        send_notification_to_all_admins(
            titre=titre,
            message=message,
            categorie='alerte',
            icon='📊'
        )
        # dashboard/views.py - Ajoutez ces fonctions

# ====================== NOTIFICATIONS ADMINISTRATEUR ======================

def send_admin_notification(admin_id, titre, message, categorie='info', icon='🔔', action_url=None, reservation_id=None):
    """Envoie une notification à un administrateur spécifique"""
    from datetime import datetime
    
    notification = {
        'admin_id': admin_id,
        'titre': titre,
        'message': message,
        'categorie': categorie,
        'icon': icon,
        'status': 'non_lu',
        'action_url': action_url,
        'reservation_id': reservation_id,
        'created_at': datetime.now()
    }
    
    db.admin_notifications.insert_one(notification)


def send_notification_to_all_admins(titre, message, categorie='info', icon='🔔', action_url=None, reservation_id=None):
    """Envoie une notification à tous les administrateurs (PyMongo direct)"""
    from datetime import datetime
    try:
        admins = list(db['utilisateurs'].find(
            {'is_staff': True, 'is_active': True}, {'_id': 1}
        ))
        if not admins:
            return
        notifications = [{
            'admin_id':       admin.get('_id'),
            'titre':          titre,
            'message':        message,
            'categorie':      categorie,
            'icon':           icon,
            'status':         'non_lu',
            'action_url':     action_url,
            'reservation_id': reservation_id,
            'created_at':     datetime.now(),
        } for admin in admins]
        if notifications:
            db.admin_notifications.insert_many(notifications)
    except Exception as _e:
        import logging; logging.getLogger(__name__).warning(f"send_notification_to_all_admins échoué: {_e}")


def notify_admins_new_reservation(employe, reservation_data):
    """Notifie tous les admins d'une nouvelle réservation"""
    titre = f"🆕 Nouvelle réservation en attente"
    message = f"{employe.get('prenom', '')} {employe.get('nom', '')} a demandé une réservation pour '{reservation_data.get('titre')}' le {reservation_data['date_debut'].strftime('%d/%m/%Y à %H:%M')}."
    send_notification_to_all_admins(
        titre=titre,
        message=message,
        categorie='reservation',
        icon='🆕',
        action_url=f'/reservations/{reservation_data.get("id")}/',
        reservation_id=reservation_data.get('id')
    )


def notify_admins_security_alert(zone, badge_id, message):
    """Notifie les admins d'une alerte de sécurité"""
    titre = f"⚠️ ALERTE SÉCURITÉ"
    message_complet = f"Tentative d'accès non autorisée détectée.\nZone: {zone}\nBadge: {badge_id}\nDétails: {message}"
    send_notification_to_all_admins(
        titre=titre,
        message=message_complet,
        categorie='alerte',
        icon='⚠️'
    )


def notify_admins_equipment_offline(equipement_nom, equipement_id):
    """Notifie les admins qu'un équipement est hors ligne"""
    titre = f"🔧 Équipement hors ligne"
    message = f"L'équipement '{equipement_nom}' est actuellement hors ligne. Une intervention est nécessaire."
    send_notification_to_all_admins(
        titre=titre,
        message=message,
        categorie='maintenance',
        icon='🔧',
        action_url=f'/equipements/{equipement_id}/'
    )


def notify_admins_reservation_cancelled(employe, reservation):
    """Notifie les admins qu'une réservation a été annulée"""
    titre = f"🗑️ Réservation annulée"
    message = f"{employe.get('prenom', '')} {employe.get('nom', '')} a annulé sa réservation '{reservation.get('titre')}'."
    send_notification_to_all_admins(
        titre=titre,
        message=message,
        categorie='reservation',
        icon='🗑️',
        action_url=f'/reservations/{reservation.get("id")}/'
    )


def notify_admins_high_occupation(zone_nom, occupation_rate):
    """Notifie les admins qu'une zone a un taux d'occupation élevé"""
    if occupation_rate >= 80:
        titre = f"📊 Taux d'occupation critique"
        message = f"La zone '{zone_nom}' a atteint {occupation_rate}% d'occupation. Une attention particulière est recommandée."
        send_notification_to_all_admins(
            titre=titre,
            message=message,
            categorie='alerte',
            icon='📊'
        )


# ====================== APIS POUR NOTIFICATIONS ADMIN ======================

@session_required
def api_admin_notifications_unread_count(request):
    if not request.session.get('is_staff', False):
        return JsonResponse({'count': 0})
    try:
        db, _ = _get_db_algiers()

        count = (
            db['notifications'].count_documents({'status': 'non_lu'}) +
            db['admin_notifications'].count_documents({'status': 'non_lu'})
        )
        return JsonResponse({'count': count})
    except Exception:
        return JsonResponse({'count': 0})

@session_required
def api_admin_mark_notification_read(request):
    """API pour marquer une notification admin comme lue"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    try:
        import json
        from bson import ObjectId
        from datetime import datetime
        
        data = json.loads(request.body)
        notification_id = data.get('notification_id')
        
        if notification_id:
            db.admin_notifications.update_one(
                {'_id': ObjectId(notification_id), 'admin_id': request.session.get('user_id', '')},
                {'$set': {'status': 'lu', 'read_at': datetime.now()}}
            )
        else:
            # Marquer toutes comme lues
            db.admin_notifications.update_many(
                {'admin_id': request.session.get('user_id', ''), 'status': 'non_lu'},
                {'$set': {'status': 'lu', 'read_at': datetime.now()}}
            )
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_admin_delete_notification(request):
    """API pour supprimer une notification admin"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    try:
        import json
        from bson import ObjectId
        
        data = json.loads(request.body)
        notification_id = data.get('notification_id')
        
        if notification_id:
            db.admin_notifications.delete_one({
                '_id': ObjectId(notification_id),
                'admin_id': request.session.get('user_id', '')
            })
        else:
            db.admin_notifications.delete_many({'admin_id': request.session.get('user_id', '')})
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_admin_send_test_notification(request):
    """API - envoyer une notification de test admin"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    from django.utils import timezone

    db, _    = _get_db_algiers()
    admin_id = _resolve_admin_id(request, db)

    db['admin_notifications'].insert_one({
        'admin_id':   admin_id,
        'titre':      '🔔 Notification de test',
        'message':    "Ceci est une notification de test pour le centre d'administration.",
        'categorie':  'info',
        'icon':       '🔔',
        'status':     'non_lu',
        'action_url': '/admin/notifications/',
        'created_at': timezone.now(),
    })
    return JsonResponse({'status': 'success', 'message': 'Notification envoyée'})
def notify_admin_new_reservation(employe, reservation_data, reservation_id):
    """Notifie les administrateurs d'une nouvelle réservation"""
    from django.contrib.auth import get_user_model
    from datetime import datetime
    
    User = get_user_model()
    admins = User.objects.filter(is_staff=True, is_active=True)
    
    for admin in admins:
        admin_notification = {
            'admin_id': admin.id,
            'titre': '🆕 Nouvelle réservation en attente',
            'message': f"{employe.get('prenom', '')} {employe.get('nom', '')} a demandé une réservation pour '{reservation_data.get('titre')}'.",
            'categorie': 'reservation',
            'icon': '🆕',
            'status': 'non_lu',
            'action_url': f'/reservations/{reservation_id}/',
            'reservation_id': reservation_id,
            'created_at': datetime.now()
        }
        db.admin_notifications.insert_one(admin_notification)
     # ====================== MOT DE PASSE OUBLIÉ ======================
# ====================== MOT DE PASSE OUBLIÉ ======================

def password_forgot(request):
    """Étape 1 : envoi du lien de réinitialisation par email."""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        if not email:
            messages.error(request, "Veuillez saisir une adresse email.")
            return render(request, 'dashboard/password_forgot.html')

        # Chercher dans MongoDB uniquement
        user_doc = db['utilisateurs'].find_one(
            {'email': {'$regex': f'^{re.escape(email)}$', '$options': 'i'},
             'is_active': True}
        )

        # Toujours afficher le même message (sécurité anti-énumération)
        msg_generique = "Si cet email existe dans notre système, un lien vous a été envoyé."

        if not user_doc:
            messages.success(request, msg_generique)
            return render(request, 'dashboard/password_forgot.html')

        # Invalider les anciens tokens MongoDB
        db['password_reset_tokens'].update_many(
            {'user_id': str(user_doc['_id']), 'used': False},
            {'$set': {'used': True}}
        )

        # Générer un nouveau token sécurisé
        import secrets
        token = secrets.token_urlsafe(48)
        expires_at = datetime.now() + timedelta(hours=1)

        db['password_reset_tokens'].insert_one({
            'user_id':    str(user_doc['_id']),
            'token':      token,
            'expires_at': expires_at,
            'used':       False,
            'created_at': datetime.now(),
            'email':      user_doc.get('email', ''),
        })

        reset_url = request.build_absolute_uri(f"/password-reset/{token}/")
        prenom = user_doc.get('first_name') or user_doc.get('prenom') or user_doc.get('username', '')

        corps_texte = (
            f"Bonjour {prenom},\n\n"
            f"Cliquez sur ce lien pour réinitialiser votre mot de passe (valable 1 heure) :\n\n"
            f"{reset_url}\n\n"
            f"Si vous n'avez pas fait cette demande, ignorez cet email.\n\n— L'équipe SIGR-CA"
        )
        corps_html = f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#0a0c10;font-family:'Segoe UI',sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#0a0c10;padding:40px 0;">
    <tr><td align="center">
      <table width="520" cellpadding="0" cellspacing="0"
             style="background:#111318;border-radius:16px;border:1px solid rgba(255,255,255,0.07);overflow:hidden;">
        <tr>
          <td style="background:linear-gradient(135deg,#1f6feb,#06b6d4);padding:32px;text-align:center;">
            <div style="font-size:36px;margin-bottom:12px;">🔐</div>
            <h1 style="margin:0;color:#fff;font-size:22px;font-weight:600;">Réinitialisation du mot de passe</h1>
            <p style="margin:8px 0 0;color:rgba(255,255,255,0.75);font-size:14px;">SIGR-CA</p>
          </td>
        </tr>
        <tr>
          <td style="padding:36px 40px;">
            <p style="color:#9ca3af;font-size:15px;margin:0 0 12px;">Bonjour <strong style="color:#f3f4f6;">{prenom}</strong>,</p>
            <p style="color:#9ca3af;font-size:15px;margin:0 0 28px;line-height:1.6;">
              Cliquez sur le bouton ci-dessous pour choisir un nouveau mot de passe.
            </p>
            <div style="text-align:center;margin:0 0 28px;">
              <a href="{reset_url}"
                 style="display:inline-block;background:linear-gradient(135deg,#1f6feb,#06b6d4);
                        color:#fff;text-decoration:none;padding:14px 36px;border-radius:10px;
                        font-size:15px;font-weight:600;">
                Réinitialiser mon mot de passe
              </a>
            </div>
            <div style="background:rgba(31,111,235,0.08);border:1px solid rgba(31,111,235,0.2);
                        border-radius:8px;padding:14px 18px;margin-bottom:24px;">
              <p style="margin:0;color:#6b7280;font-size:13px;">
                ⏱ Lien valable <strong style="color:#f59e0b;">1 heure</strong> uniquement.<br>
                🔒 Si vous n'avez pas fait cette demande, ignorez cet email.
              </p>
            </div>
            <p style="color:#6b7280;font-size:12px;margin:0;word-break:break-all;">
              Lien alternatif : <a href="{reset_url}" style="color:#3b82f6;">{reset_url}</a>
            </p>
          </td>
        </tr>
        <tr>
          <td style="padding:20px 40px;border-top:1px solid rgba(255,255,255,0.05);text-align:center;">
            <p style="color:#4b5563;font-size:12px;margin:0;">© SIGR-CA — Email automatique, ne pas répondre.</p>
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body></html>"""

        try:
            import ssl, smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from django.conf import settings as _s

            msg_email = MIMEMultipart('alternative')
            msg_email['Subject'] = "SIGR-CA — Réinitialisation de votre mot de passe"
            msg_email['From']    = _s.DEFAULT_FROM_EMAIL
            msg_email['To']      = user_doc['email']
            msg_email.attach(MIMEText(corps_texte, 'plain', 'utf-8'))
            msg_email.attach(MIMEText(corps_html,  'html',  'utf-8'))

            ctx = ssl.create_default_context()
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.ehlo()
                server.starttls(context=ctx)
                server.ehlo()
                server.login(_s.EMAIL_HOST_USER, _s.EMAIL_HOST_PASSWORD)
                server.sendmail(_s.EMAIL_HOST_USER, user_doc['email'], msg_email.as_string())

            logger.info(f"Email reset envoyé à {user_doc['email']}")
        except Exception as e:
            logger.error(f"Erreur envoi email reset: {e}")
            messages.error(request, f"Erreur lors de l'envoi de l'email : {e}")
            return render(request, 'dashboard/password_forgot.html')

        messages.success(request, msg_generique)
        return render(request, 'dashboard/password_forgot.html')

    return render(request, 'dashboard/password_forgot.html')


def password_reset_confirm(request, token):
    """Étape 2 : saisie du nouveau mot de passe via le lien."""
    # Vérifier le token dans MongoDB
    token_doc = db['password_reset_tokens'].find_one({'token': token, 'used': False})

    if not token_doc:
        messages.error(request, "Lien invalide ou déjà utilisé.")
        return redirect('password_forgot')

    if token_doc['expires_at'] < datetime.now():
        messages.error(request, "Ce lien a expiré. Veuillez en demander un nouveau.")
        db['password_reset_tokens'].update_one({'token': token}, {'$set': {'used': True}})
        return redirect('password_forgot')

    if request.method == 'POST':
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        errors = []
        if not password1:
            errors.append("Le nouveau mot de passe est obligatoire.")
        if password1 != password2:
            errors.append("Les deux mots de passe ne correspondent pas.")
        if len(password1) < 8:
            errors.append("Le mot de passe doit contenir au moins 8 caractères.")
        if password1.isdigit():
            errors.append("Le mot de passe ne peut pas être uniquement numérique.")
        if password1.lower() in ['password', 'motdepasse', '12345678', 'azertyuiop']:
            errors.append("Ce mot de passe est trop commun.")

        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'dashboard/password_reset_form.html', {'token': token})

        # Hasher avec bcrypt et mettre à jour dans MongoDB
        import bcrypt
        from bson import ObjectId
        hashed = bcrypt.hashpw(password1.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        db['utilisateurs'].update_one(
            {'_id': ObjectId(token_doc['user_id'])},
            {'$set': {'password': hashed, 'updated_at': datetime.now()}}
        )

        # Invalider le token
        db['password_reset_tokens'].update_one(
            {'token': token},
            {'$set': {'used': True, 'used_at': datetime.now()}}
        )

        # Invalider toutes les sessions actives (MongoDB)
        try:
            db['dashboard_usersession'].update_many(
                {'user_id': token_doc['user_id']},
                {'$set': {'is_active': False, 'logout_time': datetime.now()}}
            )
        except Exception as _se:
            logger.warning(f"Invalidation sessions échouée: {_se}")

        logger.info(f"Mot de passe réinitialisé pour user_id={token_doc['user_id']}")
        return redirect('password_reset_done')

    return render(request, 'dashboard/password_reset_form.html', {'token': token})


def password_reset_done(request):
    """Étape 3 : confirmation après réinitialisation réussie."""
    return render(request, 'dashboard/password_reset_done.html')
# ====================== GESTION HIÉRARCHIQUE DES RESSOURCES ======================

@session_required
def gestion_hierarchique(request):
    """Gestion hiérarchique des ressources (domaines → sites → bâtiments → étages → salles)"""
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')
    
    # Récupérer la hiérarchie complète
    domaines = list(db.domainesp.find()) if 'domaines' in db.list_collection_names() else []
    for d in domaines:
        d['id'] = str(d['_id'])
        d['sites'] = list(db.sites.find({'domaine_id': d['_id']}))
        for s in d['sites']:
            s['id'] = str(s['_id'])
            s['batiments'] = list(db.batiments.find({'site_id': s['_id']}))
            for b in s['batiments']:
                b['id'] = str(b['_id'])
                b['etages'] = list(db.etages.find({'batiment_id': b['_id']}))
                for e in b['etages']:
                    e['id'] = str(e['_id'])
                    e['salles'] = list(db.bureaux.find({'etage_id': e['_id']}))
                    for salle in e['salles']:
                        salle['id'] = str(salle['_id'])
    
    # Statistiques
    stats = {
        'total_domaines': len(domaines),
        'total_sites': db.sites.count_documents({}),
        'total_batiments': db.batiments.count_documents({}),
        'total_etages': db.etages.count_documents({}),
        'total_salles': db.bureaux.count_documents({}),
    }
    
    return render(request, 'dashboard/gestion_hierarchique.html', {
        'domaines': domaines,
        'stats': stats,
    })


@session_required
def api_hierarchie_ajouter(request):
    """API pour ajouter un élément dans la hiérarchie"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        niveau = data.get('niveau')  # domaine, site, batiment, etage
        parent_id = data.get('parent_id')
        nom = data.get('nom', '').strip()
        code = data.get('code', '').strip()
        description = data.get('description', '')
        
        if not nom:
            return JsonResponse({'error': 'Le nom est requis'}, status=400)
        
        collection_name = f"{niveau}s"  # domaines, sites, batiments, etages
        if collection_name not in db.list_collection_names():
            db.create_collection(collection_name)
        
        doc = {
            'nom': nom,
            'code': code,
            'description': description,
            'created_at': datetime.now(),
            'created_by': request.session.get('username', ''),
        }
        
        if parent_id and niveau != 'domaine':
            doc[f'{get_parent_field(niveau)}_id'] = ObjectId(parent_id)
        
        result = db[collection_name].insert_one(doc)
        
        return JsonResponse({
            'status': 'success',
            'id': str(result.inserted_id),
            'message': f'{niveau.capitalize()} "{nom}" ajouté(e)'
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_hierarchie_supprimer(request, niveau, element_id):
    """API pour supprimer un élément de la hiérarchie"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        collection_name = f"{niveau}s"
        
        # Vérifier les dépendances
        if niveau == 'domaine':
            has_children = db.sites.count_documents({'domaine_id': ObjectId(element_id)}) > 0
        elif niveau == 'site':
            has_children = db.batiments.count_documents({'site_id': ObjectId(element_id)}) > 0
        elif niveau == 'batiment':
            has_children = db.etages.count_documents({'batiment_id': ObjectId(element_id)}) > 0
        elif niveau == 'etage':
            has_children = db.bureaux.count_documents({'etage_id': ObjectId(element_id)}) > 0
        else:
            has_children = False
        
        if has_children:
            return JsonResponse({
                'status': 'error',
                'message': 'Impossible de supprimer : des éléments enfants existent'
            }, status=400)
        
        result = db[collection_name].delete_one({'_id': ObjectId(element_id)})
        
        if result.deleted_count > 0:
            return JsonResponse({'status': 'success', 'message': 'Supprimé'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Élément non trouvé'}, status=404)
            
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def get_parent_field(niveau):
    """Retourne le nom du champ parent selon le niveau"""
    mapping = {
        'site': 'domaine_id',
        'batiment': 'site_id',
        'etage': 'batiment_id',
    }
    return mapping.get(niveau, 'parent_id')


# ====================== GESTION DES INDISPONIBILITÉS PLANIFIÉES ======================

@session_required
def gestion_indisponibilites(request):
    """Gestion des indisponibilités planifiées (maintenance, réservations bloquées)"""
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')

    from datetime import datetime, timedelta

    # Créer la collection si elle n'existe pas
    if 'indisponibilites' not in db.list_collection_names():
        db.create_collection('indisponibilites')

    # Récupérer toutes les indisponibilités
    now = datetime.now()
    indispos_raw = list(db.indisponibilites.find().sort('date_debut', -1))

    # Liste affichée dans le template (peut garder les datetime pour |date)
    indispos = []
    # Liste JSON-safe pour json_script (uniquement des types sérialisables)
    indispos_json = []

    for i in indispos_raw:
        indispo_id = str(i['_id'])

        # Statut calculé
        if i.get('date_debut') and i.get('date_fin'):
            if i['date_debut'] <= now <= i['date_fin']:
                statut = 'en_cours'
            elif i['date_debut'] > now:
                statut = 'a_venir'
            else:
                statut = 'passee'
        else:
            statut = 'a_venir'

        # Nom de la ressource
        ressource_nom = ''
        if i.get('ressource_type') == 'salle':
            salle = db.bureaux.find_one({'_id': i.get('ressource_id')})
            ressource_nom = salle['nom'] if salle else 'Inconnue'
        elif i.get('ressource_type') == 'materiel':
            materiel = db.materiels.find_one({'_id': i.get('ressource_id')})
            ressource_nom = materiel['nom'] if materiel else 'Inconnu'

        # ---- Objet pour l'affichage HTML (Django sait formater les datetime) ----
        indispos.append({
            'id':            indispo_id,
            'titre':         i.get('titre', ''),
            'description':   i.get('description', ''),
            'type_indispo':  i.get('type_indispo', 'maintenance'),
            'date_debut':    i.get('date_debut'),
            'date_fin':      i.get('date_fin'),
            'statut':        statut,
            'ressource_nom': ressource_nom,
            'ressource_id':  str(i.get('ressource_id')) if i.get('ressource_id') else '',
            'ressource_type': i.get('ressource_type', ''),
        })

        # ---- Objet 100% JSON-safe pour json_script (pas d'ObjectId, dates en ISO) ----
        indispos_json.append({
            'id':            indispo_id,
            'titre':         i.get('titre', ''),
            'description':   i.get('description', ''),
            'type_indispo':  i.get('type_indispo', 'maintenance'),
            'date_debut':    i['date_debut'].isoformat() if i.get('date_debut') else '',
            'date_fin':      i['date_fin'].isoformat()   if i.get('date_fin')   else '',
            'statut':        statut,
            'ressource_nom': ressource_nom,
            'ressource_id':  str(i.get('ressource_id')) if i.get('ressource_id') else '',
            'ressource_type': i.get('ressource_type', ''),
        })

    # Ressources construites une seule fois
    salles = list(db.bureaux.find())
    for s in salles:
        s['id'] = str(s['_id'])
        s['type'] = 'salle'
        s['type_icon'] = '🚪'
        s.pop('_id', None)  # éviter tout ObjectId résiduel

    materiels = list(db.materiels.find()) if 'materiels' in db.list_collection_names() else []
    for m in materiels:
        m['id'] = str(m['_id'])
        m['type'] = 'materiel'
        m['type_icon'] = get_materiel_icon(m.get('categorie', 'autre'))
        m.pop('_id', None)

    ressources = salles + materiels

    # Stats calculées une seule fois
    stats = {
        'en_cours': db.indisponibilites.count_documents({
            'date_debut': {'$lte': now},
            'date_fin':   {'$gte': now},
        }),
        'a_venir': db.indisponibilites.count_documents({'date_debut': {'$gt': now}}),
        'passees': db.indisponibilites.count_documents({'date_fin': {'$lt': now}}),
        'total':   db.indisponibilites.count_documents({}),
    }

    return render(request, 'dashboard/gestion_indisponibilites.html', {
        'indisponibilites':      indispos,
        'indisponibilites_json': indispos_json,  # <-- pour json_script
        'ressources':            ressources,
        'stats':                 stats,
        'now':                   now,
    })

@session_required
def api_indisponibilite_ajouter(request):
    """API pour ajouter une indisponibilité planifiée"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
 
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
 
    try:
        data = json.loads(request.body)
 
        ressource_id   = data.get('ressource_id')
        ressource_type = data.get('ressource_type')
        titre          = data.get('titre', '').strip()
        description    = data.get('description', '')
        date_debut     = datetime.fromisoformat(data.get('date_debut'))
        date_fin       = datetime.fromisoformat(data.get('date_fin'))
        type_indispo   = data.get('type_indispo', 'maintenance')
        recurrence     = data.get('recurrence', 'none')
        recurrence_end = data.get('recurrence_end')
 
        if not ressource_id or not ressource_type:
            return JsonResponse({'error': 'Ressource non spécifiée'}, status=400)
 
        if not titre:
            return JsonResponse({'error': 'Le titre est obligatoire'}, status=400)
 
        if date_fin <= date_debut:
            return JsonResponse({'error': 'La date de fin doit être après la date de début'}, status=400)
 
        # Vérifier les conflits
        conflit = db.indisponibilites.find_one({
            'ressource_id': ObjectId(ressource_id),
            '$or': [
                {'date_debut': {'$lt': date_fin,  '$gte': date_debut}},
                {'date_fin':   {'$gt': date_debut, '$lte': date_fin}},
                {'date_debut': {'$lte': date_debut}, 'date_fin': {'$gte': date_fin}},
            ],
        })
 
        if conflit:
            return JsonResponse({
                'error': 'Un conflit existe avec une autre indisponibilité sur cette période'
            }, status=400)
 
        indispo = {
            'ressource_id':   ObjectId(ressource_id),
            'ressource_type': ressource_type,
            'titre':          titre,
            'description':    description,
            'date_debut':     date_debut,
            'date_fin':       date_fin,
            'type_indispo':   type_indispo,
            'recurrence':     recurrence if recurrence != 'none' else None,
            'created_at':     datetime.now(),
            'created_by':     request.session.get('username', ''),
        }
 
        if recurrence_end and recurrence != 'none':
            indispo['recurrence_end'] = datetime.fromisoformat(recurrence_end)
 
        result = db.indisponibilites.insert_one(indispo)
 
        if recurrence != 'none' and recurrence_end:
            generate_recurring_indisponibilities(indispo, result.inserted_id)
 
        # Notifier les admins
        from dashboard.views import send_notification_to_all_admins
        ressource_nom = get_ressource_name(ressource_id, ressource_type)
        # FIX: nb_notifies initialisé avant utilisation
        nb_notifies = send_notification_to_all_admins(
            titre=f"🔧 Indisponibilité planifiée",
            message=(
                f"{ressource_nom} sera indisponible du "
                f"{date_debut.strftime('%d/%m/%Y %H:%M')} au "
                f"{date_fin.strftime('%d/%m/%Y %H:%M')}"
            ),
            categorie='maintenance',
            icon='🔧',
        )
        nb_notifies = nb_notifies or 0
 
        return JsonResponse({
            'status':      'success',
            'success':     True,
            'id':          str(result.inserted_id),
            'nb_notifies': nb_notifies,
            'message':     f"Indisponibilité créée. {nb_notifies} employé(s) notifié(s).",
        })
 
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
@session_required
def api_indisponibilite_modifier(request, indispo_id):
    """API pour modifier une indisponibilité"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
 
    # FIX: accepter PUT (aligné avec le fetch du frontend)
    if request.method != 'PUT':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
 
    try:
        data = json.loads(request.body)
 
        update_data = {
            'titre':       data.get('titre'),
            'description': data.get('description'),
            'type_indispo': data.get('type_indispo'),
            'updated_at':  datetime.now(),
            'updated_by':  request.session.get('username', ''),
        }
 
        if data.get('date_debut'):
            update_data['date_debut'] = datetime.fromisoformat(data['date_debut'])
        if data.get('date_fin'):
            update_data['date_fin'] = datetime.fromisoformat(data['date_fin'])
 
        result = db.indisponibilites.update_one(
            {'_id': ObjectId(indispo_id)},
            {'$set': update_data},
        )
 
        if result.modified_count > 0:
            return JsonResponse({'status': 'success', 'success': True, 'message': 'Indisponibilité modifiée.'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Non modifié'}, status=400)
 
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
@session_required
def api_indisponibilite_supprimer(request, indispo_id):
    """API pour supprimer une indisponibilité"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
 
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
 
    try:
        result = db.indisponibilites.delete_one({'_id': ObjectId(indispo_id)})
 
        if result.deleted_count > 0:
            return JsonResponse({'status': 'success', 'success': True, 'message': 'Indisponibilité supprimée.'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Non trouvé'}, status=404)
 
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
@session_required
def api_ressources_disponibles(request):
    """API pour récupérer les ressources disponibles sur une période"""
    date_debut_str = request.GET.get('date_debut')
    date_fin_str   = request.GET.get('date_fin')
    type_ressource = request.GET.get('type', 'all')
 
    if not date_debut_str or not date_fin_str:
        return JsonResponse({'ressources': []})
 
    try:
        date_debut = datetime.fromisoformat(date_debut_str)
        date_fin   = datetime.fromisoformat(date_fin_str)
 
        indispos = list(db.indisponibilites.find({
            'date_debut': {'$lt': date_fin},
            'date_fin':   {'$gt': date_debut},
        }))
 
        ressources_indispo_ids = {str(i['ressource_id']) for i in indispos}
        ressources_disponibles = []
 
        if type_ressource in ['all', 'salle']:
            for s in db.bureaux.find({'statut': 'actif'}):
                if str(s['_id']) not in ressources_indispo_ids:
                    ressources_disponibles.append({
                        'id':       str(s['_id']),
                        'nom':      s['nom'],
                        'type':     'salle',
                        'capacite': s.get('capacite_max', 10),
                        'icone':    '🚪',
                    })
 
        if type_ressource in ['all', 'materiel']:
            if 'materiels' in db.list_collection_names():
                for m in db.materiels.find({'statut': 'disponible'}):
                    if str(m['_id']) not in ressources_indispo_ids:
                        ressources_disponibles.append({
                            'id':        str(m['_id']),
                            'nom':       m['nom'],
                            'type':      'materiel',
                            'categorie': m.get('categorie', 'autre'),
                            'icone':     get_materiel_icon(m.get('categorie', 'autre')),
                        })
 
        return JsonResponse({'ressources': ressources_disponibles})
 
    except Exception as e:
        return JsonResponse({'ressources': [], 'error': str(e)})
 
def generate_recurring_indisponibilities(parent_indispo, parent_id):
    """Génère les occurrences récurrentes d'une indisponibilité"""
    recurrence     = parent_indispo.get('recurrence')
    recurrence_end = parent_indispo.get('recurrence_end')
    date_debut     = parent_indispo['date_debut']
    date_fin       = parent_indispo['date_fin']
    duration       = date_fin - date_debut
 
    if not recurrence_end:
        return
 
    current_start = date_debut
    occurrences   = []
 
    while current_start <= recurrence_end:
        if current_start != date_debut:  # Ne pas dupliquer l'original
            occurrence = {k: v for k, v in parent_indispo.items() if k != '_id'}  # FIX: exclure _id
            occurrence['parent_id']  = parent_id
            occurrence['date_debut'] = current_start
            occurrence['date_fin']   = current_start + duration
            occurrence['created_at'] = datetime.now()
            occurrences.append(occurrence)
 
        if recurrence == 'daily':
            current_start += timedelta(days=1)
        elif recurrence == 'weekly':
            current_start += timedelta(weeks=1)
        elif recurrence == 'monthly':
            month = current_start.month + 1
            year  = current_start.year + (1 if month > 12 else 0)
            month = month if month <= 12 else 1
            current_start = current_start.replace(year=year, month=month)
 
    if occurrences:
        db.indisponibilites.insert_many(occurrences)

# ====================== EXPORT CSV — INDISPONIBILITÉS ======================

@session_required
def api_export_indisponibilites_csv(request):
    import csv
    from datetime import datetime
    from django.http import HttpResponse

    if not request.session.get('is_staff', False):
        return HttpResponse("Non autorisé", status=403)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="indisponibilites_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    )
    response.write('\ufeff')  # BOM UTF-8 pour Excel

    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Titre', 'Type', 'Statut', 'Ressource',
        'Date début', 'Date fin', 'Durée (h)',
        'Récurrence', 'Description', 'Créé par', 'Créé le',
    ])

    now = datetime.now()
    type_labels = {
        'maintenance': 'Maintenance',
        'reservation_bloquee': 'Réservation bloquée',
        'fermeture': 'Fermeture',
    }

    for i in db.indisponibilites.find().sort('date_debut', -1):
        date_debut = i.get('date_debut')
        date_fin   = i.get('date_fin')

        if date_debut and date_fin:
            if date_debut <= now <= date_fin:   statut = 'En cours'
            elif date_debut > now:              statut = 'À venir'
            else:                               statut = 'Passée'
            duree = round((date_fin - date_debut).total_seconds() / 3600, 1)
        else:
            statut = '—'
            duree  = ''

        ressource_nom = '—'
        if i.get('ressource_type') == 'salle':
            s = db.bureaux.find_one({'_id': i.get('ressource_id')})
            ressource_nom = s['nom'] if s else 'Inconnue'
        elif i.get('ressource_type') == 'materiel':
            m = db.materiels.find_one({'_id': i.get('ressource_id')}) if 'materiels' in db.list_collection_names() else None
            ressource_nom = m['nom'] if m else 'Inconnu'

        writer.writerow([
            i.get('titre', ''),
            type_labels.get(i.get('type_indispo', ''), i.get('type_indispo', '')),
            statut,
            ressource_nom,
            date_debut.strftime('%d/%m/%Y %H:%M') if date_debut else '—',
            date_fin.strftime('%d/%m/%Y %H:%M')   if date_fin   else '—',
            duree,
            i.get('recurrence') or 'Aucune',
            i.get('description', ''),
            i.get('created_by', '—'),
            i['created_at'].strftime('%d/%m/%Y %H:%M') if i.get('created_at') else '—',
        ])

    return response


# ====================== EXPORT PDF — INDISPONIBILITÉS ======================

@session_required
def api_export_indisponibilites_pdf(request):
    import io
    from datetime import datetime
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )

    if not request.session.get('is_staff', False):
        return HttpResponse("Non autorisé", status=403)

    now    = datetime.now()
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm,    bottomMargin=2*cm,
    )
    styles   = getSampleStyleSheet()
    elements = []

    title_s = ParagraphStyle('T', parent=styles['Heading1'],
                             fontSize=18, textColor=colors.HexColor('#1f6feb'), spaceAfter=4)
    sub_s   = ParagraphStyle('S', parent=styles['Normal'],
                             fontSize=10, textColor=colors.grey, spaceAfter=10)
    section_s = ParagraphStyle('Sec', parent=styles['Heading2'],
                               fontSize=13, spaceBefore=14, spaceAfter=6,
                               textColor=colors.HexColor('#1f6feb'))
    small   = ParagraphStyle('Sm', fontSize=8)
    footer_s= ParagraphStyle('F',  fontSize=8, textColor=colors.grey)

    # En-tête
    elements.append(Paragraph("RAPPORT DES INDISPONIBILITÉS PLANIFIÉES", title_s))
    elements.append(Paragraph("Système Intégré de Gestion des Ressources — SIGR-CA", sub_s))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1f6feb')))
    elements.append(Spacer(1, 0.4*cm))

    # Résumé KPIs
    en_cours = db.indisponibilites.count_documents({'date_debut': {'$lte': now}, 'date_fin': {'$gte': now}})
    a_venir  = db.indisponibilites.count_documents({'date_debut': {'$gt': now}})
    passees  = db.indisponibilites.count_documents({'date_fin':   {'$lt': now}})
    total    = db.indisponibilites.count_documents({})

    def kpi_cell(label, value, color):
        return [
            Paragraph(label, ParagraphStyle('kl', fontSize=9, textColor=colors.grey, alignment=1)),
            Paragraph(str(value), ParagraphStyle('kv', fontSize=22, fontName='Helvetica-Bold',
                                                 textColor=colors.HexColor(color), alignment=1)),
        ]

    kpi_data = [
        [Paragraph('En cours', ParagraphStyle('kl', fontSize=9, textColor=colors.grey, alignment=1)),
         Paragraph('À venir',  ParagraphStyle('kl', fontSize=9, textColor=colors.grey, alignment=1)),
         Paragraph('Passées',  ParagraphStyle('kl', fontSize=9, textColor=colors.grey, alignment=1)),
         Paragraph('Total',    ParagraphStyle('kl', fontSize=9, textColor=colors.grey, alignment=1))],
        [Paragraph(str(en_cours), ParagraphStyle('kv', fontSize=22, fontName='Helvetica-Bold', textColor=colors.HexColor('#f85149'), alignment=1)),
         Paragraph(str(a_venir),  ParagraphStyle('kv', fontSize=22, fontName='Helvetica-Bold', textColor=colors.HexColor('#f59e0b'), alignment=1)),
         Paragraph(str(passees),  ParagraphStyle('kv', fontSize=22, fontName='Helvetica-Bold', textColor=colors.grey, alignment=1)),
         Paragraph(str(total),    ParagraphStyle('kv', fontSize=22, fontName='Helvetica-Bold', textColor=colors.HexColor('#1f6feb'), alignment=1))],
    ]
    t_kpi = Table(kpi_data, colWidths=[6.5*cm]*4)
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f0f6ff')),
        ('BOX',        (0,0), (-1,-1), 1,   colors.HexColor('#1f6feb')),
        ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#cce0ff')),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('PADDING',    (0,0), (-1,-1), 10),
    ]))
    elements.append(t_kpi)
    elements.append(Spacer(1, 0.6*cm))

    # Tableau détail
    elements.append(Paragraph("Détail de toutes les indisponibilités", section_s))

    indispos = list(db.indisponibilites.find().sort('date_debut', -1))
    type_labels = {'maintenance': 'Maintenance', 'reservation_bloquee': 'Réservation bloquée', 'fermeture': 'Fermeture'}

    if not indispos:
        elements.append(Paragraph("Aucune indisponibilité enregistrée.", styles['Normal']))
    else:
        hdr = [Paragraph(h, ParagraphStyle('H', fontSize=8, fontName='Helvetica-Bold', textColor=colors.white))
               for h in ["Titre", "Type", "Ressource", "Date début", "Date fin", "Durée (h)", "Statut", "Créé par"]]
        rows = [hdr]

        for i in indispos:
            date_debut = i.get('date_debut')
            date_fin   = i.get('date_fin')
            if date_debut and date_fin:
                if date_debut <= now <= date_fin:   statut_str = 'En cours';  s_color = colors.HexColor('#f85149')
                elif date_debut > now:              statut_str = 'À venir';   s_color = colors.HexColor('#f59e0b')
                else:                               statut_str = 'Passée';    s_color = colors.grey
                duree = round((date_fin - date_debut).total_seconds() / 3600, 1)
            else:
                statut_str = '—'; s_color = colors.grey; duree = '—'

            ressource_nom = '—'
            if i.get('ressource_type') == 'salle':
                s = db.bureaux.find_one({'_id': i.get('ressource_id')})
                ressource_nom = s['nom'] if s else 'Inconnue'
            elif i.get('ressource_type') == 'materiel':
                m = db.materiels.find_one({'_id': i.get('ressource_id')}) if 'materiels' in db.list_collection_names() else None
                ressource_nom = m['nom'] if m else 'Inconnu'

            t_color = {'maintenance': colors.HexColor('#f59e0b'),
                       'reservation_bloquee': colors.HexColor('#f85149'),
                       'fermeture': colors.HexColor('#1f6feb')}.get(i.get('type_indispo',''), colors.grey)

            rows.append([
                Paragraph(i.get('titre','—'),     ParagraphStyle('tb', fontSize=8, fontName='Helvetica-Bold')),
                Paragraph(type_labels.get(i.get('type_indispo',''), ''), ParagraphStyle('ty', fontSize=8, textColor=t_color)),
                Paragraph(ressource_nom,           small),
                Paragraph(date_debut.strftime('%d/%m/%Y\n%H:%M') if date_debut else '—', ParagraphStyle('dt', fontName='Courier', fontSize=8)),
                Paragraph(date_fin.strftime('%d/%m/%Y\n%H:%M')   if date_fin   else '—', ParagraphStyle('dt', fontName='Courier', fontSize=8)),
                Paragraph(str(duree),              ParagraphStyle('du', fontSize=8, textColor=colors.HexColor('#1f6feb'))),
                Paragraph(statut_str,              ParagraphStyle('st', fontSize=8, textColor=s_color, fontName='Helvetica-Bold')),
                Paragraph(i.get('created_by','—'), small),
            ])

        t = Table(rows, colWidths=[5.5*cm, 3.2*cm, 3.5*cm, 2.8*cm, 2.8*cm, 1.8*cm, 2.2*cm, 4.2*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#1f6feb')),
            ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#d0d0d0')),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#f7f9ff')]),
            ('PADDING',       (0,0), (-1,-1), 6),
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ]))
        elements.append(t)

    # Pied de page
    elements.append(Spacer(1, 0.6*cm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(Paragraph(
        f"Rapport généré le {now.strftime('%d/%m/%Y à %H:%M')} "
        f"par {request.user.get_full_name() or request.session.get('username', '')} — SIGR-CA",
        footer_s
    ))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="indisponibilites_{now.strftime("%Y%m%d_%H%M")}.pdf"'
    return response


def get_ressource_name(ressource_id, ressource_type):
    """Récupère le nom d'une ressource par son ID"""
    if ressource_type == 'salle':
        salle = db.bureaux.find_one({'_id': ObjectId(ressource_id)})
        return salle['nom'] if salle else 'Salle inconnue'
    elif ressource_type == 'materiel':
        materiel = db.materiels.find_one({'_id': ObjectId(ressource_id)})
        return materiel['nom'] if materiel else 'Matériel inconnu'
    return 'Ressource inconnue'


def get_materiel_icon(categorie):
    """Retourne l'icône correspondant à la catégorie du matériel"""
    icons = {
        'informatique': '💻',
        'mobilier': '🪑',
        'audiovisuel': '📽️',
        'imprimante': '🖨️',
        'securite': '🔒',
        'vehicule': '🚗',
        'outillage': '🔧',
        'autre': '📦'
    }
    return icons.get(categorie, '📦')
 # dashboard/views.py - Assurez-vous que ces fonctions sont au bon niveau d'indentation

@session_required
def api_hierarchie_modifier(request, niveau, element_id):
    """API pour modifier un élément de la hiérarchie"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    if request.method != 'PUT':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        collection_name = f"{niveau}s"
        
        update_data = {
            'nom': data.get('nom'),
            'code': data.get('code'),
            'description': data.get('description'),
            'updated_at': datetime.now(),
            'updated_by': request.session.get('username', ''),
        }
        
        result = db[collection_name].update_one(
            {'_id': ObjectId(element_id)},
            {'$set': update_data}
        )
        
        if result.modified_count > 0:
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Non modifié'}, status=400)
            
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        # dashboard/views.py - Ajouter cette API

@session_required
def api_smart_suggestions(request):
    """API pour les suggestions intelligentes de créneaux"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        resource_id = request.GET.get('resource_id')
        date_debut_str = request.GET.get('date_debut')
        date_fin_str = request.GET.get('date_fin')
        
        if not resource_id or not date_debut_str or not date_fin_str:
            return JsonResponse({'error': 'Paramètres manquants'}, status=400)
        
        date_debut = datetime.fromisoformat(date_debut_str)
        date_fin = datetime.fromisoformat(date_fin_str)
        
        # Récupérer l'employé si connecté
        employe_id = None
        if not request.session.get('is_staff', False):
            employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
            if employe:
                employe_id = employe['_id']
        
        engine = get_suggestion_engine(db)
        suggestions = engine.suggest_alternative_slots_advanced(
            ObjectId(resource_id), date_debut, date_fin, employe_id
        )
        
        return JsonResponse({
            'status': 'success',
            'suggestions': suggestions,
            'total': len(suggestions)
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_user_preferences_save(request):
    """Sauvegarder les préférences utilisateur pour les suggestions"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
        if not employe:
            employe = db.employees.find_one({'django_username': request.session.get('username', '')})
        
        if not employe:
            return JsonResponse({'error': 'Employé non trouvé'}, status=404)
        
        preferences = {
            'preferred_days': data.get('preferred_days', []),
            'preferred_hours_start': data.get('preferred_hours_start', []),
            'preferred_duration': data.get('preferred_duration', 60),
            'avoid_overlap': data.get('avoid_overlap', True),
            'preferred_rooms': data.get('preferred_rooms', [])
        }
        
        db.employees.update_one(
            {'_id': employe['_id']},
            {'$set': {'preferences_reservation': preferences}}
        )
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        # dashboard/views.py - Ajouter ces fonctions

# ====================== CONFIGURATION DES PLAGES HORAIRES ======================

@session_required
def horaires_activite(request):
    """Configuration des plages horaires d'activité globales"""
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')
    
    # Récupérer la configuration existante
    config = db.system_config.find_one({'type': 'horaires_activite'})
    if not config:
        config = {
            'default': {
                'monday': {'enabled': True, 'start': '08:00', 'end': '18:00'},
                'tuesday': {'enabled': True, 'start': '08:00', 'end': '18:00'},
                'wednesday': {'enabled': True, 'start': '08:00', 'end': '18:00'},
                'thursday': {'enabled': True, 'start': '08:00', 'end': '18:00'},
                'friday': {'enabled': True, 'start': '08:00', 'end': '17:00'},
                'saturday': {'enabled': False, 'start': '09:00', 'end': '13:00'},
                'sunday': {'enabled': False, 'start': '09:00', 'end': '13:00'},
            },
            'exceptions': []  # Périodes exceptionnelles (fermetures, nocturnes)
        }
    
    # Récupérer les zones avec règles spécifiques
    zones_rules = list(db.zones_horaires.find()) if 'zones_horaires' in db.list_collection_names() else []
    for z in zones_rules:
        z['id'] = str(z['_id'])
    
    return render(request, 'dashboard/horaires_activite.html', {
        'config': config,
        'zones_rules': zones_rules,
        'jours': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        'jours_labels': {
            'monday': 'Lundi', 'tuesday': 'Mardi', 'wednesday': 'Mercredi',
            'thursday': 'Jeudi', 'friday': 'Vendredi', 'saturday': 'Samedi', 'sunday': 'Dimanche'
        }
    })


@session_required
def api_horaires_save(request):
    """Sauvegarder la configuration des horaires"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        config = {
            'type': 'horaires_activite',
            'default': data.get('default', {}),
            'exceptions': data.get('exceptions', []),
            'updated_at': datetime.now(),
            'updated_by': request.session.get('username', '')
        }
        
        db.system_config.update_one(
            {'type': 'horaires_activite'},
            {'$set': config},
            upsert=True
        )
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@session_required
def api_zone_horaire_save(request):
    """Sauvegarder les horaires spécifiques d'une zone"""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        zone_id = data.get('zone_id')
        
        zone_rule = {
            'zone_id': ObjectId(zone_id),
            'zone_nom': data.get('zone_nom'),
            'horaires': data.get('horaires', {}),
            'is_active': data.get('is_active', True),
            'updated_at': datetime.now(),
            'updated_by': request.session.get('username', '')
        }
        
        if 'zones_horaires' not in db.list_collection_names():
            db.create_collection('zones_horaires')
        
        db.zones_horaires.update_one(
            {'zone_id': ObjectId(zone_id)},
            {'$set': zone_rule},
            upsert=True
        )
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def is_access_allowed_by_schedule(zone_id, timestamp):
    """Vérifie si l'accès est autorisé selon les horaires (utilisé par api_verify_access)"""
    # Récupérer la configuration globale
    global_config = db.system_config.find_one({'type': 'horaires_activite'})
    
    # Récupérer les règles spécifiques à la zone
    zone_rule = db.zones_horaires.find_one({'zone_id': zone_id})
    
    day_name = timestamp.strftime('%A').lower()
    current_time = timestamp.strftime('%H:%M')
    
    # Priorité aux règles spécifiques de la zone
    if zone_rule and zone_rule.get('is_active'):
        day_config = zone_rule.get('horaires', {}).get(day_name, {})
        if day_config.get('enabled', False):
            return day_config.get('start', '00:00') <= current_time <= day_config.get('end', '23:59')
        return False
    
    # Sinon, utiliser la configuration globale
    if global_config:
        day_config = global_config.get('default', {}).get(day_name, {})
        if day_config.get('enabled', False):
            return day_config.get('start', '00:00') <= current_time <= day_config.get('end', '23:59')
        return False
    
    # Par défaut, accès autorisé 8h-18h en semaine
    if day_name in ['saturday', 'sunday']:
        return False
    return '08:00' <= current_time <= '18:00'
    # dashboard/views.py - Ajoutez cette fonction

@session_required
def preferences_reservation(request):
    """Page des préférences de réservation pour les suggestions IA"""
    if request.session.get('is_staff', False):
        return redirect('dashboard')
    
    # Récupérer l'employé
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        employe = db.employees.find_one({'django_username': request.session.get('username', '')})
    
    if not employe:
        return redirect('login')
    
    employe['id'] = str(employe['_id'])
    
    # Récupérer les préférences existantes
    preferences = employe.get('preferences_reservation', {
        'preferred_days': ['monday', 'tuesday', 'wednesday', 'thursday'],
        'preferred_hours_start': [9, 10, 11, 14, 15],
        'preferred_duration': 60,
        'avoid_overlap': True,
        'preferred_rooms': []
    })
    
    # Récupérer toutes les salles pour les préférences
    salles = list(db.bureaux.find())
    for s in salles:
        s['id'] = str(s['_id'])
    
    return render(request, 'dashboard/preferences_reservation.html', {
        'employe': employe,
        'preferences': preferences,
        'salles': salles,
        'jours': [
            {'value': 'monday', 'label': 'Lundi'},
            {'value': 'tuesday', 'label': 'Mardi'},
            {'value': 'wednesday', 'label': 'Mercredi'},
            {'value': 'thursday', 'label': 'Jeudi'},
            {'value': 'friday', 'label': 'Vendredi'},
        ],
        'heures': [
            8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18
        ]
    })
    # dashboard/views_rfid.py
# Enrôlement RFID & provisionnement QR — à coller à la fin de dashboard/views.py

# ====================== ENRÔLEMENT RFID & PROVISIONNEMENT QR ======================

@session_required
def enrolement_badges(request):
    """Page principale de gestion des badges RFID et QR code des employés."""
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')

    employes = list(db.employees.find().sort('nom', 1))
    for e in employes:
        e['id'] = str(e['_id'])
        e['a_badge'] = bool(e.get('badge_id'))
        e['type_badge'] = e.get('badge_type', 'RFID')

    stats = {
        'total': len(employes),
        'avec_badge': sum(1 for e in employes if e.get('badge_id')),
        'sans_badge': sum(1 for e in employes if not e.get('badge_id')),
        'rfid': sum(1 for e in employes if e.get('badge_type') == 'RFID' and e.get('badge_id')),
        'qr': sum(1 for e in employes if e.get('badge_type') == 'QR' and e.get('badge_id')),
    }

    return render(request, 'dashboard/enrolement_badges.html', {
        'employes': employes,
        'stats': stats,
    })


@session_required
@require_http_methods(["POST"])
def api_badge_affecter(request, employe_id):
    """Affecter ou modifier le badge RFID/QR d'un employé."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    try:
        data = json.loads(request.body)
        badge_id   = data.get('badge_id', '').strip()
        badge_type = data.get('badge_type', 'RFID').upper()

        if not badge_id:
            return JsonResponse({'error': 'Identifiant de badge requis'}, status=400)

        if badge_type not in ('RFID', 'QR'):
            return JsonResponse({'error': 'Type de badge invalide (RFID ou QR)'}, status=400)

        employe = db.employees.find_one({'_id': ObjectId(employe_id)})
        if not employe:
            return JsonResponse({'error': 'Employé introuvable'}, status=404)

        # Vérifier l'unicité du badge_id (sauf si c'est le même employé)
        existant = db.employees.find_one({
            'badge_id': badge_id,
            '_id': {'$ne': ObjectId(employe_id)}
        })
        if existant:
            nom_existant = f"{existant.get('prenom','')} {existant.get('nom','')}".strip()
            return JsonResponse({
                'error': f"Ce badge est déjà attribué à {nom_existant}"
            }, status=400)

        ancien_badge = employe.get('badge_id')

        # Mettre à jour MongoDB
        db.employees.update_one(
            {'_id': ObjectId(employe_id)},
            {'$set': {
                'badge_id': badge_id,
                'badge_type': badge_type,
                'badge_updated_at': datetime.now(),
                'badge_updated_by': request.session.get('username', ''),
            }}
        )

        # Mettre à jour badge_rfid dans MongoDB directement
        if employe.get('django_user_id'):
            try:
                db['utilisateurs'].update_one(
                    {'id': int(employe['django_user_id'])},
                    {'$set': {'badge_rfid': badge_id}}
                )
            except Exception as _e:
                logger.warning(f"Mise à jour badge_rfid échouée: {_e}")

        # Journaliser l'action
        db.acces_logs.insert_one({
            'action': 'badge_affecte',
            'employe_id': ObjectId(employe_id),
            'badge_id': badge_id,
            'badge_type': badge_type,
            'ancien_badge': ancien_badge,
            'fait_par': request.session.get('username', ''),
            'timestamp': datetime.now(),
        })

        # Envoyer email de notification à l'employé
        if employe.get('email'):
            try:
                from dashboard.utils_email import email_badge_rfid_affecte
                email_badge_rfid_affecte(employe, badge_id, badge_type)
            except Exception as e:
                logger.warning(f"Email badge non envoyé: {e}")

        action = "modifié" if ancien_badge else "affecté"
        return JsonResponse({
            'success': True,
            'message': f"Badge {badge_type} {action} avec succès.",
            'badge_id': badge_id,
            'badge_type': badge_type,
        })

    except Exception as e:
        logger.error(f"Erreur api_badge_affecter: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@session_required
@require_http_methods(["POST"])
def api_badge_supprimer(request, employe_id):
    """Révoquer le badge d'un employé."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    try:
        employe = db.employees.find_one({'_id': ObjectId(employe_id)})
        if not employe:
            return JsonResponse({'error': 'Employé introuvable'}, status=404)

        ancien_badge = employe.get('badge_id')
        if not ancien_badge:
            return JsonResponse({'error': 'Cet employé n\'a pas de badge'}, status=400)

        db.employees.update_one(
            {'_id': ObjectId(employe_id)},
            {'$unset': {'badge_id': '', 'badge_type': ''},
             '$set': {'badge_revoked_at': datetime.now(), 'badge_revoked_by': request.session.get('username', '')}}
        )

        if employe.get('django_user_id'):
            try:
                db['utilisateurs'].update_one(
                    {'id': int(employe['django_user_id'])},
                    {'$unset': {'badge_rfid': ''}}
                )
            except Exception as _e:
                logger.warning(f"Révocation badge_rfid échouée: {_e}")

        db.acces_logs.insert_one({
            'action': 'badge_revoque',
            'employe_id': ObjectId(employe_id),
            'badge_id': ancien_badge,
            'fait_par': request.session.get('username', ''),
            'timestamp': datetime.now(),
        })

        return JsonResponse({'success': True, 'message': 'Badge révoqué avec succès.'})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@session_required
def api_badge_verifier(request):
    """Vérifie si un badge_id est déjà utilisé (pour validation en temps réel)."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    badge_id   = request.GET.get('badge_id', '').strip()
    employe_id = request.GET.get('employe_id', '')

    if not badge_id:
        return JsonResponse({'disponible': False, 'message': 'Badge ID vide'})

    query = {'badge_id': badge_id}
    if employe_id:
        query['_id'] = {'$ne': ObjectId(employe_id)}

    existant = db.employees.find_one(query)
    if existant:
        nom = f"{existant.get('prenom','')} {existant.get('nom','')}".strip()
        return JsonResponse({'disponible': False, 'message': f'Attribué à {nom}'})

    return JsonResponse({'disponible': True, 'message': 'Badge disponible'})


@session_required
def api_generer_qr_employe(request, employe_id):
    """Génère et affecte automatiquement un QR code unique à un employé."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    try:
        import secrets
        employe = db.employees.find_one({'_id': ObjectId(employe_id)})
        if not employe:
            return JsonResponse({'error': 'Employé introuvable'}, status=404)

        # Générer un QR code unique
        qr_id = f"QR-{employe.get('nom','EMP')[:3].upper()}-{secrets.token_hex(4).upper()}"

        # S'assurer de l'unicité
        while db.employees.find_one({'badge_id': qr_id}):
            qr_id = f"QR-{employe.get('nom','EMP')[:3].upper()}-{secrets.token_hex(4).upper()}"

        db.employees.update_one(
            {'_id': ObjectId(employe_id)},
            {'$set': {
                'badge_id': qr_id,
                'badge_type': 'QR',
                'badge_updated_at': datetime.now(),
                'badge_updated_by': request.session.get('username', ''),
            }}
        )

        # Générer l'image QR
        import qrcode, base64
        from io import BytesIO
        qr_data = json.dumps({
            'employe_id': employe_id,
            'badge_id': qr_id,
            'nom': f"{employe.get('prenom','')} {employe.get('nom','')}",
            'type': 'SIGR-CA-ACCESS',
        })
        qr_img = qrcode.make(qr_data)
        buf = BytesIO()
        qr_img.save(buf, format='PNG')
        qr_b64 = base64.b64encode(buf.getvalue()).decode()

        # Stocker le QR en base
        db.employees.update_one(
            {'_id': ObjectId(employe_id)},
            {'$set': {'qr_code_base64': qr_b64}}
        )

        # Email de notification
        if employe.get('email'):
            try:
                from dashboard.utils_email import email_badge_rfid_affecte
                email_badge_rfid_affecte(employe, qr_id, 'QR')
            except Exception as e:
                logger.warning(f"Email QR non envoyé: {e}")

        return JsonResponse({
            'success': True,
            'badge_id': qr_id,
            'qr_base64': qr_b64,
            'message': f'QR code généré : {qr_id}',
        })

    except Exception as e:
        logger.error(f"Erreur génération QR: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@session_required
def api_export_badges_csv(request):
    """Export CSV de tous les badges."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    import csv
    from django.http import HttpResponse

    employes = list(db.employees.find().sort('nom', 1))
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="badges_sigr_ca_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')  # BOM UTF-8 pour Excel

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Nom', 'Prénom', 'Email', 'Département', 'Badge ID', 'Type', 'Statut'])
    for e in employes:
        writer.writerow([
            e.get('nom', ''),
            e.get('prenom', ''),
            e.get('email', ''),
            e.get('departement', ''),
            e.get('badge_id', ''),
            e.get('badge_type', ''),
            'Actif' if e.get('badge_id') else 'Sans badge',
        ])
    return response
    # dashboard/views_indispo.py
# À coller à la fin de dashboard/views.py
# Remplace les fonctions gestion_indisponibilites, api_indisponibilite_ajouter,
# api_indisponibilite_modifier, api_indisponibilite_supprimer

# ====================== INDISPONIBILITÉS PLANIFIÉES (COMPLET) ======================

def _verifier_indisponibilite(ressource_id, ressource_type, date_debut, date_fin, exclure_id=None):
    """
    Vérifie si une ressource est sous indisponibilité sur un créneau.
    Retourne le document d'indisponibilité si conflit, None sinon.
    """
    query = {
        'ressource_type': ressource_type,
        'date_debut': {'$lt': date_fin},
        'date_fin':   {'$gt': date_debut},
    }
    try:
        query['ressource_id'] = ObjectId(ressource_id)
    except Exception:
        query['ressource_id'] = ressource_id

    if exclure_id:
        try:
            query['_id'] = {'$ne': ObjectId(exclure_id)}
        except Exception:
            pass

    return db.indisponibilites.find_one(query)


@session_required
def gestion_indisponibilites(request):
    """Page de gestion des indisponibilités planifiées."""
    if not request.session.get('is_staff', False):
        return redirect('employe_espace')

    now = datetime.now()

    # Créer la collection si besoin
    if 'indisponibilites' not in db.list_collection_names():
        db.create_collection('indisponibilites')

    indispos = list(db.indisponibilites.find().sort('date_debut', -1))
    for i in indispos:
        i['id'] = str(i['_id'])
        # Résoudre le nom de la ressource
        try:
            if i.get('ressource_type') == 'salle':
                r = db.bureaux.find_one({'_id': ObjectId(str(i['ressource_id']))})
                i['ressource_nom'] = r['nom'] if r else 'Inconnue'
            elif i.get('ressource_type') == 'materiel':
                r = db.materiels.find_one({'_id': ObjectId(str(i['ressource_id']))})
                i['ressource_nom'] = r['nom'] if r else 'Inconnu'
            else:
                i['ressource_nom'] = str(i.get('ressource_id', '?'))
        except Exception:
            i['ressource_nom'] = str(i.get('ressource_id', '?'))

        # Statut calculé
        if i['date_fin'] < now:
            i['statut_calc'] = 'passee'
        elif i['date_debut'] <= now <= i['date_fin']:
            i['statut_calc'] = 'en_cours'
        else:
            i['statut_calc'] = 'a_venir'

        # Compter les réservations impactées
        try:
            i['reservations_impactees'] = db.reservations.count_documents({
                'bureau_id': i.get('ressource_id'),
                'statut': {'$in': ['confirmee', 'en_attente']},
                'date_debut': {'$lt': i['date_fin']},
                'date_fin':   {'$gt': i['date_debut']},
            })
        except Exception:
            i['reservations_impactees'] = 0

    # Ressources pour le formulaire
    salles = list(db.bureaux.find())
    for s in salles:
        s['id'] = str(s['_id'])
        s['type'] = 'salle'

    materiels = list(db.materiels.find()) if 'materiels' in db.list_collection_names() else []
    for m in materiels:
        m['id'] = str(m['_id'])
        m['type'] = 'materiel'

    stats = {
        'en_cours': db.indisponibilites.count_documents({'date_debut': {'$lte': now}, 'date_fin': {'$gte': now}}),
        'a_venir':  db.indisponibilites.count_documents({'date_debut': {'$gt': now}}),
        'passees':  db.indisponibilites.count_documents({'date_fin': {'$lt': now}}),
        'total':    db.indisponibilites.count_documents({}),
    }

    return render(request, 'dashboard/gestion_indisponibilites.html', {
        'indisponibilites': indispos,
        'ressources': salles + materiels,
        'stats': stats,
    })


@session_required
@require_http_methods(["POST"])
def api_indisponibilite_ajouter(request):
    """Créer une indisponibilité et notifier les employés impactés."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    try:
        data = json.loads(request.body)
        ressource_id   = data.get('ressource_id', '').strip()
        ressource_type = data.get('ressource_type', 'salle')
        titre          = data.get('titre', '').strip()
        description    = data.get('description', '')
        type_indispo   = data.get('type_indispo', 'maintenance')
        recurrence     = data.get('recurrence', 'none')
        recurrence_end = data.get('recurrence_end')

        if not ressource_id:
            return JsonResponse({'error': 'Ressource non spécifiée'}, status=400)
        if not titre:
            return JsonResponse({'error': 'Titre requis'}, status=400)

        try:
            date_debut = datetime.fromisoformat(data.get('date_debut'))
            date_fin   = datetime.fromisoformat(data.get('date_fin'))
        except Exception:
            return JsonResponse({'error': 'Format de date invalide'}, status=400)

        if date_fin <= date_debut:
            return JsonResponse({'error': 'La date de fin doit être après la date de début'}, status=400)

        # Vérifier conflit avec d'autres indisponibilités
        conflit = _verifier_indisponibilite(ressource_id, ressource_type, date_debut, date_fin)
        if conflit:
            return JsonResponse({'error': 'Conflit avec une indisponibilité existante sur cette période'}, status=400)

        indispo = {
            'ressource_id':   ObjectId(ressource_id),
            'ressource_type': ressource_type,
            'titre':          titre,
            'description':    description,
            'type_indispo':   type_indispo,
            'date_debut':     date_debut,
            'date_fin':       date_fin,
            'recurrence':     recurrence if recurrence != 'none' else None,
            'created_at':     datetime.now(),
            'created_by':     request.session.get('username', ''),
        }
        if recurrence_end and recurrence != 'none':
            indispo['recurrence_end'] = datetime.fromisoformat(recurrence_end)

        result = db.indisponibilites.insert_one(indispo)

        # ── Notifier les employés dont les réservations sont impactées ──
        reservations_impactees = list(db.reservations.find({
            'bureau_id': ObjectId(ressource_id),
            'statut': {'$in': ['confirmee', 'en_attente']},
            'date_debut': {'$lt': date_fin},
            'date_fin':   {'$gt': date_debut},
        }))

        ressource_nom = titre
        try:
            if ressource_type == 'salle':
                r = db.bureaux.find_one({'_id': ObjectId(ressource_id)})
                if r:
                    ressource_nom = r['nom']
        except Exception:
            pass

        nb_notifies = 0
        for resa in reservations_impactees:
            try:
                emp_id = resa.get('employe_id')
                employe = None
                if emp_id:
                    employe = db.employees.find_one({'_id': ObjectId(str(emp_id))})
                if employe and employe.get('email'):
                    from dashboard.utils_email import email_maintenance_ressource
                    email_maintenance_ressource(
                        employe['email'], ressource_nom,
                        date_debut, date_fin,
                        motif=description or titre
                    )
                    nb_notifies += 1
                    # Notification in-app
                    try:
                        if employe.get('_id'):
                            db.notifications.insert_one({
                                'employe_id': str(employe['_id']),
                                'titre':      f'🔧 Maintenance planifiée — {ressource_nom}',
                                'message':    f"La ressource '{ressource_nom}' sera indisponible du "
                                              f"{date_debut.strftime('%d/%m/%Y %H:%M')} au "
                                              f"{date_fin.strftime('%d/%m/%Y %H:%M')}. Motif : {description or titre}",
                                'categorie':  'alerte',
                                'icon':       '🔧',
                                'status':     'non_lu',
                                'created_at': datetime.now(),
                            })
                    except Exception:
                        pass
            except Exception as e:
                logger.warning(f"Notification indispo non envoyée: {e}")

        if recurrence and recurrence != 'none':
            _generate_recurring_indisponibilities(indispo, result.inserted_id)

        return JsonResponse({
            'success': True,
            'id': str(result.inserted_id),
            'nb_notifies': nb_notifies,
            'message': f"Indisponibilité créée. {nb_notifies} employé(s) notifié(s).",
        })

    except Exception as e:
        logger.error(f"Erreur api_indisponibilite_ajouter: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@session_required
@require_http_methods(["PUT", "POST"])
def api_indisponibilite_modifier(request, indispo_id):
    """Modifier une indisponibilité existante."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    try:
        data = json.loads(request.body)
        indispo = db.indisponibilites.find_one({'_id': ObjectId(indispo_id)})
        if not indispo:
            return JsonResponse({'error': 'Indisponibilité introuvable'}, status=404)

        date_debut = datetime.fromisoformat(data.get('date_debut'))
        date_fin   = datetime.fromisoformat(data.get('date_fin'))

        if date_fin <= date_debut:
            return JsonResponse({'error': 'La date de fin doit être après la date de début'}, status=400)

        conflit = _verifier_indisponibilite(
            str(indispo['ressource_id']), indispo['ressource_type'],
            date_debut, date_fin, exclure_id=indispo_id
        )
        if conflit:
            return JsonResponse({'error': 'Conflit avec une autre indisponibilité'}, status=400)

        db.indisponibilites.update_one(
            {'_id': ObjectId(indispo_id)},
            {'$set': {
                'titre':       data.get('titre', indispo['titre']),
                'description': data.get('description', indispo.get('description', '')),
                'type_indispo': data.get('type_indispo', indispo.get('type_indispo', 'maintenance')),
                'date_debut':  date_debut,
                'date_fin':    date_fin,
                'updated_at':  datetime.now(),
                'updated_by':  request.session.get('username', ''),
            }}
        )

        return JsonResponse({'success': True, 'message': 'Indisponibilité modifiée.'})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@session_required
@require_http_methods(["DELETE", "POST"])
def api_indisponibilite_supprimer(request, indispo_id):
    """Supprimer une indisponibilité."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    try:
        result = db.indisponibilites.delete_one({'_id': ObjectId(indispo_id)})
        if result.deleted_count == 0:
            return JsonResponse({'error': 'Indisponibilité introuvable'}, status=404)
        return JsonResponse({'success': True, 'message': 'Indisponibilité supprimée.'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@session_required
def api_indisponibilite_check(request):
    """
    API appelée lors de la création d'une réservation pour vérifier
    si la ressource est disponible (pas en maintenance).
    """
    ressource_id   = request.GET.get('ressource_id', '')
    ressource_type = request.GET.get('ressource_type', 'salle')
    date_debut_str = request.GET.get('date_debut', '')
    date_fin_str   = request.GET.get('date_fin', '')

    if not all([ressource_id, date_debut_str, date_fin_str]):
        return JsonResponse({'disponible': True})

    try:
        date_debut = datetime.fromisoformat(date_debut_str)
        date_fin   = datetime.fromisoformat(date_fin_str)
    except Exception:
        return JsonResponse({'disponible': True})

    conflit = _verifier_indisponibilite(ressource_id, ressource_type, date_debut, date_fin)
    if conflit:
        return JsonResponse({
            'disponible': False,
            'raison': conflit.get('titre', 'Indisponibilité planifiée'),
            'date_debut': conflit['date_debut'].strftime('%d/%m/%Y %H:%M'),
            'date_fin':   conflit['date_fin'].strftime('%d/%m/%Y %H:%M'),
            'type':       conflit.get('type_indispo', 'maintenance'),
        })

    return JsonResponse({'disponible': True})


def _generate_recurring_indisponibilities(parent, parent_id):
    """Génère les occurrences récurrentes d'une indisponibilité."""
    recurrence = parent.get('recurrence')
    recurrence_end = parent.get('recurrence_end')
    if not recurrence or not recurrence_end:
        return

    delta_map = {'daily': timedelta(days=1), 'weekly': timedelta(weeks=1), 'monthly': timedelta(days=30)}
    delta = delta_map.get(recurrence)
    if not delta:
        return

    duree = parent['date_fin'] - parent['date_debut']
    current = parent['date_debut'] + delta
    occurrences = []

    while current <= recurrence_end and len(occurrences) < 52:
        occ = dict(parent)
        occ.pop('_id', None)
        occ['date_debut']  = current
        occ['date_fin']    = current + duree
        occ['parent_id']   = parent_id
        occ['is_recurring'] = True
        occurrences.append(occ)
        current += delta

    if occurrences:
        db.indisponibilites.insert_many(occurrences)
        

@session_required
def api_materiel_disponibilite(request, materiel_id):
    """Verifie si un materiel est libre sur un creneau donne."""
    from datetime import datetime
    from bson import ObjectId

    debut_str = request.GET.get('debut', '').strip()
    fin_str = request.GET.get('fin', '').strip()

    if not (debut_str and fin_str):
        return JsonResponse({'disponible': False, 'erreur': 'parametres manquants'})

    try:
        date_debut = datetime.strptime(debut_str, '%Y-%m-%dT%H:%M')
        date_fin = datetime.strptime(fin_str, '%Y-%m-%dT%H:%M')
    except ValueError:
        return JsonResponse({'disponible': False, 'erreur': 'dates invalides'})

    if date_fin <= date_debut:
        return JsonResponse({'disponible': False, 'erreur': 'la fin doit etre apres le debut'})

    try:
        mat_oid = ObjectId(materiel_id)
    except Exception:
        mat_oid = materiel_id

    conflit = db.reservations.find_one({
        '$or': [{'materiel_id': materiel_id}, {'materiel_id': mat_oid}],
        'statut': {'$in': ['confirmee', 'en_attente']},
        'date_debut': {'$lt': date_fin},
        'date_fin': {'$gt': date_debut},
    })

    if conflit:
        return JsonResponse({
            'disponible': False,
            'motif': "deja reserve de " + conflit['date_debut'].strftime('%d/%m %H:%M') + " a " + conflit['date_fin'].strftime('%H:%M'),
        })

    try:
        if 'indisponibilites' in db.list_collection_names():
            indispo = db.indisponibilites.find_one({
                'ressource_type': 'materiel',
                'ressource_id': mat_oid,
                'date_debut': {'$lt': date_fin},
                'date_fin': {'$gt': date_debut},
            })
            if indispo:
                return JsonResponse({
                    'disponible': False,
                    'motif': "en maintenance (" + indispo.get('titre', 'planifiee') + ")",
                })
    except Exception:
        pass

    return JsonResponse({'disponible': True})

# ====================== API : SUGGESTIONS CRÉNEAUX IA ======================
@session_required
def api_suggestions_creneaux(request):
    """
    Suggère des créneaux libres pour une ressource donnée.
    GET params : resource_id, resource_type, date (YYYY-MM-DD), duree (min), nb_participants
    """
    resource_id   = request.GET.get('resource_id')
    resource_type = request.GET.get('resource_type', 'salle')
    date_str      = request.GET.get('date')
    duree_min     = int(request.GET.get('duree', 60))
    nb_part       = int(request.GET.get('nb_participants', 1))

    if not resource_id or not date_str:
        return JsonResponse({'suggestions': [], 'error': 'Paramètres manquants'})

    try:
        jour = datetime.fromisoformat(date_str + 'T00:00:00')
    except Exception:
        return JsonResponse({'suggestions': [], 'error': 'Date invalide'})

    # ── Vérification capacité (salles uniquement) ──────────────────────────────
    if resource_type == 'salle':
        try:
            salle = db.bureaux.find_one({'_id': ObjectId(resource_id)})
            if not salle and str(resource_id).isdigit():
                salle = db.bureaux.find_one({'id': int(resource_id)})
            if salle and salle.get('capacite_max') and nb_part > salle['capacite_max']:
                return JsonResponse({
                    'suggestions': [],
                    'warning': f"Capacité insuffisante ({salle['capacite_max']} places max pour {nb_part} participants)",
                })
        except Exception:
            pass

    # ── Génération des créneaux 8h–19h, pas de 30 min ─────────────────────────
    suggestions   = []
    heure_debut   = 8
    heure_fin     = 19
    pas           = 30  # minutes
    duree         = timedelta(minutes=duree_min)

    current = jour.replace(hour=heure_debut, minute=0, second=0, microsecond=0)
    limite  = jour.replace(hour=heure_fin,   minute=0, second=0, microsecond=0)

    while current + duree <= limite:
        creneau_debut = current
        creneau_fin   = current + duree

        check = check_ressource_disponibilite(
            resource_id, resource_type, creneau_debut, creneau_fin
        )

        if check['disponible']:
            # Score de pertinence
            score = 100
            if current.minute == 0:
                score += 20
            if 9 <= current.hour <= 11:   # matinée productive
                score += 15
            if current.hour == 14:        # début d'après-midi
                score += 10

            suggestions.append({
                'debut':       creneau_debut.isoformat(),
                'fin':         creneau_fin.isoformat(),
                'debut_label': creneau_debut.strftime('%H:%M'),
                'fin_label':   creneau_fin.strftime('%H:%M'),
                'duree_min':   duree_min,
                'score':       score,
            })

        current += timedelta(minutes=pas)

    # Trier par score puis ré-afficher par heure, garder les 6 meilleurs
    suggestions.sort(key=lambda s: -s['score'])
    suggestions = suggestions[:6]
    suggestions.sort(key=lambda s: s['debut'])

    return JsonResponse({'suggestions': suggestions})

    # ============================================================
# IA — APIs ML réelles
# ============================================================
from dashboard.ai_engine import (
    OccupationPredictor, PersonalRecommender, AnomalyDetector, train_all_models
)


@session_required
def api_predict_occupation(request, resource_id):
    """Heatmap 7j × 12h des taux d'occupation prédits."""
    predictor = OccupationPredictor()
    heatmap = predictor.predict_week(resource_id)
    if heatmap is None:
        return JsonResponse({
            'error': 'Modèle non entraîné. Lance: python manage.py train_ai',
            'heatmap': []
        })
    return JsonResponse({'heatmap': heatmap})


@session_required
def api_recommandations(request):
    """Top 3 ressources recommandées pour l'utilisateur connecté."""
    employe = db.employees.find_one({'django_user_id': request.session.get('user_id', '')})
    if not employe:
        return JsonResponse({'recommendations': []})

    recommender = PersonalRecommender()
    recos = recommender.recommend(str(employe['_id']), top_n=3)

    # Enrichir avec les noms
    enriched = []
    for r in recos:
        try:
            res_oid = ObjectId(r['resource_id'])
        except Exception:
            continue
        bureau   = db.bureaux.find_one({'_id': res_oid})
        materiel = db.materiels.find_one({'_id': res_oid}) if not bureau else None
        if bureau:
            enriched.append({**r, 'nom': bureau['nom'], 'icon': '🚪', 'type': 'salle'})
        elif materiel:
            enriched.append({**r, 'nom': materiel['nom'], 'icon': '📦', 'type': 'materiel'})

    return JsonResponse({'recommendations': enriched})


@session_required
def api_anomalies(request):
    """Liste des réservations anormales (admin uniquement)."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Accès refusé'}, status=403)

    detector = AnomalyDetector()
    anomalies = detector.detect_recent(days=14)

    # Nettoyer pour JSON
    clean = []
    for a in anomalies:
        clean.append({
            'reservation_id':  a['reservation_id'],
            'employe_id':      a['employe_id'],
            'duree_min':       int(a['duree_min']),
            'nb_participants': int(a['nb_participants']),
            'heure':           int(a['heure']),
            'date_debut':      a['date_debut'].isoformat() if hasattr(a['date_debut'], 'isoformat') else str(a['date_debut']),
            'score':           round(float(a['anomaly_score']), 3),
        })
    return JsonResponse({'anomalies': clean})


@session_required
def api_train_models(request):
    """Déclenche le réentraînement (admin uniquement)."""
    if not request.session.get('is_staff', False):
        return JsonResponse({'error': 'Accès refusé'}, status=403)

    results = train_all_models()
    return JsonResponse({'success': True, 'results': results})
    from .approval_service import ApprovalService
from .models import ApprovalRequest, ApprovalAuditLog, ApprovalDelegation

def _client_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    return xff.split(',')[0].strip() if xff else request.META.get('REMOTE_ADDR')

# Liste des demandes à traiter par le manager connecté
def approbation_inbox(request):
    user_id = str(request.session.get('user_id', ''))
    try:
        demandes = ApprovalRequest.objects.filter(
            approbateur_id=user_id, statut='en_attente'
        ).order_by('-created_at')
    except Exception as _e:
        logger.warning(f"approbation_inbox ORM échoué: {_e}")
        demandes = []
    return render(request, 'dashboard/approbation_inbox.html',
                  {'demandes': demandes})

def reservation_approuver(request, reservation_id):
    if request.method != 'POST':
        return redirect('reservation_detail', reservation_id=reservation_id)
    ar = ApprovalRequest.objects.filter(
        reservation_id=reservation_id, statut='en_attente'
    ).order_by('niveau').first()
    if not ar:
        messages.error(request, "Aucune demande en attente.")
        return redirect('reservation_detail', reservation_id=reservation_id)
    res = ApprovalService.approuver(ar.id, get_session_user(request),
            commentaire=request.POST.get('commentaire',''),
            ip=_client_ip(request))
    messages.success(request, "Approbation enregistrée.")
    return redirect('reservation_detail', reservation_id=reservation_id)

def reservation_rejeter(request, reservation_id):
    if request.method != 'POST':
        return redirect('reservation_detail', reservation_id=reservation_id)
    ar = ApprovalRequest.objects.filter(
        reservation_id=reservation_id, statut='en_attente'
    ).order_by('niveau').first()
    if ar:
        ApprovalService.rejeter(ar.id, get_session_user(request),
            commentaire=request.POST.get('commentaire',''),
            ip=_client_ip(request))
    messages.success(request, "Réservation rejetée.")
    return redirect('reservation_detail', reservation_id=reservation_id)

def delegation_creer(request):
    if request.method == 'POST':
        ApprovalDelegation.objects.create(
            delegant_id=str(request.session.get('user_id', '')),
            delegant_nom=request.user.get_full_name() or request.session.get('username', ''),
            delegataire_id=request.POST['delegataire_id'],
            delegataire_nom=request.POST['delegataire_nom'],
            delegataire_email=request.POST['delegataire_email'],
            motif=request.POST.get('motif',''),
            date_debut=request.POST['date_debut'],
            date_fin=request.POST['date_fin'],
        )
        messages.success(request, "Délégation créée.")
        return redirect('delegation_creer')
    return render(request, 'dashboard/delegation.html')

def reservation_audit(request, reservation_id):
    logs = ApprovalAuditLog.objects.filter(reservation_id=reservation_id)
    return render(request, 'dashboard/reservation_audit.html',
                  {'logs': logs, 'reservation_id': reservation_id})
from .queue_service import QueueService
from .models import WaitingQueue

def queue_rejoindre(request, resource_id):
    """L'utilisateur rejoint la file après détection conflit."""
    if request.method != 'POST':
        return redirect('reservation_ajouter')

    from datetime import datetime
    d1 = datetime.fromisoformat(request.POST['date_debut'])
    d2 = datetime.fromisoformat(request.POST['date_fin'])

    wq = QueueService.ajouter(
        resource_id=resource_id,
        resource_nom=request.POST.get('resource_nom',''),
        user=get_session_user(request),
        date_debut=d1, date_fin=d2,
        titre=request.POST.get('titre',''),
        nb_participants=int(request.POST.get('nb_participants', 1)),
        flexible_minutes=int(request.POST.get('flexible_minutes', 30)),
    )
    messages.success(request,
        f"Ajouté à la file d'attente, position {wq.position}.")
    return redirect('employe_mes_reservations')

def queue_alternatives(request, resource_id):
    """API JSON : créneaux alternatifs."""
    from datetime import datetime
    d1 = datetime.fromisoformat(request.GET['date_debut'])
    d2 = datetime.fromisoformat(request.GET['date_fin'])
    propositions = QueueService.proposer_alternatives(resource_id, d1, d2)
    return JsonResponse({'propositions': [
        {'date_debut': p['date_debut'].isoformat(),
         'date_fin':   p['date_fin'].isoformat(),
         'decalage_min': p['decalage_min']} for p in propositions]})

def queue_confirmer(request, queue_id):
    """L'utilisateur transforme sa place en file en réservation."""
    try:
        wq = WaitingQueue.objects.get(id=queue_id, user_id=str(request.session.get('user_id', '')))
    except Exception:
        messages.error(request, "File d'attente introuvable.")
        return redirect('employe_mes_reservations')
    if wq.statut != 'notifie':
        messages.error(request, "Notification expirée.")
        return redirect('employe_mes_reservations')

    # Créer la réservation (Mongo)
    from config.mongo import db
    reservation_data = {
        'titre': wq.titre or f"Réservation {wq.resource_nom}",
        'resource_id': wq.resource_id,
        'employe_id': wq.user_id,
        'employe_nom': wq.user_nom,
        'date_debut': wq.date_debut_souhaitee,
        'date_fin':   wq.date_fin_souhaitee,
        'nb_participants': wq.nb_participants,
        'statut': 'en_attente',
        'created_at': timezone.now(),
        'created_by': wq.user_id,
    }
    result = db.reservations.insert_one(reservation_data)

    wq.statut = 'converti'
    wq.save()

    # Lancer le workflow d'approbation
    from .approval_service import ApprovalService
    ApprovalService.creer_workflow(str(result.inserted_id),
                                   reservation_data, get_session_user(request))

    messages.success(request, "Réservation créée depuis la file d'attente.")
    return redirect('employe_mes_reservations')

def queue_quitter(request, queue_id):
    try:
        wq = WaitingQueue.objects.get(id=queue_id, user_id=str(request.session.get('user_id', '')))
        wq.statut = 'annule'
        wq.save()
    except Exception as _e:
        logger.warning(f"queue_quitter échoué: {_e}")
    messages.info(request, "Vous avez quitté la file d'attente.")
    return redirect('employe_mes_reservations')
# dashboard/views.py - Ajoutez à la fin du fichier
@session_required
def api_stats_compare(request):
    """API pour comparer deux périodes"""
    if not request.session.get('is_staff', False): return JsonResponse({'error': 'Non autorisé'}, status=403)
    from datetime import datetime, timedelta
    type_period = request.GET.get('type', 'month')
    now = datetime.now()
    if type_period == 'week':
        period1_start, period1_end = now - timedelta(days=14), now - timedelta(days=7)
        period2_start, period2_end = now - timedelta(days=7), now
        label1, label2 = "Semaine précédente", "Semaine actuelle"
    elif type_period == 'quarter':
        period1_start, period1_end = now - timedelta(days=90), now - timedelta(days=45)
        period2_start, period2_end = now - timedelta(days=45), now
        label1, label2 = "Trimestre précédent", "Trimestre actuel"
    elif type_period == 'year':
        period1_start, period1_end = now - timedelta(days=730), now - timedelta(days=365)
        period2_start, period2_end = now - timedelta(days=365), now
        label1, label2 = "Année précédente", "Année actuelle"
    else:
        period1_start, period1_end = now - timedelta(days=60), now - timedelta(days=30)
        period2_start, period2_end = now - timedelta(days=30), now
        label1, label2 = "Mois précédent", "Mois actuel"
    return JsonResponse({'period1': {'total': db.acces_logs.count_documents({'timestamp': {'$gte': period1_start, '$lt': period1_end}}), 'label': label1}, 'period2': {'total': db.acces_logs.count_documents({'timestamp': {'$gte': period2_start, '$lt': period2_end}}), 'label': label2}})

@session_required
def api_stats_export_excel(request):
    """Export Excel avec openpyxl"""
    if not request.session.get('is_staff', False): return JsonResponse({'error': 'Non autorisé'}, status=403)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    from datetime import datetime, timedelta
    from io import BytesIO
    from django.http import HttpResponse
    days = int(request.GET.get('days', 30))
    start_date = datetime.now() - timedelta(days=days)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Accès quotidiens"
    ws1.append(['Date', 'Accès autorisés', 'Accès refusés', 'Total', 'Taux succès (%)'])
    for i in range(days, -1, -1):
        day_start = (datetime.now() - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        a = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'AUTORISE'})
        r = db.acces_logs.count_documents({'timestamp': {'$gte': day_start, '$lt': day_end}, 'resultat': 'REFUSE'})
        ws1.append([day_start.strftime('%d/%m/%Y'), a, r, a + r, round(a/(a+r)*100, 1) if (a+r) > 0 else 0])
    header_font, header_fill = Font(bold=True, color="FFFFFF"), PatternFill(start_color="1f6feb", fill_type="solid")
    for col in range(1, 6):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center")
        ws1.column_dimensions[chr(64 + col)].width = 15
    buffer = BytesIO()
    wb.save(buffer); buffer.seek(0)
    return HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="statistiques_{datetime.now().strftime("%Y%m%d")}.xlsx"'})
    # ─── Chemin du logo (à adapter si votre app ne s'appelle pas "dashboard") ────
LOGO_PATH = os.path.join(settings.STATIC_ROOT, 'img', 'logo.png')


def _logo_img(width=4.5 * cm, height=4 * cm):
    """Retourne un objet Image ReportLab si le logo existe, sinon None."""
    if os.path.exists(LOGO_PATH):
        return RLImage(LOGO_PATH, width=width, height=height)
    return None


def _photo_img_from_resource(photo_val, max_w=7 * cm, max_h=5 * cm):
    """
    Construit un objet Image ReportLab depuis :
      - une URL  http(s)://...
      - un data URL  data:image/...;base64,...
    Retourne None si impossible.
    """
    if not photo_val:
        return None
    try:
        if photo_val.startswith('data:image'):
            # base64 data URL → bytes
            header, b64data = photo_val.split(',', 1)
            import base64
            img_bytes = base64.b64decode(b64data)
            buf = io.BytesIO(img_bytes)
        elif photo_val.startswith('http'):
            req = urllib.request.Request(photo_val, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=4) as resp:
                img_bytes = resp.read()
            buf = io.BytesIO(img_bytes)
        else:
            return None
        return RLImage(buf, width=max_w, height=max_h)
    except Exception:
        return None


def _build_header(elements, title_text, subtitle_text, styles):
    """En-tête commun : logo à gauche, titre à droite, ligne bleue."""
    title_style = ParagraphStyle(
        'HdrTitle', parent=styles['Heading1'],
        fontSize=16, textColor=colors.HexColor('#1f6feb'),
        spaceAfter=2, leading=20,
    )
    sub_style = ParagraphStyle(
        'HdrSub', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey,
    )

    logo = _logo_img()
    if logo:
        hdr_table = Table(
            [[logo, [Paragraph(title_text, title_style), Paragraph(subtitle_text, sub_style)]]],
            colWidths=[5 * cm, 12.5 * cm],
        )
        hdr_table.setStyle(TableStyle([
            ('VALIGN',  (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(hdr_table)
    else:
        elements.append(Paragraph(title_text, title_style))
        elements.append(Paragraph(subtitle_text, sub_style))

    elements.append(Spacer(1, 0.3 * cm))
    elements.append(HRFlowable(width='100%', thickness=2, color=colors.HexColor('#1f6feb')))
    elements.append(Spacer(1, 0.4 * cm))

def send_rappel_retour_ressource():
    """
    À appeler chaque jour (cron ou management command).
    Envoie un rappel à l'employé dont la réservation d'une ressource
    (matériel) se termine demain — pour qu'il la rende.
    """
    from datetime import datetime, timedelta

    maintenant  = datetime.now()
    demain_debut = (maintenant + timedelta(days=1)).replace(hour=0,  minute=0,  second=0,  microsecond=0)
    demain_fin   = (maintenant + timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=0)

    # Réservations de matériel qui se terminent demain
    reservations = list(db.reservations.find({
        'resource_type': 'materiel',
        'statut':        'confirmee',
        'date_fin':      {'$gte': demain_debut, '$lte': demain_fin},
    }))

    logger.info(f"[RAPPEL] {len(reservations)} réservation(s) matériel se terminent demain.")

    for resa in reservations:
        employe_id = resa.get('employe_id')
        if not employe_id:
            continue

        try:
            employe = db.employees.find_one({'_id': employe_id if isinstance(employe_id, ObjectId) else ObjectId(str(employe_id))})
        except Exception:
            continue

        if not employe:
            continue

        prenom      = employe.get('prenom', '')
        nom         = employe.get('nom', '')
        titre       = resa.get('titre', 'Sans titre')
        materiel_nom = resa.get('materiel_nom') or resa.get('bureau_nom', 'Matériel')
        date_fin    = resa['date_fin']
        resa_id     = str(resa['_id'])

        message_texte = (
            f"Bonjour {prenom} {nom},\n\n"
            f"Votre réservation du matériel « {materiel_nom} » ('{titre}') "
            f"se termine demain le {date_fin.strftime('%d/%m/%Y à %H:%M')}.\n\n"
            f"Merci de le restituer avant cette date.\n\n"
            f"SIGR-CA"
        )

        # ── Notification en base ──
        db.notifications.insert_one({
            'employe_id':       str(employe['_id']),
            'destinataire':     employe.get('email', ''),
            'type_notification': 'email',
            'categorie':        'rappel_retour',
            'icon':             '📦',
            'titre':            f"⏰ Rappel : retour de « {materiel_nom} » demain",
            'message':          message_texte,
            'statut':           'non_lu',
            'action_url':       '/employe/reservations/',
            'reservation_id':   resa_id,
            'created_at':       datetime.now(),
        })

        # ── Email ──
        if employe.get('email'):
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                send_mail(
                    f"⏰ Rappel retour matériel — {materiel_nom}",
                    message_texte,
                    settings.DEFAULT_FROM_EMAIL,
                    [employe['email']],
                    fail_silently=True,
                )
                logger.info(f"[RAPPEL] Email envoyé à {employe['email']} pour '{materiel_nom}'")
            except Exception as e:
                logger.warning(f"[RAPPEL] Email échoué pour {employe.get('email')}: {e}")
        else:
            logger.warning(f"[RAPPEL] Pas d'email pour {prenom} {nom}")

